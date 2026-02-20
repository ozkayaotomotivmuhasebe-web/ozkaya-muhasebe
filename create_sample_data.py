"""
Örnek Excel dosyaları oluşturma aracı
Muhasebe Takip Sistemi'ne veri aktarmak için kullanabilirsiniz
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from pathlib import Path

def create_excel_file(filename, sheet_data):
    """Excel dosyası oluştur"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_data['sheet_name']
    
    # Başlık stili
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıkları yaz
    for col_num, header in enumerate(sheet_data['headers'], 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Verileri yaz
    for row_num, row_data in enumerate(sheet_data['data'], 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Sütun genişliklerini ayarla
    for col_num, width in enumerate(sheet_data.get('widths', [20]*len(sheet_data['headers'])), 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Dosyayı kaydet
    wb.save(filename)
    print(f"✓ {filename} oluşturuldu")

def main():
    """Örnek Excel dosyalarını oluştur"""
    
    # Sample data klasörü oluştur
    sample_dir = Path("sample_data")
    sample_dir.mkdir(exist_ok=True)
    
    # 1. CARELERİ İÇE AKTARMA (Müşteriler/Tedarikçiler)
    today = datetime.now().date()
    carileri_data = {
        'sheet_name': 'Cariler',
        'headers': ['Ad', 'Tür', 'Vergi Numarası', 'E-mail', 'Telefon', 'Şehir', 'Adres'],
        'data': [
            ['ABC Ticaret Ltd.', 'MÜŞTERİ', '1234567890', 'info@abc.com', '5321234567', 'İstanbul', 'Kadıköy Mah. 123. Cad.'],
            ['XYZ İnşaat', 'TEDARİKÇİ', '0987654321', 'sales@xyz.com', '5329876543', 'Ankara', 'Çankaya Mah. 456. Cad.'],
            ['DEF Perakende', 'HER İKİSİ', '1112223334', 'info@def.com', '5338889999', 'İzmir', 'Alsancak Mah. 789. Cad.'],
            ['GHI Hizmetler', 'MÜŞTERİ', '5556667778', 'contact@ghi.com', '5345556666', 'Bursa', 'Osmangazi Mah. 111. Cad.'],
            ['JKL Üretim', 'TEDARİKÇİ', '9998887776', 'sales@jkl.com', '5357778888', 'Gaziantep', 'Şahinbey Mah. 222. Cad.'],
        ],
        'widths': [20, 15, 15, 20, 15, 15, 30]
    }
    
    create_excel_file(str(sample_dir / "01_carileri_aktar.xlsx"), carileri_data)
    
    # 2. BANKA HESAPLARI İÇE AKTARMA
    banka_data = {
        'sheet_name': 'Banka Hesapları',
        'headers': ['Banka Adı', 'Hesap Numarası', 'IBAN', 'Şube', 'Bakiye', 'Para Birimi', 'Limit'],
        'data': [
            ['Ziraat Bankası', '1234567890', 'TR330006100519786457841326', 'İstanbul', '50000', 'TRY', '10000'],
            ['Garanti BBVA', '9876543210', 'TR820001001458028343219026', 'Ankara', '30000', 'TRY', '15000'],
            ['İş Bankası', '5555555555', 'TR440010000000123456789012', 'İzmir', '75000', 'TRY', '20000'],
            ['Vakıfbank', '1111111111', 'TR380001000050000001234567', 'Bursa', '25000', 'TRY', '5000'],
            ['Akbank', '2222222222', 'TR350006500001234567890123', 'Gaziantep', '100000', 'TRY', '25000'],
        ],
        'widths': [18, 15, 30, 15, 12, 12, 12]
    }
    
    create_excel_file(str(sample_dir / "02_banka_hesaplari_aktar.xlsx"), banka_data)
    
    # 3. İŞLEMLER İÇE AKTARMA
    islem_data = {
        'sheet_name': 'İşlemler',
        'headers': ['Tarih', 'Cari Adı', 'Banka', 'İşlem Tipi', 'Tutar', 'Açıklama'],
        'data': [
            [today.strftime("%d.%m.%Y"), 'ABC Ticaret Ltd.', 'Ziraat Bankası', 'Gelen', '5000', 'Ürün satışı'],
            [(today - timedelta(days=1)).strftime("%d.%m.%Y"), 'XYZ İnşaat', 'Garanti BBVA', 'Giden', '2500', 'İnşaat malzemeleri'],
            [(today - timedelta(days=2)).strftime("%d.%m.%Y"), 'DEF Perakende', 'İş Bankası', 'Gelen', '7500', 'Kargo ücretleri'],
            [(today - timedelta(days=3)).strftime("%d.%m.%Y"), 'GHI Hizmetler', 'Vakıfbank', 'Giden', '1200', 'Hizmet ücreti'],
            [(today - timedelta(days=4)).strftime("%d.%m.%Y"), 'JKL Üretim', 'Akbank', 'Gelen', '15000', 'Toplu satış'],
        ],
        'widths': [12, 20, 18, 12, 12, 25]
    }
    
    create_excel_file(str(sample_dir / "03_islemler_aktar.xlsx"), islem_data)
    
    # 4. FATURALAR İÇE AKTARMA
    fatura_data = {
        'sheet_name': 'Faturalar',
        'headers': ['Fatura No', 'Tarih', 'Vade Tarihi', 'Cari Adı', 'Tip', 'Tutar', 'KDV %', 'KDV Tutar', 'Toplam', 'Banka'],
        'data': [
            ['F-2024-001', today.strftime("%d.%m.%Y"), (today + timedelta(days=30)).strftime("%d.%m.%Y"), 'ABC Ticaret Ltd.', 'GIDEN', '1000', '20', '200', '1200', 'Ziraat Bankası'],
            ['F-2024-002', (today - timedelta(days=5)).strftime("%d.%m.%Y"), (today + timedelta(days=25)).strftime("%d.%m.%Y"), 'XYZ İnşaat', 'GELEN', '5000', '20', '1000', '6000', 'Garanti BBVA'],
            ['F-2024-003', (today - timedelta(days=10)).strftime("%d.%m.%Y"), (today + timedelta(days=20)).strftime("%d.%m.%Y"), 'DEF Perakende', 'GIDEN', '2500', '18', '450', '2950', 'İş Bankası'],
            ['F-2024-004', (today - timedelta(days=15)).strftime("%d.%m.%Y"), (today + timedelta(days=15)).strftime("%d.%m.%Y"), 'GHI Hizmetler', 'GELEN', '3000', '20', '600', '3600', 'Vakıfbank'],
            ['F-2024-005', (today - timedelta(days=20)).strftime("%d.%m.%Y"), (today + timedelta(days=10)).strftime("%d.%m.%Y"), 'JKL Üretim', 'GIDEN', '8000', '20', '1600', '9600', 'Akbank'],
        ],
        'widths': [12, 12, 12, 20, 8, 10, 8, 12, 10, 18]
    }
    
    create_excel_file(str(sample_dir / "04_faturalar_aktar.xlsx"), fatura_data)
    
    # 5. KREDİ KARTI İŞLEMLERİ İÇE AKTARMA
    kredi_data = {
        'sheet_name': 'Kredi Kartları',
        'headers': ['Kart Adı', 'Son 4 Hanesi', 'Banka', 'Limit', 'Limit Kullanımı', 'Kesim Günü', 'Vade Günü'],
        'data': [
            ['Kişisel Kredi Kartı 1', '1234', 'Ziraat Bankası', '50000', '15000', '1', '25'],
            ['İşletme Kredi Kartı', '5678', 'Garanti BBVA', '100000', '45000', '5', '29'],
            ['Yedek Kredi Kartı', '9012', 'İş Bankası', '30000', '8000', '10', '30'],
        ],
        'widths': [20, 13, 15, 12, 18, 12, 11]
    }
    
    create_excel_file(str(sample_dir / "05_kredi_kartlari_aktar.xlsx"), kredi_data)
    
    print("\n" + "="*50)
    print("✓ Tüm örnek dosyalar başarıyla oluşturuldu!")
    print("="*50)
    print("\nOluşturulan dosyalar 'sample_data' klasöründe:")
    print("  1. 01_carileri_aktar.xlsx - Müşteri/Tedarikçi verisi")
    print("  2. 02_banka_hesaplari_aktar.xlsx - Banka hesapları")
    print("  3. 03_islemler_aktar.xlsx - Banka işlemleri")
    print("  4. 04_faturalar_aktar.xlsx - Fatura verileri")
    print("  5. 05_kredi_kartlari_aktar.xlsx - Kredi kartı verileri")
    print("\nBu dosyaları Excel'de düzenleyebilir veya")
    print("uygulamaya içe aktarabilirsiniz.\n")

if __name__ == "__main__":
    main()
