from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QTabWidget, 
                           QLabel, QPushButton, QMessageBox, QTableWidget, QTableWidgetItem,
                           QFrame, QHeaderView, QComboBox, QDateEdit, QLineEdit, QTextBrowser,
                           QScrollArea, QSizePolicy, QCheckBox, QGroupBox, QSpinBox, QDialog,
                           QSplitter, QTextEdit)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from PyQt5.QtCore import Qt, QDate, QTimer, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor
from src.database.models import User, TransactionType, PaymentMethod, Transaction
from src.services.invoice_service import InvoiceService
from src.services.cari_service import CariService
from src.services.bank_service import BankService
from src.services.transaction_service import TransactionService
from src.services.credit_card_service import CreditCardService
from src.services.loan_service import LoanService
from src.services.report_service import ReportService
from src.ui.dialogs.user_management_dialog import UserManagementDialog
from src.services.auth_service import AuthService
from src.services.user_settings_service import UserSettingsService
from src.services.google_sheets_service import GoogleSheetsService
from src.utils.app_icon import get_app_icon
from src.utils.helpers import format_currency_tr, format_tr
from src.ui.kira_takip import KiraTakipWidget
from datetime import datetime, date, timedelta
from collections import defaultdict
from pathlib import Path
import shutil
import json
import config

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    Font = Alignment = PatternFill = Border = Side = get_column_letter = None


class GoogleSheetsWorker(QThread):
    finished_signal = pyqtSignal(str, bool, str, dict)

    def __init__(self, operation: str, user_id: int, spreadsheet_id: str, sheet_mappings=None):
        super().__init__()
        self.operation = operation
        self.user_id = user_id
        self.spreadsheet_id = spreadsheet_id
        self.sheet_mappings = sheet_mappings or {}

    def run(self):
        try:
            service = GoogleSheetsService()
            if self.operation == 'test':
                success, message = service.test_connection(self.spreadsheet_id)
                self.finished_signal.emit(self.operation, success, message, {})
                return

            if self.operation == 'sync':
                success, message, stats = service.sync_from_sheets(
                    self.user_id,
                    self.spreadsheet_id,
                    self.sheet_mappings
                )
                self.finished_signal.emit(self.operation, success, message, stats or {})
                return

            self.finished_signal.emit(self.operation, False, "Geçersiz işlem türü", {})
        except Exception as e:
            self.finished_signal.emit(self.operation, False, str(e), {})


class MainWindow(QMainWindow):
    """Ana uygulama penceresi"""
    
    def __init__(self, user: User):
        super().__init__()
        self.user = user
        self.auto_backup_timer = None
        self.gsheets_worker = None
        app_icon = get_app_icon()
        if not app_icon.isNull():
            self.setWindowIcon(app_icon)
        self.setWindowTitle(f"{config.APP_NAME} - {user.full_name} ({user.role.upper()})")
        self.setGeometry(0, 0, config.WINDOW_WIDTH, config.WINDOW_HEIGHT)
        
        self.setStyleSheet("""
            QMainWindow { background-color: #fafafa; }
            QTabWidget::pane { border: 1px solid #ddd; }
            QTabBar::tab { background-color: #e0e0e0; padding: 8px 12px; min-width: 70px; min-height: 28px; }
            QTabBar::tab:selected { background-color: white; }
            QTabBar::scroller { width: 28px; }
            QTabBar QToolButton { background-color: #d0d0d0; border: 1px solid #bbb; border-radius: 2px; }
            QTabBar QToolButton:hover { background-color: #b0b0b0; }
            QLabel { font-size: 10pt; }
            QLineEdit, QComboBox, QDateEdit, QTextEdit, QSpinBox, QDoubleSpinBox {
                padding: 6px 8px;
                font-size: 10pt;
                min-height: 30px;
            }
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 14px;
                font-weight: bold;
                min-height: 30px;
            }
            QPushButton:hover { background-color: #1976D2; }
            QTableWidget {
                border: 1px solid #ddd;
                font-size: 10pt;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 6px;
                font-size: 9pt;
            }
        """)
        
        self.init_ui()
        self.setup_auto_backup_scheduler()
        self.setup_google_sheets_timer()
        self.setup_dashboard_refresh_timer()
    
    def init_ui(self):
        """Arayüz öğelerini başlat"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Üst yenileme çubuğu
        top_bar = QHBoxLayout()
        top_bar.setSpacing(8)

        btn_refresh = QPushButton("🔄 Yenile")
        btn_refresh.setMinimumHeight(30)
        btn_refresh.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #546E7A; }
        """)
        btn_refresh.clicked.connect(self.manual_refresh_all)
        top_bar.addWidget(btn_refresh)

        self.last_refresh_label = QLabel("Son yenileme: -")
        self.last_refresh_label.setStyleSheet("color: #666;")
        top_bar.addWidget(self.last_refresh_label)
        top_bar.addStretch()

        layout.addLayout(top_bar)

        # Tab widget
        self.tabs = QTabWidget()
        self.tabs.setUsesScrollButtons(True)
        self.tabs.tabBar().setExpanding(False)
        self.tabs.tabBar().setElideMode(Qt.ElideNone)
        self.dashboard_tab = None
        self._loaded_tabs = set()
        
        # Dashboard (izinli ise)
        if self.user.can_view_dashboard:
            self.dashboard_tab = self.create_dashboard_tab()
            self.tabs.addTab(self.dashboard_tab, "📊 Dashboard")
        
        # Rol bazlı sekmeler
        if self.user.role == 'admin':
            if self.user.can_view_transactions:
                self.tabs.addTab(self.create_transactions_tab(), "💰 İşlemler")
            if self.user.can_view_invoices:
                self.tabs.addTab(self.create_invoices_tab(), "📄 Faturalar")
            if self.user.can_view_caris:
                self.tabs.addTab(self.create_caris_tab(), "📋 Cari Hesaplar")
            if self.user.can_view_cari_extract:
                self.tabs.addTab(self.create_cari_extract_tab(), "📑 Cari Ekstre")
            if self.user.can_view_banks:
                self.tabs.addTab(self.create_bank_tab(), "🏦 Banka Hesapları")
            if self.user.can_view_credit_cards:
                self.tabs.addTab(self.create_credit_cards_tab(), "💳 Kredi Kartları")
            if self.user.can_view_loans:
                self.tabs.addTab(self.create_loans_tab(), "📊 Krediler")
            if getattr(self.user, 'can_view_kira_takip', True):
                self.tabs.addTab(self.create_kira_takip_tab(), "🏠 Kira Takip")
            if self.user.can_view_reports:
                self.tabs.addTab(self.create_reports_tab(), "📊 Raporlar")
            if self.user.can_view_payroll:
                self.tabs.addTab(self.create_payroll_tab(), "💼 Maaş Bordro")
            if self.user.can_view_employees:
                self.tabs.addTab(self.create_employees_tab(), "👥 Çalışanlar")
            if self.user.can_view_bulk_payroll:
                self.tabs.addTab(self.create_bulk_payroll_tab(), "⚙️ Toplu Bordro Hesapla")
            if self.user.can_view_payroll_records:
                self.tabs.addTab(self.create_payroll_records_tab(), "📚 Bordro Kayıtları")
            if self.user.can_view_admin_panel:
                self.tabs.addTab(self.create_admin_panel_tab(), "👨‍💼 Admin Panel")
        else:
            # Normal kullanıcılar sadece izin verilenleri görebilir
            if self.user.can_view_transactions:
                self.tabs.addTab(self.create_transactions_tab(), "💰 İşlemler")
            if self.user.can_view_invoices:
                self.tabs.addTab(self.create_invoices_tab(), "📄 Faturalar")
            if self.user.can_view_caris:
                self.tabs.addTab(self.create_caris_tab(), "📋 Cari Hesaplar")
            if self.user.can_view_cari_extract:
                self.tabs.addTab(self.create_cari_extract_tab(), "📑 Cari Ekstre")
            if self.user.can_view_banks:
                self.tabs.addTab(self.create_bank_tab(), "🏦 Banka Hesapları")
            if self.user.can_view_credit_cards:
                self.tabs.addTab(self.create_credit_cards_tab(), "💳 Kredi Kartları")
            if self.user.can_view_loans:
                self.tabs.addTab(self.create_loans_tab(), "📊 Krediler")
            if getattr(self.user, 'can_view_kira_takip', True):
                self.tabs.addTab(self.create_kira_takip_tab(), "🏠 Kira Takip")
            if self.user.can_view_reports:
                self.tabs.addTab(self.create_reports_tab(), "📊 Raporlar")
            if self.user.can_view_payroll:
                self.tabs.addTab(self.create_payroll_tab(), "💼 Maaş Bordro")
            if self.user.can_view_employees:
                self.tabs.addTab(self.create_employees_tab(), "👥 Çalışanlar")
            if self.user.can_view_bulk_payroll:
                self.tabs.addTab(self.create_bulk_payroll_tab(), "⚙️ Toplu Bordro Hesapla")
            if self.user.can_view_payroll_records:
                self.tabs.addTab(self.create_payroll_records_tab(), "📚 Bordro Kayıtları")
        
        if self.user.can_view_settings:
            self.tabs.addTab(self.create_settings_tab(), "⚙️ Ayarlar")

        self.tabs.currentChanged.connect(self.on_tab_changed)
        layout.addWidget(self.tabs)
        central_widget.setLayout(layout)
        
        self.showMaximized()
        QTimer.singleShot(0, lambda: self._refresh_current_tab_data(force=True))

    def _resize_table(self, table, stretch_col: int = None, min_row_height: int = 36):
        """Tablo kolon/satırlarını içeriğe göre ayarla.
        stretch_col: geri kalan boşluğu dolduracak sütun indexi (None = son sütun uzar)
        """
        if table is None:
            return

        table.setWordWrap(True)
        row_count = table.rowCount()

        # Büyük tablolarda otomatik içerik hesaplaması eski bilgisayarlarda ciddi yavaşlığa neden olur.
        if row_count <= 200:
            # Fusion EXE'de setMinimumSectionSize ÖNCE ayarlanmalı, yoksa resizeRowsToContents sıfırlar
            table.verticalHeader().setMinimumSectionSize(min_row_height)
            table.verticalHeader().setDefaultSectionSize(min_row_height)
            table.resizeColumnsToContents()
            table.resizeRowsToContents()
            # Ekstra güvence: herhangi bir satır hâlâ küçükse zorla
            for i in range(row_count):
                if table.rowHeight(i) < min_row_height:
                    table.setRowHeight(i, min_row_height)
        else:
            table.verticalHeader().setMinimumSectionSize(28)
            table.verticalHeader().setDefaultSectionSize(32)

        header = table.horizontalHeader()
        if stretch_col is not None:
            header.setStretchLastSection(False)
            for col in range(table.columnCount()):
                if col == stretch_col:
                    header.setSectionResizeMode(col, QHeaderView.Stretch)
                else:
                    header.setSectionResizeMode(col, QHeaderView.Interactive)
        else:
            header.setStretchLastSection(True)

    def manual_refresh_all(self):
        """Butonla tum verileri yenile"""
        self.refresh_all_data()
        if hasattr(self, "last_refresh_label"):
            self.last_refresh_label.setText(
                f"Son yenileme: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
            )

    def on_tab_changed(self, index):
        """Sekme değişince yalnızca aktif ekranı yenile."""
        self._refresh_current_tab_data(force=True)

    def _refresh_current_tab_data(self, force=False):
        """Görünmeyen sekmeleri zorlamadan yalnızca aktif sekmenin verisini yükle."""
        if not hasattr(self, "tabs") or self.tabs.count() == 0:
            return

        try:
            tab_text = self.tabs.tabText(self.tabs.currentIndex())
            if not force and tab_text in self._loaded_tabs:
                return

            if "Dashboard" in tab_text:
                self.refresh_dashboard()
            elif "İşlemler" in tab_text and hasattr(self, 'table_transactions'):
                self.refresh_transactions_table()
            elif "Faturalar" in tab_text and hasattr(self, 'table_invoices'):
                self.refresh_invoice_table()
            elif "Cari Hesaplar" in tab_text and hasattr(self, 'table_caris'):
                self.refresh_cari_table()
            elif "Cari Ekstre" in tab_text and hasattr(self, '_reload_cari_extract_combo'):
                self._reload_cari_extract_combo()
            elif "Banka Hesapları" in tab_text and hasattr(self, 'table_banks'):
                self.refresh_bank_table()
            elif "Kredi Kartları" in tab_text and hasattr(self, 'table_credit_cards'):
                self.refresh_credit_cards_table()
            elif "Krediler" in tab_text and hasattr(self, 'table_loans'):
                self.refresh_loan_stats()
                self.refresh_loans_table()

            self._loaded_tabs.add(tab_text)
        except Exception as e:
            print(f"Sekme yenileme hatası: {e}")

    def _auto_refresh_dashboard(self):
        """Dashboard'u arka planda yenile"""
        if not getattr(self.user, "can_view_dashboard", False):
            return
        try:
            self.refresh_dashboard()
            if hasattr(self, "last_refresh_label"):
                self.last_refresh_label.setText(
                    f"Son yenileme: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
                )
        except Exception as e:
            print(f"Dashboard otomatik yenileme hatasi: {e}")

    def setup_dashboard_refresh_timer(self):
        """Dashboard icin otomatik yenileme zamanlayicisi"""
        self.dashboard_refresh_timer = QTimer(self)
        self.dashboard_refresh_timer.timeout.connect(self._auto_refresh_dashboard)
        self.dashboard_refresh_timer.start(60 * 1000)
    
    def _get_payment_method_display_text(self, payment_method_value, payment_type=None):
        """Ödeme metodu için görüntü metni döndür - önce payment_type'a bak, yoksa enum'a"""
        # Eğer payment_type varsa (orijinal değer Excel'den), onu göster
        if payment_type and payment_type.strip():
            return payment_type.strip()
        
        # Yoksa payment_method enum'unu Türçkleştir
        payment_display = {
            'NAKIT': 'Nakit',
            'BANKA': 'Banka Hesabı',
            'KREDI_KARTI': 'Kredi Kartları',
            'CARI': 'Cari Hesap',
            'TRANSFER': 'Transfer'
        }
        return payment_display.get(payment_method_value, payment_method_value)

    def _get_dashboard_card_defs(self):
        return [
            ("total_invoices", "Toplam Fatura", "#2196F3"),
            ("total_cari", "Toplam Cari", "#4CAF50"),
            ("total_income", "Toplam Gelir", "#009688"),
            ("total_expense", "Toplam Gider", "#f44336"),
            ("bank_total_balance", "Banka Toplam Bakiye", "#3F51B5"),
            ("credit_card_total_debt", "Kredi Kartı Toplam Borç", "#FF9800"),
            ("loan_total_debt", "Kredi Toplam Borç", "#E91E63"),
            ("borrow_total_debt", "Ödünç Borçlar Toplamı", "#795548"),
            ("kira_monthly", "Kira (Bu Ay Tahsilat)", "#1565C0"),
            ("overdue_invoices", "\u26d4 Gecikmi\u015f Fatura", "#B71C1C"),
        ]

    def _get_dashboard_card_keys(self):
        default_keys = [key for key, _, _ in self._get_dashboard_card_defs()]
        keys = UserSettingsService.get_json_setting(self.user.id, "dashboard_cards", None)
        if not keys:
            UserSettingsService.set_json_setting(self.user.id, "dashboard_cards", default_keys)
            return default_keys
        # Sadece gecerli (tanimli) kartlari dondur, kullanicinin kaldirdigini geri ekleme
        return [k for k in keys if k in default_keys]

    def _set_dashboard_card_value(self, key, value):
        if not hasattr(self, "dashboard_cards"):
            return
        card = self.dashboard_cards.get(key)
        if not card:
            return
        labels = card.findChildren(QLabel)
        if len(labels) > 1:
            labels[1].setText(value)
    
    def create_dashboard_tab(self) -> QWidget:
        """Dashboard sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Başlık
        title = QLabel(f"Hoşgeldiniz, {self.user.username}!")
        title.setFont(QFont("SegoeUI", 24, QFont.Bold))
        title.setStyleSheet("color: #333;")
        layout.addWidget(title)
        
        # Hızlı Eylem Butonları
        quick_actions_layout = QHBoxLayout()
        
        # Google Sheets Sync Butonu
        gsheets_url = UserSettingsService.get_setting(self.user.id, "google_sheets_url", None)
        if gsheets_url:
            btn_gsheets_sync = QPushButton("🔄 Google Sheets Senkronize")
            btn_gsheets_sync.setMinimumHeight(32)
            btn_gsheets_sync.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 12px;
                    font-weight: bold;
                    font-size: 10pt;
                }
                QPushButton:hover { background-color: #45a049; }
            """)
            btn_gsheets_sync.clicked.connect(self.sync_from_google_sheets)
            quick_actions_layout.addWidget(btn_gsheets_sync)
        
        quick_actions_layout.addStretch()
        layout.addLayout(quick_actions_layout)
        
        layout.addSpacing(10)
        
        # Dashboard cards
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(15)

        self.dashboard_cards = {}
        selected_keys = set(self._get_dashboard_card_keys())
        for key, title_text, color in self._get_dashboard_card_defs():
            if key not in selected_keys:
                continue
            card = self.create_stat_card(title_text, "0", color)
            self.dashboard_cards[key] = card
            cards_layout.addWidget(card)

        layout.addLayout(cards_layout)
        
        # İki tablo yan yana
        tables_layout = QHBoxLayout()
        tables_layout.setSpacing(20)
        
        # Sol: Bugün Yapılacak Ödemeler
        left_layout = QVBoxLayout()
        left_title = QLabel("📅 Bugün Yapılacak Ödemeler")
        left_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        left_title.setStyleSheet("color: #d32f2f; margin-bottom: 5px;")
        left_layout.addWidget(left_title)
        
        self.table_recent = QTableWidget()
        self.table_recent.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_recent.setColumnCount(6)
        self.table_recent.setHorizontalHeaderLabels(["Tarih", "Müşteri", "Ürün", "İşlem", "Miktar", "İşlemler"])
        self.table_recent.horizontalHeader().setStretchLastSection(False)
        self.table_recent.setColumnWidth(0, 90)
        self.table_recent.setColumnWidth(1, 180)
        self.table_recent.setColumnWidth(2, 180)
        self.table_recent.setColumnWidth(3, 120)
        self.table_recent.setColumnWidth(4, 100)
        self.table_recent.setColumnWidth(5, 140)
        self.table_recent.setStyleSheet("QTableWidget { border: 1px solid #ddd; border-radius: 4px; }")
        self.table_recent.setMinimumHeight(400)
        left_layout.addWidget(self.table_recent)
        
        # Sağ: Ödünç Borçlar
        right_layout = QVBoxLayout()
        right_title = QLabel("Ödünç Borçlar")
        right_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        right_title.setStyleSheet("color: #d32f2f; margin-bottom: 5px;")
        right_layout.addWidget(right_title)
        
        self.table_pending = QTableWidget()
        self.table_pending.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_pending.setColumnCount(3)
        self.table_pending.setHorizontalHeaderLabels(["Cari", "Kalan Borç", "İşlemler"])
        self.table_pending.horizontalHeader().setStretchLastSection(False)
        self.table_pending.setColumnWidth(0, 260)
        self.table_pending.setColumnWidth(1, 140)
        self.table_pending.setColumnWidth(2, 160)
        self.table_pending.setStyleSheet("QTableWidget { border: 2px solid #ffcdd2; border-radius: 4px; }")
        self.table_pending.setMinimumHeight(400)
        right_layout.addWidget(self.table_pending)
        
        tables_layout.addLayout(left_layout, 1)
        tables_layout.addLayout(right_layout, 1)
        
        layout.addLayout(tables_layout, 1)
        widget.setLayout(layout)
        
        self.refresh_dashboard()
        return widget
    
    def create_stat_card(self, title: str, value: str, color: str) -> QFrame:
        """İstatistik kartı oluştur"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 8px;
                padding: 0px;
            }}
        """)
        card.setMinimumHeight(120)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(5)
        
        lbl_title = QLabel(title)
        lbl_title.setFont(QFont("Segoe UI", 11))
        lbl_title.setStyleSheet("color: white; font-weight: bold;")
        layout.addWidget(lbl_title)
        
        lbl_value = QLabel(value)
        lbl_value.setFont(QFont("Segoe UI", 32, QFont.Bold))
        lbl_value.setStyleSheet("color: white;")
        layout.addWidget(lbl_value)
        
        card.setLayout(layout)
        return card
    
    def refresh_dashboard(self):
        """Dashboard verileri yenile"""
        try:
            from src.database.db import SessionLocal
            from src.database.models import Transaction, TransactionType, BankAccount, CreditCard, Loan, Cari
            from sqlalchemy import func
            
            session = SessionLocal()
            
            # Transaction tablosundan fatura istatistiklerini al
            all_transactions = session.query(Transaction).filter_by(user_id=self.user.id).all()
            
            # Fatura işlemlerini filtrele
            invoice_transactions = [t for t in all_transactions if t.transaction_type in [
                TransactionType.KESILEN_FATURA, 
                TransactionType.GELEN_FATURA
            ]]
            
            # Cari hesapları al
            caris = CariService.get_caris(self.user.id)
            
            # İstatistikleri güncelle
            total_invoices = len(invoice_transactions)
            total_cari = len(caris) if caris else 0

            total_income = sum(
                t.amount for t in all_transactions
                if t.transaction_type in [TransactionType.GELIR, TransactionType.KESILEN_FATURA]
            )
            total_expense = sum(
                t.amount for t in all_transactions
                if t.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]
            )

            bank_total_balance = session.query(BankAccount).filter(
                BankAccount.user_id == self.user.id,
                BankAccount.is_active == True
            ).with_entities(func.sum(BankAccount.balance)).scalar() or 0.0

            credit_card_total_debt = session.query(CreditCard).filter(
                CreditCard.user_id == self.user.id,
                CreditCard.is_active == True
            ).with_entities(func.sum(CreditCard.current_debt)).scalar() or 0.0

            loan_summary = LoanService.get_loans_summary(self.user.id) or {}
            loan_total_debt = loan_summary.get('toplam_kalan', 0.0)

            borrow_candidates = session.query(Cari).filter(
                Cari.user_id == self.user.id,
                Cari.is_active == True
            ).all()
            borrow_total_debt = sum(c.balance for c in borrow_candidates if self._is_borrow_cari(c))

            self._set_dashboard_card_value("total_invoices", str(total_invoices))
            self._set_dashboard_card_value("total_cari", str(total_cari))
            self._set_dashboard_card_value("total_income", format_currency_tr(total_income))
            self._set_dashboard_card_value("total_expense", format_currency_tr(total_expense))
            self._set_dashboard_card_value("bank_total_balance", format_currency_tr(bank_total_balance))
            self._set_dashboard_card_value("credit_card_total_debt", format_currency_tr(credit_card_total_debt))
            self._set_dashboard_card_value("loan_total_debt", format_currency_tr(loan_total_debt))
            self._set_dashboard_card_value("borrow_total_debt", format_currency_tr(borrow_total_debt))

            # ── Gecikmiş Kesilen Fatura toplam tutarı ──────────────────────────
            from datetime import date as _date_type
            _today = _date_type.today()
            overdue_amount = sum(
                (t.amount - (getattr(t, 'paid_amount', 0.0) or 0.0))
                for t in all_transactions
                if t.transaction_type == TransactionType.KESILEN_FATURA
                and getattr(t, 'due_date', None) is not None
                and (t.due_date - _today).days < 0
                and not (getattr(t, 'paid_amount', 0.0) or 0.0) >= t.amount
            )
            self._set_dashboard_card_value("overdue_invoices", format_currency_tr(overdue_amount))

            # ── Kira Takip: Bu ay beklenen tahsilat ──────────────────────
            try:
                import json
                from pathlib import Path
                _kira_file = Path("data") / f"kira_takip_data_{self.user.id}.json"
                _kira_monthly = 0.0
                if _kira_file.exists():
                    _kira_data = json.loads(_kira_file.read_text(encoding="utf-8"))
                    _cur_month = date.today().month
                    _cur_year  = date.today().year
                    for _tab in _kira_data.get("tabs", []):
                        if _tab.get("year", _cur_year) != _cur_year:
                            continue
                        _payments = {}
                        for _pid, _pv in _tab.get("payments", {}).items():
                            _payments[int(_pid)] = {int(_m): _s for _m, _s in _pv.items()}
                        for _c in _tab.get("contracts", []):
                            _cid = int(_c.get("id", -1))
                            _durum = _payments.get(_cid, {}).get(_cur_month, "ODENMEDI")
                            if _durum == "ODENDI":
                                _kira_monthly += float(_c.get("tutar", 0))
                            elif isinstance(_durum, str) and _durum.startswith("KISMI:"):
                                try:
                                    _kira_monthly += float(_durum[6:].replace(".","").replace(",","."))
                                except Exception:
                                    pass
                self._set_dashboard_card_value("kira_monthly", format_currency_tr(_kira_monthly))
            except Exception:
                pass

            # Bugün yapılacak ödemeleri göster (yalnızca bugün tarihli işlemler + vadesi bugün olan krediler)
            from datetime import date
            today = date.today()

            def _normalize_date(value):
                if value is None:
                    return None
                if isinstance(value, datetime):
                    return value.date()
                if isinstance(value, date):
                    return value
                if hasattr(value, "toPyDate"):
                    return value.toPyDate()
                if isinstance(value, str):
                    text = value.strip()
                    if not text:
                        return None
                    try:
                        return datetime.fromisoformat(text).date()
                    except ValueError:
                        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                            try:
                                return datetime.strptime(text, fmt).date()
                            except ValueError:
                                continue
                return None

            def _parse_due_day(value):
                if value is None:
                    return None
                if isinstance(value, int):
                    return value
                if isinstance(value, float):
                    return int(value)
                text = str(value).strip()
                if not text:
                    return None
                if text.endswith(".0"):
                    text = text[:-2]
                digits = "".join(ch for ch in text if ch.isdigit())
                if not digits:
                    return None
                try:
                    return int(digits)
                except ValueError:
                    return None
            
            pending_rows = []

            transaction_pending_payments = [
                t for t in all_transactions
                if t.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]
                and _normalize_date(t.transaction_date) == today
            ]
            for trans in transaction_pending_payments:
                trans_date = _normalize_date(trans.transaction_date) or today
                pending_rows.append({
                    'row_type': 'transaction',
                    'date': trans_date,
                    'customer': trans.cari.name if trans.cari else '-',
                    'product': trans.description[:30] if trans.description else '-',
                    'operation': {
                        'GELIR': 'Gelir',
                        'GIDER': 'Gider',
                        'KESILEN_FATURA': 'Kesilen Fatura',
                        'GELEN_FATURA': 'Gelen Fatura',
                        'KREDI_ODEME': 'Kredi Ödeme',
                        'KREDI_KARTI_ODEME': 'KK Ödeme'
                    }.get(trans.transaction_type.value, trans.transaction_type.value),
                    'amount': trans.amount,
                    'transaction_id': trans.id,
                    'loan_id': None
                })

            loans_due_today = session.query(Loan).filter(
                Loan.user_id == self.user.id,
                Loan.is_active == True,
                Loan.status == 'AKTIF'
            ).all()
            for loan in loans_due_today:
                due_day = _parse_due_day(loan.due_day)
                if due_day is None or due_day != today.day:
                    continue
                remaining_amount = self._get_loan_remaining_amount(loan)
                if remaining_amount <= 0:
                    continue
                installment_amount = loan.monthly_payment if loan.monthly_payment and loan.monthly_payment > 0 else remaining_amount
                installment_amount = min(installment_amount, remaining_amount)
                pending_rows.append({
                    'row_type': 'loan_due',
                    'date': today,
                    'customer': loan.loan_name,
                    'product': f"{loan.bank_name} - {loan.loan_type}",
                    'operation': 'Kredi Taksidi',
                    'amount': installment_amount,
                    'transaction_id': None,
                    'loan_id': loan.id,
                    'credit_card_id': None
                })

            credit_cards_due_today = session.query(CreditCard).filter(
                CreditCard.user_id == self.user.id,
                CreditCard.is_active == True,
                CreditCard.current_debt > 0
            ).all()
            for card in credit_cards_due_today:
                due_day = _parse_due_day(card.due_day)
                if due_day is None or due_day != today.day:
                    continue
                pending_rows.append({
                    'row_type': 'credit_card_due',
                    'date': today,
                    'customer': card.card_name,
                    'product': card.bank_name,
                    'operation': 'Kredi Kartı Ödeme',
                    'amount': card.current_debt,
                    'transaction_id': None,
                    'loan_id': None,
                    'credit_card_id': card.id
                })

            # Tarihe göre sırala (en yakın önce)
            recent = sorted(pending_rows, key=lambda x: x['date'])[:15]
            self.table_recent.setRowCount(len(recent))
            
            for i, row in enumerate(recent):
                # Tarih - bugünse vurgula
                date_item = QTableWidgetItem(str(row['date']))
                if row['date'] == today:
                    date_item.setBackground(Qt.yellow)
                elif row['date'] < today:
                    date_item.setBackground(QColor(255, 200, 200))  # Açık kırmızı - gecikmiş
                self.table_recent.setItem(i, 0, date_item)
                
                # Müşteri (Cari)
                self.table_recent.setItem(i, 1, QTableWidgetItem(row['customer']))
                
                # Ürün (Açıklama)
                self.table_recent.setItem(i, 2, QTableWidgetItem(row['product']))
                
                # İşlem (İşlem Türü)
                self.table_recent.setItem(i, 3, QTableWidgetItem(row['operation']))
                
                # Miktar
                self.table_recent.setItem(i, 4, QTableWidgetItem(f"{format_tr(row['amount'])} ₺"))

                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)
                action_layout.setSpacing(5)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                if row['row_type'] == 'transaction':
                    btn_edit = QPushButton("✏️ Düzenle")
                    btn_edit.setMinimumHeight(24)
                    btn_edit.setStyleSheet("""
                        QPushButton {
                            background-color: #2196F3;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #1976D2; }
                    """)
                    btn_edit.clicked.connect(lambda checked, tid=row['transaction_id']: self.edit_transaction(tid))
                    action_layout.addWidget(btn_edit)

                    btn_delete = QPushButton("🗑️ Sil")
                    btn_delete.setMinimumHeight(24)
                    btn_delete.setStyleSheet("""
                        QPushButton {
                            background-color: #f44336;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #da190b; }
                    """)
                    btn_delete.clicked.connect(lambda checked, tid=row['transaction_id']: self.delete_transaction(tid))
                    action_layout.addWidget(btn_delete)
                elif row['row_type'] == 'loan_due':
                    btn_preview = QPushButton("📑 Dökümü Aç")
                    btn_preview.setMinimumHeight(24)
                    btn_preview.setStyleSheet("""
                        QPushButton {
                            background-color: #4CAF50;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #45a049; }
                    """)
                    btn_preview.clicked.connect(lambda checked, lid=row['loan_id']: self.show_loan_statement(lid))
                    action_layout.addWidget(btn_preview)
                elif row['row_type'] == 'credit_card_due':
                    btn_preview = QPushButton("📑 Dökümü Aç")
                    btn_preview.setMinimumHeight(24)
                    btn_preview.setStyleSheet("""
                        QPushButton {
                            background-color: #4CAF50;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #45a049; }
                    """)
                    btn_preview.clicked.connect(
                        lambda checked, cid=row['credit_card_id']: self.show_credit_card_statement(cid)
                    )
                    action_layout.addWidget(btn_preview)

                self.table_recent.setCellWidget(i, 5, action_widget)
            
            # Ödünç borçlar - cari tipi "ÖDÜNÇ PARA"
            borrow_candidates = session.query(Cari).filter(
                Cari.user_id == self.user.id,
                Cari.is_active == True
            ).order_by(Cari.name.asc()).all()
            borrow_caris = [c for c in borrow_candidates if self._is_borrow_cari(c)]
            self.table_pending.setRowCount(len(borrow_caris))

            for i, cari in enumerate(borrow_caris):
                self.table_pending.setItem(i, 0, QTableWidgetItem(cari.name))
                self.table_pending.setItem(i, 1, QTableWidgetItem(f"{format_tr(cari.balance)} ₺"))

                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)
                action_layout.setSpacing(5)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                btn_statement = QPushButton("📑 Dökümü Aç")
                btn_statement.setMinimumHeight(24)
                btn_statement.setStyleSheet("""
                    QPushButton {
                        background-color: #4CAF50;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #45a049; }
                """)
                btn_statement.clicked.connect(lambda checked, cid=cari.id: self.show_cari_extract_for(cid))
                action_layout.addWidget(btn_statement)

                self.table_pending.setCellWidget(i, 2, action_widget)
            
            self._resize_table(self.table_recent)
            self._resize_table(self.table_pending)
            session.close()
        except Exception as e:
            print(f"Dashboard hata: {e}")
            import traceback
            traceback.print_exc()

    def _is_borrow_cari(self, cari):
        """Cari tipinin ödünç para olup olmadığını kontrol et"""
        raw = (getattr(cari, "cari_type", "") or "").casefold().strip()
        compact = "".join(ch for ch in raw if not ch.isspace())
        if compact == "ödünçpara" or compact == "oduncpara":
            return True
        return "ödünçpara" in compact or "oduncpara" in compact

    def show_cari_extract_for(self, cari_id):
        """Cari ekstresi ekranını seçili cari ile aç"""
        try:
            target_index = None
            if hasattr(self, "tabs"):
                for i in range(self.tabs.count()):
                    if "Cari Ekstre" in self.tabs.tabText(i):
                        target_index = i
                        break
                if target_index is not None:
                    self.tabs.setCurrentIndex(target_index)

            if hasattr(self, "_reload_cari_extract_combo"):
                self._reload_cari_extract_combo()

            idx = -1
            if hasattr(self, "cari_extract_combo"):
                idx = self.cari_extract_combo.findData(cari_id)
                if idx < 0:
                    idx = self.cari_extract_combo.findData(str(cari_id))
                if idx < 0:
                    try:
                        from src.database.db import SessionLocal
                        from src.database.models import Cari

                        session = SessionLocal()
                        cari = session.query(Cari).filter(Cari.id == cari_id).first()
                        session.close()
                        if cari and cari.name:
                            for i in range(self.cari_extract_combo.count()):
                                if cari.name in self.cari_extract_combo.itemText(i):
                                    idx = i
                                    break
                    except Exception:
                        pass

                if idx >= 0:
                    self.cari_extract_combo.setCurrentIndex(idx)

            if hasattr(self, "show_cari_extract"):
                self.show_cari_extract()
        except Exception as e:
            QMessageBox.warning(self, "Uyarı", f"Cari ekstre açılamadı: {e}")
    
    def create_transactions_tab(self) -> QWidget:
        """İşlemler sekmesi - Excel tarzı"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Başlık
        title = QLabel("💰 Tüm İşlemler")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title.setStyleSheet("color: #333;")
        layout.addWidget(title)
        
        # Filtre ve buton paneli
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        # Sütun genişliklerini kaydet butonu
        btn_save_columns = QPushButton("💾 Kaydet")
        btn_save_columns.setMinimumHeight(30)
        btn_save_columns.setToolTip("Sütun genişliklerini kaydet")
        btn_save_columns.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 10px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_columns.clicked.connect(self.save_transaction_column_widths)
        filter_layout.addWidget(btn_save_columns)

        # Tarih filtreleri
        filter_layout.addWidget(QLabel("Başlangıç:"))
        self.start_date_filter = QDateEdit()
        self.start_date_filter.setCalendarPopup(True)
        self.start_date_filter.setDisplayFormat("dd.MM.yyyy")
        self.start_date_filter.setDate(QDate(QDate.currentDate().year(), QDate.currentDate().month(), 1))
        self.start_date_filter.setMinimumHeight(30)
        filter_layout.addWidget(self.start_date_filter)
        
        filter_layout.addWidget(QLabel("Bitiş:"))
        self.end_date_filter = QDateEdit()
        self.end_date_filter.setCalendarPopup(True)
        self.end_date_filter.setDisplayFormat("dd.MM.yyyy")
        self.end_date_filter.setDate(QDate.currentDate())
        self.end_date_filter.setMinimumHeight(30)
        filter_layout.addWidget(self.end_date_filter)

        self.start_date_filter.dateChanged.connect(self.apply_transaction_filter)
        self.end_date_filter.dateChanged.connect(self.apply_transaction_filter)
        
        # Filtre uygula butonu
        btn_filter = QPushButton("🔍 Filtrele")
        btn_filter.setMinimumHeight(30)
        btn_filter.clicked.connect(self.apply_transaction_filter)
        filter_layout.addWidget(btn_filter)

        # Tüm dönem butonu
        btn_all_period = QPushButton("📅 Tüm Dönem")
        btn_all_period.setMinimumHeight(30)
        btn_all_period.setToolTip("Tüm tarihlerdeki işlemleri göster")
        btn_all_period.clicked.connect(self.show_all_transactions)
        filter_layout.addWidget(btn_all_period)
        
        # Arama alanı - Tüm sütunlarda ara
        filter_layout.addWidget(QLabel("🔍 Ara:"))
        self.search_customer_input = QLineEdit()
        self.search_customer_input.setPlaceholderText("Tarih, müşteri, açıklama, tutar, konu vb. yazın...")
        self.search_customer_input.setMinimumHeight(30)
        self.search_customer_input.setMaximumWidth(350)
        self.search_customer_input.textChanged.connect(self.search_customer_transactions)
        filter_layout.addWidget(self.search_customer_input)
        
        filter_layout.addStretch()
        
        # Yeni işlem butonu
        btn_new_transaction = QPushButton("➕ Yeni İşlem")
        btn_new_transaction.setMinimumHeight(35)
        btn_new_transaction.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
        """)
        btn_new_transaction.clicked.connect(self.show_new_transaction_dialog)
        filter_layout.addWidget(btn_new_transaction)
        
        # Yenile
        btn_refresh_transactions = QPushButton("🔄 Yenile")
        btn_refresh_transactions.setMinimumHeight(35)
        btn_refresh_transactions.clicked.connect(self.refresh_transactions_table)
        filter_layout.addWidget(btn_refresh_transactions)
        
        btn_import_transactions = QPushButton("📥 Excel'den Aktar")
        btn_import_transactions.setMinimumHeight(35)
        btn_import_transactions.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #546E7A; }
        """)
        btn_import_transactions.clicked.connect(self.import_transactions_from_excel)
        filter_layout.addWidget(btn_import_transactions)

        btn_export_transactions = QPushButton("📤 Excel'e Aktar")
        btn_export_transactions.setMinimumHeight(35)
        btn_export_transactions.setStyleSheet("""
            QPushButton {
                background-color: #1D6F42;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #155232; }
        """)
        btn_export_transactions.clicked.connect(self.export_transactions_to_excel)
        filter_layout.addWidget(btn_export_transactions)

        # Toplu Sil butonu
        self.btn_bulk_delete_toggle = QPushButton("☑️ Toplu Sil")
        self.btn_bulk_delete_toggle.setMinimumHeight(35)
        self.btn_bulk_delete_toggle.setStyleSheet("""
            QPushButton {
                background-color: #FF5722;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #E64A19; }
        """)
        self.btn_bulk_delete_toggle.clicked.connect(self.toggle_bulk_delete_mode)
        filter_layout.addWidget(self.btn_bulk_delete_toggle)

        self.btn_bulk_delete_confirm = QPushButton("🗑️ Seçilenleri Sil")
        self.btn_bulk_delete_confirm.setMinimumHeight(35)
        self.btn_bulk_delete_confirm.setStyleSheet("""
            QPushButton {
                background-color: #b71c1c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #7f0000; }
        """)
        self.btn_bulk_delete_confirm.clicked.connect(self.delete_selected_transactions)
        self.btn_bulk_delete_confirm.setVisible(False)
        filter_layout.addWidget(self.btn_bulk_delete_confirm)

        self.btn_bulk_delete_cancel = QPushButton("❌ İptal")
        self.btn_bulk_delete_cancel.setMinimumHeight(35)
        self.btn_bulk_delete_cancel.setStyleSheet("""
            QPushButton {
                background-color: #9E9E9E;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #616161; }
        """)
        self.btn_bulk_delete_cancel.clicked.connect(self.toggle_bulk_delete_mode)
        self.btn_bulk_delete_cancel.setVisible(False)
        filter_layout.addWidget(self.btn_bulk_delete_cancel)

        layout.addLayout(filter_layout)
        
        # İşlemler tablosu - Excel tarzı
        self.table_transactions = QTableWidget()
        self.table_transactions.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_transactions.setColumnCount(9)
        self.table_transactions.setHorizontalHeaderLabels([
            "TARİH", "TÜR", "MÜŞTERİ ÜNVANI", "AÇIKLAMA", 
            "ÖDEME ŞEKLİ", "KONU", "ÖDEYEN KİŞİ", "TUTAR", "İŞLEMLER"
        ])
        
        # Tablo stilleri
        self.table_transactions.setAlternatingRowColors(True)
        self.table_transactions.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_transactions.setSelectionMode(QTableWidget.SingleSelection)
        self.table_transactions.horizontalHeader().setStretchLastSection(False)
        
        # Kolon genişlikleri - optimize edilmiş
        self.table_transactions.setColumnWidth(0, 90)   # Tarih
        self.table_transactions.setColumnWidth(1, 140)  # Tür
        self.table_transactions.setColumnWidth(2, 180)  # Müşteri
        self.table_transactions.setColumnWidth(3, 220)  # Açıklama
        self.table_transactions.setColumnWidth(4, 100)  # Ödeme Şekli
        self.table_transactions.setColumnWidth(5, 100)  # Konu
        self.table_transactions.setColumnWidth(6, 100)  # Ödeyen
        self.table_transactions.setColumnWidth(7, 110)  # Tutar
        self.table_transactions.setColumnWidth(8, 200)  # İşlemler

        # Kayıtlı sütun genişliklerini uygula
        self.load_transaction_column_widths()

        self.table_transactions.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                gridline-color: #e0e0e0;
                background-color: white;
            }
            QHeaderView::section {
                background-color: #2c5f8d;
                color: white;
                padding: 8px;
                border: 1px solid #1e4564;
                font-weight: bold;
            self._resize_table(self.table_recent)
            self._resize_table(self.table_pending)
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #cce5ff;
                color: black;
            }
        """)
        
        layout.addWidget(self.table_transactions)
        
        widget.setLayout(layout)
        return widget

    # ── Genel sütun genişliği kaydet / yükle ──────────────────────────────
    def save_column_widths(self, table: QTableWidget, key: str, silent: bool = False):
        """Herhangi bir tablonun sütun genişliklerini JSON dosyasına kaydet."""
        try:
            save_path = Path("data") / "column_widths.json"
            save_path.parent.mkdir(parents=True, exist_ok=True)
            data = {}
            if save_path.exists():
                with open(save_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            widths = {str(col): table.columnWidth(col) for col in range(table.columnCount())}
            data[key] = widths
            with open(save_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            if not silent:
                QMessageBox.information(self, "Kaydet", "Sütun genişlikleri başarıyla kaydedildi.")
        except Exception as e:
            if not silent:
                QMessageBox.warning(self, "Hata", f"Kayıt sırasında hata oluştu:\n{e}")

    def load_column_widths(self, table: QTableWidget, key: str):
        """JSON dosyasından kaydedilmiş sütun genişliklerini tabloya uygula."""
        try:
            save_path = Path("data") / "column_widths.json"
            if not save_path.exists():
                return
            with open(save_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            widths = data.get(key, {})
            for col_str, width in widths.items():
                col = int(col_str)
                if 0 <= col < table.columnCount():
                    table.setColumnWidth(col, width)
        except Exception:
            pass

    def _make_save_col_btn(self, table: QTableWidget, key: str) -> QPushButton:
        """Sütun genişliği kaydet butonu oluştur (tekrar kullanılabilir)."""
        btn = QPushButton("💾 Kaydet")
        btn.setMinimumHeight(30)
        btn.setToolTip("Sütun genişliklerini kaydet")
        btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 10px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn.clicked.connect(lambda: self.save_column_widths(table, key))
        return btn

    # ── transactions (eski uyumluluk sarmalayıcılar) ────────────────────
    def save_transaction_column_widths(self):
        self.save_column_widths(self.table_transactions, "transactions")

    def load_transaction_column_widths(self):
        self.load_column_widths(self.table_transactions, "transactions")

    def refresh_transactions_table(self):
        """İşlemler tablosunu yenile - aktif tarih filtreleriyle"""
        self.apply_transaction_filter()

    def _refresh_transactions_table_full(self):
        """Tüm işlemleri yükler (dahili kullanım)"""
        try:
            self.table_transactions.setUpdatesEnabled(False)
            transactions = TransactionService.get_all_transactions(self.user.id)
            transactions = sorted(transactions, key=lambda x: x.transaction_date, reverse=True)
            self.table_transactions.setRowCount(len(transactions))

            for i, trans in enumerate(transactions):
                self.table_transactions.setRowHidden(i, False)

                date_item = QTableWidgetItem(str(trans.transaction_date))
                date_item.setData(Qt.UserRole, trans.id)
                self.table_transactions.setItem(i, 0, date_item)

                type_text = trans.transaction_type.value if trans.transaction_type else ""
                type_item = QTableWidgetItem(type_text)
                if trans.transaction_type in [TransactionType.GELIR, TransactionType.KESILEN_FATURA]:
                    type_item.setBackground(Qt.green)
                elif trans.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]:
                    type_item.setBackground(Qt.red)
                self.table_transactions.setItem(i, 1, type_item)

                self.table_transactions.setItem(i, 2, QTableWidgetItem(trans.customer_name))
                self.table_transactions.setItem(i, 3, QTableWidgetItem(trans.description))

                payment_text = self._get_payment_method_display_text(trans.payment_method.value, trans.payment_type) if trans.payment_method else ""
                self.table_transactions.setItem(i, 4, QTableWidgetItem(payment_text))
                self.table_transactions.setItem(i, 5, QTableWidgetItem(trans.subject or ""))
                self.table_transactions.setItem(i, 6, QTableWidgetItem(trans.person or ""))

                amount_item = QTableWidgetItem(f"{format_tr(trans.amount)} ₺")
                amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table_transactions.setItem(i, 7, amount_item)

                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(4, 3, 4, 3)
                action_layout.setSpacing(6)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                action_widget.setMinimumHeight(36)

                btn_edit = QPushButton("Düzenle")
                btn_edit.setMinimumWidth(72)
                btn_edit.setMinimumHeight(28)
                btn_edit.setMaximumWidth(90)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 3px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #1976D2; }
                """)
                btn_edit.setProperty('transaction_id', trans.id)
                btn_edit.clicked.connect(lambda checked, tid=trans.id: self.edit_transaction(tid))
                action_layout.addWidget(btn_edit)

                btn_delete = QPushButton("Sil")
                btn_delete.setMinimumWidth(48)
                btn_delete.setMinimumHeight(24)
                btn_delete.setMaximumWidth(60)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 3px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.setProperty('transaction_id', trans.id)
                btn_delete.clicked.connect(lambda checked, tid=trans.id: self.delete_transaction(tid))
                action_layout.addWidget(btn_delete)

                self.table_transactions.setCellWidget(i, 8, action_widget)

            self._resize_table(self.table_transactions, stretch_col=3, min_row_height=48)
            self.table_transactions.setColumnWidth(8, 160)
            self.load_transaction_column_widths()
        except Exception as e:
            print(f"İşlemler yükleme hatası: {e}")
            QMessageBox.critical(self, "Hata", f"İşlemler yüklenirken hata: {str(e)}")
        finally:
            self.table_transactions.setUpdatesEnabled(True)
    
    def apply_transaction_filter(self):
        """Tarih filtresini uygula"""
        try:
            self.table_transactions.setUpdatesEnabled(False)
            start_date = self.start_date_filter.date().toPyDate()
            end_date = self.end_date_filter.date().toPyDate()
            
            transactions = TransactionService.get_all_transactions(
                self.user.id, start_date, end_date
            )
            # Yeni işlemler üstte olsun (descending sort)
            transactions = sorted(transactions, key=lambda x: x.transaction_date, reverse=True)
            
            # Tabloyu güncelle (aynı kod yukarıdaki gibi)
            self.table_transactions.setRowCount(len(transactions))
            for i, trans in enumerate(transactions):
                _d_item = QTableWidgetItem(str(trans.transaction_date))
                _d_item.setData(Qt.UserRole, trans.id)
                self.table_transactions.setItem(i, 0, _d_item)
                type_item = QTableWidgetItem(trans.transaction_type.value)
                if trans.transaction_type in [TransactionType.GELIR, TransactionType.KESILEN_FATURA]:
                    type_item.setBackground(Qt.green)
                elif trans.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]:
                    type_item.setBackground(Qt.red)
                self.table_transactions.setItem(i, 1, type_item)
                self.table_transactions.setItem(i, 2, QTableWidgetItem(trans.customer_name))
                self.table_transactions.setItem(i, 3, QTableWidgetItem(trans.description))
                payment_text = self._get_payment_method_display_text(trans.payment_method.value, trans.payment_type)
                self.table_transactions.setItem(i, 4, QTableWidgetItem(payment_text))
                self.table_transactions.setItem(i, 5, QTableWidgetItem(trans.subject or ""))
                self.table_transactions.setItem(i, 6, QTableWidgetItem(trans.person or ""))
                amount_item = QTableWidgetItem(f"{format_tr(trans.amount)} ₺")
                amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table_transactions.setItem(i, 7, amount_item)
                
                # İşlemler butonları
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(4, 3, 4, 3)
                action_layout.setSpacing(6)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                action_widget.setMinimumHeight(36)
                
                btn_edit = QPushButton("Düzenle")
                btn_edit.setMinimumWidth(72)
                btn_edit.setMinimumHeight(28)
                btn_edit.setMaximumWidth(90)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 3px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #1976D2; }
                """)
                btn_edit.clicked.connect(lambda checked, tid=trans.id: self.edit_transaction(tid))
                action_layout.addWidget(btn_edit)
                
                btn_delete = QPushButton("Sil")
                btn_delete.setMinimumWidth(48)
                btn_delete.setMinimumHeight(24)
                btn_delete.setMaximumWidth(60)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 3px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.clicked.connect(lambda checked, tid=trans.id: self.delete_transaction(tid))
                action_layout.addWidget(btn_delete)
                
                self.table_transactions.setCellWidget(i, 8, action_widget)
            
            self._resize_table(self.table_transactions, stretch_col=3, min_row_height=48)
            self.table_transactions.setColumnWidth(8, 160)
            self.load_transaction_column_widths()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Filtreleme hatası: {str(e)}")
        finally:
            self.table_transactions.setUpdatesEnabled(True)

    def show_all_transactions(self):
        """Tarih filtrelerini kaldırarak tüm işlemleri göster"""
        self.start_date_filter.blockSignals(True)
        self.end_date_filter.blockSignals(True)
        self.start_date_filter.setDate(QDate(2000, 1, 1))
        self.end_date_filter.setDate(QDate.currentDate())
        self.start_date_filter.blockSignals(False)
        self.end_date_filter.blockSignals(False)
        self.apply_transaction_filter()
    
    def search_customer_transactions(self):
        """Tüm sütunlarda ara"""
        search_text = self.search_customer_input.text().strip().lower()
        
        if not search_text:
            # Eğer arama boşsa tümünü göster
            self.refresh_transactions_table()
            return
        
        try:
            start_date = self.start_date_filter.date().toPyDate()
            end_date = self.end_date_filter.date().toPyDate()
            transactions = TransactionService.get_all_transactions(self.user.id, start_date, end_date)
            # Yeni işlemler üstte olsun (descending sort)
            transactions = sorted(transactions, key=lambda x: x.transaction_date, reverse=True)
            
            # Tüm sütunlarda ara: tarih, tip, müşteri, açıklama, ödeme yöntemi, konu, kişi, tutar
            filtered = []
            for t in transactions:
                search_in = [
                    str(t.transaction_date),  # Tarih
                    t.transaction_type.value if t.transaction_type else "",  # İşlem türü
                    t.customer_name or "",  # Müşteri adı
                    t.description or "",  # Açıklama
                    t.payment_method.value if t.payment_method else "",  # Ödeme yöntemi
                    t.subject or "",  # Konu
                    t.person or "",  # Kişi
                    str(format_tr(t.amount)),  # Tutar (formatlı)
                    str(t.amount)  # Tutar (normal)
                ]
                # Eğer arama metni herhangi bir sütunda varsa ekle
                if any(search_text in field.lower() for field in search_in):
                    filtered.append(t)
            
            # Tabloyu güncelle
            self.table_transactions.setRowCount(len(filtered))
            for i, trans in enumerate(filtered):
                _s_item = QTableWidgetItem(str(trans.transaction_date))
                _s_item.setData(Qt.UserRole, trans.id)
                self.table_transactions.setItem(i, 0, _s_item)
                type_item = QTableWidgetItem(trans.transaction_type.value)
                if trans.transaction_type in [TransactionType.GELIR, TransactionType.KESILEN_FATURA]:
                    type_item.setBackground(Qt.green)
                elif trans.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]:
                    type_item.setBackground(Qt.red)
                self.table_transactions.setItem(i, 1, type_item)
                self.table_transactions.setItem(i, 2, QTableWidgetItem(trans.customer_name))
                self.table_transactions.setItem(i, 3, QTableWidgetItem(trans.description))
                payment_text = self._get_payment_method_display_text(trans.payment_method.value, trans.payment_type)
                self.table_transactions.setItem(i, 4, QTableWidgetItem(payment_text))
                self.table_transactions.setItem(i, 5, QTableWidgetItem(trans.subject or ""))
                self.table_transactions.setItem(i, 6, QTableWidgetItem(trans.person or ""))
                amount_item = QTableWidgetItem(f"{format_tr(trans.amount)} ₺")
                amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table_transactions.setItem(i, 7, amount_item)
                
                # İşlemler butonları
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(4, 3, 4, 3)
                action_layout.setSpacing(6)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                action_widget.setMinimumHeight(36)
                
                btn_edit = QPushButton("Düzenle")
                btn_edit.setMinimumWidth(130)
                btn_edit.setMinimumHeight(36)
                btn_edit.setMaximumWidth(140)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 5px;
                        padding: 7px 14px;
                        font-size: 12pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #1976D2; }
                """)
                btn_edit.setProperty('transaction_id', trans.id)
                btn_edit.clicked.connect(lambda checked, tid=trans.id: self.edit_transaction(tid))
                action_layout.addWidget(btn_edit)
                
                btn_delete = QPushButton("Sil")
                btn_delete.setMinimumWidth(100)
                btn_delete.setMinimumHeight(36)
                btn_delete.setMaximumWidth(110)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 5px;
                        padding: 7px 14px;
                        font-size: 12pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.setProperty('transaction_id', trans.id)
                btn_delete.clicked.connect(lambda checked, tid=trans.id: self.delete_transaction(tid))
                action_layout.addWidget(btn_delete)
                
                self.table_transactions.setCellWidget(i, 8, action_widget)
            
            self._resize_table(self.table_transactions, stretch_col=3, min_row_height=48)
            self.table_transactions.setColumnWidth(8, 200)
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Arama hatası: {str(e)}")
    
    def show_new_transaction_dialog(self):
        """Yeni işlem ekleme dialog'unu göster"""
        from src.ui.dialogs.transaction_dialog import TransactionDialog
        dialog = TransactionDialog(self.user.id, self)
        if dialog.exec_():
            self.refresh_all_data()
    
    def edit_transaction(self, transaction_id):
        """İşlem düzenle"""
        from src.ui.dialogs.transaction_dialog import TransactionDialog
        dialog = TransactionDialog(self.user.id, self, transaction_id=transaction_id)
        if dialog.exec_():
            self.refresh_all_data()
    
    def delete_transaction(self, transaction_id):
        """İşlem sil"""
        reply = QMessageBox.question(
            self, "Onay", 
            "Bu işlemi silmek istediğinize emin misiniz?\n\nİlgili hesaplar otomatik olarak güncellenecek.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            success, msg = TransactionService.delete_transaction(transaction_id)
            if success:
                QMessageBox.information(self, "Başarılı", "İşlem silindi ve hesaplar güncellendi")
                self.refresh_all_data()
            else:
                QMessageBox.critical(self, "Hata", f"İşlem silinemedi: {msg}")

    def toggle_bulk_delete_mode(self):
        """Toplu silme modunu aç/kapat"""
        self._bulk_delete_mode = not getattr(self, '_bulk_delete_mode', False)
        if self._bulk_delete_mode:
            self.table_transactions.setSelectionMode(QTableWidget.MultiSelection)
            self.btn_bulk_delete_toggle.setVisible(False)
            self.btn_bulk_delete_confirm.setVisible(True)
            self.btn_bulk_delete_cancel.setVisible(True)
            self.table_transactions.clearSelection()
        else:
            self.table_transactions.setSelectionMode(QTableWidget.SingleSelection)
            self.btn_bulk_delete_toggle.setVisible(True)
            self.btn_bulk_delete_confirm.setVisible(False)
            self.btn_bulk_delete_cancel.setVisible(False)
            self.table_transactions.clearSelection()

    def delete_selected_transactions(self):
        """Seçili işlemleri toplu sil"""
        selected_rows = set(idx.row() for idx in self.table_transactions.selectedIndexes())
        if not selected_rows:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek istediğiniz işlemleri seçin.")
            return
        reply = QMessageBox.question(
            self, "Toplu Silme Onayı",
            f"{len(selected_rows)} adet işlemi silmek istediğinize emin misiniz?\n\nBu işlem geri alınamaz!",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        errors = []
        success_count = 0
        for row in selected_rows:
            item = self.table_transactions.item(row, 0)
            if item:
                tid = item.data(Qt.UserRole)
                if tid:
                    ok, msg = TransactionService.delete_transaction(tid)
                    if ok:
                        success_count += 1
                    else:
                        errors.append(msg)
        self.toggle_bulk_delete_mode()
        self.refresh_all_data()
        if errors:
            QMessageBox.warning(self, "Sonuç", f"{success_count} işlem silindi.\n{len(errors)} işlem silinemedi.")
        else:
            QMessageBox.information(self, "Başarılı", f"{success_count} işlem başarıyla silindi.")

    def create_invoices_tab(self) -> QWidget:
        """Faturalar sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("📄 Fatura Yönetimi")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        layout.addWidget(title)
        
        btn_layout = QHBoxLayout()
        btn_new = QPushButton("➕ Yeni Fatura")
        btn_new.clicked.connect(self.show_new_invoice_dialog)
        btn_layout.addWidget(btn_new)
        
        btn_refresh = QPushButton("🔄 Yenile")
        btn_refresh.clicked.connect(self.refresh_invoice_table)
        btn_layout.addWidget(btn_refresh)

        btn_save_inv = QPushButton("💾 Kaydet")
        btn_save_inv.setMinimumHeight(30)
        btn_save_inv.setToolTip("Sütun genişliklerini kaydet")
        btn_save_inv.setStyleSheet("""
            QPushButton {
                background-color: #FF9800; color: white;
                border: none; border-radius: 4px;
                padding: 4px 10px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_inv.clicked.connect(lambda: self.save_column_widths(self.table_invoices, "invoices"))
        btn_layout.addWidget(btn_save_inv)

        self.invoice_search_input = QLineEdit()
        self.invoice_search_input.setPlaceholderText("Fatura no veya cari adı ara...")
        self.invoice_search_input.textChanged.connect(self.filter_invoice_table)
        btn_layout.addWidget(self.invoice_search_input)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.table_invoices = QTableWidget()
        self.table_invoices.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_invoices.setColumnCount(7)
        self.table_invoices.setHorizontalHeaderLabels(
            ["Fatura No", "Cari", "Tutar", "Durum", "Tarih", "Vade Tarihi", "Vade Durumu"]
        )
        self.table_invoices.horizontalHeader().setStretchLastSection(False)
        self.table_invoices.setColumnWidth(0, 120)
        self.table_invoices.setColumnWidth(1, 220)
        self.table_invoices.setColumnWidth(2, 120)
        self.table_invoices.setColumnWidth(3, 80)
        self.table_invoices.setColumnWidth(4, 100)
        self.table_invoices.setColumnWidth(5, 100)
        self.table_invoices.setColumnWidth(6, 250)
        self.table_invoices.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_invoices.customContextMenuRequested.connect(self._invoice_context_menu)
        self.load_column_widths(self.table_invoices, "invoices")
        layout.addWidget(self.table_invoices)
        
        widget.setLayout(layout)
        return widget
    
    def refresh_invoice_table(self):
        """Fatura tablosunu yenile — vade takibi + kısmi ödeme desteği."""
        try:
            from src.database.db import SessionLocal
            from src.database.models import Transaction, TransactionType
            from src.services.transaction_service import TransactionService
            from datetime import date as date_type
            from PyQt5.QtCore import Qt
            from PyQt5.QtGui import QColor, QBrush

            # Önce otomatik ödeme tespitini çalıştır
            try:
                TransactionService.auto_detect_paid_invoices(self.user.id)
            except Exception:
                pass

            session = SessionLocal()
            invoices = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_type.in_([
                    TransactionType.KESILEN_FATURA,
                    TransactionType.GELEN_FATURA
                ])
            ).order_by(Transaction.transaction_date.desc()).all()

            # Session açıkken tüm gerekli alanları çek
            inv_data = []
            for inv in invoices:
                inv_data.append({
                    'id': inv.id,
                    'invoice_number': getattr(inv, 'invoice_number', None),
                    'cari_name': inv.cari.name if inv.cari else '-',
                    'amount': inv.amount,
                    'transaction_type': inv.transaction_type,
                    'transaction_date': inv.transaction_date,
                    'due_date': getattr(inv, 'due_date', None),
                    'is_paid': getattr(inv, 'is_paid', False) or False,
                    'paid_date': getattr(inv, 'paid_date', None),
                    'paid_amount': getattr(inv, 'paid_amount', 0.0) or 0.0,
                })
            session.close()

            self.table_invoices.setRowCount(len(inv_data))
            today = date_type.today()

            for i, inv in enumerate(inv_data):
                inv_id = inv['id']
                total_amount = inv['amount']
                paid_amount = inv['paid_amount']
                invoice_no = inv['invoice_number'] or f"F-{inv_id}"

                # Fatura No (UserRole'de meta sakla)
                item0 = QTableWidgetItem(invoice_no)
                item0.setData(Qt.UserRole,      inv_id)
                item0.setData(Qt.UserRole + 1,  paid_amount)
                item0.setData(Qt.UserRole + 2,  total_amount)
                self.table_invoices.setItem(i, 0, item0)

                # Cari adı
                self.table_invoices.setItem(i, 1, QTableWidgetItem(inv['cari_name']))

                # Tutar
                self.table_invoices.setItem(i, 2, QTableWidgetItem(f"{format_tr(total_amount)} TL"))

                # Durum (Fatura türü)
                is_kesilen = inv['transaction_type'] == TransactionType.KESILEN_FATURA
                status = "Kesilen" if is_kesilen else "Gelen"
                self.table_invoices.setItem(i, 3, QTableWidgetItem(status))

                # Tarih
                self.table_invoices.setItem(i, 4, QTableWidgetItem(str(inv['transaction_date'])))

                # Vade Tarihi ve Vade Durumu
                due_date   = inv['due_date']
                is_fully_paid = total_amount > 0 and paid_amount >= total_amount
                is_partial    = 0 < paid_amount < total_amount
                paid_date  = inv['paid_date']

                self.table_invoices.setItem(i, 5, QTableWidgetItem(
                    str(due_date) if (is_kesilen and due_date) else "-"
                ))

                row_color = None
                if paid_amount > total_amount and total_amount > 0:
                    # Fazla ödeme durumu (nadiren olur, overflow başka faturaya aktarılamadıysa)
                    fazla = paid_amount - total_amount
                    paid_str = f" ({paid_date})" if paid_date else ""
                    vade_item = QTableWidgetItem(f"✅ Ödendi{paid_str}  💰 Fazla: {format_tr(fazla)} TL")
                    vade_item.setForeground(QBrush(QColor("#1565C0")))
                    self.table_invoices.setItem(i, 6, vade_item)
                    row_color = "#E3F2FD"
                elif is_fully_paid:
                    paid_str = f" ({paid_date})" if paid_date else ""
                    vade_item = QTableWidgetItem(f"\u2705 \u00d6dendi{paid_str}")
                    vade_item.setForeground(QBrush(QColor("#1B5E20")))
                    self.table_invoices.setItem(i, 6, vade_item)
                    row_color = "#E8F5E9"
                elif is_partial and is_kesilen:
                    kalan = total_amount - paid_amount
                    vade_item = QTableWidgetItem(
                        f"\U0001f7e0 K\u0131smi: {format_tr(paid_amount)} / {format_tr(total_amount)} TL"
                        f"  (Kalan: {format_tr(kalan)} TL)"
                    )
                    vade_item.setForeground(QBrush(QColor("#E65100")))
                    self.table_invoices.setItem(i, 6, vade_item)
                    row_color = "#FFF8E1"
                elif is_kesilen and due_date:
                    days_left = (due_date - today).days
                    if days_left < 0:
                        vade_text  = f"\u26d4 {abs(days_left)} g\u00fcn gecikmi\u015f"
                        row_color  = "#FFEBEE"
                        text_color = "#B71C1C"
                    elif days_left == 0:
                        vade_text  = "\U0001f534 Bug\u00fcn \u00f6deme g\u00fcn\u00fc!"
                        row_color  = "#FFF3E0"
                        text_color = "#E65100"
                    elif days_left <= 5:
                        vade_text  = f"\U0001f7e1 {days_left} g\u00fcn sonra \u00f6deme"
                        row_color  = "#FFFDE7"
                        text_color = "#F57F17"
                    else:
                        vade_text  = f"\U0001f7e2 {days_left} g\u00fcn kald\u0131"
                        row_color  = None
                        text_color = "#2E7D32"
                    vade_item = QTableWidgetItem(vade_text)
                    vade_item.setForeground(QBrush(QColor(text_color)))
                    self.table_invoices.setItem(i, 6, vade_item)
                elif is_kesilen:
                    self.table_invoices.setItem(i, 6, QTableWidgetItem("Vade yok"))
                else:
                    self.table_invoices.setItem(i, 6, QTableWidgetItem("-"))

                # Satır arka plan rengi
                if row_color:
                    for col in range(self.table_invoices.columnCount()):
                        existing = self.table_invoices.item(i, col)
                        if existing:
                            existing.setBackground(QBrush(QColor(row_color)))

            self._resize_table(self.table_invoices, stretch_col=1)
            self.load_column_widths(self.table_invoices, "invoices")
            self.filter_invoice_table()
        except Exception as e:
            print(f"Fatura yükleme hatası: {e}")
            import traceback
            traceback.print_exc()

    def _invoice_context_menu(self, pos):
        """Fatura tablosunda sağ tık menüsü — ödeme durumu değiştirme (tam/kısmi)."""
        from PyQt5.QtWidgets import QMenu, QAction, QMessageBox
        from PyQt5.QtCore import Qt
        from src.services.transaction_service import TransactionService

        row = self.table_invoices.rowAt(pos.y())
        if row < 0:
            return

        item0 = self.table_invoices.item(row, 0)
        if not item0:
            return
        transaction_id = item0.data(Qt.UserRole)
        if not transaction_id:
            return

        paid_amount  = item0.data(Qt.UserRole + 1) or 0.0
        total_amount = item0.data(Qt.UserRole + 2) or 0.0
        is_paid      = total_amount > 0 and paid_amount >= total_amount
        is_partial   = 0 < paid_amount < total_amount

        menu = QMenu(self)
        if not is_paid:
            menu.addAction("✅  Tam Ödendi İşaretle")
        if not is_paid:
            menu.addAction("🟠  Ödeme Gir (tarih sırasıyla)")
        if is_partial:
            menu.addAction("🟠  Ödemeyi Güncelle (tarih sırasıyla)")
        if is_paid or is_partial:
            menu.addAction("↩️  Ödemeyi Sıfırla")

        if menu.isEmpty():
            return

        action = menu.exec_(self.table_invoices.viewport().mapToGlobal(pos))
        if action is None:
            return

        def _refresh():
            self.refresh_invoice_table()
            self.refresh_dashboard()

        label = action.text()

        if "✅" in label:  # Tam ödendi
            ok, msg = TransactionService.mark_invoice_as_paid(transaction_id, paid=True)
            if ok:
                _refresh()
            else:
                QMessageBox.warning(self, "Hata", msg)

        elif "↩️" in label:  # Sıfırla
            ok, msg = TransactionService.set_partial_payment(transaction_id, 0.0)
            if ok:
                _refresh()
            else:
                QMessageBox.warning(self, "Hata", msg)

        else:  # Ödeme gir / güncelle — tarih sırasıyla tüm faturalara dağıt
            from PyQt5.QtWidgets import QInputDialog
            cari_item = self.table_invoices.item(row, 1)
            cari_name = cari_item.text() if cari_item else ""
            amount_str, ok = QInputDialog.getText(
                self, "Ödeme Girişi",
                f"Cari: {cari_name}\n"
                f"Bu cari için toplam tahsilat tutarını girin (TL).\n"
                f"Ödeme, fatura tarih sırasına göre EN ESKİ faturadan başlayarak uygulanır.",
                text=""
            )
            if not ok or not amount_str.strip():
                return
            try:
                entered = float(amount_str.strip().replace(',', '.'))
            except ValueError:
                QMessageBox.warning(self, "Hata", "Geçerli bir sayı girin.")
                return
            ok2, msg2, applied_list = TransactionService.apply_cari_payment(transaction_id, entered)
            if ok2:
                _refresh()
                QMessageBox.information(self, "Ödeme Uygulandı", msg2)
            else:
                QMessageBox.warning(self, "Hata", msg2)

    def filter_invoice_table(self):
        if not hasattr(self, 'table_invoices'):
            return
        query = ""
        if hasattr(self, 'invoice_search_input'):
            query = self.invoice_search_input.text().casefold().strip()

        for row in range(self.table_invoices.rowCount()):
            invoice_item = self.table_invoices.item(row, 0)
            cari_item = self.table_invoices.item(row, 1)
            invoice_no = invoice_item.text().casefold() if invoice_item else ""
            cari_name = cari_item.text().casefold() if cari_item else ""
            match = not query or query in invoice_no or query in cari_name
            self.table_invoices.setRowHidden(row, not match)
    
    def show_new_invoice_dialog(self):
        """Yeni fatura dialog'u - İşlemler sekmesine yönlendir"""
        # İşlemler sekmesine geç (index 1)
        self.tabs.setCurrentIndex(1)
        
        # Bilgi mesajı göster
        QMessageBox.information(
            self, 
            "Fatura Ekleme", 
            "Fatura eklemek için aşağıdaki 'Yeni İşlem' butonunu kullanarak:\n\n"
            "• Kesilen Fatura (Müşteriye kestiniz)\n"
            "• Gelen Fatura (Tedarikçiden geldi)\n\n"
            "türlerinden birini seçerek fatura ekleyebilirsiniz."
        )
    
    def create_caris_tab(self) -> QWidget:
        """Cari Hesaplar sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("📋 Cari Hesap Yönetimi")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        layout.addWidget(title)
        
        btn_layout = QHBoxLayout()
        
        # Arama kutusu
        btn_layout.addWidget(QLabel("🔍 Ara:"))
        self.cari_list_search_input = QLineEdit()
        self.cari_list_search_input.setPlaceholderText("Cari adı veya telefon ara...")
        self.cari_list_search_input.setMinimumHeight(30)
        self.cari_list_search_input.setMaximumWidth(250)
        self.cari_list_search_input.textChanged.connect(self.filter_cari_table)
        btn_layout.addWidget(self.cari_list_search_input)
        
        btn_layout.addSpacing(20)
        
        btn_new = QPushButton("➕ Yeni Cari")
        btn_new.clicked.connect(self.show_new_cari_dialog)
        btn_layout.addWidget(btn_new)
        
        btn_refresh = QPushButton("🔄 Yenile")
        btn_refresh.clicked.connect(self.refresh_cari_table)
        btn_layout.addWidget(btn_refresh)
        
        btn_import_caris = QPushButton("📥 Excel'den Aktar")
        btn_import_caris.clicked.connect(self.import_caris_from_excel)
        btn_layout.addWidget(btn_import_caris)

        btn_save_cari = QPushButton("💾 Kaydet")
        btn_save_cari.setMinimumHeight(30)
        btn_save_cari.setToolTip("Sütun genişliklerini kaydet")
        btn_save_cari.setStyleSheet("""
            QPushButton {
                background-color: #FF9800; color: white;
                border: none; border-radius: 4px;
                padding: 4px 10px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_cari.clicked.connect(lambda: self.save_column_widths(self.table_caris, "caris"))
        btn_layout.addWidget(btn_save_cari)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.table_caris = QTableWidget()
        self.table_caris.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_caris.setColumnCount(5)
        self.table_caris.setHorizontalHeaderLabels(["Ad", "Tip", "Bakiye", "Telefon", "İşlemler"])
        self.table_caris.horizontalHeader().setStretchLastSection(False)
        self.table_caris.setColumnWidth(0, 300)
        self.table_caris.setColumnWidth(1, 150)
        self.table_caris.setColumnWidth(2, 130)
        self.table_caris.setColumnWidth(3, 150)
        self.table_caris.setColumnWidth(4, 140)
        self.load_column_widths(self.table_caris, "caris")
        layout.addWidget(self.table_caris)

        borrow_title = QLabel("💳 Ödünç Para Carileri")
        borrow_title.setFont(QFont("Segoe UI", 13, QFont.Bold))
        layout.addWidget(borrow_title)

        borrow_btn_layout = QHBoxLayout()
        btn_new_borrow = QPushButton("➕ Ödünç Cari")
        btn_new_borrow.clicked.connect(self.show_new_borrow_cari_dialog)
        borrow_btn_layout.addWidget(btn_new_borrow)

        btn_refresh_borrow = QPushButton("🔄 Yenile")
        btn_refresh_borrow.clicked.connect(self.refresh_borrow_cari_table)
        borrow_btn_layout.addWidget(btn_refresh_borrow)

        btn_save_borrow = QPushButton("💾 Kaydet")
        btn_save_borrow.setMinimumHeight(30)
        btn_save_borrow.setToolTip("Sütun genişliklerini kaydet")
        btn_save_borrow.setStyleSheet("""
            QPushButton {
                background-color: #FF9800; color: white;
                border: none; border-radius: 4px;
                padding: 4px 10px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_borrow.clicked.connect(lambda: self.save_column_widths(self.table_caris_borrow, "caris_borrow"))
        borrow_btn_layout.addWidget(btn_save_borrow)
        borrow_btn_layout.addStretch()
        layout.addLayout(borrow_btn_layout)

        self.table_caris_borrow = QTableWidget()
        self.table_caris_borrow.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_caris_borrow.setColumnCount(5)
        self.table_caris_borrow.setHorizontalHeaderLabels(["Ad", "Tip", "Bakiye", "Telefon", "İşlemler"])
        self.table_caris_borrow.horizontalHeader().setStretchLastSection(False)
        self.table_caris_borrow.setColumnWidth(0, 300)
        self.table_caris_borrow.setColumnWidth(1, 150)
        self.table_caris_borrow.setColumnWidth(2, 130)
        self.table_caris_borrow.setColumnWidth(3, 150)
        self.table_caris_borrow.setColumnWidth(4, 140)
        self.load_column_widths(self.table_caris_borrow, "caris_borrow")
        layout.addWidget(self.table_caris_borrow)
        
        widget.setLayout(layout)
        return widget

    def show_new_borrow_cari_dialog(self):
        """Yeni ödünç cari dialog'u"""
        from src.ui.dialogs.cari_dialog import CariDialog
        dialog = CariDialog(self.user.id, self, default_type="ÖDÜNÇ PARA")
        if dialog.exec_():
            self.refresh_cari_table()
    
    def refresh_cari_table(self):
        """Cari tablosunu yenile"""
        try:
            self.table_caris.setUpdatesEnabled(False)
            caris = CariService.get_caris(self.user.id)
            self.table_caris.setRowCount(len(caris) if caris else 0)
            
            if caris:
                for i, cari in enumerate(caris):
                    self.table_caris.setItem(i, 0, QTableWidgetItem(cari.name))
                    self.table_caris.setItem(i, 1, QTableWidgetItem(cari.cari_type))
                    self.table_caris.setItem(i, 2, QTableWidgetItem(format_tr(cari.balance)))
                    self.table_caris.setItem(i, 3, QTableWidgetItem(cari.phone or ""))

                    action_widget = QWidget()
                    action_layout = QHBoxLayout(action_widget)
                    action_layout.setContentsMargins(5, 2, 5, 2)
                    action_layout.setSpacing(5)
                    action_layout.setAlignment(Qt.AlignCenter)
                    action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                    btn_edit = QPushButton("✏️ Düzenle")
                    btn_edit.setMinimumHeight(24)
                    btn_edit.setStyleSheet("""
                        QPushButton {
                            background-color: #2196F3;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #0b7dda; }
                    """)
                    btn_edit.clicked.connect(lambda checked, cid=cari.id: self.show_edit_cari_dialog(cid))
                    action_layout.addWidget(btn_edit)

                    btn_delete = QPushButton("🗑️ Sil")
                    btn_delete.setMinimumHeight(24)
                    btn_delete.setStyleSheet("""
                        QPushButton {
                            background-color: #f44336;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #da190b; }
                    """)
                    btn_delete.clicked.connect(lambda checked, cid=cari.id: self.delete_cari(cid))
                    action_layout.addWidget(btn_delete)

                    self.table_caris.setCellWidget(i, 4, action_widget)

            self._resize_table(self.table_caris, stretch_col=0)
            self.load_column_widths(self.table_caris, "caris")
            
            if hasattr(self, 'cari_extract_combo'):
                self.cari_extract_combo.clear()
                self.cari_extract_combo.addItem("-- Cari Seçiniz --", None)
                if caris:
                    for cari in caris:
                        self.cari_extract_combo.addItem(f"{cari.name} ({cari.cari_type})", cari.id)

            if hasattr(self, 'cari_list_search_input'):
                self.filter_cari_table(self.cari_list_search_input.text())

            if hasattr(self, 'cari_extract_combo'):
                self._reload_cari_extract_combo()

            self.refresh_borrow_cari_table(caris)
        except Exception as e:
            print(f"Cari yükleme hatası: {e}")
            import traceback
            traceback.print_exc()
        finally:
            self.table_caris.setUpdatesEnabled(True)

    def refresh_borrow_cari_table(self, caris=None):
        """Ödünç cari tablosunu yenile"""
        try:
            if caris is None:
                caris = CariService.get_caris(self.user.id)
            borrow_caris = [c for c in (caris or []) if c.cari_type == "ÖDÜNÇ PARA"]
            if not hasattr(self, "table_caris_borrow"):
                return

            self.table_caris_borrow.setRowCount(len(borrow_caris))
            for i, cari in enumerate(borrow_caris):
                self.table_caris_borrow.setItem(i, 0, QTableWidgetItem(cari.name))
                self.table_caris_borrow.setItem(i, 1, QTableWidgetItem(cari.cari_type))
                self.table_caris_borrow.setItem(i, 2, QTableWidgetItem(format_tr(cari.balance)))
                self.table_caris_borrow.setItem(i, 3, QTableWidgetItem(cari.phone or ""))

                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)
                action_layout.setSpacing(5)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                btn_edit = QPushButton("✏️ Düzenle")
                btn_edit.setMinimumHeight(24)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #0b7dda; }
                """)
                btn_edit.clicked.connect(lambda checked, cid=cari.id: self.show_edit_cari_dialog(cid))
                action_layout.addWidget(btn_edit)

                btn_delete = QPushButton("🗑️ Sil")
                btn_delete.setMinimumHeight(24)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.clicked.connect(lambda checked, cid=cari.id: self.delete_cari(cid))
                action_layout.addWidget(btn_delete)

                self.table_caris_borrow.setCellWidget(i, 4, action_widget)

            self._resize_table(self.table_caris_borrow, stretch_col=0)
            self.load_column_widths(self.table_caris_borrow, "caris_borrow")
        except Exception as e:
            print(f"Ödünç cari yükleme hatası: {e}")
    
    def filter_cari_table(self, search_text=None):
        """Cari tablosunu arama metnine göre filtrele"""
        if search_text is None:
            search_text = self.cari_list_search_input.text() if hasattr(self, 'cari_list_search_input') else ""

        search_text = str(search_text).casefold().strip()
        
        for row in range(self.table_caris.rowCount()):
            # Ad ve telefon kolonlarını kontrol et
            name_item = self.table_caris.item(row, 0)
            phone_item = self.table_caris.item(row, 3)
            
            name_text = name_item.text().casefold() if name_item else ""
            phone_text = phone_item.text().casefold() if phone_item else ""

            name_match = search_text in name_text
            phone_match = search_text in phone_text
            
            # Eğer arama metni boşsa veya eşleşme varsa göster
            if not search_text or name_match or phone_match:
                self.table_caris.setRowHidden(row, False)
            else:
                self.table_caris.setRowHidden(row, True)
    
    def show_new_cari_dialog(self):
        """Yeni cari dialog'u"""
        from src.ui.dialogs.cari_dialog import CariDialog
        dialog = CariDialog(self.user.id, self)
        if dialog.exec_():
            self.refresh_all_data()

    def show_edit_cari_dialog(self, cari_id):
        """Cari düzenleme dialog'u"""
        from src.ui.dialogs.cari_dialog import CariDialog
        dialog = CariDialog(self.user.id, self, cari_id=cari_id)
        if dialog.exec_():
            self.refresh_all_data()

    def delete_cari(self, cari_id):
        """Cari sil (pasif et)"""
        reply = QMessageBox.question(
            self,
            "Onay",
            "Bu cariyi silmek istediğinize emin misiniz?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            success, msg = CariService.delete_cari(cari_id)
            if success:
                QMessageBox.information(self, "Başarılı", "Cari silindi")
                self.refresh_all_data()
            else:
                QMessageBox.critical(self, "Hata", msg)
    
    def create_bank_tab(self) -> QWidget:
        """Banka Hesapları sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        title = QLabel("🏦 Banka Hesapları")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        layout.addWidget(title)
        
        # İstatistikler
        stats_layout = QHBoxLayout()
        try:
            stats = BankService.get_bank_statistics(self.user.id)
            
            stats_layout.addWidget(self.create_stat_card("Toplam Hesap", str(stats['total_accounts']), "#1976D2"))
            stats_layout.addWidget(self.create_stat_card("Toplam Bakiye", f"{stats['total_balance']:,.0f} ₺", "#4CAF50"))
            stats_layout.addWidget(self.create_stat_card("Ek Hesap Limiti", f"{stats['total_overdraft']:,.0f} ₺", "#FF9800"))
            stats_layout.addWidget(self.create_stat_card("Kullanılabilir", f"{stats['total_available']:,.0f} ₺", "#9C27B0"))
        except Exception as e:
            print(f"Banka stats hatası: {e}")
        
        layout.addLayout(stats_layout)
        layout.addSpacing(15)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_new = QPushButton("➕ Yeni Hesap")
        btn_new.setMinimumHeight(35)
        btn_new.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1565C0; }
        """)
        btn_new.clicked.connect(self.show_new_bank_dialog)
        btn_layout.addWidget(btn_new)
        
        btn_refresh = QPushButton("🔄 Yenile")
        btn_refresh.setMinimumHeight(35)
        btn_refresh.clicked.connect(self.refresh_bank_table)
        btn_layout.addWidget(btn_refresh)
        
        btn_import_banks = QPushButton("📥 Excel'den Aktar")
        btn_import_banks.setMinimumHeight(35)
        btn_import_banks.clicked.connect(self.import_banks_from_excel)
        btn_layout.addWidget(btn_import_banks)

        btn_save_banks = QPushButton("💾 Kaydet")
        btn_save_banks.setMinimumHeight(35)
        btn_save_banks.setToolTip("Sütun genişliklerini kaydet")
        btn_save_banks.setStyleSheet("""
            QPushButton {
                background-color: #FF9800; color: white;
                border: none; border-radius: 4px;
                padding: 8px 16px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_banks.clicked.connect(lambda: self.save_column_widths(self.table_banks, "banks"))
        btn_layout.addWidget(btn_save_banks)

        # Arama kutucuğu
        btn_layout.addWidget(QLabel("🔍 Ara:"))
        self.bank_search_input = QLineEdit()
        self.bank_search_input.setPlaceholderText("Banka adı, hesap no, para birimi vb. yazın...")
        self.bank_search_input.setMinimumHeight(35)
        self.bank_search_input.setMaximumWidth(350)
        self.bank_search_input.textChanged.connect(self.search_banks_table)
        btn_layout.addWidget(self.bank_search_input)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.table_banks = QTableWidget()
        self.table_banks.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_banks.setColumnCount(6)
        self.table_banks.setHorizontalHeaderLabels(["Banka", "Hesap No", "Bakiye", "Ek Hesap Limiti", "Para Birimi", "İşlemler"])
        self.table_banks.horizontalHeader().setStretchLastSection(False)
        self.table_banks.setColumnWidth(0, 180)
        self.table_banks.setColumnWidth(1, 220)
        self.table_banks.setColumnWidth(2, 120)
        self.table_banks.setColumnWidth(3, 130)
        self.table_banks.setColumnWidth(4, 100)
        self.table_banks.setColumnWidth(5, 220)
        self.load_column_widths(self.table_banks, "banks")
        layout.addWidget(self.table_banks)
        
        widget.setLayout(layout)
        return widget
    
    def refresh_bank_table(self):
        """Banka tablosunu yenile"""
        try:
            accounts = BankService.get_accounts(self.user.id)
            self.table_banks.setRowCount(len(accounts) if accounts else 0)
            
            if accounts:
                for i, acc in enumerate(accounts):
                    # Tüm satırları göster (arama gizlemişse açmak için)
                    self.table_banks.setRowHidden(i, False)
                    
                    self.table_banks.setItem(i, 0, QTableWidgetItem(acc.bank_name))
                    self.table_banks.setItem(i, 1, QTableWidgetItem(acc.account_number))
                    self.table_banks.setItem(i, 2, QTableWidgetItem(f"{format_tr(acc.balance)} ₺"))
                    
                    # Ek hesap limiti
                    overdraft = getattr(acc, 'overdraft_limit', 0.0)
                    self.table_banks.setItem(i, 3, QTableWidgetItem(f"{format_tr(overdraft)} ₺"))
                    
                    self.table_banks.setItem(i, 4, QTableWidgetItem(acc.currency))

                    action_widget = QWidget()
                    action_layout = QHBoxLayout(action_widget)
                    action_layout.setContentsMargins(5, 2, 5, 2)
                    action_layout.setSpacing(5)
                    action_layout.setAlignment(Qt.AlignCenter)
                    action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                    btn_edit = QPushButton("✏️ Düzenle")
                    btn_edit.setMinimumHeight(24)
                    btn_edit.setStyleSheet("""
                        QPushButton {
                            background-color: #2196F3;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #0b7dda; }
                    """)
                    btn_edit.clicked.connect(lambda checked, bid=acc.id: self.show_edit_bank_dialog(bid))
                    action_layout.addWidget(btn_edit)

                    btn_delete = QPushButton("🗑️ Sil")
                    btn_delete.setMinimumHeight(24)
                    btn_delete.setStyleSheet("""
                        QPushButton {
                            background-color: #f44336;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #da190b; }
                    """)
                    btn_delete.clicked.connect(lambda checked, bid=acc.id: self.delete_bank(bid))
                    action_layout.addWidget(btn_delete)

                    btn_statement = QPushButton("📑 Dökümü Aç")
                    btn_statement.setMinimumHeight(24)
                    btn_statement.setStyleSheet("""
                        QPushButton {
                            background-color: #607D8B;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #546E7A; }
                    """)
                    btn_statement.clicked.connect(lambda checked, bid=acc.id: self.show_bank_statement(bid))
                    action_layout.addWidget(btn_statement)

                    self.table_banks.setCellWidget(i, 5, action_widget)
            
            self._resize_table(self.table_banks, stretch_col=1)
            self.load_column_widths(self.table_banks, "banks")
        except Exception as e:
            print(f"Banka yükleme hatası: {e}")

    def search_banks_table(self):
        """Banka tablosunda ara"""
        search_text = self.bank_search_input.text().strip().lower() if hasattr(self, 'bank_search_input') else ""
        
        if not search_text:
            self.refresh_bank_table()
            return
        
        try:
            accounts = BankService.get_accounts(self.user.id)
            self.table_banks.setRowCount(len(accounts) if accounts else 0)
            
            matched_count = 0
            if accounts:
                for i, acc in enumerate(accounts):
                    # Tüm sütunlarda ara: banka, hesap no, bakiye, ek hesap limiti, para birimi
                    search_in = [
                        acc.bank_name or "",
                        acc.account_number or "",
                        str(format_tr(acc.balance)),
                        str(acc.balance),
                        str(format_tr(getattr(acc, 'overdraft_limit', 0.0))),
                        str(getattr(acc, 'overdraft_limit', 0.0)),
                        acc.currency or ""
                    ]
                    
                    if not any(search_text in field.lower() for field in search_in):
                        self.table_banks.setRowHidden(i, True)
                        continue
                    
                    matched_count += 1
                    self.table_banks.setRowHidden(i, False)
                    
                    self.table_banks.setItem(i, 0, QTableWidgetItem(acc.bank_name))
                    self.table_banks.setItem(i, 1, QTableWidgetItem(acc.account_number))
                    self.table_banks.setItem(i, 2, QTableWidgetItem(f"{format_tr(acc.balance)} ₺"))
                    
                    overdraft = getattr(acc, 'overdraft_limit', 0.0)
                    self.table_banks.setItem(i, 3, QTableWidgetItem(f"{format_tr(overdraft)} ₺"))
                    
                    self.table_banks.setItem(i, 4, QTableWidgetItem(acc.currency))

                    action_widget = QWidget()
                    action_layout = QHBoxLayout(action_widget)
                    action_layout.setContentsMargins(5, 2, 5, 2)
                    action_layout.setSpacing(5)
                    action_layout.setAlignment(Qt.AlignCenter)
                    action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                    btn_edit = QPushButton("✏️ Düzenle")
                    btn_edit.setMinimumHeight(24)
                    btn_edit.setStyleSheet("""
                        QPushButton {
                            background-color: #2196F3;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #0b7dda; }
                    """)
                    btn_edit.clicked.connect(lambda checked, bid=acc.id: self.show_edit_bank_dialog(bid))
                    action_layout.addWidget(btn_edit)

                    btn_delete = QPushButton("🗑️ Sil")
                    btn_delete.setMinimumHeight(24)
                    btn_delete.setStyleSheet("""
                        QPushButton {
                            background-color: #f44336;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #da190b; }
                    """)
                    btn_delete.clicked.connect(lambda checked, bid=acc.id: self.delete_bank(bid))
                    action_layout.addWidget(btn_delete)

                    btn_statement = QPushButton("📑 Dökümü Aç")
                    btn_statement.setMinimumHeight(24)
                    btn_statement.setStyleSheet("""
                        QPushButton {
                            background-color: #607D8B;
                            color: white;
                            border: none;
                            border-radius: 3px;
                            padding: 4px 8px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QPushButton:hover { background-color: #546E7A; }
                    """)
                    btn_statement.clicked.connect(lambda checked, bid=acc.id: self.show_bank_statement(bid))
                    action_layout.addWidget(btn_statement)
                    
                    self.table_banks.setCellWidget(i, 5, action_widget)
            
            self._resize_table(self.table_banks, stretch_col=1)
            self.load_column_widths(self.table_banks, "banks")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Arama hatası: {str(e)}")
    
    def show_new_bank_dialog(self):
        """Yeni banka hesabı dialog'u"""
        from src.ui.dialogs.bank_dialog import BankDialog
        dialog = BankDialog(self.user.id, self)
        if dialog.exec_():
            self.refresh_all_data()

    def show_edit_bank_dialog(self, account_id):
        """Banka hesabı düzenleme dialog'u"""
        from src.ui.dialogs.bank_dialog import BankDialog
        dialog = BankDialog(self.user.id, self, account_id=account_id)
        if dialog.exec_():
            self.refresh_all_data()

    def delete_bank(self, account_id):
        """Banka hesabı sil (pasif et)"""
        reply = QMessageBox.question(
            self,
            "Onay",
            "Bu banka hesabını silmek istediğinize emin misiniz?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            success, msg = BankService.delete_account(account_id)
            if success:
                QMessageBox.information(self, "Başarılı", "Banka hesabı silindi")
                self.refresh_all_data()
            else:
                QMessageBox.critical(self, "Hata", msg)
    
    def create_admin_panel_tab(self) -> QWidget:
        """Admin Panel sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel("👨‍💼 Admin Panel")
        title.setFont(QFont("Segoe UI", 18, QFont.Bold))
        title.setStyleSheet("color: #333;")
        layout.addWidget(title)
        
        layout.addWidget(QLabel("━━━━━━━━━━━━━━━━━"))
        layout.addSpacing(10)
        
        # Kullanıcı Yönetimi
        user_mgmt_title = QLabel("Kullanıcı Yönetimi")
        user_mgmt_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        layout.addWidget(user_mgmt_title)
        
        btn_user_mgmt = QPushButton("👥 Kullanıcıları Yönet")
        btn_user_mgmt.setMinimumHeight(40)
        btn_user_mgmt.setStyleSheet("""
            QPushButton {
                background-color: #667eea;
            }
            QPushButton:hover { background-color: #5568d3; }
        """)
        btn_user_mgmt.clicked.connect(self.show_user_management)
        layout.addWidget(btn_user_mgmt)
        
        layout.addSpacing(20)
        layout.addWidget(QLabel("━━━━━━━━━━━━━━━━━"))
        layout.addSpacing(10)
        
        # Sistem İstatistikleri
        stats_title = QLabel("Sistem İstatistikleri")
        stats_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        layout.addWidget(stats_title)
        
        try:
            from src.services.admin_service import AdminService
            users = AdminService.get_all_users()
            total_users = len(users) if users else 0
            admin_users = len([u for u in users if u.role == 'admin']) if users else 0
            active_users = len([u for u in users if u.is_active]) if users else 0
            
            stats_info = QLabel(f"""
Toplam Kullanıcı: {total_users}
Admin Sayısı: {admin_users}
Aktif Kullanıcı: {active_users}
Pasif Kullanıcı: {total_users - active_users}
            """.strip())
            stats_info.setFont(QFont("Segoe UI", 11))
            stats_info.setStyleSheet("color: #555; background-color: #f5f5f5; padding: 10px; border-radius: 4px;")
            layout.addWidget(stats_info)
        except Exception as e:
            print(f"Admin stats hata: {e}")
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_kira_takip_tab(self) -> QWidget:
        """Kira Takip sekmesi"""
        return KiraTakipWidget(user_id=self.user.id)

    def show_user_management(self):
        """Kullanıcı yönetimi dialog'unu aç"""
        dialog = UserManagementDialog(self)
        dialog.exec_()
    
    def create_settings_tab(self) -> QWidget:
        """Ayarlar sekmesi"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel("⚙️ Ayarlar")
        title.setFont(QFont("Segoe UI", 18, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("━━━━━━━━━━━━━━━━━"))
        
        # Kullanıcı Bilgileri ve Dashboard Kutucukları yan yana
        user_dashboard_layout = QHBoxLayout()
        user_dashboard_layout.setSpacing(30)
        
        # Sol: Kullanıcı Bilgileri
        user_info_layout = QVBoxLayout()
        info_title = QLabel("👤 Kullanıcı Bilgileri")
        info_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        user_info_layout.addWidget(info_title)
        
        user_info_layout.addSpacing(5)
        user_info_layout.addWidget(QLabel(f"Ad: {self.user.full_name}"))
        user_info_layout.addWidget(QLabel(f"Email: {self.user.email}"))
        user_info_layout.addWidget(QLabel(f"Kullanıcı Adı: {self.user.username}"))
        user_info_layout.addStretch()
        
        user_dashboard_layout.addLayout(user_info_layout)
        
        # Sağ: Dashboard Kutucukları
        dashboard_layout = QVBoxLayout()
        dashboard_title = QLabel("📊 Dashboard Kutucukları")
        dashboard_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        dashboard_layout.addWidget(dashboard_title)

        dashboard_layout.addSpacing(5)
        
        self.dashboard_card_checks = {}
        selected_keys = set(self._get_dashboard_card_keys())
        for key, title_text, _ in self._get_dashboard_card_defs():
            chk = QCheckBox(title_text)
            chk.setChecked(key in selected_keys)
            self.dashboard_card_checks[key] = chk
            dashboard_layout.addWidget(chk)

        dashboard_layout.addSpacing(10)
        
        btn_save_dashboard = QPushButton("💾 Dashboard Ayarlarını Kaydet")
        btn_save_dashboard.setMinimumHeight(36)
        btn_save_dashboard.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #546E7A; }
        """)
        btn_save_dashboard.clicked.connect(self.save_dashboard_preferences)
        dashboard_layout.addWidget(btn_save_dashboard)
        
        user_dashboard_layout.addLayout(dashboard_layout)
        
        layout.addLayout(user_dashboard_layout)
        
        layout.addSpacing(15)
        layout.addWidget(QLabel("━━━━━━━━━━━━━━━━━"))
        layout.addSpacing(15)
        
        # Veri Yedekleme
        backup_title = QLabel("💾 Veri Yedekleme")
        backup_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        layout.addWidget(backup_title)
        
        btn_db_backup = QPushButton("🗄️ Veritabanını Yedekle")
        btn_db_backup.setMinimumHeight(40)
        btn_db_backup.clicked.connect(self.backup_database)
        layout.addWidget(btn_db_backup)
        
        btn_excel_backup = QPushButton("📥 Tüm Verileri Excel'e Yedekle")
        btn_excel_backup.setMinimumHeight(40)
        btn_excel_backup.clicked.connect(self.export_all_data_to_excel)
        layout.addWidget(btn_excel_backup)

        self.auto_backup_enabled_check = QCheckBox("🔁 Otomatik Yedekleme Aç")
        self.auto_backup_enabled_check.setChecked(
            UserSettingsService.get_json_setting(self.user.id, "auto_backup_enabled", False)
        )
        layout.addWidget(self.auto_backup_enabled_check)

        self.auto_backup_period_combo = QComboBox()
        self.auto_backup_period_combo.addItem("Gün", "day")
        self.auto_backup_period_combo.addItem("Hafta", "week")
        self.auto_backup_period_combo.addItem("Ay", "month")
        self.auto_backup_period_combo.addItem("Yıl", "year")

        saved_period = UserSettingsService.get_setting(self.user.id, "auto_backup_period", "day")
        period_index = max(0, self.auto_backup_period_combo.findData(saved_period))
        self.auto_backup_period_combo.setCurrentIndex(period_index)

        auto_backup_period_layout = QHBoxLayout()
        auto_backup_period_layout.addWidget(QLabel("Periyot:"))
        auto_backup_period_layout.addWidget(self.auto_backup_period_combo)
        auto_backup_period_layout.addStretch()
        layout.addLayout(auto_backup_period_layout)

        btn_save_auto_backup = QPushButton("💾 Otomatik Yedeklemeyi Kaydet")
        btn_save_auto_backup.setMinimumHeight(36)
        btn_save_auto_backup.clicked.connect(self.save_auto_backup_preferences)
        layout.addWidget(btn_save_auto_backup)
        
        layout.addSpacing(20)
        layout.addWidget(QLabel("━━━━━━━━━━━━━━━━━"))
        layout.addSpacing(20)
        
        # Google Sheets Senkronizasyonu
        gsheets_title = QLabel("📊 Google Sheets Senkronizasyonu")
        gsheets_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        layout.addWidget(gsheets_title)
        
        layout.addSpacing(5)
        
        gsheets_info = QLabel("Google Sheets'ten tek yönlü veri çekme özelliği")
        gsheets_info.setStyleSheet("color: #666; font-size: 10pt;")
        layout.addWidget(gsheets_info)
        
        layout.addSpacing(10)
        
        # Google Sheets URL input
        url_label = QLabel("Sheets URL:")
        url_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(url_label)
        
        self.gsheets_url_input = QLineEdit()
        self.gsheets_url_input.setPlaceholderText("https://docs.google.com/spreadsheets/d/...")
        self.gsheets_url_input.setMinimumHeight(32)
        saved_url = UserSettingsService.get_setting(self.user.id, "google_sheets_url", "")
        self.gsheets_url_input.setText(saved_url or "")
        layout.addWidget(self.gsheets_url_input)
        

        layout.addSpacing(10)

        # Sayfa isimleri
        sheets_label = QLabel("Sayfa İsimleri:")
        sheets_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(sheets_label)

        # Cari Sayfası
        cari_sheet_layout = QHBoxLayout()
        cari_sheet_layout.addWidget(QLabel("Cari Sayfası:"))
        self.gsheets_caris_sheet_input = QLineEdit()
        self.gsheets_caris_sheet_input.setPlaceholderText("Cariler")
        self.gsheets_caris_sheet_input.setMinimumHeight(32)
        saved_caris = UserSettingsService.get_setting(self.user.id, "google_sheets_caris_sheet", "Cariler")
        self.gsheets_caris_sheet_input.setText(saved_caris or "Cariler")
        cari_sheet_layout.addWidget(self.gsheets_caris_sheet_input)
        layout.addLayout(cari_sheet_layout)

        # İşlem Sayfası
        trans_sheet_layout = QHBoxLayout()
        trans_sheet_layout.addWidget(QLabel("İşlem Sayfası:"))
        self.gsheets_trans_sheet_input = QLineEdit()
        self.gsheets_trans_sheet_input.setPlaceholderText("İşlemler")
        self.gsheets_trans_sheet_input.setMinimumHeight(32)
        saved_trans = UserSettingsService.get_setting(self.user.id, "google_sheets_trans_sheet", "İşlemler")
        self.gsheets_trans_sheet_input.setText(saved_trans or "İşlemler")
        trans_sheet_layout.addWidget(self.gsheets_trans_sheet_input)
        layout.addLayout(trans_sheet_layout)

        # Gelen Fatura
        gelen_fatura_layout = QHBoxLayout()
        gelen_fatura_layout.addWidget(QLabel("Gelen Fatura Sayfası:"))
        self.gsheets_gelen_fatura_sheet_input = QLineEdit()
        self.gsheets_gelen_fatura_sheet_input.setPlaceholderText("Gelen Fatura")
        self.gsheets_gelen_fatura_sheet_input.setMinimumHeight(32)
        saved_gelen_fatura = UserSettingsService.get_setting(self.user.id, "google_sheets_gelen_fatura_sheet", "Gelen Fatura")
        self.gsheets_gelen_fatura_sheet_input.setText(saved_gelen_fatura or "Gelen Fatura")
        gelen_fatura_layout.addWidget(self.gsheets_gelen_fatura_sheet_input)
        layout.addLayout(gelen_fatura_layout)

        # Kesilen Fatura
        kesilen_fatura_layout = QHBoxLayout()
        kesilen_fatura_layout.addWidget(QLabel("Kesilen Fatura Sayfası:"))
        self.gsheets_kesilen_fatura_sheet_input = QLineEdit()
        self.gsheets_kesilen_fatura_sheet_input.setPlaceholderText("Kesilen Fatura")
        self.gsheets_kesilen_fatura_sheet_input.setMinimumHeight(32)
        saved_kesilen_fatura = UserSettingsService.get_setting(self.user.id, "google_sheets_kesilen_fatura_sheet", "Kesilen Fatura")
        self.gsheets_kesilen_fatura_sheet_input.setText(saved_kesilen_fatura or "Kesilen Fatura")
        kesilen_fatura_layout.addWidget(self.gsheets_kesilen_fatura_sheet_input)
        layout.addLayout(kesilen_fatura_layout)

        # Gider
        gider_layout = QHBoxLayout()
        gider_layout.addWidget(QLabel("Gider Sayfası:"))
        self.gsheets_gider_sheet_input = QLineEdit()
        self.gsheets_gider_sheet_input.setPlaceholderText("Gider")
        self.gsheets_gider_sheet_input.setMinimumHeight(32)
        saved_gider = UserSettingsService.get_setting(self.user.id, "google_sheets_gider_sheet", "Gider")
        self.gsheets_gider_sheet_input.setText(saved_gider or "Gider")
        gider_layout.addWidget(self.gsheets_gider_sheet_input)
        layout.addLayout(gider_layout)

        # Gelir
        gelir_layout = QHBoxLayout()
        gelir_layout.addWidget(QLabel("Gelir Sayfası:"))
        self.gsheets_gelir_sheet_input = QLineEdit()
        self.gsheets_gelir_sheet_input.setPlaceholderText("Gelir")
        self.gsheets_gelir_sheet_input.setMinimumHeight(32)
        saved_gelir = UserSettingsService.get_setting(self.user.id, "google_sheets_gelir_sheet", "Gelir")
        self.gsheets_gelir_sheet_input.setText(saved_gelir or "Gelir")
        gelir_layout.addWidget(self.gsheets_gelir_sheet_input)
        layout.addLayout(gelir_layout)

        layout.addSpacing(10)
        
        # Otomatik senkronizasyon
        self.gsheets_auto_sync_check = QCheckBox("🔄 Otomatik Aktarım (Her 10 dakikada)")
        self.gsheets_auto_sync_check.setChecked(
            UserSettingsService.get_json_setting(self.user.id, "google_sheets_auto_sync", False)
        )
        layout.addWidget(self.gsheets_auto_sync_check)
        
        layout.addSpacing(10)
        
        # Butonlar
        gsheets_buttons_layout = QHBoxLayout()
        gsheets_buttons_layout.setSpacing(10)
        
        btn_gsheets_test = QPushButton("🔗 Bağlantıyı Test Et")
        btn_gsheets_test.setMinimumHeight(38)
        btn_gsheets_test.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1976D2; }
        """)
        btn_gsheets_test.clicked.connect(self.test_google_sheets_connection)
        self.gsheets_test_button = btn_gsheets_test
        gsheets_buttons_layout.addWidget(btn_gsheets_test)
        
        btn_gsheets_sync = QPushButton("🔄 Şimdi Google Sheets'ten Aktar")
        btn_gsheets_sync.setMinimumHeight(38)
        btn_gsheets_sync.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
        """)
        btn_gsheets_sync.clicked.connect(self.sync_from_google_sheets)
        self.gsheets_sync_button = btn_gsheets_sync
        gsheets_buttons_layout.addWidget(btn_gsheets_sync)
        
        layout.addLayout(gsheets_buttons_layout)
        
        layout.addSpacing(10)
        
        btn_gsheets_save = QPushButton("💾 Google Sheets Ayarlarını Kaydet")
        btn_gsheets_save.setMinimumHeight(40)
        btn_gsheets_save.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 16px;
                font-weight: bold;
                font-size: 11pt;
            }
            QPushButton:hover { background-color: #546E7A; }
        """)
        btn_gsheets_save.clicked.connect(self.save_google_sheets_settings)
        layout.addWidget(btn_gsheets_save)
        
        layout.addSpacing(10)
        
        # Son senkronizasyon zamanı
        last_sync = UserSettingsService.get_setting(self.user.id, "google_sheets_last_sync", None)
        if last_sync:
            try:
                last_sync_dt = datetime.fromisoformat(last_sync)
                last_sync_text = f"Son senkronizasyon: {last_sync_dt.strftime('%d.%m.%Y %H:%M')}"
            except:
                last_sync_text = "Son senkronizasyon: -"
        else:
            last_sync_text = "Son senkronizasyon: Henüz yapılmadı"
        
        self.gsheets_last_sync_label = QLabel(last_sync_text)
        self.gsheets_last_sync_label.setStyleSheet("color: #888; font-size: 9pt; font-style: italic; padding: 5px;")
        layout.addWidget(self.gsheets_last_sync_label)
        
        layout.addSpacing(20)
        layout.addWidget(QLabel("━━━━━━━━━━━━━━━━━"))
        layout.addSpacing(20)
        
        # Butonlar
        btn_change_pwd = QPushButton("🔒 Şifre Değiştir")
        btn_change_pwd.setMinimumHeight(40)
        btn_change_pwd.clicked.connect(self.show_change_password_dialog)
        layout.addWidget(btn_change_pwd)
        
        layout.addSpacing(20)
        
        btn_logout = QPushButton("🚪 Çıkış Yap")
        btn_logout.setMinimumHeight(40)
        btn_logout.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
            }
            QPushButton:hover { background-color: #da190b; }
        """)
        btn_logout.clicked.connect(self.logout)
        layout.addWidget(btn_logout)
        
        layout.addStretch()
        widget.setLayout(layout)
        scroll.setWidget(widget)
        return scroll

    def save_dashboard_preferences(self):
        """Dashboard kutucuk tercihlerini kaydet"""
        if not hasattr(self, "dashboard_card_checks"):
            return

        selected = [
            key for key, chk in self.dashboard_card_checks.items()
            if chk.isChecked()
        ]

        if not selected:
            QMessageBox.warning(self, "Uyari", "En az bir kutucuk secmelisiniz!")
            return

        UserSettingsService.set_json_setting(self.user.id, "dashboard_cards", selected)
        self.rebuild_dashboard_tab()
        QMessageBox.information(self, "Basarili", "Dashboard ayarlari guncellendi")

    def rebuild_dashboard_tab(self):
        """Dashboard sekmesini yeniden olustur"""
        if not self.user.can_view_dashboard:
            return

        index = self.tabs.indexOf(self.dashboard_tab) if self.dashboard_tab else -1
        if index == -1:
            index = 0
        else:
            self.tabs.removeTab(index)

        self.dashboard_tab = self.create_dashboard_tab()
        self.tabs.insertTab(index, self.dashboard_tab, "📊 Dashboard")
        self.tabs.setCurrentIndex(index)
    
    def show_change_password_dialog(self):
        """Şifre değiştir dialog'u"""
        QMessageBox.information(self, "Bilgi", "Şifre değiştirme özelliği yakında eklenecek")

    def save_auto_backup_preferences(self):
        """Otomatik yedekleme ayarlarini kaydet"""
        enabled = self.auto_backup_enabled_check.isChecked()
        period = self.auto_backup_period_combo.currentData() or "day"

        UserSettingsService.set_json_setting(self.user.id, "auto_backup_enabled", enabled)
        UserSettingsService.set_setting(self.user.id, "auto_backup_period", period)

        self.setup_auto_backup_scheduler()
        QMessageBox.information(self, "Basarili", "Otomatik yedekleme ayarlari kaydedildi")

    def _get_auto_backup_delta(self):
        period = UserSettingsService.get_setting(self.user.id, "auto_backup_period", "day")
        period_map = {
            "day": timedelta(days=1),
            "week": timedelta(weeks=1),
            "month": timedelta(days=30),
            "year": timedelta(days=365)
        }
        return period_map.get(period, timedelta(days=1))

    def _get_auto_backup_dir(self) -> Path:
        backup_dir = Path(config.DATABASE_DIR) / "auto_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        return backup_dir

    def _create_automatic_db_backup(self):
        db_path = Path(config.DATABASE_DIR) / "muhasebe.db"
        if not db_path.exists():
            raise FileNotFoundError("Veritabani dosyasi bulunamadi")

        backup_dir = self._get_auto_backup_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        target_path = backup_dir / f"muhasebe_auto_backup_{timestamp}.db"

        shutil.copy2(db_path, target_path)

        for suffix in ["-wal", "-shm"]:
            extra_src = Path(str(db_path) + suffix)
            if extra_src.exists():
                extra_dst = backup_dir / f"{target_path.name}{suffix}"
                shutil.copy2(extra_src, extra_dst)

        return target_path

    def _is_auto_backup_due(self, now: datetime) -> bool:
        last_run_raw = UserSettingsService.get_setting(self.user.id, "auto_backup_last_run", None)
        if not last_run_raw:
            return True

        try:
            last_run = datetime.fromisoformat(last_run_raw)
        except Exception:
            return True

        return (now - last_run) >= self._get_auto_backup_delta()

    def run_auto_backup_if_due(self):
        enabled = UserSettingsService.get_json_setting(self.user.id, "auto_backup_enabled", False)
        if not enabled:
            return

        now = datetime.now()
        if not self._is_auto_backup_due(now):
            return

        try:
            backup_path = self._create_automatic_db_backup()
            UserSettingsService.set_setting(self.user.id, "auto_backup_last_run", now.isoformat())
            self.statusBar().showMessage(f"Otomatik yedek alindi: {backup_path.name}", 5000)
        except Exception as e:
            self.statusBar().showMessage(f"Otomatik yedekleme hatasi: {e}", 7000)

    def setup_auto_backup_scheduler(self):
        if self.auto_backup_timer is None:
            self.auto_backup_timer = QTimer(self)
            self.auto_backup_timer.timeout.connect(self.run_auto_backup_if_due)

        enabled = UserSettingsService.get_json_setting(self.user.id, "auto_backup_enabled", False)
        if enabled:
            self.run_auto_backup_if_due()
            self.auto_backup_timer.start(60 * 60 * 1000)
        else:
            self.auto_backup_timer.stop()
    
    def backup_database(self):
        """Veritabanini dosya olarak yedekle"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            db_path = Path(config.DATABASE_DIR) / "muhasebe.db"
            if not db_path.exists():
                QMessageBox.warning(self, "Uyarı", "Veritabani dosyasi bulunamadi!")
                return
            
            default_name = f"muhasebe_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Veritabani Yedegi Kaydet",
                default_name,
                "SQLite DB (*.db)"
            )
            
            if not file_path:
                return
            
            shutil.copy2(db_path, file_path)
            copied_extra = []
            for suffix in ["-wal", "-shm"]:
                extra_src = Path(str(db_path) + suffix)
                if extra_src.exists():
                    extra_dst = Path(file_path).with_suffix(Path(file_path).suffix + suffix)
                    shutil.copy2(extra_src, extra_dst)
                    copied_extra.append(extra_dst.name)
            
            msg = f"Yedek olusturuldu:\n{file_path}"
            if copied_extra:
                msg += f"\nEk dosyalar: {', '.join(copied_extra)}"
            QMessageBox.information(self, "Basarili", msg)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Veritabani yedeklenirken hata:\n{str(e)}")
    
    def _safe_value(self, value):
        if hasattr(value, "value"):
            return value.value
        return value
    
    def _write_sheet(self, ws, headers, rows):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 16
    
    def export_all_data_to_excel(self):
        """Tum verileri Excel'e yedekle"""
        try:
            from datetime import datetime, date
            from PyQt5.QtWidgets import QFileDialog
            from openpyxl import Workbook
            from src.database.db import SessionLocal
            from src.database.models import (
                User, BankAccount, Cari, Invoice, InvoiceLineItem,
                BankTransaction, Report, CreditCard, Transaction
            )
            
            default_name = f"tum_veriler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Yedegi Kaydet",
                default_name,
                "Excel Dosyasi (*.xlsx)"
            )
            if not file_path:
                return
            
            wb = Workbook()
            session = SessionLocal()
            
            # Users
            ws = wb.active
            ws.title = "Users"
            users = session.query(User).all()
            rows = []
            for u in users:
                rows.append([
                    u.id, u.username, u.email, u.full_name, u.role,
                    u.can_view_dashboard, u.can_view_transactions, u.can_view_invoices,
                    u.can_view_caris, u.can_view_cari_extract, u.can_view_banks,
                    u.can_view_credit_cards, u.can_view_loans, u.can_view_reports,
                    u.can_view_payroll, u.can_view_employees, u.can_view_bulk_payroll,
                    u.can_view_payroll_records, u.can_view_settings, u.can_view_admin_panel,
                    u.is_active,
                    u.created_at, u.last_login
                ])
            self._write_sheet(ws, [
                "id", "username", "email", "full_name", "role",
                "can_view_dashboard", "can_view_transactions", "can_view_invoices",
                "can_view_caris", "can_view_cari_extract", "can_view_banks",
                "can_view_credit_cards", "can_view_loans", "can_view_reports",
                "can_view_payroll", "can_view_employees", "can_view_bulk_payroll",
                "can_view_payroll_records", "can_view_settings", "can_view_admin_panel",
                "is_active",
                "created_at", "last_login"
            ], rows)
            
            # Bank Accounts
            ws = wb.create_sheet("BankAccounts")
            accounts = session.query(BankAccount).all()
            rows = []
            for a in accounts:
                rows.append([
                    a.id, a.user_id, a.bank_name, a.account_number, a.iban,
                    a.branch, a.balance, a.overdraft_limit, a.currency,
                    a.is_active, a.created_at
                ])
            self._write_sheet(ws, [
                "id", "user_id", "bank_name", "account_number", "iban",
                "branch", "balance", "overdraft_limit", "currency",
                "is_active", "created_at"
            ], rows)
            
            # Caris
            ws = wb.create_sheet("Caris")
            caris = session.query(Cari).all()
            rows = []
            for c in caris:
                rows.append([
                    c.id, c.user_id, c.name, c.cari_type, c.tax_number, c.email,
                    c.phone, c.address, c.city, c.balance, c.is_active, c.created_at
                ])
            self._write_sheet(ws, [
                "id", "user_id", "name", "cari_type", "tax_number", "email",
                "phone", "address", "city", "balance", "is_active", "created_at"
            ], rows)
            
            # Invoices
            ws = wb.create_sheet("Invoices")
            invoices = session.query(Invoice).all()
            rows = []
            for inv in invoices:
                rows.append([
                    inv.id, inv.user_id, inv.cari_id, inv.bank_account_id,
                    inv.invoice_number, inv.invoice_date, inv.due_date,
                    self._safe_value(inv.invoice_type), inv.amount, inv.tax_rate,
                    inv.tax_amount, inv.total_amount, inv.description,
                    inv.paid_amount, inv.status, inv.created_at, inv.updated_at
                ])
            self._write_sheet(ws, [
                "id", "user_id", "cari_id", "bank_account_id",
                "invoice_number", "invoice_date", "due_date", "invoice_type",
                "amount", "tax_rate", "tax_amount", "total_amount",
                "description", "paid_amount", "status", "created_at", "updated_at"
            ], rows)
            
            # Invoice Line Items
            ws = wb.create_sheet("InvoiceItems")
            items = session.query(InvoiceLineItem).all()
            rows = []
            for it in items:
                rows.append([
                    it.id, it.invoice_id, it.description, it.quantity,
                    it.unit_price, it.total
                ])
            self._write_sheet(ws, [
                "id", "invoice_id", "description", "quantity", "unit_price", "total"
            ], rows)
            
            # Bank Transactions
            ws = wb.create_sheet("BankTransactions")
            btx = session.query(BankTransaction).all()
            rows = []
            for bt in btx:
                rows.append([
                    bt.id, bt.user_id, bt.bank_account_id, bt.invoice_id,
                    bt.transaction_date, bt.amount, bt.transaction_type,
                    bt.category, bt.description, bt.reference
                ])
            self._write_sheet(ws, [
                "id", "user_id", "bank_account_id", "invoice_id",
                "transaction_date", "amount", "transaction_type",
                "category", "description", "reference"
            ], rows)
            
            # Reports
            ws = wb.create_sheet("Reports")
            reports = session.query(Report).all()
            rows = []
            for r in reports:
                rows.append([
                    r.id, r.user_id, r.report_type, r.title, r.generated_at,
                    r.start_date, r.end_date, r.data
                ])
            self._write_sheet(ws, [
                "id", "user_id", "report_type", "title", "generated_at",
                "start_date", "end_date", "data"
            ], rows)
            
            # Credit Cards
            ws = wb.create_sheet("CreditCards")
            cards = session.query(CreditCard).all()
            rows = []
            for card in cards:
                rows.append([
                    card.id, card.user_id, card.card_name, card.card_number_last4,
                    card.card_holder, card.bank_name, card.card_limit, card.current_debt,
                    card.available_limit, card.closing_day, card.due_day,
                    card.is_active, card.created_at
                ])
            self._write_sheet(ws, [
                "id", "user_id", "card_name", "card_number_last4",
                "card_holder", "bank_name", "card_limit", "current_debt",
                "available_limit", "closing_day", "due_day", "is_active", "created_at"
            ], rows)
            
            # Transactions
            ws = wb.create_sheet("Transactions")
            transactions = session.query(Transaction).all()
            rows = []
            for t in transactions:
                rows.append([
                    t.id, t.user_id, t.transaction_date,
                    self._safe_value(t.transaction_type), self._safe_value(t.payment_method),
                    t.cari_id, t.bank_account_id, t.credit_card_id,
                    t.customer_name, t.description, t.subject, t.payment_type,
                    t.amount, t.person, t.notes, t.created_at, t.updated_at
                ])
            self._write_sheet(ws, [
                "id", "user_id", "transaction_date", "transaction_type", "payment_method",
                "cari_id", "bank_account_id", "credit_card_id",
                "customer_name", "description", "subject", "payment_type",
                "amount", "person", "notes", "created_at", "updated_at"
            ], rows)
            
            session.close()
            wb.save(file_path)
            QMessageBox.information(self, "Basarili", f"Excel yedegi olusturuldu:\n{file_path}")
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kutuphanesi bulunamadi. Lutfen 'pip install openpyxl' komutunu calistirin.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel yedegi olusturulurken hata:\n{str(e)}")

    def _normalize_header(self, value):
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = text.replace(" ", "_")
        text = text.replace("-", "_")
        text = text.replace("(", "").replace(")", "")
        return text

    def _read_excel_rows(self, file_path):
        from openpyxl import load_workbook
        headers = []
        rows = []
        if file_path.lower().endswith('.xls'):
            try:
                import xlrd
            except ImportError:
                QMessageBox.critical(self, "Hata", "'xlrd' kütüphanesi bulunamadı. Lütfen 'pip install xlrd' komutunu çalıştırın.")
                return []
            wb = xlrd.open_workbook(file_path, on_demand=True)
            ws = wb.sheet_by_index(0)
            for i in range(ws.nrows):
                row = ws.row_values(i)
                if i == 0:
                    headers = [self._normalize_header(h) for h in row]
                    continue
                if not row or all(cell == '' or cell is None for cell in row):
                    continue
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = cell
                rows.append(row_dict)
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                QMessageBox.critical(self, "Hata", "openpyxl kütüphanesi bulunamadı. Lütfen 'pip install openpyxl' komutunu çalıştırın.")
                return []
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i == 1:
                    headers = [self._normalize_header(h) for h in row]
                    continue
                if not row or all(cell is None for cell in row):
                    continue
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = cell
                rows.append(row_dict)
        return rows

    def _get_row_value(self, row, keys, default=None):
        for key in keys:
            if key in row and row[key] is not None:
                return row[key]
        return default

    def _parse_float(self, value, default=0.0):
        if value is None or value == "":
            return default
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return float(text)
        except Exception:
            return default

    def _parse_date(self, value):
        from datetime import datetime, date
        if value is None:
            return None
        if isinstance(value, date):
            return value
        text = str(value).strip()
        for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"]:
            try:
                return datetime.strptime(text, fmt).date()
            except Exception:
                continue
        return None

    def export_transactions_to_excel(self):
        """Mevcut işlemleri Excel dosyasına aktar"""
        from PyQt5.QtWidgets import QFileDialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from src.utils.helpers import format_tr
        import os

        try:
            transactions = TransactionService.get_all_transactions(self.user.id)
            transactions = sorted(transactions, key=lambda x: x.transaction_date, reverse=True)

            if not transactions:
                QMessageBox.information(self, "Bilgi", "Aktaracak işlem bulunamadı.")
                return

            # Dosya kaydetme dialogı
            from datetime import date
            default_name = f"Islemler_{date.today().strftime('%Y%m%d')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Excel Dosyası Kaydet", default_name,
                "Excel Dosyası (*.xlsx)"
            )
            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "İşlemler"

            # Başlık stili
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="1D6F42")
            header_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="AAAAAA")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            headers = ["Tarih", "Tür", "Müşteri", "Açıklama", "Ödeme Şekli", "Konu", "Ödeyen Kişi", "Tutar (₺)"]
            col_widths = [14, 18, 22, 30, 18, 20, 20, 16]

            for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

            ws.row_dimensions[1].height = 22

            # Veri satırları
            number_align = Alignment(horizontal="right")
            for row_idx, trans in enumerate(transactions, start=2):
                payment_text = self._get_payment_method_display_text(
                    trans.payment_method.value, trans.payment_type
                ) if trans.payment_method else ""
                values = [
                    str(trans.transaction_date),
                    trans.transaction_type.value if trans.transaction_type else "",
                    trans.customer_name or "",
                    trans.description or "",
                    payment_text,
                    trans.subject or "",
                    trans.person or "",
                    trans.amount,
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = border
                    if col == 8:  # Tutar sütunu
                        cell.number_format = '#,##0.00'
                        cell.alignment = number_align
                # Satır rengi (gelir/gider)
                if trans.transaction_type and trans.transaction_type.value in ["Gelir", "Kesilen Fatura"]:
                    row_fill = PatternFill("solid", fgColor="E8F5E9")
                elif trans.transaction_type and trans.transaction_type.value in ["Gider", "Gelen Fatura"]:
                    row_fill = PatternFill("solid", fgColor="FFEBEE")
                else:
                    row_fill = PatternFill("solid", fgColor="FFFFFF")
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).fill = row_fill

            # Toplam satırı
            total_row = len(transactions) + 2
            ws.cell(row=total_row, column=7, value="TOPLAM:").font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=sum(t.amount for t in transactions))
            ws.cell(row=total_row, column=8).number_format = '#,##0.00'
            ws.cell(row=total_row, column=8).font = Font(bold=True)
            ws.cell(row=total_row, column=8).alignment = number_align

            wb.save(file_path)
            QMessageBox.information(self, "Başarılı",
                f"{len(transactions)} işlem başarıyla aktarıldı.\n{file_path}")

            # Dosyayı otomatik aç
            os.startfile(file_path)

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel aktarım hatası:\n{str(e)}")

    def import_transactions_from_excel(self):
        """Gelişmiş Banka/Kredi Kartı ekstresini ithal et - Sütun eşleştirmesi, hızlı kurallar, vb."""
        from src.ui.dialogs.advanced_bank_import_dialog import AdvancedBankImportDialog
        
        dialog = AdvancedBankImportDialog(self.user.id, self)
        if dialog.exec_():
            self.refresh_transactions_table()
            # Otomatik oluşturulan banka hesaplarını ve kredi kartlarını göster
            if hasattr(self, 'table_banks'):
                self.refresh_bank_table()
            if hasattr(self, 'table_credit_cards'):
                self.refresh_credit_cards_table()

    def import_caris_from_excel(self):
        """Excel'den toplu cari aktar"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Cari Excel Dosyasi Sec",
                "",
                "Excel Dosyasi (*.xls *.xlsx"
            )
            if not file_path:
                return
            
            rows = self._read_excel_rows(file_path)
            if not rows:
                QMessageBox.warning(self, "Uyari", "Excel dosyasinda veri bulunamadi")
                return
            
            success = 0
            errors = []
            for idx, row in enumerate(rows, start=2):
                name = self._get_row_value(row, ["name", "cari_adi", "ad"])
                if not name:
                    errors.append(f"Satir {idx}: cari adi eksik")
                    continue
                cari_type = self._get_row_value(row, ["cari_type", "tip"], "MÜŞTERİ")
                phone = self._get_row_value(row, ["phone", "telefon"], None)
                email = self._get_row_value(row, ["email", "e_posta"], None)
                address = self._get_row_value(row, ["address", "adres"], None)
                balance = self._parse_float(self._get_row_value(row, ["balance", "bakiye"], 0.0), 0.0)
                
                ok, msg = CariService.create_cari(
                    self.user.id,
                    str(name),
                    str(cari_type),
                    email=email,
                    phone=phone,
                    address=address,
                    balance=balance
                )
                if ok:
                    success += 1
                else:
                    errors.append(f"Satir {idx}: {msg}")
            
            self.refresh_all_data()
            msg = f"Aktarim tamamlandi. Basarili: {success}"
            if errors:
                msg += f"\nHata: {len(errors)} (ilk 5)\n" + "\n".join(errors[:5])
            QMessageBox.information(self, "Bilgi", msg)
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kutuphanesi bulunamadi. Lutfen 'pip install openpyxl' komutunu calistirin.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel aktarim hatasi:\n{str(e)}")

    def import_banks_from_excel(self):
        """Excel'den toplu banka aktar"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Banka Excel Dosyasi Sec",
                "",
                "Excel Dosyasi (*.xls *.xlsx)"
            )
            if not file_path:
                return
            
            rows = self._read_excel_rows(file_path)
            if not rows:
                QMessageBox.warning(self, "Uyari", "Excel dosyasinda veri bulunamadi")
                return
            
            success = 0
            errors = []
            for idx, row in enumerate(rows, start=2):
                bank_name = self._get_row_value(row, ["bank_name", "banka_adi", "banka"])
                account_number = self._get_row_value(row, ["account_number", "hesap_no"])
                if not bank_name or not account_number:
                    errors.append(f"Satir {idx}: banka_adi veya hesap_no eksik")
                    continue
                
                iban = self._get_row_value(row, ["iban"], None)
                branch = self._get_row_value(row, ["branch", "sube"], None)
                currency = self._get_row_value(row, ["currency", "para_birimi"], "TRY")
                balance = self._parse_float(self._get_row_value(row, ["balance", "bakiye"], 0.0), 0.0)
                overdraft = self._parse_float(self._get_row_value(row, ["overdraft_limit", "ek_hesap_limiti"], 0.0), 0.0)
                
                ok, msg = BankService.create_account(
                    self.user.id,
                    str(bank_name),
                    str(account_number),
                    iban=iban,
                    currency=str(currency),
                    balance=balance,
                    branch=branch,
                    overdraft_limit=overdraft
                )
                if ok:
                    success += 1
                else:
                    errors.append(f"Satir {idx}: {msg}")
            
            self.refresh_all_data()
            msg = f"Aktarim tamamlandi. Basarili: {success}"
            if errors:
                msg += f"\nHata: {len(errors)} (ilk 5)\n" + "\n".join(errors[:5])
            QMessageBox.information(self, "Bilgi", msg)
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kutuphanesi bulunamadi. Lutfen 'pip install openpyxl' komutunu calistirin.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel aktarim hatasi:\n{str(e)}")

    def import_credit_cards_from_excel(self):
        """Excel'den toplu kredi karti aktar"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Kredi Karti Excel Dosyasi Sec",
                "",
                "Excel Dosyasi (*.xls *.xlsx)"
            )
            if not file_path:
                return
            
            rows = self._read_excel_rows(file_path)
            if not rows:
                QMessageBox.warning(self, "Uyari", "Excel dosyasinda veri bulunamadi")
                return
            
            success = 0
            errors = []
            for idx, row in enumerate(rows, start=2):
                card_name = self._get_row_value(row, ["card_name", "kart_adi", "kart"])
                card_number = self._get_row_value(row, ["card_number_last4", "son4", "card_number"], None)
                card_holder = self._get_row_value(row, ["card_holder", "kart_sahibi"], None)
                bank_name = self._get_row_value(row, ["bank_name", "banka_adi", "banka"], None)
                card_limit = self._parse_float(self._get_row_value(row, ["card_limit", "limit"], None), None)
                
                if not (card_name and card_number and card_holder and bank_name and card_limit is not None):
                    errors.append(f"Satir {idx}: zorunlu alan eksik")
                    continue
                
                card_number_str = str(card_number)
                last4 = card_number_str[-4:]
                closing_day = self._get_row_value(row, ["closing_day", "kesim_gunu"], 1)
                due_day = self._get_row_value(row, ["due_day", "odeme_gunu"], 15)
                
                result, msg = CreditCardService.create_card(
                    self.user.id,
                    str(card_name),
                    str(last4),
                    str(card_holder),
                    str(bank_name),
                    float(card_limit),
                    closing_day=int(closing_day),
                    due_day=int(due_day)
                )
                if result:
                    success += 1
                else:
                    errors.append(f"Satir {idx}: {msg}")
            
            self.refresh_all_data()
            msg = f"Aktarim tamamlandi. Basarili: {success}"
            if errors:
                msg += f"\nHata: {len(errors)} (ilk 5)\n" + "\n".join(errors[:5])
            QMessageBox.information(self, "Bilgi", msg)
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kutuphanesi bulunamadi. Lutfen 'pip install openpyxl' komutunu calistirin.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel aktarim hatasi:\n{str(e)}")
    
    def refresh_all_data(self):
        """Aktif ekranı ve dashboard'u gereksiz yük bindirmeden yenile."""
        try:
            self._loaded_tabs.clear()
            self.refresh_dashboard()
            self._refresh_current_tab_data(force=True)

            # Aktif rapor ekranı varsa yenile
            if hasattr(self, '_current_report_key') and self._current_report_key:
                try:
                    self._generate_sidebar_report(self._current_report_key)
                except Exception:
                    pass
        except Exception as e:
            print(f"Veri yenileme hatası: {e}")
            import traceback
            traceback.print_exc()
    
    def logout(self):
        """Çıkış yap"""
        reply = QMessageBox.question(self, "Çıkış", "Çıkmak istediğinizden emin misiniz?",
                                    QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.close()
    
    def create_credit_cards_tab(self) -> QWidget:
        """Kredi Kartları sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        title = QLabel("💳 Kredi Kartı Yönetimi")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        layout.addWidget(title)
        
        # İstatistikler
        stats_layout = QHBoxLayout()
        try:
            stats = CreditCardService.get_card_statistics(self.user.id)
            
            stats_layout.addWidget(self.create_stat_card("Toplam Kart", str(stats['total_cards']), "#9C27B0"))
            stats_layout.addWidget(self.create_stat_card("Toplam Limit", f"{stats['total_limit']:,.0f} ₺", "#2196F3"))
            stats_layout.addWidget(self.create_stat_card("Toplam Borç", f"{stats['total_debt']:,.0f} ₺", "#f44336"))
            stats_layout.addWidget(self.create_stat_card("Kullanılabilir", f"{stats['total_available']:,.0f} ₺", "#4CAF50"))
        except Exception as e:
            print(f"Kart stats hatası: {e}")
        
        layout.addLayout(stats_layout)
        layout.addSpacing(15)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_new = QPushButton("➕ Yeni Kredi Kartı")
        btn_new.setMinimumHeight(35)
        btn_new.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #7B1FA2; }
        """)
        btn_new.clicked.connect(self.show_new_credit_card_dialog)
        btn_layout.addWidget(btn_new)
        
        btn_refresh = QPushButton("🔄 Yenile")
        btn_refresh.setMinimumHeight(35)
        btn_refresh.clicked.connect(self.refresh_credit_cards_table)
        btn_layout.addWidget(btn_refresh)
        
        btn_import_cards = QPushButton("📥 Excel'den Aktar")
        btn_import_cards.setMinimumHeight(35)
        btn_import_cards.clicked.connect(self.import_credit_cards_from_excel)
        btn_layout.addWidget(btn_import_cards)

        btn_save_cards = QPushButton("💾 Kaydet")
        btn_save_cards.setMinimumHeight(35)
        btn_save_cards.setToolTip("Sütun genişliklerini kaydet")
        btn_save_cards.setStyleSheet("""
            QPushButton {
                background-color: #FF9800; color: white;
                border: none; border-radius: 4px;
                padding: 8px 16px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_cards.clicked.connect(lambda: self.save_column_widths(self.table_credit_cards, "credit_cards"))
        btn_layout.addWidget(btn_save_cards)

        # Arama kutucuğu
        btn_layout.addWidget(QLabel("🔍 Ara:"))
        self.credit_card_search_input = QLineEdit()
        self.credit_card_search_input.setPlaceholderText("Kart adı, banka, son 4 hane vb. yazın...")
        self.credit_card_search_input.setMinimumHeight(35)
        self.credit_card_search_input.setMaximumWidth(350)
        self.credit_card_search_input.textChanged.connect(self.search_credit_cards_table)
        btn_layout.addWidget(self.credit_card_search_input)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Tablo
        self.table_credit_cards = QTableWidget()
        self.table_credit_cards.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_credit_cards.setColumnCount(8)
        self.table_credit_cards.setHorizontalHeaderLabels([
            "Kart Adı", "Banka", "Son 4 Hane", "Limit", "Borç", "Kullanılabilir", "Durum", "İşlemler"
        ])
        self.table_credit_cards.horizontalHeader().setStretchLastSection(False)
        self.table_credit_cards.setColumnWidth(0, 180)
        self.table_credit_cards.setColumnWidth(1, 130)
        self.table_credit_cards.setColumnWidth(2, 90)
        self.table_credit_cards.setColumnWidth(3, 110)
        self.table_credit_cards.setColumnWidth(4, 110)
        self.table_credit_cards.setColumnWidth(5, 110)
        self.table_credit_cards.setColumnWidth(6, 80)
        self.table_credit_cards.setColumnWidth(7, 180)
        self.load_column_widths(self.table_credit_cards, "credit_cards")
        layout.addWidget(self.table_credit_cards)
        
        widget.setLayout(layout)
        return widget
    
    def refresh_credit_cards_table(self):
        """Kredi kartları tablosunu yenile"""
        try:
            cards = CreditCardService.get_all_cards(self.user.id)
            # Hızlı arama için id→card map
            card_map = {c.id: c for c in cards}

            self.table_credit_cards.setRowCount(len(cards))
            
            for i, card in enumerate(cards):
                # Tüm satırları göster (arama gizlemişse açmak için)
                self.table_credit_cards.setRowHidden(i, False)
                
                is_child = card.parent_card_id is not None

                # Kart adı — ek kart ise 🔗 işareti ekle
                name_display = f"🔗 {card.card_name}" if is_child else card.card_name
                self.table_credit_cards.setItem(i, 0, QTableWidgetItem(name_display))
                self.table_credit_cards.setItem(i, 1, QTableWidgetItem(card.bank_name))
                self.table_credit_cards.setItem(i, 2, QTableWidgetItem(f"****{card.card_number_last4}"))

                # Limit — ek kart ise "Paylaşımlı" göster
                if is_child:
                    parent = card_map.get(card.parent_card_id)
                    parent_name = parent.card_name if parent else "?"
                    limit_item = QTableWidgetItem(f"Paylaşımlı ({parent_name})")
                    limit_item.setForeground(Qt.darkMagenta)
                else:
                    limit_item = QTableWidgetItem(f"{format_tr(card.card_limit)} ₺")
                self.table_credit_cards.setItem(i, 3, limit_item)

                self.table_credit_cards.setItem(i, 4, QTableWidgetItem(f"{format_tr(card.current_debt)} ₺"))
                self.table_credit_cards.setItem(i, 5, QTableWidgetItem(f"{format_tr(card.available_limit)} ₺"))
                
                status = "Aktif" if card.is_active else "Pasif"
                self.table_credit_cards.setItem(i, 6, QTableWidgetItem(status))

                # Ek kart satırlarına hafif mor arka plan ver
                if is_child:
                    _shared_bg = QColor(243, 229, 245)
                    for col in range(7):
                        item = self.table_credit_cards.item(i, col)
                        if item:
                            item.setBackground(_shared_bg)
                
                # Butonlar
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)
                action_layout.setSpacing(5)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                btn_extract = QPushButton("📑 Dökümü Aç")
                btn_extract.setMinimumHeight(25)
                btn_extract.setStyleSheet("""
                    QPushButton {
                        background-color: #4CAF50;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #45a049; }
                """)
                btn_extract.clicked.connect(lambda checked, cid=card.id: self.show_credit_card_statement(cid))
                action_layout.addWidget(btn_extract)

                btn_edit = QPushButton("✏️ Düzenle")
                btn_edit.setMinimumHeight(25)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #0b7dda; }
                """)
                btn_edit.clicked.connect(lambda checked, cid=card.id: self.show_edit_credit_card_dialog(cid))
                action_layout.addWidget(btn_edit)
                
                btn_delete = QPushButton("🗑️ Sil")
                btn_delete.setMinimumHeight(25)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.clicked.connect(lambda checked, cid=card.id: self.delete_credit_card(cid))
                action_layout.addWidget(btn_delete)
                
                self.table_credit_cards.setCellWidget(i, 7, action_widget)
            
            self._resize_table(self.table_credit_cards, stretch_col=0)
            self.load_column_widths(self.table_credit_cards, "credit_cards")
        except Exception as e:
            print(f"Kredi kartı yükleme hatası: {e}")
            QMessageBox.critical(self, "Hata", f"Kredi kartları yüklenirken hata: {str(e)}")

    def search_credit_cards_table(self):
        """Kredi kartları tablosunda ara"""
        search_text = self.credit_card_search_input.text().strip().lower() if hasattr(self, 'credit_card_search_input') else ""
        
        if not search_text:
            self.refresh_credit_cards_table()
            return
        
        try:
            cards = CreditCardService.get_all_cards(self.user.id)
            card_map = {c.id: c for c in cards}

            self.table_credit_cards.setRowCount(len(cards))
            
            matched_count = 0
            for i, card in enumerate(cards):
                # Tüm sütunlarda ara: kart adı, banka, son 4 hane, limit, borç, durum
                search_in = [
                    card.card_name or "",
                    card.bank_name or "",
                    card.card_number_last4 or "",
                    str(format_tr(card.card_limit)),
                    str(card.card_limit),
                    str(format_tr(card.current_debt)),
                    str(card.current_debt),
                    "Aktif" if card.is_active else "Pasif"
                ]
                
                if not any(search_text in field.lower() for field in search_in):
                    self.table_credit_cards.setRowHidden(i, True)
                    continue
                
                matched_count += 1
                self.table_credit_cards.setRowHidden(i, False)
                
                is_child = card.parent_card_id is not None

                # Kart adı — ek kart ise 🔗 işareti ekle
                name_display = f"🔗 {card.card_name}" if is_child else card.card_name
                self.table_credit_cards.setItem(i, 0, QTableWidgetItem(name_display))
                self.table_credit_cards.setItem(i, 1, QTableWidgetItem(card.bank_name))
                self.table_credit_cards.setItem(i, 2, QTableWidgetItem(f"****{card.card_number_last4}"))

                # Limit — ek kart ise "Paylaşımlı" göster
                if is_child:
                    parent = card_map.get(card.parent_card_id)
                    parent_name = parent.card_name if parent else "?"
                    limit_item = QTableWidgetItem(f"Paylaşımlı ({parent_name})")
                    limit_item.setForeground(Qt.darkMagenta)
                else:
                    limit_item = QTableWidgetItem(f"{format_tr(card.card_limit)} ₺")
                self.table_credit_cards.setItem(i, 3, limit_item)

                self.table_credit_cards.setItem(i, 4, QTableWidgetItem(f"{format_tr(card.current_debt)} ₺"))
                self.table_credit_cards.setItem(i, 5, QTableWidgetItem(f"{format_tr(card.available_limit)} ₺"))
                
                status = "Aktif" if card.is_active else "Pasif"
                self.table_credit_cards.setItem(i, 6, QTableWidgetItem(status))

                # Ek kart satırlarına hafif mor arka plan ver
                if is_child:
                    _shared_bg = QColor(243, 229, 245)
                    for col in range(7):
                        item = self.table_credit_cards.item(i, col)
                        if item:
                            item.setBackground(_shared_bg)
                
                # Butonlar
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)
                action_layout.setSpacing(5)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                btn_extract = QPushButton("📑 Dökümü Aç")
                btn_extract.setMinimumHeight(25)
                btn_extract.setStyleSheet("""
                    QPushButton {
                        background-color: #4CAF50;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #45a049; }
                """)
                btn_extract.clicked.connect(lambda checked, cid=card.id: self.show_credit_card_statement(cid))
                action_layout.addWidget(btn_extract)

                btn_edit = QPushButton("✏️ Düzenle")
                btn_edit.setMinimumHeight(25)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #0b7dda; }
                """)
                btn_edit.clicked.connect(lambda checked, cid=card.id: self.show_edit_credit_card_dialog(cid))
                action_layout.addWidget(btn_edit)
                
                btn_delete = QPushButton("🗑️ Sil")
                btn_delete.setMinimumHeight(25)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.clicked.connect(lambda checked, cid=card.id: self.delete_credit_card(cid))
                action_layout.addWidget(btn_delete)
                
                self.table_credit_cards.setCellWidget(i, 7, action_widget)
            
            self._resize_table(self.table_credit_cards, stretch_col=0)
            self.load_column_widths(self.table_credit_cards, "credit_cards")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Arama hatası: {str(e)}")
    
    def show_new_credit_card_dialog(self):
        """Yeni kredi kartı dialog"""
        from src.ui.dialogs.credit_card_dialog import CreditCardDialog
        dialog = CreditCardDialog(self.user.id, self)
        if dialog.exec_():
            self.refresh_all_data()

    def show_edit_credit_card_dialog(self, card_id):
        """Kredi kartı düzenleme dialog"""
        from src.ui.dialogs.credit_card_dialog import CreditCardDialog
        dialog = CreditCardDialog(self.user.id, self, card_id=card_id)
        if dialog.exec_():
            self.refresh_all_data()
    
    def delete_credit_card(self, card_id):
        """Kredi kartı sil"""
        reply = QMessageBox.question(self, "Onay", "Bu kredi kartını silmek istediğinize emin misiniz?",
                                    QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            success, msg = CreditCardService.delete_card(card_id)
            if success:
                QMessageBox.information(self, "Başarılı", "Kredi kartı silindi")
                self.refresh_all_data()
            else:
                QMessageBox.critical(self, "Hata", msg)
    
    def create_loans_tab(self) -> QWidget:
        """Krediler sekmesi"""
        from src.services.loan_service import LoanService
        from src.ui.dialogs.loan_dialog import LoanDialog
        
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        title = QLabel("📊 Kredi Yönetimi")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        layout.addWidget(title)
        
        # İstatistikler
        stats_layout = QHBoxLayout()
        self.loan_stats_cards = {}

        total_card = self.create_stat_card("Toplam Kredi", "0 ₺", "#FF9800")
        paid_card = self.create_stat_card("Toplam Ödenen", "0 ₺", "#4CAF50")
        remaining_card = self.create_stat_card("Kalan Borç", "0 ₺", "#f44336")
        active_card = self.create_stat_card("Aktif Krediler", "0", "#2196F3")

        self.loan_stats_cards['toplam_kredi'] = total_card
        self.loan_stats_cards['toplam_odenen'] = paid_card
        self.loan_stats_cards['toplam_kalan'] = remaining_card
        self.loan_stats_cards['akif_kredi_sayisi'] = active_card

        stats_layout.addWidget(total_card)
        stats_layout.addWidget(paid_card)
        stats_layout.addWidget(remaining_card)
        stats_layout.addWidget(active_card)
        
        layout.addLayout(stats_layout)
        layout.addSpacing(15)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_new = QPushButton("➕ Yeni Kredi")
        btn_new.setMinimumHeight(35)
        btn_new.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #E68900; }
        """)
        btn_new.clicked.connect(lambda: self.show_new_loan_dialog())
        btn_layout.addWidget(btn_new)

        btn_preview = QPushButton("🔍 Ön İzleme")
        btn_preview.setMinimumHeight(35)
        btn_preview.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
        """)
        btn_preview.clicked.connect(self.show_selected_loan_statement)
        btn_layout.addWidget(btn_preview)

        btn_import_excel = QPushButton("📥 Excelden Aktar")
        btn_import_excel.setMinimumHeight(35)
        btn_import_excel.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #546E7A; }
        """)
        btn_import_excel.clicked.connect(self.import_loans_from_excel)
        btn_layout.addWidget(btn_import_excel)

        btn_sample_excel = QPushButton("🧾 Örnek Excel")
        btn_sample_excel.setMinimumHeight(35)
        btn_sample_excel.setStyleSheet("""
            QPushButton {
                background-color: #795548;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #6D4C41; }
        """)
        btn_sample_excel.clicked.connect(self.export_loan_template_excel)
        btn_layout.addWidget(btn_sample_excel)
        
        btn_bulk_delete = QPushButton("🗑️ Toplu Sil")
        btn_bulk_delete.setMinimumHeight(35)
        btn_bulk_delete.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #d32f2f; }
        """)
        btn_bulk_delete.clicked.connect(self.delete_loans_bulk)
        btn_layout.addWidget(btn_bulk_delete)

        btn_refresh = QPushButton("🔄 Yenile")
        btn_refresh.setMinimumHeight(35)
        btn_refresh.clicked.connect(self.refresh_loans_table)
        btn_layout.addWidget(btn_refresh)

        btn_save_loans = QPushButton("💾 Kaydet")
        btn_save_loans.setMinimumHeight(35)
        btn_save_loans.setToolTip("Sütun genişliklerini kaydet")
        btn_save_loans.setStyleSheet("""
            QPushButton {
                background-color: #FF9800; color: white;
                border: none; border-radius: 4px;
                padding: 8px 16px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_loans.clicked.connect(lambda: self.save_column_widths(self.table_loans, "loans"))
        btn_layout.addWidget(btn_save_loans)

        # Arama kutucuğu
        btn_layout.addWidget(QLabel("🔍 Ara:"))
        self.loan_search_input = QLineEdit()
        self.loan_search_input.setPlaceholderText("Kredi adı, banka, tutar vb. yazın...")
        self.loan_search_input.setMinimumHeight(35)
        self.loan_search_input.setMaximumWidth(300)
        self.loan_search_input.textChanged.connect(self.search_loans_table)
        btn_layout.addWidget(self.loan_search_input)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Tablo
        self.table_loans = QTableWidget()
        self.table_loans.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_loans.setColumnCount(13)
        self.table_loans.setHorizontalHeaderLabels([
            "Kredi Adı", "Banka", "Firma", "Tip", "Çekilen Tutar", "Toplam", "Ödenen", "Kalan", "Aylık", "Durum", "Kalan Taksit", "Ödeme Günü", "İşlemler"
        ])
        self.table_loans.horizontalHeader().setStretchLastSection(False)
        self.table_loans.setColumnWidth(0, 160)
        self.table_loans.setColumnWidth(1, 120)
        self.table_loans.setColumnWidth(2, 120)
        self.table_loans.setColumnWidth(3, 100)
        self.table_loans.setColumnWidth(4, 120)
        self.table_loans.setColumnWidth(5, 110)
        self.table_loans.setColumnWidth(6, 110)
        self.table_loans.setColumnWidth(7, 110)
        self.table_loans.setColumnWidth(8, 100)
        self.table_loans.setColumnWidth(9, 100)
        self.table_loans.setColumnWidth(10, 100)
        self.table_loans.setColumnWidth(11, 100)
        self.table_loans.setColumnWidth(12, 260)
        self.table_loans.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_loans.setSelectionMode(QTableWidget.ExtendedSelection)
        self.load_column_widths(self.table_loans, "loans")
        layout.addWidget(self.table_loans)
        
        widget.setLayout(layout)
        return widget

    def _set_loan_stat_card_value(self, key, value):
        if not hasattr(self, "loan_stats_cards"):
            return
        card = self.loan_stats_cards.get(key)
        if not card:
            return
        labels = card.findChildren(QLabel)
        if len(labels) > 1:
            labels[1].setText(value)

    def refresh_loan_stats(self):
        """Krediler sekmesi üst istatistik kartlarını yenile"""
        from src.services.loan_service import LoanService

        try:
            stats = LoanService.get_loans_summary(self.user.id) or {}
            self._set_loan_stat_card_value('toplam_kredi', f"{stats.get('toplam_kredi', 0):,.0f} ₺")
            self._set_loan_stat_card_value('toplam_odenen', f"{stats.get('toplam_odenen', 0):,.0f} ₺")
            self._set_loan_stat_card_value('toplam_kalan', f"{stats.get('toplam_kalan', 0):,.0f} ₺")
            self._set_loan_stat_card_value('akif_kredi_sayisi', str(stats.get('akif_kredi_sayisi', 0)))
        except Exception as e:
            print(f"Kredi stats yenileme hatası: {e}")

    def _get_loan_total_repayment(self, loan):
        total_repayment = float(getattr(loan, 'remaining_balance', 0) or 0)
        loan_amount = float(getattr(loan, 'loan_amount', 0) or 0)
        return max(total_repayment, loan_amount)

    def _get_loan_remaining_amount(self, loan):
        """Kalan borcu formülle hesapla: geri ödenecek - ödenen"""
        total_repayment = self._get_loan_total_repayment(loan)
        total_paid = float(getattr(loan, 'total_paid', 0) or 0)
        return max(0.0, total_repayment - total_paid)
    
    def refresh_loans_table(self):
        """Kredi tablosunu yenile"""
        from src.services.loan_service import LoanService
        
        try:
            loans = LoanService.get_loans(self.user.id, active_only=True)
            self.table_loans.setRowCount(len(loans))
            self.refresh_loan_stats()
            
            for i, loan in enumerate(loans):
                # Tüm satırları göster (arama gizlemişse açmak için)
                self.table_loans.setRowHidden(i, False)
                
                remaining_amount = self._get_loan_remaining_amount(loan)
                name_item = QTableWidgetItem(loan.loan_name)
                name_item.setData(Qt.UserRole, loan.id)
                self.table_loans.setItem(i, 0, name_item)
                self.table_loans.setItem(i, 1, QTableWidgetItem(loan.bank_name))
                self.table_loans.setItem(i, 2, QTableWidgetItem(loan.company_name or ""))
                self.table_loans.setItem(i, 3, QTableWidgetItem(loan.loan_type))
                self.table_loans.setItem(i, 4, QTableWidgetItem(f"{format_tr(loan.loan_amount)}"))
                total_repayment = self._get_loan_total_repayment(loan)
                self.table_loans.setItem(i, 5, QTableWidgetItem(f"{format_tr(total_repayment)} ₺"))
                self.table_loans.setItem(i, 6, QTableWidgetItem(f"{format_tr(loan.total_paid)} ₺"))
                self.table_loans.setItem(i, 7, QTableWidgetItem(f"{format_tr(remaining_amount)} ₺"))
                self.table_loans.setItem(i, 8, QTableWidgetItem(f"{format_tr(loan.monthly_payment)} ₺"))
                
                status = loan.status
                self.table_loans.setItem(i, 9, QTableWidgetItem(status))
                
                # Kalan taksit sayısı hesapla
                remaining_installments = 0
                if loan.total_installments and loan.paid_installments:
                    remaining_installments = max(0, loan.total_installments - loan.paid_installments)
                elif loan.total_installments:
                    remaining_installments = loan.total_installments
                
                self.table_loans.setItem(i, 10, QTableWidgetItem(str(remaining_installments)))
                self.table_loans.setItem(i, 11, QTableWidgetItem(f"{loan.due_day}"))
                
                # Butonlar
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)
                action_layout.setSpacing(5)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                
                btn_edit = QPushButton("✏️ Düzenle")
                btn_edit.setMinimumHeight(25)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #0b7dda; }
                """)
                btn_edit.clicked.connect(lambda checked, lid=loan.id: self.show_edit_loan_dialog(lid))
                action_layout.addWidget(btn_edit)
                
                btn_delete = QPushButton("🗑️ Sil")
                btn_delete.setMinimumHeight(25)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.clicked.connect(lambda checked, lid=loan.id: self.delete_loan(lid))
                action_layout.addWidget(btn_delete)

                btn_statement = QPushButton("📑 Dökümü Aç")
                btn_statement.setMinimumHeight(25)
                btn_statement.setStyleSheet("""
                    QPushButton {
                        background-color: #607D8B;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #546E7A; }
                """)
                btn_statement.clicked.connect(lambda checked, lid=loan.id: self.show_loan_statement(lid))
                action_layout.addWidget(btn_statement)
                
                self.table_loans.setCellWidget(i, 12, action_widget)
            
            self._resize_table(self.table_loans, stretch_col=0)
            self.load_column_widths(self.table_loans, "loans")
        except Exception as e:
            print(f"Kredi yükleme hatası: {e}")
            QMessageBox.critical(self, "Hata", f"Krediler yüklenirken hata: {str(e)}")

    def search_loans_table(self):
        """Kredi tablosunda ara"""
        search_text = self.loan_search_input.text().strip().lower() if hasattr(self, 'loan_search_input') else ""
        
        if not search_text:
            self.refresh_loans_table()
            return
        
        from src.services.loan_service import LoanService
        
        try:
            loans = LoanService.get_loans(self.user.id, active_only=True)
            self.table_loans.setRowCount(len(loans))
            self.refresh_loan_stats()
            
            matched_count = 0
            for i, loan in enumerate(loans):
                # Tüm sütunlarda ara
                search_in = [
                    loan.loan_name or "",  # Kredi adı
                    loan.bank_name or "",  # Banka
                    loan.company_name or "",  # Firma
                    loan.loan_type or "",  # Kredi türü
                    str(format_tr(loan.loan_amount)),  # Tutar (formatlı)
                    str(loan.loan_amount),  # Tutar (normal)
                    str(format_tr(loan.total_paid)),  # Ödenen (formatlı)
                    str(loan.total_paid),  # Ödenen (normal)
                    loan.status or ""  # Durum
                ]
                
                # Eğer arama metni herhangi bir sütunda varsa göster
                if not any(search_text in field.lower() for field in search_in):
                    self.table_loans.setRowHidden(i, True)
                    continue
                
                matched_count += 1
                self.table_loans.setRowHidden(i, False)
                
                remaining_amount = self._get_loan_remaining_amount(loan)
                name_item = QTableWidgetItem(loan.loan_name)
                name_item.setData(Qt.UserRole, loan.id)
                self.table_loans.setItem(i, 0, name_item)
                self.table_loans.setItem(i, 1, QTableWidgetItem(loan.bank_name))
                self.table_loans.setItem(i, 2, QTableWidgetItem(loan.company_name or ""))
                self.table_loans.setItem(i, 3, QTableWidgetItem(loan.loan_type))
                self.table_loans.setItem(i, 4, QTableWidgetItem(f"{format_tr(loan.loan_amount)}"))
                total_repayment = self._get_loan_total_repayment(loan)
                self.table_loans.setItem(i, 5, QTableWidgetItem(f"{format_tr(total_repayment)} ₺"))
                self.table_loans.setItem(i, 6, QTableWidgetItem(f"{format_tr(loan.total_paid)} ₺"))
                self.table_loans.setItem(i, 7, QTableWidgetItem(f"{format_tr(remaining_amount)} ₺"))
                self.table_loans.setItem(i, 8, QTableWidgetItem(f"{format_tr(loan.monthly_payment)} ₺"))
                
                status = loan.status
                self.table_loans.setItem(i, 9, QTableWidgetItem(status))
                
                # Kalan taksit sayısı hesapla
                remaining_installments = 0
                if loan.total_installments and loan.paid_installments:
                    remaining_installments = max(0, loan.total_installments - loan.paid_installments)
                elif loan.total_installments:
                    remaining_installments = loan.total_installments
                
                self.table_loans.setItem(i, 10, QTableWidgetItem(str(remaining_installments)))
                self.table_loans.setItem(i, 11, QTableWidgetItem(f"{loan.due_day}"))
                
                # Butonlar
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)
                action_layout.setSpacing(5)
                action_layout.setAlignment(Qt.AlignCenter)
                action_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                
                btn_edit = QPushButton("✏️ Düzenle")
                btn_edit.setMinimumHeight(25)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #0b7dda; }
                """)
                btn_edit.clicked.connect(lambda checked, lid=loan.id: self.show_edit_loan_dialog(lid))
                action_layout.addWidget(btn_edit)
                
                btn_delete = QPushButton("🗑️ Sil")
                btn_delete.setMinimumHeight(25)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #da190b; }
                """)
                btn_delete.clicked.connect(lambda checked, lid=loan.id: self.delete_loan(lid))
                action_layout.addWidget(btn_delete)

                btn_statement = QPushButton("📑 Dökümü Aç")
                btn_statement.setMinimumHeight(25)
                btn_statement.setStyleSheet("""
                    QPushButton {
                        background-color: #607D8B;
                        color: white;
                        border: none;
                        border-radius: 3px;
                        padding: 4px 8px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #546E7A; }
                """)
                btn_statement.clicked.connect(lambda checked, lid=loan.id: self.show_loan_statement(lid))
                action_layout.addWidget(btn_statement)
                
                self.table_loans.setCellWidget(i, 12, action_widget)
            
            self._resize_table(self.table_loans, stretch_col=0)
            self.load_column_widths(self.table_loans, "loans")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Arama hatası: {str(e)}")

    def get_selected_loan_id(self):
        """Seçili kredi satırının ID bilgisini getir"""
        if not hasattr(self, 'table_loans'):
            return None
        row = self.table_loans.currentRow()
        if row < 0:
            return None
        item = self.table_loans.item(row, 0)
        if not item:
            return None
        return item.data(Qt.UserRole)

    def _parse_excel_float(self, value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(" ", "").replace("₺", "").replace("TL", "")
        if not text or text in ("-", "--", "---", "N/A", "n/a", "-"):
            return None
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                normalized = text.replace(".", "").replace(",", ".")
            else:
                normalized = text.replace(",", "")
        elif "," in text:
            normalized = text.replace(",", ".")
        else:
            normalized = text
        try:
            return float(normalized)
        except ValueError:
            return None

    def _parse_excel_date(self, value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return None

    def export_loan_template_excel(self):
        try:
            from PyQt5.QtWidgets import QFileDialog
            from openpyxl import Workbook

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Örnek Kredi Excel",
                "kredi_ornek.xlsx",
                "Excel Dosyası (*.xlsx)"
            )
            if not file_path:
                return
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            headers = [
                "Kredi Adı", "Banka", "Firma Adı", "Tip", "Çekilen Tutar", "Toplam Borç",
                "Başlangıç Tarihi", "Ödeme Günü", "Aylık Taksit", "Faiz Oranı",
                "Bitiş Tarihi", "Toplam Taksit", "Not"
            ]

            wb = Workbook()
            ws = wb.active
            ws.title = "Krediler"
            ws.append(headers)
            ws.append([
                "İş Bankası Konut Kredisi", "İş Bankası", "ÖZKAYA LTD.", "KONUT", 1500000, 1650000,
                "01.02.2026", 15, 25000, 2.85,
                "01.02.2036", 120, "Örnek kredi kaydı"
            ])

            wb.save(file_path)
            QMessageBox.information(self, "Başarılı", f"Örnek Excel oluşturuldu:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Örnek Excel oluşturulamadı:\n{str(e)}")

    def import_loans_from_excel(self):
        try:
            from PyQt5.QtWidgets import QFileDialog
            from openpyxl import load_workbook

            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Kredi Excel Seç",
                "",
                "Excel Dosyası (*.xlsx *.xls)"
            )
            if not file_path:
                return

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            def normalize_header(value):
                text = str(value or "").casefold().strip()
                for src, dst in {
                    "ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c"
                }.items():
                    text = text.replace(src, dst)
                return "".join(ch for ch in text if ch.isalnum())

            header_map = {
                "krediadı": "loan_name",
                "krediadi": "loan_name",
                "banka": "bank_name",
                "bank": "bank_name",
                "firmaadi": "company_name",
                "firma": "company_name",
                "sirket": "company_name",
                "tip": "loan_type",
                "kreditipi": "loan_type",
                "cekilentutar": "loan_amount",
                "kreditutari": "loan_amount",
                "toplamborc": "remaining_balance",
                "kalanborc": "remaining_balance",
                "baslangictarihi": "start_date",
                "odemegunu": "due_day",
                "ayliktaksit": "monthly_payment",
                "faizorani": "interest_rate",
                "bitistarihi": "end_date",
                "toplamtaksit": "total_installments",
                "not": "notes",
            }

            headers = [normalize_header(cell.value) for cell in ws[1]]
            column_keys = []
            for h in headers:
                column_keys.append(header_map.get(h))

            success_count = 0
            errors = []

            for row_idx in range(2, ws.max_row + 1):
                row_values = [cell.value for cell in ws[row_idx]]
                if not any(v is not None and str(v).strip() for v in row_values):
                    continue

                data = {}
                for key, value in zip(column_keys, row_values):
                    if key:
                        data[key] = value

                loan_name = str(data.get("loan_name", "") or "").strip()
                bank_name = str(data.get("bank_name", "") or "").strip()
                company_name = str(data.get("company_name", "") or "").strip() or None
                loan_type = str(data.get("loan_type", "") or "").strip()
                loan_amount = self._parse_excel_float(data.get("loan_amount"))
                start_date = self._parse_excel_date(data.get("start_date"))
                due_day_raw = self._parse_excel_float(data.get("due_day"))

                if not loan_name or not bank_name or not loan_type or loan_amount is None or start_date is None or due_day_raw is None:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik")
                    continue

                remaining_balance = self._parse_excel_float(data.get("remaining_balance"))
                monthly_payment = self._parse_excel_float(data.get("monthly_payment")) or 0.0
                interest_rate = self._parse_excel_float(data.get("interest_rate")) or 0.0
                end_date = self._parse_excel_date(data.get("end_date"))
                total_installments = self._parse_excel_float(data.get("total_installments"))
                notes = str(data.get("notes", "") or "").strip() or None

                loan, message = LoanService.create_loan(
                    self.user.id,
                    loan_name,
                    bank_name,
                    company_name=company_name,
                    loan_type=loan_type,
                    loan_amount=float(loan_amount),
                    start_date=start_date,
                    due_day=int(due_day_raw),
                    interest_rate=float(interest_rate),
                    monthly_payment=float(monthly_payment),
                    remaining_balance=float(remaining_balance) if remaining_balance is not None else None,
                    end_date=end_date,
                    total_installments=int(total_installments) if total_installments is not None else None,
                    notes=notes
                )

                if loan:
                    success_count += 1
                else:
                    errors.append(f"Satır {row_idx}: {message}")

            self.refresh_loans_table()

            if errors:
                error_text = "\n".join(errors[:8])
                if len(errors) > 8:
                    error_text += "\n..."
                QMessageBox.warning(
                    self,
                    "Excel Aktarım Tamamlandı",
                    f"Başarılı: {success_count}\nHatalı: {len(errors)}\n\nHatalar:\n{error_text}"
                )
            else:
                QMessageBox.information(
                    self,
                    "Excel Aktarım Tamamlandı",
                    f"Başarılı: {success_count}"
                )
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel aktarımı başarısız:\n{str(e)}")

    def show_selected_loan_statement(self):
        """Seçili kredi için ön izleme dökümü aç"""
        loan_id = self.get_selected_loan_id()
        if not loan_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce krediler tablosundan bir kayıt seçiniz.")
            return
        self.show_loan_statement(loan_id)

    def _extract_loan_id_from_notes(self, notes_text):
        """Transaction notundan loan_id bilgisini al"""
        if not notes_text:
            return None
        notes = str(notes_text).strip()
        if not notes.startswith("loan_id:"):
            return None
        try:
            return int(notes.split(":", 1)[1].strip())
        except (ValueError, TypeError):
            return None

    def show_loan_statement(self, loan_id):
        """Seçili kredi için ödeme dökümünü aç"""
        from PyQt5.QtWidgets import QDialog
        from src.database.db import SessionLocal
        from src.database.models import Loan

        session = SessionLocal()
        try:
            loan = session.query(Loan).filter(
                Loan.id == loan_id,
                Loan.user_id == self.user.id,
                Loan.is_active == True
            ).first()

            if not loan:
                QMessageBox.warning(self, "Uyarı", "Kredi kaydı bulunamadı.")
                return

            current_remaining = self._get_loan_remaining_amount(loan)

            all_loan_payments = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_type == TransactionType.KREDI_ODEME
            ).order_by(Transaction.transaction_date.asc(), Transaction.id.asc()).all()

            loan_payments = []
            for trans in all_loan_payments:
                if self._extract_loan_id_from_notes(trans.notes) == loan.id:
                    loan_payments.append(trans)

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Kredi Dökümü - {loan.loan_name}")
            dialog.resize(980, 620)

            main_layout = QVBoxLayout(dialog)

            # Kalan taksit sayısı hesapla
            remaining_installments = 0
            if loan.total_installments and loan.paid_installments:
                remaining_installments = max(0, loan.total_installments - loan.paid_installments)
            elif loan.total_installments:
                remaining_installments = loan.total_installments

            summary_label = QLabel(
                f"<b>Kredi:</b> {loan.loan_name} &nbsp;&nbsp; "
                f"<b>Banka:</b> {loan.bank_name} &nbsp;&nbsp; "
                f"<b>Tip:</b> {loan.loan_type}<br>"
                f"<b>Toplam:</b> {format_tr(self._get_loan_total_repayment(loan))} ₺ &nbsp;&nbsp; "
                f"<b>Toplam Ödenen:</b> {format_tr(loan.total_paid)} ₺ &nbsp;&nbsp; "
                f"<b>Güncel Kalan:</b> {format_tr(current_remaining)} ₺<br>"
                f"<b>Kalan Taksit:</b> {remaining_installments} &nbsp;&nbsp; "
                f"<b>Ödeme Günü:</b> Ayın {loan.due_day}. günü"
            )
            summary_label.setWordWrap(True)
            main_layout.addWidget(summary_label)

            table = QTableWidget()
            table.setColumnCount(6)
            table.setHorizontalHeaderLabels([
                "Ödeme Tarihi", "Ödeme Yöntemi", "Kaynak", "Açıklama", "Ödenen", "Kalan Bakiye"
            ])
            table.setRowCount(len(loan_payments))

            running_balance = self._get_loan_total_repayment(loan)
            for row, trans in enumerate(loan_payments):
                running_balance = max(0, running_balance - trans.amount)

                payment_method = trans.payment_method.value if trans.payment_method else "-"
                source_text = "-"
                if trans.bank_account:
                    source_text = f"{trans.bank_account.bank_name} ({trans.bank_account.account_number})"

                table.setItem(row, 0, QTableWidgetItem(str(trans.transaction_date)))
                table.setItem(row, 1, QTableWidgetItem(payment_method))
                table.setItem(row, 2, QTableWidgetItem(source_text))
                table.setItem(row, 3, QTableWidgetItem(trans.description or ""))
                table.setItem(row, 4, QTableWidgetItem(f"{format_tr(trans.amount)} ₺"))
                table.setItem(row, 5, QTableWidgetItem(f"{format_tr(running_balance)} ₺"))

            table.setColumnWidth(0, 110)
            table.setColumnWidth(1, 120)
            table.setColumnWidth(2, 210)
            table.setColumnWidth(3, 290)
            table.setColumnWidth(4, 110)
            table.setColumnWidth(5, 120)
            table.horizontalHeader().setStretchLastSection(False)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self._resize_table(table)

            main_layout.addWidget(table)

            info_label = QLabel(f"Toplam {len(loan_payments)} ödeme kaydı listeleniyor.")
            main_layout.addWidget(info_label)

            close_btn = QPushButton("Kapat")
            close_btn.clicked.connect(dialog.accept)
            main_layout.addWidget(close_btn, alignment=Qt.AlignRight)

            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kredi dökümü açılırken hata: {str(e)}")
        finally:
            session.close()

    def show_credit_card_statement(self, card_id):
        """Seçili kredi kartı için döküm ekranını aç"""
        from PyQt5.QtWidgets import QDialog
        from src.database.db import SessionLocal
        from src.database.models import CreditCard

        session = SessionLocal()
        try:
            card = session.query(CreditCard).filter(
                CreditCard.id == card_id,
                CreditCard.user_id == self.user.id,
                CreditCard.is_active == True
            ).first()

            if not card:
                QMessageBox.warning(self, "Uyarı", "Kredi kartı kaydı bulunamadı.")
                return

            transactions = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.credit_card_id == card.id
            ).order_by(Transaction.transaction_date.asc(), Transaction.id.asc()).all()

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Kredi Kartı Dökümü - {card.card_name}")
            dialog.resize(980, 620)

            main_layout = QVBoxLayout(dialog)

            summary_label = QLabel(
                f"<b>Kart:</b> {card.card_name} &nbsp;&nbsp; "
                f"<b>Banka:</b> {card.bank_name} &nbsp;&nbsp; "
                f"<b>Limit:</b> {format_tr(card.card_limit)} ₺<br>"
                f"<b>Güncel Borç:</b> {format_tr(card.current_debt)} ₺ &nbsp;&nbsp; "
                f"<b>Kullanılabilir:</b> {format_tr(card.available_limit)} ₺ &nbsp;&nbsp; "
                f"<b>Kesim:</b> {card.closing_day} &nbsp;&nbsp; "
                f"<b>Son Ödeme:</b> {card.due_day}"
            )
            summary_label.setWordWrap(True)
            main_layout.addWidget(summary_label)

            table = QTableWidget()
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels([
                "Tarih", "İşlem Türü", "Açıklama", "Tutar", "Bakiye"
            ])
            table.setRowCount(len(transactions))

            running_debt = 0.0
            for row, trans in enumerate(transactions):
                if trans.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]:
                    running_debt += trans.amount
                elif trans.transaction_type == TransactionType.KREDI_KARTI_ODEME:
                    running_debt = max(0.0, running_debt - trans.amount)

                table.setItem(row, 0, QTableWidgetItem(str(trans.transaction_date)))
                table.setItem(row, 1, QTableWidgetItem(trans.transaction_type.value))
                table.setItem(row, 2, QTableWidgetItem(trans.description or ""))
                table.setItem(row, 3, QTableWidgetItem(f"{format_tr(trans.amount)} ₺"))
                table.setItem(row, 4, QTableWidgetItem(f"{format_tr(running_debt)} ₺"))

            table.setColumnWidth(0, 110)
            table.setColumnWidth(1, 140)
            table.setColumnWidth(2, 360)
            table.setColumnWidth(3, 120)
            table.setColumnWidth(4, 120)
            table.horizontalHeader().setStretchLastSection(False)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self._resize_table(table)

            main_layout.addWidget(table)

            info_label = QLabel(f"Toplam {len(transactions)} işlem listeleniyor.")
            main_layout.addWidget(info_label)

            close_btn = QPushButton("Kapat")
            close_btn.clicked.connect(dialog.accept)
            main_layout.addWidget(close_btn, alignment=Qt.AlignRight)

            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kredi kartı dökümü açılırken hata: {str(e)}")
        finally:
            session.close()

    def show_bank_statement(self, account_id):
        """Seçili banka hesabı için döküm ekranını aç"""
        from PyQt5.QtWidgets import QDialog
        from src.database.db import SessionLocal
        from src.database.models import BankAccount, BankTransaction

        session = SessionLocal()
        try:
            account = session.query(BankAccount).filter(
                BankAccount.id == account_id,
                BankAccount.user_id == self.user.id,
                BankAccount.is_active == True
            ).first()

            if not account:
                QMessageBox.warning(self, "Uyarı", "Banka hesabı bulunamadı.")
                return

            transactions = session.query(BankTransaction).filter(
                BankTransaction.user_id == self.user.id,
                BankTransaction.bank_account_id == account.id
            ).order_by(BankTransaction.transaction_date.asc(), BankTransaction.id.asc()).all()

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Banka Dökümü - {account.bank_name}")
            dialog.resize(980, 620)

            main_layout = QVBoxLayout(dialog)

            summary_label = QLabel(
                f"<b>Banka:</b> {account.bank_name} &nbsp;&nbsp; "
                f"<b>Hesap:</b> {account.account_number} &nbsp;&nbsp; "
                f"<b>Para Birimi:</b> {account.currency}<br>"
                f"<b>Bakiye:</b> {format_tr(account.balance)} ₺ &nbsp;&nbsp; "
                f"<b>Ek Hesap:</b> {format_tr(getattr(account, 'overdraft_limit', 0.0))} ₺"
            )
            summary_label.setWordWrap(True)
            main_layout.addWidget(summary_label)

            table = QTableWidget()
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels([
                "Tarih", "İşlem Türü", "Açıklama", "Tutar", "Bakiye"
            ])
            table.setRowCount(len(transactions))

            running_balance = 0.0
            for row, trans in enumerate(transactions):
                if trans.transaction_type == "INCOME":
                    running_balance += trans.amount
                else:
                    running_balance -= trans.amount

                table.setItem(row, 0, QTableWidgetItem(str(trans.transaction_date)))
                table.setItem(row, 1, QTableWidgetItem(trans.transaction_type))
                table.setItem(row, 2, QTableWidgetItem(trans.description or ""))
                table.setItem(row, 3, QTableWidgetItem(f"{format_tr(trans.amount)} ₺"))
                table.setItem(row, 4, QTableWidgetItem(f"{format_tr(running_balance)} ₺"))

            table.setColumnWidth(0, 110)
            table.setColumnWidth(1, 140)
            table.setColumnWidth(2, 360)
            table.setColumnWidth(3, 120)
            table.setColumnWidth(4, 120)
            table.horizontalHeader().setStretchLastSection(False)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self._resize_table(table)

            main_layout.addWidget(table)

            info_label = QLabel(f"Toplam {len(transactions)} işlem listeleniyor.")
            main_layout.addWidget(info_label)

            close_btn = QPushButton("Kapat")
            close_btn.clicked.connect(dialog.accept)
            main_layout.addWidget(close_btn, alignment=Qt.AlignRight)

            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Banka dökümü açılırken hata: {str(e)}")
        finally:
            session.close()
    
    def show_new_loan_dialog(self):
        """Yeni kredi dialog"""
        from src.ui.dialogs.loan_dialog import LoanDialog
        dialog = LoanDialog(self.user.id, self)
        if dialog.exec_():
            self.refresh_loans_table()
    
    def show_edit_loan_dialog(self, loan_id):
        """Kredi düzenleme dialog"""
        from src.ui.dialogs.loan_dialog import LoanDialog
        dialog = LoanDialog(self.user.id, self, loan_id)
        if dialog.exec_():
            self.refresh_loans_table()
    
    def delete_loan(self, loan_id):
        """Kredi sil"""
        from src.services.loan_service import LoanService
        
        reply = QMessageBox.question(self, "Onay", "Bu krediyi silmek istediğinize emin misiniz?",
                                    QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            success, msg = LoanService.delete_loan(loan_id)
            if success:
                QMessageBox.information(self, "Başarılı", "Kredi silindi")
                self.refresh_loans_table()
            else:
                QMessageBox.critical(self, "Hata", msg)

    def delete_loans_bulk(self):
        """Seçili kredileri toplu sil"""
        from src.services.loan_service import LoanService

        selected_rows = self.table_loans.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek istediğiniz kredileri seçin.\n(Ctrl veya Shift ile çoklu seçim yapabilirsiniz.)")
            return

        loan_ids = []
        loan_names = []
        for index in selected_rows:
            row = index.row()
            name_item = self.table_loans.item(row, 0)
            bank_item = self.table_loans.item(row, 1)
            loan_id = name_item.data(Qt.UserRole) if name_item else None
            if loan_id:
                loan_ids.append(loan_id)
                name = name_item.text() if name_item else ""
                bank = bank_item.text() if bank_item else ""
                loan_names.append(f"{name} ({bank})")

        if not loan_ids:
            QMessageBox.warning(self, "Uyarı", "Seçili satırlarda kredi bulunamadı.")
            return

        names_text = "\n".join(loan_names[:10])
        if len(loan_names) > 10:
            names_text += f"\n... ve {len(loan_names) - 10} tane daha"

        reply = QMessageBox.question(
            self, "Toplu Silme Onayı",
            f"{len(loan_ids)} kredi silinecek:\n\n{names_text}\n\nEmin misiniz?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        success_count = 0
        errors = []
        for loan_id in loan_ids:
            ok, msg = LoanService.delete_loan(loan_id)
            if ok:
                success_count += 1
            else:
                errors.append(msg)

        self.refresh_loans_table()
        if errors:
            QMessageBox.warning(self, "Toplu Silme Tamamlandı",
                f"Silinen: {success_count}\nHatalı: {len(errors)}\n\n" + "\n".join(errors[:5]))
        else:
            QMessageBox.information(self, "Başarılı", f"{success_count} kredi başarıyla silindi.")
    
    def create_cari_extract_tab(self) -> QWidget:
        """Cari Ekstre sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        title = QLabel("📑 Cari Hesap Ekstresi")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        layout.addWidget(title)
        
        # Cari seçimi ve arama
        filter_layout = QHBoxLayout()
        
        # Müşteri arama kutusu
        filter_layout.addWidget(QLabel("Müşteri Ara:"))
        self.cari_extract_search_input = QLineEdit()
        self.cari_extract_search_input.setPlaceholderText("Müşteri adı yazın...")
        self.cari_extract_search_input.setMinimumHeight(35)
        self.cari_extract_search_input.setMaximumWidth(200)
        self.cari_extract_search_input.textChanged.connect(self.filter_cari_combo)
        filter_layout.addWidget(self.cari_extract_search_input)
        
        filter_layout.addWidget(QLabel("Cari Hesap:"))
        
        self.cari_extract_combo = QComboBox()
        self.cari_extract_combo.setMinimumHeight(35)
        self.cari_extract_combo.setMinimumWidth(250)
        self.cari_extract_combo.addItem("-- Cari Seçiniz --", None)
        
        # Tüm carileri sakla (filtreleme için)
        self.all_caris = []
        self._reload_cari_extract_combo()
        filter_layout.addWidget(self.cari_extract_combo)
        
        btn_show_extract = QPushButton("📊 Ekstre Göster")
        btn_show_extract.setMinimumHeight(35)
        btn_show_extract.clicked.connect(self.show_cari_extract)
        filter_layout.addWidget(btn_show_extract)
        
        btn_export_excel = QPushButton("📥 Excel'e Aktar")
        btn_export_excel.setMinimumHeight(35)
        btn_export_excel.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
        """)
        btn_export_excel.clicked.connect(self.export_cari_extract_to_excel)
        filter_layout.addWidget(btn_export_excel)

        btn_save_extract = QPushButton("💾 Kaydet")
        btn_save_extract.setMinimumHeight(35)
        btn_save_extract.setToolTip("Sütun genişliklerini kaydet")
        btn_save_extract.setStyleSheet("""
            QPushButton {
                background-color: #FF9800; color: white;
                border: none; border-radius: 4px;
                padding: 8px 14px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)
        btn_save_extract.clicked.connect(lambda: self.save_column_widths(self.table_cari_extract, "cari_extract"))
        filter_layout.addWidget(btn_save_extract)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # Ekstre tablosu
        self.table_cari_extract = QTableWidget()
        self.table_cari_extract.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_cari_extract.setColumnCount(6)
        self.table_cari_extract.setHorizontalHeaderLabels([
            "Tarih", "İşlem Türü", "Açıklama", "Borç", "Alacak", "Bakiye"
        ])
        self.table_cari_extract.horizontalHeader().setStretchLastSection(False)
        self.table_cari_extract.setColumnWidth(0, 100)
        self.table_cari_extract.setColumnWidth(1, 150)
        self.table_cari_extract.setColumnWidth(2, 300)
        self.table_cari_extract.setColumnWidth(3, 120)
        self.table_cari_extract.setColumnWidth(4, 120)
        self.table_cari_extract.setColumnWidth(5, 120)
        self.load_column_widths(self.table_cari_extract, "cari_extract")
        layout.addWidget(self.table_cari_extract)
        
        widget.setLayout(layout)
        return widget
    
    def _reload_cari_extract_combo(self):
        """Cari ekstre combobox verisini veritabanından yenile"""
        try:
            selected_id = self.cari_extract_combo.currentData() if hasattr(self, 'cari_extract_combo') else None
            self.all_caris = CariService.get_caris(self.user.id) or []
            search_text = self.cari_extract_search_input.text() if hasattr(self, 'cari_extract_search_input') else ""
            self.filter_cari_combo(search_text, selected_id)
        except Exception as e:
            print(f"Cari ekstre listesi yenileme hatası: {e}")

    def filter_cari_combo(self, search_text=None, preferred_id=None):
        """Arama kutusuna göre cari combobox'ı filtrele"""
        if search_text is None:
            search_text = self.cari_extract_search_input.text() if hasattr(self, 'cari_extract_search_input') else ""

        search_text = str(search_text).casefold().strip()
        
        # Combobox'ı temizle
        self.cari_extract_combo.clear()
        self.cari_extract_combo.addItem("-- Cari Seçiniz --", None)
        
        # Eğer arama metni boşsa, tüm carileri göster
        if not search_text:
            if self.all_caris:
                for cari in self.all_caris:
                    self.cari_extract_combo.addItem(f"{cari.name} ({cari.cari_type})", cari.id)
        else:
            # Arama metnini içeren carileri göster
            if self.all_caris:
                for cari in self.all_caris:
                    if search_text in (cari.name or "").casefold():
                        self.cari_extract_combo.addItem(f"{cari.name} ({cari.cari_type})", cari.id)

        if preferred_id:
            idx = self.cari_extract_combo.findData(preferred_id)
            if idx >= 0:
                self.cari_extract_combo.setCurrentIndex(idx)
    
    def show_cari_extract(self):
        """Cari ekstre göster"""
        cari_id = self.cari_extract_combo.currentData()
        if not cari_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir cari hesap seçiniz!")
            return
        
        try:
            from src.database.db import SessionLocal
            from src.database.models import Cari, Transaction
            
            session = SessionLocal()
            cari = session.query(Cari).filter(Cari.id == cari_id).first()
            transactions = session.query(Transaction).filter(
                Transaction.cari_id == cari_id
            ).order_by(Transaction.transaction_date.asc()).all()
            session.close()
            
            self.table_cari_extract.setRowCount(len(transactions))
            running_balance = 0
            
            for i, trans in enumerate(transactions):
                self.table_cari_extract.setItem(i, 0, QTableWidgetItem(str(trans.transaction_date)))
                self.table_cari_extract.setItem(i, 1, QTableWidgetItem(trans.transaction_type.value))
                self.table_cari_extract.setItem(i, 2, QTableWidgetItem(trans.description))
                
                # Borç/Alacak hesaplama
                if trans.transaction_type.value in [
                    'GIDER', 'KESILEN_FATURA',
                    'KREDI_ODEME', 'KREDI_KARTI_ODEME',
                    'EK_HESAP_FAIZLERI', 'KREDI_DOSYA_MASRAFI', 'EKSPERTIZ_UCRETI'
                ]:
                    debt = trans.amount
                    credit = 0
                    running_balance += trans.amount
                elif trans.transaction_type.value in [
                    'GELIR', 'GELEN_FATURA', 'KREDI_CEKIMI'
                ]:
                    debt = 0
                    credit = trans.amount
                    running_balance -= trans.amount
                else:
                    debt = 0
                    credit = 0
                
                self.table_cari_extract.setItem(i, 3, QTableWidgetItem(f"{format_tr(debt)} ₺" if debt > 0 else "-"))
                self.table_cari_extract.setItem(i, 4, QTableWidgetItem(f"{format_tr(credit)} ₺" if credit > 0 else "-"))
                self.table_cari_extract.setItem(i, 5, QTableWidgetItem(f"{format_tr(running_balance)} ₺"))

            # En alta TOPLAM / KALAN BAKİYE satırı ekle
            total_row = len(transactions)
            self.table_cari_extract.setRowCount(total_row + 1)
            from PyQt5.QtGui import QFont, QColor
            from PyQt5.QtCore import Qt as _Qt
            bold_font = QFont()
            bold_font.setBold(True)
            bold_font.setPointSize(10)
            summary_bg = QColor("#1565C0")  # koyu mavi
            summary_fg = QColor("#FFFFFF")
            labels = ["", "", "KALAN BAKİYE", "", "", f"{format_tr(running_balance)} ₺"]
            for col, text in enumerate(labels):
                item = QTableWidgetItem(text)
                item.setFont(bold_font)
                item.setBackground(summary_bg)
                item.setForeground(summary_fg)
                item.setTextAlignment(_Qt.AlignCenter)
                self.table_cari_extract.setItem(total_row, col, item)

            self._resize_table(self.table_cari_extract, stretch_col=2)
            self.load_column_widths(self.table_cari_extract, "cari_extract")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ekstre yüklenirken hata: {str(e)}")
    
    def export_cari_extract_to_excel(self):
        """Cari ekstresini Excel'e aktar"""
        cari_id = self.cari_extract_combo.currentData()
        if not cari_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir cari hesap seçiniz!")
            return
        
        # Tabloda veri var mı kontrol et
        if self.table_cari_extract.rowCount() == 0:
            QMessageBox.warning(self, "Uyarı", "Önce 'Ekstre Göster' butonuna basarak verileri yükleyiniz!")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from datetime import datetime
            from src.database.db import SessionLocal
            from src.database.models import Cari
            import os
            
            # Cari bilgisini al
            session = SessionLocal()
            cari = session.query(Cari).filter(Cari.id == cari_id).first()
            cari_name = cari.name if cari else "Bilinmeyen"
            session.close()
            
            # Excel dosyası oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Cari Ekstre"
            
            # Başlık bilgileri
            ws['A1'] = 'CARİ HESAP EKSTRESİ'
            ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A1:F1')
            ws.row_dimensions[1].height = 30
            
            ws['A2'] = f'Cari: {cari_name}'
            ws['A2'].font = Font(size=12, bold=True)
            ws.merge_cells('A2:C2')
            
            ws['D2'] = f'Tarih: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            ws['D2'].font = Font(size=10)
            ws['D2'].alignment = Alignment(horizontal='right')
            ws.merge_cells('D2:F2')
            
            # Boş satır
            ws.row_dimensions[3].height = 5
            
            # Başlıklar
            headers = ['Tarih', 'İşlem Türü', 'Açıklama', 'Borç', 'Alacak', 'Bakiye']
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            # Verileri ekle
            for row in range(self.table_cari_extract.rowCount()):
                excel_row = row + 5
                for col in range(6):
                    item = self.table_cari_extract.item(row, col)
                    if item:
                        cell = ws.cell(row=excel_row, column=col+1)
                        value = item.text()
                        
                        # Sayısal değerleri düzenle
                        if col >= 3:  # Borç, Alacak, Bakiye kolonları
                            value = value.replace(' ₺', '').replace('.', '').replace(',', '.')
                            if value != '-':
                                try:
                                    cell.value = float(value)
                                    cell.number_format = '#,##0.00 "₺"'
                                except:
                                    cell.value = value
                            else:
                                cell.value = '-'
                        else:
                            cell.value = value
                        
                        cell.border = border
                        cell.alignment = Alignment(
                            horizontal='right' if col >= 3 else 'left',
                            vertical='center'
                        )
            
            # Dosya kaydetme dialog
            from PyQt5.QtWidgets import QFileDialog
            
            # Varsayılan dosya adı
            default_filename = f"Cari_Ekstre_{cari_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                default_filename,
                "Excel Dosyası (*.xlsx)"
            )
            
            if file_path:
                wb.save(file_path)
                QMessageBox.information(
                    self, 
                    "Başarılı", 
                    f"Cari ekstre başarıyla Excel'e aktarıldı:\n{file_path}"
                )
                
                # Dosyayı aç
                reply = QMessageBox.question(
                    self,
                    "Dosyayı Aç",
                    "Excel dosyasını şimdi açmak ister misiniz?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    os.startfile(file_path)
        
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kütüphanesi bulunamadı. Lütfen 'pip install openpyxl' komutunu çalıştırın.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel'e aktarılırken hata oluştu:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def create_reports_tab(self) -> QWidget:
        """Gelişmiş Raporlar sekmesi — sol kenar çubuğu + KPI kartları + modern HTML"""
        self._current_report_key = None

        wrapper = QWidget()
        main_layout = QHBoxLayout(wrapper)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ── SOL KENAR ÇUBUĞU ──────────────────────────────────────────────
        sidebar = QWidget()
        sidebar.setMinimumWidth(190)
        sidebar.setMaximumWidth(260)
        sidebar.setStyleSheet("background-color: #1a2332;")
        sb_layout = QVBoxLayout(sidebar)
        sb_layout.setContentsMargins(0, 0, 0, 0)
        sb_layout.setSpacing(0)

        sb_title = QLabel("  📊  RAPORLAR")
        sb_title.setStyleSheet("""
            QLabel {
                background-color: #0d1520;
                color: #90CAF9;
                padding: 14px 10px;
                font-size: 10pt;
                font-weight: bold;
                letter-spacing: 1px;
            }
        """)
        sb_layout.addWidget(sb_title)

        STYLE_NORMAL = """
            QPushButton {
                background-color: transparent;
                color: #90A4AE;
                border: none;
                border-left: 3px solid transparent;
                text-align: left;
                padding: 8px 10px;
                font-size: 9pt;
            }
            QPushButton:hover {
                background-color: #263545;
                color: #E3F2FD;
                border-left: 3px solid #546E7A;
            }
        """
        STYLE_ACTIVE = """
            QPushButton {
                background-color: #0D47A1;
                color: white;
                border: none;
                border-left: 3px solid #42A5F5;
                text-align: left;
                padding: 8px 10px;
                font-size: 9pt;
                font-weight: bold;
            }
        """

        self._report_btn_styles = (STYLE_NORMAL, STYLE_ACTIVE)
        self.report_sidebar_buttons = {}

        report_menu = [
            ("genel",       "🏠",  "Genel Finansal Özet"),
            ("gelir_gider", "💰",  "Gelir-Gider Raporu"),
            ("cari",        "👥",  "Cari Bakiyeleri"),
            ("banka",       "🏦",  "Banka Hesapları"),
            ("nakit_kasasi","💵",  "Nakit Kasası"),
            ("kredi_karti", "💳",  "Kredi Kartları"),
            ("kredi",       "📋",  "Krediler"),
            ("aylik",       "📅",  "Aylık Karşılaştırma"),
            ("top_cari",    "🏆",  "En Aktif Cariler"),
            ("odeme",       "💸",  "Ödeme Dağılımı"),
            ("haftalik",    "📈",  "Haftalık Trend"),
            ("maas",        "👷",  "Maaş Ödemeleri"),
            ("konu_gider",  "🏷️",  "Konuya Göre Giderler"),
            ("kira_takip",    "🏠",  "Kira Takip Raporu"),
            ("fatura_vade",   "📆",  "Fatura Vade Takibi"),
            ("kredi_bitis",   "⏰",  "Kredi Bitiş Sıralaması"),
            ("kredi_bu_ay",   "💳",  "Bu Ay Ödenecek Krediler"),
        ]

        for key, icon, label in report_menu:
            btn = QPushButton(f"  {icon}  {label}")
            btn.setStyleSheet(STYLE_NORMAL)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setMinimumHeight(38)
            btn.clicked.connect(lambda checked=False, k=key: self._generate_sidebar_report(k))
            sb_layout.addWidget(btn)
            self.report_sidebar_buttons[key] = btn

        sb_layout.addStretch()

        # ── SAĞ PANEL ─────────────────────────────────────────────────────
        right = QWidget()
        right.setStyleSheet("background-color: #F0F4F8;")
        r_layout = QVBoxLayout(right)
        r_layout.setContentsMargins(16, 12, 16, 12)
        r_layout.setSpacing(8)

        # Başlık + aksiyon butonları
        top_row = QHBoxLayout()
        self.report_page_title = QLabel("📊  Raporlar ve Analizler")
        self.report_page_title.setFont(QFont("Segoe UI", 15, QFont.Bold))
        self.report_page_title.setStyleSheet("color: #1A237E;")
        top_row.addWidget(self.report_page_title)
        top_row.addStretch()

        btn_refresh = QPushButton("🔄  Yenile")
        btn_refresh.setMinimumHeight(34)
        btn_refresh.setStyleSheet("""
            QPushButton { background-color:#546E7A; color:white; border:none;
                border-radius:4px; padding:6px 14px; font-weight:bold; }
            QPushButton:hover { background-color:#455A64; }
        """)
        btn_refresh.clicked.connect(lambda: self._generate_sidebar_report(self._current_report_key) if self._current_report_key else None)
        top_row.addWidget(btn_refresh)

        btn_print = QPushButton("🖨️  Yazdır")
        btn_print.setMinimumHeight(34)
        btn_print.setStyleSheet("""
            QPushButton { background-color:#37474F; color:white; border:none;
                border-radius:4px; padding:6px 14px; font-weight:bold; }
            QPushButton:hover { background-color:#263238; }
        """)
        btn_print.clicked.connect(self._print_report)
        top_row.addWidget(btn_print)

        btn_export = QPushButton("📥  Excel'e Aktar")
        btn_export.setMinimumHeight(34)
        btn_export.setStyleSheet("""
            QPushButton { background-color:#2E7D32; color:white; border:none;
                border-radius:4px; padding:6px 14px; font-weight:bold; }
            QPushButton:hover { background-color:#1B5E20; }
        """)
        btn_export.clicked.connect(self.export_report_to_excel)
        top_row.addWidget(btn_export)
        r_layout.addLayout(top_row)

        # Filtre çubuğu
        filter_bar = QFrame()
        filter_bar.setStyleSheet("""
            QFrame { background-color:white; border:1px solid #CFD8DC;
                     border-radius:6px; }
        """)
        fb_layout = QHBoxLayout(filter_bar)
        fb_layout.setContentsMargins(10, 6, 10, 6)
        fb_layout.setSpacing(8)

        lbl_s = QLabel("📅 Başlangıç:")
        lbl_s.setStyleSheet("border:none; background:transparent;")
        fb_layout.addWidget(lbl_s)

        self.report_start_date = QDateEdit()
        self.report_start_date.setCalendarPopup(True)
        self.report_start_date.setDate(QDate.currentDate().addMonths(-1))
        self.report_start_date.setDisplayFormat("dd.MM.yyyy")
        self.report_start_date.setMinimumHeight(30)
        fb_layout.addWidget(self.report_start_date)

        lbl_dash = QLabel("—")
        lbl_dash.setStyleSheet("border:none; background:transparent;")
        fb_layout.addWidget(lbl_dash)

        lbl_e = QLabel("Bitiş:")
        lbl_e.setStyleSheet("border:none; background:transparent;")
        fb_layout.addWidget(lbl_e)

        self.report_end_date = QDateEdit()
        self.report_end_date.setCalendarPopup(True)
        self.report_end_date.setDate(QDate.currentDate())
        self.report_end_date.setDisplayFormat("dd.MM.yyyy")
        self.report_end_date.setMinimumHeight(30)
        fb_layout.addWidget(self.report_end_date)

        QBTN = """
            QPushButton { background-color:#E3F2FD; color:#1565C0;
                border:1px solid #90CAF9; border-radius:3px;
                padding:2px 10px; font-size:9pt; }
            QPushButton:hover { background-color:#BBDEFB; }
        """
        for lbl, rtype in [("Bu Ay", "this_month"), ("Geçen Ay", "last_month"), ("Bu Yıl", "this_year")]:
            qb = QPushButton(lbl)
            qb.setMinimumHeight(28)
            qb.setStyleSheet(QBTN)
            qb.clicked.connect(lambda checked=False, rt=rtype: self._set_date_range(rt))
            fb_layout.addWidget(qb)

        # Konu filtresi (sadece konu_gider raporu için görünür)
        self.report_konu_label = QLabel("🏷️ Konu:")
        self.report_konu_label.setStyleSheet("border:none; background:transparent;")
        self.report_konu_label.setVisible(False)
        fb_layout.addWidget(self.report_konu_label)

        self.report_konu_filter = QComboBox()
        self.report_konu_filter.setMinimumHeight(30)
        self.report_konu_filter.setMinimumWidth(180)
        self.report_konu_filter.setVisible(False)
        self.report_konu_filter.currentIndexChanged.connect(
            lambda: self._generate_sidebar_report("konu_gider") if self._current_report_key == "konu_gider" else None
        )
        fb_layout.addWidget(self.report_konu_filter)

        fb_layout.addStretch()
        r_layout.addWidget(filter_bar)

        # Rapor gösterim alanı
        self.report_display = QTextBrowser()
        self.report_display.setStyleSheet("""
            QTextBrowser {
                background-color: white;
                border: 1px solid #CFD8DC;
                border-radius: 6px;
                padding: 4px;
                font-family: Segoe UI, Arial, sans-serif;
                font-size: 10pt;
            }
        """)
        self.report_display.setOpenExternalLinks(False)
        self.report_display.setHtml(self._report_welcome_html())
        r_layout.addWidget(self.report_display)

        main_layout.addWidget(sidebar)
        main_layout.addWidget(right, 1)
        return wrapper
    
    def _set_date_range(self, range_type):
        """Hızlı tarih aralığı ayarlama"""
        today = QDate.currentDate()
        if range_type == 'this_month':
            self.report_start_date.setDate(QDate(today.year(), today.month(), 1))
            self.report_end_date.setDate(today)
        elif range_type == 'last_month':
            last = today.addMonths(-1)
            self.report_start_date.setDate(QDate(last.year(), last.month(), 1))
            self.report_end_date.setDate(QDate(today.year(), today.month(), 1).addDays(-1))
        elif range_type == 'this_year':
            self.report_start_date.setDate(QDate(today.year(), 1, 1))
            self.report_end_date.setDate(today)
        # Aktif raporu yenile
        if self._current_report_key:
            self._generate_sidebar_report(self._current_report_key)

    def _generate_sidebar_report(self, key: str):
        """Kenar çubuğu butonundan rapor oluştur"""
        if not key:
            return
        self._current_report_key = key

        # Buton stillerini güncelle
        NORMAL, ACTIVE = self._report_btn_styles
        for k, btn in self.report_sidebar_buttons.items():
            btn.setStyleSheet(ACTIVE if k == key else NORMAL)

        start_date = self.report_start_date.date().toPyDate()
        end_date   = self.report_end_date.date().toPyDate()
        detail     = 'Özet'

        # Başlık güncelle
        titles = {
            "genel":       "🏠  Genel Finansal Özet",
            "gelir_gider": "💰  Gelir-Gider Raporu",
            "cari":        "👥  Cari Bakiyeleri",
            "banka":       "🏦  Banka Hesapları",
            "nakit_kasasi":"💵  Nakit Kasası",
            "kredi_karti": "💳  Kredi Kartları",
            "kredi":       "📋  Krediler",
            "aylik":       "📅  Aylık Karşılaştırma",
            "top_cari":    "🏆  En Aktif Cariler",
            "odeme":       "💸  Ödeme Dağılımı",
            "haftalik":    "📈  Haftalık Trend",
            "maas":        "👷  Maaş Ödemeleri",
            "konu_gider":  "🏷️  Konuya Göre Giderler",
            "kira_takip": "🏠  Kira Takip Raporu",
            "kredi_bitis": "⏰  Kredi Bitiş Sıralaması",
            "kredi_bu_ay": "💳  Bu Ay Ödenecek Krediler",
        }
        self.report_page_title.setText(titles.get(key, "📊  Raporlar"))

        # Konu filtresi görünürlüğü
        if hasattr(self, 'report_konu_filter'):
            is_konu = key == "konu_gider"
            self.report_konu_label.setVisible(is_konu)
            self.report_konu_filter.setVisible(is_konu)

        self.report_display.setHtml("<p style='color:#666; padding:20px;'>⏳ Rapor oluşturuluyor...</p>")

        try:
            if key == "genel":
                data = ReportService.generate_comprehensive_report(self.user.id)
                html = self._format_comprehensive_report(data)
            elif key == "gelir_gider":
                data = ReportService.generate_income_expense_report(self.user.id)
                html = self._format_income_expense_report(data, detail, start_date, end_date)
            elif key == "cari":
                data = ReportService.generate_cari_balance_report(self.user.id)
                html = self._format_cari_balance_report(data)
            elif key == "banka":
                data = ReportService.generate_bank_summary_report(self.user.id)
                html = self._format_bank_summary_report(data)
            elif key == "nakit_kasasi":
                html = self._generate_nakit_kasasi_report(start_date, end_date)
            elif key == "kredi_karti":
                data = ReportService.generate_credit_card_summary(self.user.id)
                html = self._format_credit_card_report(data)
            elif key == "kredi":
                data = ReportService.generate_loan_summary_report(self.user.id)
                html = self._format_loan_summary_report(data)
            elif key == "aylik":
                html = self._generate_monthly_comparison_report(start_date, end_date)
            elif key == "top_cari":
                html = self._generate_top_caris_report(start_date, end_date, detail)
            elif key == "odeme":
                html = self._generate_payment_method_distribution(start_date, end_date)
            elif key == "haftalik":
                html = self._generate_weekly_trend_report(start_date, end_date)
            elif key == "maas":
                html = self._generate_payroll_report(start_date, end_date)
            elif key == "konu_gider":
                html = self._generate_konu_gider_report(start_date, end_date)
            elif key == "kira_takip":
                html = self._generate_kira_takip_report()
            elif key == "fatura_vade":
                html = self._generate_fatura_vade_report()
            elif key == "kredi_bitis":
                html = self._generate_kredi_bitis_report()
            elif key == "kredi_bu_ay":
                html = self._generate_kredi_bu_ay_report()
            else:
                html = "<p>Bilinmeyen rapor türü.</p>"
            self.report_display.setHtml(html)
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulurken hata:\n{str(e)}")
            traceback.print_exc()

    # Geriye uyumluluk için eski generate_report metodu da çalışsın
    def generate_report(self):
        report_type = getattr(self, "report_type_combo", None)
        if report_type:
            key_map = {
                "Kapsamlı Genel Rapor":          "genel",
                "Gelir-Gider Raporu":             "gelir_gider",
                "Cari Bakiye Raporu":             "cari",
                "Banka Özet Raporu":              "banka",
                "Kredi Kartı Özet Raporu":        "kredi_karti",
                "Kredi Özet Raporu":              "kredi",
                "Aylık Karşılaştırma Raporu":     "aylik",
                "En Çok İşlem Yapılan Cariler":   "top_cari",
                "Ödeme Yöntemi Dağılımı":         "odeme",
                "Haftalık Trend Analizi":         "haftalik",
            }
            self._generate_sidebar_report(key_map.get(report_type.currentText(), "genel"))

    def _print_report(self):
        """Raporu yaz"""
        try:
            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                self.report_display.print_(printer)
        except Exception as e:
            QMessageBox.warning(self, "Yazdırma Hatası", str(e))

    def _report_welcome_html(self) -> str:
        return """
        <div style='padding:40px; text-align:center;'>
            <p style='font-size:48pt;'>📊</p>
            <p style='font-size:18pt; font-weight:bold; color:#1A237E;'>Raporlar ve Analizler</p>
            <p style='font-size:11pt; color:#546E7A;'>
                Sol menüden bir rapor türü seçerek başlayın.<br><br>
                📅 Tarih aralığını ayarlayabilir,<br>
                🖨️ Yazdırabilir veya<br>
                📥 Excel'e aktarabilirsiniz.
            </p>
        </div>
        """

    def _rh(self, icon: str, title: str, subtitle: str, accent: str = "#1A237E") -> str:
        """Rapor başlık HTML bloğu"""
        from datetime import datetime
        now = datetime.now().strftime("%d.%m.%Y %H:%M")
        return f"""
        <table width='100%' cellpadding='0' cellspacing='0'>
        <tr><td style='background-color:{accent}; padding:18px 20px; border-radius:6px;'>
            <span style='font-size:22pt;'>{icon}</span>
            <span style='font-size:14pt; font-weight:bold; color:white; margin-left:10px;'>{title}</span><br>
            <span style='font-size:9pt; color:#B0BEC5;'>{subtitle} &nbsp;|&nbsp; Oluşturulma: {now}</span>
        </td></tr>
        </table>
        <br>
        """

    def _kpi_row(self, items: list) -> str:
        """KPI kartları satırı. items = [(ikon, başlık, değer, renk), ...]"""
        cols = "".join(f"""
            <td width='{100//len(items)}%' style='padding:4px;'>
                <table width='100%' cellpadding='10' cellspacing='0'
                       style='background-color:{color}; border-radius:6px;'>
                <tr><td align='center'>
                    <div style='font-size:18pt;'>{icon}</div>
                    <div style='font-size:8.5pt; color:#CFD8DC; font-weight:bold;'>{label}</div>
                    <div style='font-size:13pt; font-weight:bold; color:white; margin-top:2px;'>{value}</div>
                </td></tr>
                </table>
            </td>
        """ for icon, label, value, color in items)
        return f"<table width='100%' cellpadding='0' cellspacing='0'><tr>{cols}</tr></table><br>"

    def _section(self, title: str, color: str = "#1565C0") -> str:
        return f"""
        <table width='100%' cellpadding='0' cellspacing='0' style='margin-top:12px;'>
        <tr><td style='background-color:{color}; padding:8px 14px; border-radius:4px;'>
            <span style='font-size:11pt; font-weight:bold; color:white;'>{title}</span>
        </td></tr></table>
        """

    def _table_header(self, cols: list, color: str = "#E3F2FD") -> str:
        cells = "".join(f"<th style='padding:8px 10px; background-color:{color}; "
                        f"border-bottom:2px solid #90CAF9; text-align:left; font-size:9pt;'>{c}</th>" for c in cols)
        return f"<table width='100%' cellpadding='0' cellspacing='0' style='border-collapse:collapse; margin-top:4px;'><tr>{cells}</tr>"

    def _tr(self, values: list, colors: list = None, bold: bool = False, bg: str = None) -> str:
        w = "font-weight:bold;" if bold else ""
        bgstyle = f"background-color:{bg};" if bg else ""
        cells = ""
        for i, v in enumerate(values):
            c = f"color:{colors[i]};" if colors and i < len(colors) and colors[i] else ""
            cells += f"<td style='padding:7px 10px; border-bottom:1px solid #ECEFF1; {w}{c}{bgstyle} font-size:9.5pt;'>{v}</td>"
        return f"<tr>{cells}</tr>"

    def _progress_bar_html(self, pct: float) -> str:
        """Metin tabanlı ilerleme çubuğu Qt HTML için"""
        pct = max(0, min(100, pct))
        filled = int(pct / 5)
        empty  = 20 - filled
        color = "#EF5350" if pct > 80 else "#FFA726" if pct > 50 else "#66BB6A"
        bar = "<span style='color:{};'>&#9608;</span>".format(color) * filled + \
              "<span style='color:#E0E0E0;'>&#9608;</span>" * empty
        return f"{bar} <span style='font-size:9pt; color:#546E7A;'>{pct:.0f}%</span>"

    def _format_comprehensive_report(self, data):
        """Genel özet raporu — modern HTML"""
        ie  = data['income_expense']
        cb  = data['cari_balance']
        bs  = data['bank_summary']
        cc  = data['credit_card_summary']
        ofh = data['overall_financial_health']

        net_profit    = ie['net_profit']
        net_worth     = ofh['net_worth']
        np_color      = "#2E7D32" if net_profit >= 0 else "#C62828"
        nw_color      = "#2E7D32" if net_worth  >= 0 else "#C62828"

        html  = self._rh("🏠", "Genel Finansal Özet", "Tüm hesapların anlık durumu", "#1A237E")

        html += self._kpi_row([
            ("💰", "Toplam Gelir",    f"{ie['total_income']:,.0f} ₺",    "#1B5E20"),
            ("📉", "Toplam Gider",    f"{ie['total_expense']:,.0f} ₺",   "#B71C1C"),
            ("📊", "Net Kar/Zarar",   f"{net_profit:,.0f} ₺",            np_color if net_profit>=0 else "#C62828"),
            ("🏦", "Banka Bakiyesi",  f"{bs['total_balance_try']:,.0f} ₺","#0D47A1"),
        ])

        # Genel mali durum
        html += self._section("💼 Genel Mali Durum", "#1A237E")
        html += self._table_header(["Kalem", "Tutar"])
        rows = [
            ("Likit Varlıklar (Banka)",  f"{ofh['liquid_assets']:,.2f} ₺",  "#1B5E20"),
            ("Alacaklar (Cari)",         f"{ofh['receivables']:,.2f} ₺",    "#1B5E20"),
            ("Ödenecek Borçlar",         f"-{ofh['payables']:,.2f} ₺",      "#C62828"),
            ("Kredi Kartı Borcu",        f"-{ofh['credit_card_debt']:,.2f} ₺","#C62828"),
        ]
        for i, (lbl, val, vc) in enumerate(rows):
            bg = "#FAFAFA" if i % 2 else "white"
            html += self._tr([lbl, val], [None, vc], bg=bg)
        # Net değer satırı
        html += self._tr(
            ["NET DEĞER", f"{net_worth:,.2f} ₺"],
            [None, nw_color], bold=True, bg="#E8EAF6"
        )
        html += "</table><br>"

        # Gelir-Gider özeti
        html += self._section("💰 Gelir-Gider Özeti", "#1B5E20")
        html += self._table_header(["Kalem", "Tutar"])
        html += self._tr(["Toplam Gelir",   f"{ie['total_income']:,.2f} ₺"],  [None, "#1B5E20"], bg="white")
        html += self._tr(["Toplam Gider",   f"{ie['total_expense']:,.2f} ₺"], [None, "#C62828"], bg="#FAFAFA")
        html += self._tr(["Net Kar/Zarar",  f"{net_profit:,.2f} ₺"],          [None, np_color],  bold=True, bg="#E8F5E9")
        html += self._tr(["Toplam İşlem",   str(ie['transaction_count'])],     [None, None],      bg="white")
        html += "</table><br>"

        # Cari hesaplar
        html += self._section("👥 Cari Hesaplar", "#E65100")
        html += self._table_header(["Kalem", "Tutar"])
        html += self._tr(["Toplam Alacak", f"{cb['total_receivable']:,.2f} ₺"], [None, "#1B5E20"], bg="white")
        html += self._tr(["Toplam Borç",   f"{cb['total_payable']:,.2f} ₺"],   [None, "#C62828"], bg="#FAFAFA")
        html += self._tr(["Net Bakiye",    f"{cb['net_balance']:,.2f} ₺"],     [None, None],      bold=True, bg="#FFF3E0")
        html += "</table><br>"

        # Kredi kartları
        html += self._section("💳 Kredi Kartları", "#4A148C")
        html += self._table_header(["Kalem", "Tutar"])
        usage = cc['overall_usage_rate'] if 'overall_usage_rate' in cc else 0
        html += self._tr(["Toplam Limit",         f"{cc['total_limit']:,.2f} ₺"],     [None, None],      bg="white")
        html += self._tr(["Toplam Borç",           f"{cc['total_debt']:,.2f} ₺"],      [None, "#C62828"],  bg="#FAFAFA")
        html += self._tr(["Kullanılabilir Limit",  f"{cc['total_available']:,.2f} ₺"], [None, "#1B5E20"],  bg="white")
        html += self._tr(["Kullanım Oranı",        f"{usage:.1f}%"],                   [None, None],      bold=True, bg="#F3E5F5")
        html += "</table>"
        return html

    def _format_income_expense_report(self, data, detail_level='Özet', start_date=None, end_date=None):
        """Gelir-Gider raporu — modern HTML"""
        net      = data['net_profit']
        nc       = "#2E7D32" if net >= 0 else "#C62828"
        donem    = f"{data['period']['start']} — {data['period']['end']}"
        subtitle = f"Dönem: {donem}"
        if start_date and end_date:
            subtitle += f"  |  Seçili: {start_date} – {end_date}"

        html  = self._rh("💰", "Gelir-Gider Raporu", subtitle, "#1B5E20")
        html += self._kpi_row([
            ("➕", "Toplam Gelir",   f"{data['total_income']:,.0f} ₺",   "#1B5E20"),
            ("➖", "Toplam Gider",   f"{data['total_expense']:,.0f} ₺",  "#B71C1C"),
            ("📊", "Net Kar/Zarar",  f"{net:,.0f} ₺",                    "#0D47A1" if net >= 0 else "#B71C1C"),
            ("🔢", "İşlem Sayısı",   str(data['transaction_count']),     "#37474F"),
        ])

        html += self._section("📋 Özet", "#1B5E20")
        html += self._table_header(["Kalem", "Tutar", "Oran"])
        total = data['total_income'] + data['total_expense'] or 1
        html += self._tr(["Toplam Gelir",  f"{data['total_income']:,.2f} ₺",  f"{data['total_income']/total*100:.1f}%"],  [None, "#1B5E20", None], bg="white")
        html += self._tr(["Toplam Gider",  f"{data['total_expense']:,.2f} ₺", f"{data['total_expense']/total*100:.1f}%"], [None, "#C62828", None], bg="#FAFAFA")
        html += self._tr(["Net Kar/Zarar", f"{net:,.2f} ₺",                   ""],                                        [None, nc, None],        bold=True, bg="#E8F5E9")
        html += "</table><br>"

        if detail_level == 'Detaylı':
            html += self._section("📄 Son 100 İşlem", "#1B5E20")
            html += self._table_header(["Tarih", "Tür", "Müşteri", "Açıklama", "Tutar"])
            from src.database.db import SessionLocal
            session = SessionLocal()
            try:
                txs = session.query(Transaction).filter(
                    Transaction.user_id == self.user.id
                ).order_by(Transaction.transaction_date.desc()).limit(100).all()
                for i, t in enumerate(txs):
                    income_types = [TransactionType.GELIR, TransactionType.KESILEN_FATURA]
                    vc = "#1B5E20" if t.transaction_type in income_types else "#C62828"
                    icon = "➕" if t.transaction_type in income_types else "➖"
                    bg  = "white" if i % 2 == 0 else "#FAFAFA"
                    html += self._tr([
                        str(t.transaction_date)[:10],
                        f"{icon} {t.transaction_type.value}",
                        t.customer_name or "—",
                        (t.description or "")[:50],
                        f"{t.amount:,.2f} ₺"
                    ], [None, vc, None, None, vc], bg=bg)
            finally:
                session.close()
            html += "</table>"
        return html

    def _format_cari_balance_report(self, data):
        """Cari bakiye raporu — modern HTML"""
        nb = data['net_balance']
        nc = "#2E7D32" if nb >= 0 else "#C62828"

        html  = self._rh("👥", "Cari Bakiye Raporu", f"Toplam {data['total_caris']} cari hesap", "#E65100")
        html += self._kpi_row([
            ("👥", "Toplam Cari",    str(data['total_caris']),                "#37474F"),
            ("📥", "Toplam Alacak",  f"{data['total_receivable']:,.0f} ₺",   "#1B5E20"),
            ("📤", "Toplam Borç",    f"{data['total_payable']:,.0f} ₺",      "#B71C1C"),
            ("⚖️", "Net Bakiye",    f"{nb:,.0f} ₺",                          "#0D47A1" if nb >= 0 else "#B71C1C"),
        ])

        html += self._section("📋 Cari Hesap Detayları", "#E65100")
        html += self._table_header(["Cari Adı", "Tür", "Bakiye", "Durum"])
        for i, cari in enumerate(data['caris']):
            bal = cari['balance']
            vc  = "#1B5E20" if bal > 0 else "#C62828" if bal < 0 else "#546E7A"
            bg  = "white" if i % 2 == 0 else "#FAFAFA"
            html += self._tr([
                cari['name'], cari['type'],
                f"{bal:,.2f} ₺", cari['status']
            ], [None, None, vc, None], bg=bg)
        html += "</table>"
        return html

    def _format_bank_summary_report(self, data):
        """Banka özet raporu — modern HTML"""
        html  = self._rh("🏦", "Banka Hesapları Raporu", f"Toplam {data['total_accounts']} hesap", "#0D47A1")
        html += self._kpi_row([
            ("🏦", "Toplam Hesap",      str(data['total_accounts']),                 "#37474F"),
            ("💵", "Toplam Bakiye TRY", f"{data['total_balance_try']:,.0f} ₺",       "#1B5E20"),
        ])
        html += self._section("🏦 Hesap Detayları", "#0D47A1")
        html += self._table_header(["Banka", "Hesap No", "Bakiye", "Para Birimi"])
        for i, bank in enumerate(data['banks']):
            bg = "white" if i % 2 == 0 else "#FAFAFA"
            vc = "#1B5E20" if bank['balance'] >= 0 else "#C62828"
            html += self._tr([
                bank['bank_name'], bank['account_number'],
                f"{bank['balance']:,.2f}", bank['currency']
            ], [None, None, vc, None], bg=bg)
        html += "</table>"
        return html

    def _format_credit_card_report(self, data):
        """Kredi kartı raporu — modern HTML"""
        usage = data.get('overall_usage_rate', 0)
        html  = self._rh("💳", "Kredi Kartları Raporu", f"Toplam {data['total_cards']} kart  |  Kullanım: {usage:.1f}%", "#4A148C")
        html += self._kpi_row([
            ("💳", "Toplam Kart",       str(data['total_cards']),                "#37474F"),
            ("🔢", "Toplam Limit",      f"{data['total_limit']:,.0f} ₺",        "#1565C0"),
            ("💸", "Toplam Borç",       f"{data['total_debt']:,.0f} ₺",         "#B71C1C"),
            ("✅", "Kullanılabilir",    f"{data['total_available']:,.0f} ₺",    "#1B5E20"),
        ])
        html += self._section("💳 Kart Detayları", "#4A148C")
        html += self._table_header(["Kart Adı", "Banka", "Limit", "Borç", "Kullanılabilir", "Kullanım"])
        for i, card in enumerate(data['cards']):
            bg = "white" if i % 2 == 0 else "#FAFAFA"
            ur = card['usage_rate']
            uc = "#C62828" if ur > 80 else "#E65100" if ur > 50 else "#2E7D32"
            html += self._tr([
                card['card_name'], card['bank'],
                f"{card['limit']:,.2f} ₺",
                f"{card['debt']:,.2f} ₺",
                f"{card['available']:,.2f} ₺",
                self._progress_bar_html(ur)
            ], [None, None, None, "#C62828", "#1B5E20", None], bg=bg)
        html += "</table>"
        return html

    def _format_loan_summary_report(self, data):
        """Kredi özet raporu — modern HTML"""
        html  = self._rh("📋", "Krediler Raporu",
                         f"Toplam {data['total_loans']} kredi  |  Aktif: {data['active_loans']}",
                         "#1A237E")
        html += self._kpi_row([
            ("📋", "Toplam Kredi",   str(data['total_loans']),                  "#37474F"),
            ("💰", "Toplam Tutar",   f"{data['total_loan_amount']:,.0f} ₺",    "#0D47A1"),
            ("✅", "Toplam Ödenen",  f"{data['total_paid']:,.0f} ₺",           "#1B5E20"),
            ("⏳", "Kalan Borç",     f"{data['total_remaining']:,.0f} ₺",      "#B71C1C"),
        ])
        html += self._section("📋 Kredi Detayları", "#1A237E")
        html += self._table_header(["Kredi", "Banka", "Tip", "Toplam", "Ödenen", "Kalan", "Durum", "İlerleme"])
        for i, loan in enumerate(data['loans']):
            bg = "white" if i % 2 == 0 else "#FAFAFA"
            pr = loan['progress_rate']
            html += self._tr([
                loan['loan_name'], loan['bank_name'], loan['loan_type'],
                f"{loan['loan_amount']:,.2f} ₺",
                f"{loan['total_paid']:,.2f} ₺",
                f"{loan['remaining_balance']:,.2f} ₺",
                loan['status'],
                self._progress_bar_html(pr)
            ], [None, None, None, None, "#1B5E20", "#C62828", None, None], bg=bg)
        html += "</table>"
        return html

    def export_report_to_excel(self):
        """Raporu Excel'e aktar"""
        _key_map = {
            "genel": "Kapsamlı Genel Rapor",
            "gelir_gider": "Gelir-Gider Raporu",
            "cari": "Cari Bakiye Raporu",
            "banka": "Banka Özet Raporu",
            "kredi_karti": "Kredi Kartı Özet Raporu",
            "kredi": "Kredi Özet Raporu",
            "maas": "Maaş Ödemeleri Raporu",
        }
        report_type = _key_map.get(getattr(self, '_current_report_key', None), "")
        if not report_type:
            QMessageBox.warning(self, "Uyarı", "Önce sol menüden bir rapor türü seçin!")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime
            from PyQt5.QtWidgets import QFileDialog
            import os
            
            # Rapor verisini al
            if report_type == "Kapsamlı Genel Rapor":
                data = ReportService.generate_comprehensive_report(self.user.id)
                filename_prefix = "Kapsamli_Genel_Rapor"
            elif report_type == "Gelir-Gider Raporu":
                data = ReportService.generate_income_expense_report(self.user.id)
                filename_prefix = "Gelir_Gider_Raporu"
            elif report_type == "Cari Bakiye Raporu":
                data = ReportService.generate_cari_balance_report(self.user.id)
                filename_prefix = "Cari_Bakiye_Raporu"
            elif report_type == "Banka Özet Raporu":
                data = ReportService.generate_bank_summary_report(self.user.id)
                filename_prefix = "Banka_Ozet_Raporu"
            elif report_type == "Kredi Kartı Özet Raporu":
                data = ReportService.generate_credit_card_summary(self.user.id)
                filename_prefix = "Kredi_Karti_Raporu"
            elif report_type == "Kredi Özet Raporu":
                data = ReportService.generate_loan_summary_report(self.user.id)
                filename_prefix = "Kredi_Ozet_Raporu"
            elif report_type == "Maaş Ödemeleri Raporu":
                self._export_payroll_report_excel()
                return
            else:
                QMessageBox.warning(self, "Uyarı", "Önce bir rapor türü seçiniz!")
                return
            
            # Excel oluştur
            wb = Workbook()
            ws = wb.active
            
            # Stil tanımlamaları
            header_font = Font(size=14, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            title_font = Font(size=12, bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Rapor türüne göre Excel formatı
            if report_type == "Kapsamlı Genel Rapor":
                self._export_comprehensive_report_excel(ws, data, header_font, header_fill, title_font, border)
            elif report_type == "Gelir-Gider Raporu":
                self._export_income_expense_excel(ws, data, header_font, header_fill, title_font, border)
            elif report_type == "Cari Bakiye Raporu":
                self._export_cari_balance_excel(ws, data, header_font, header_fill, title_font, border)
            elif report_type == "Banka Özet Raporu":
                self._export_bank_summary_excel(ws, data, header_font, header_fill, title_font, border)
            elif report_type == "Kredi Kartı Özet Raporu":
                self._export_credit_card_excel(ws, data, header_font, header_fill, title_font, border)
            elif report_type == "Kredi Özet Raporu":
                self._export_loan_summary_excel(ws, data, header_font, header_fill, title_font, border)
            
            # Dosya kaydetme dialog
            default_filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                default_filename,
                "Excel Dosyası (*.xlsx)"
            )
            
            if file_path:
                wb.save(file_path)
                QMessageBox.information(
                    self,
                    "Başarılı",
                    f"Rapor başarıyla Excel'e aktarıldı:\n{file_path}"
                )
                
                # Dosyayı aç
                reply = QMessageBox.question(
                    self,
                    "Dosyayı Aç",
                    "Excel dosyasını şimdi açmak ister misiniz?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    os.startfile(file_path)
        
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kütüphanesi bulunamadı. Lütfen 'pip install openpyxl' komutunu çalıştırın.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel'e aktarılırken hata oluştu:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def _export_comprehensive_report_excel(self, ws, data, header_font, header_fill, title_font, border):
        """Kapsamlı rapor Excel export"""
        ws.title = "Kapsamlı Rapor"
        
        # Başlık
        ws['A1'] = 'KAPSAMLI GENEL RAPOR'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = f'Rapor Tarihi: {data["report_date"][:19]}'
        ws['A2'].font = Font(size=10)
        ws.merge_cells('A2:D2')
        
        row = 4
        
        # Genel Mali Durum
        ws[f'A{row}'] = 'GENEL MALİ DURUM'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        health = data['overall_financial_health']
        items = [
            ('Likid Varlıklar (Banka)', health['liquid_assets']),
            ('Alacaklar', health['receivables']),
            ('Borçlar', health['payables']),
            ('Kredi Kartı Borcu', health['credit_card_debt']),
            ('NET DEĞER', health['net_worth'])
        ]
        
        for label, value in items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = format_currency_tr(value)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            if label == 'NET DEĞER':
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True, size=12)
                ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            row += 1
        
        row += 1
        
        # Gelir-Gider
        ws[f'A{row}'] = 'GELİR-GİDER ÖZETİ'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ie = data['income_expense']
        ws[f'A{row}'] = 'Toplam Gelir'
        ws[f'B{row}'] = f"{ie['total_income']:,.2f} ₺"
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
        
        ws[f'A{row}'] = 'Toplam Gider'
        ws[f'B{row}'] = f"{ie['total_expense']:,.2f} ₺"
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
        
        ws[f'A{row}'] = 'Net Kar/Zarar'
        ws[f'B{row}'] = f"{ie['net_profit']:,.2f} ₺"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _export_income_expense_excel(self, ws, data, header_font, header_fill, title_font, border):
        """Gelir-Gider raporu Excel export"""
        ws.title = "Gelir-Gider"
        
        ws['A1'] = 'GELİR-GİDER RAPORU'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:C1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = f'Dönem: {data["period"]["start"]} - {data["period"]["end"]}'
        ws.merge_cells('A2:C2')
        
        row = 4
        headers = [('Toplam Gelir', data['total_income']),
                   ('Toplam Gider', data['total_expense']),
                   ('Net Kar/Zarar', data['net_profit']),
                   ('İşlem Sayısı', data['transaction_count'])]
        
        for label, value in headers:
            ws[f'A{row}'] = label
            if isinstance(value, (int, float)) and label != 'İşlem Sayısı':
                ws[f'B{row}'] = format_currency_tr(value)
            else:
                ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            if label == 'Net Kar/Zarar':
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws[f'B{row}'].font = Font(bold=True, s=12)
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _export_cari_balance_excel(self, ws, data, header_font, header_fill, title_font, border):
        """Cari bakiye raporu Excel export"""
        ws.title = "Cari Bakiye"
        
        ws['A1'] = 'CARİ BAKİYE RAPORU'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        row = 3
        ws[f'A{row}'] = 'Toplam Cari Sayısı'
        ws[f'B{row}'] = data['total_caris']
        row += 1
        ws[f'A{row}'] = 'Toplam Alacak'
        ws[f'B{row}'] = f"{data['total_receivable']:,.2f} ₺"
        ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        row += 1
        ws[f'A{row}'] = 'Toplam Borç'
        ws[f'B{row}'] = f"{data['total_payable']:,.2f} ₺"
        ws[f'B{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        row += 1
        ws[f'A{row}'] = 'Net Bakiye'
        ws[f'B{row}'] = f"{data['net_balance']:,.2f} ₺"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        
        row += 2
        ws[f'A{row}'] = 'CARİ DETAYLARI'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Başlıklar
        headers = ['Cari Adı', 'Tür', 'Bakiye', 'Durum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = border
        row += 1
        
        # Veriler
        for cari in data['caris']:
            ws[f'A{row}'] = cari['name']
            ws[f'B{row}'] = cari['type']
            ws[f'C{row}'] = f"{cari['balance']:,.2f} ₺"
            ws[f'D{row}'] = cari['status']
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            
            if cari['balance'] > 0:
                ws[f'C{row}'].font = Font(color='006100')
            elif cari['balance'] < 0:
                ws[f'C{row}'].font = Font(color='9C0006')
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
    
    def _export_bank_summary_excel(self, ws, data, header_font, header_fill, title_font, border):
        """Banka özet raporu Excel export"""
        ws.title = "Banka Özeti"
        
        ws['A1'] = 'BANKA ÖZET RAPORU'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        row = 3
        ws[f'A{row}'] = 'Toplam Hesap Sayısı'
        ws[f'B{row}'] = data['total_accounts']
        row += 1
        ws[f'A{row}'] = 'Toplam Bakiye (TRY)'
        ws[f'B{row}'] = f"{data['total_balance_try']:,.2f} ₺"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        row += 2
        ws[f'A{row}'] = 'HESAP DETAYLARI'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ['Banka Adı', 'Hesap No', 'Bakiye', 'Para Birimi']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = border
        row += 1
        
        for bank in data['banks']:
            ws[f'A{row}'] = bank['bank_name']
            ws[f'B{row}'] = bank['account_number']
            ws[f'C{row}'] = f"{bank['balance']:,.2f}"
            ws[f'D{row}'] = bank['currency']
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
    
    def _export_credit_card_excel(self, ws, data, header_font, header_fill, title_font, border):
        """Kredi kartı raporu Excel export"""
        ws.title = "Kredi Kartları"
        
        ws['A1'] = 'KREDİ KARTI ÖZET RAPORU'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        row = 3
        summary = [
            ('Toplam Kart Sayısı', data['total_cards'], None),
            ('Toplam Limit', f"{data['total_limit']:,.2f} ₺", None),
            ('Toplam Borç', f"{data['total_debt']:,.2f} ₺", 'FFC7CE'),
            ('Kullanılabilir Limit', f"{data['total_available']:,.2f} ₺", 'C6EFCE'),
            ('Genel Kullanım Oranı', f"{data['overall_usage_rate']:.1f}%", None)
        ]
        
        for label, value, fill in summary:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            if fill:
                ws[f'B{row}'].fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
            if 'Toplam Borç' in label or 'Kullanılabilir' in label:
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        ws[f'A{row}'] = 'KART DETAYLARI'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        headers = ['Kart Adı', 'Banka', 'Limit', 'Borç', 'Kullanılabilir', 'Kullanım %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = border
        row += 1
        
        for card in data['cards']:
            ws[f'A{row}'] = card['card_name']
            ws[f'B{row}'] = card['bank']
            ws[f'C{row}'] = f"{card['limit']:,.2f} ₺"
            ws[f'D{row}'] = f"{card['debt']:,.2f} ₺"
            ws[f'E{row}'] = f"{card['available']:,.2f} ₺"
            ws[f'F{row}'] = f"{card['usage_rate']:.1f}%"
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
            
            # Kullanım oranına göre renklendirme
            if card['usage_rate'] > 80:
                ws[f'F{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif card['usage_rate'] > 50:
                ws[f'F{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12

    def _export_loan_summary_excel(self, ws, data, header_font, header_fill, title_font, border):
        """Kredi özet raporu Excel export"""
        ws.title = "Kredi Özeti"

        ws['A1'] = 'KREDİ ÖZET RAPORU'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30

        row = 3
        summary = [
            ('Toplam Kredi Sayısı', data['total_loans'], None),
            ('Toplam Kredi Tutarı', f"{data['total_loan_amount']:,.2f} ₺", None),
            ('Toplam Ödenen', f"{data['total_paid']:,.2f} ₺", 'C6EFCE'),
            ('Toplam Kalan Borç', f"{data['total_remaining']:,.2f} ₺", 'FFC7CE'),
            ('Aktif Kredi Sayısı', data['active_loans'], None)
        ]

        for label, value, fill in summary:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            if fill:
                ws[f'B{row}'].fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
            row += 1

        row += 1
        ws[f'A{row}'] = 'KREDİ DETAYLARI'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        headers = ['Kredi', 'Banka', 'Tip', 'Toplam', 'Ödenen', 'Kalan', 'Durum', 'İlerleme %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = border
        row += 1

        for loan in data['loans']:
            ws[f'A{row}'] = loan['loan_name']
            ws[f'B{row}'] = loan['bank_name']
            ws[f'C{row}'] = loan['loan_type']
            ws[f'D{row}'] = f"{loan['loan_amount']:,.2f} ₺"
            ws[f'E{row}'] = f"{loan['total_paid']:,.2f} ₺"
            ws[f'F{row}'] = f"{loan['remaining_balance']:,.2f} ₺"
            ws[f'G{row}'] = loan['status']
            ws[f'H{row}'] = f"{loan['progress_rate']:.1f}%"

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border

            row += 1

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12

    def _export_payroll_report_excel(self):
        """Maaş ödemeleri raporunu Excel'e aktar"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from PyQt5.QtWidgets import QFileDialog
            import os

            start_date = self.report_start_date.date().toPyDate()
            end_date   = self.report_end_date.date().toPyDate()

            all_records = self._load_saved_payroll_records()
            from datetime import date as _date
            filtered = []
            for r in all_records:
                try:
                    rd = _date(int(r.get('year', 0)), int(r.get('month', 0)), 1)
                    if start_date <= rd <= end_date:
                        filtered.append(r)
                except Exception:
                    continue

            wb = Workbook()
            ws = wb.active
            ws.title = "Maaş Ödemeleri"

            h_font  = Font(size=11, bold=True, color='FFFFFF')
            h_fill  = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
            border  = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'),  bottom=Side(style='thin'))

            ws['A1'] = 'MAAŞ ÖDEMELERİ RAPORU'
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:H1')
            ws.row_dimensions[1].height = 28

            ws['A2'] = f'Dönem: {start_date} — {end_date}'
            ws['A2'].font = Font(size=10)
            ws.merge_cells('A2:H2')

            headers = ['Dönem', 'Çalışan', 'Brüt (₺)', 'Net (₺)',
                       'SGK Kesinti (₺)', 'Gelir Vergisi (₺)', 'Damga Vergisi (₺)', 'Toplam Kesinti (₺)']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = h_font
                cell.fill = h_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            row = 5
            for r in sorted(filtered,
                    key=lambda x: (int(x.get('year', 0)), int(x.get('month', 0)),
                                   str(x.get('employee', '')))):
                donem = f"{int(r.get('month', 0)):02d}/{r.get('year', '')}"
                vals  = [donem,
                         r.get('employee', ''),
                         float(r.get('gross_total', 0) or 0),
                         float(r.get('net_salary', 0) or 0),
                         float(r.get('sgk_deduction', 0) or 0),
                         float(r.get('income_tax', 0) or 0),
                         float(r.get('stamp_tax', 0) or 0),
                         float(r.get('total_deductions', 0) or 0)]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = border
                    if col >= 3:
                        cell.number_format = '#,##0.00'
                row += 1

            for col, w in enumerate([12, 24, 14, 14, 16, 16, 16, 16], 1):
                ws.column_dimensions[get_column_letter(col)].width = w

            fname = f"Maas_Odemeleri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path, _ = QFileDialog.getSaveFileName(self, "Excel Kaydet", fname, "Excel (*.xlsx)")
            if path:
                wb.save(path)
                QMessageBox.information(self, "Başarılı", f"Excel kaydedildi:\n{path}")
                reply = QMessageBox.question(self, "Aç", "Dosyayı şimdi açmak ister misiniz?",
                                             QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.Yes:
                    os.startfile(path)
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl bulunamadı. 'pip install openpyxl' çalıştırın.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel aktarımı başarısız:\n{str(e)}")

    def _generate_monthly_comparison_report(self, start_date, end_date):
        """Aylık karşılaştırma raporu - modern stil"""
        from src.database.db import SessionLocal
        session = SessionLocal()
        try:
            monthly_data = defaultdict(lambda: {'gelir': 0, 'gider': 0})
            transactions = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).all()
            for t in transactions:
                mk = t.transaction_date.strftime('%Y-%m')
                if t.transaction_type in [TransactionType.GELIR, TransactionType.KESILEN_FATURA]:
                    monthly_data[mk]['gelir'] += t.amount
                elif t.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]:
                    monthly_data[mk]['gider'] += t.amount

            sorted_months = sorted(monthly_data.keys())
            total_gelir = sum(monthly_data[m]['gelir'] for m in sorted_months)
            total_gider = sum(monthly_data[m]['gider'] for m in sorted_months)
            net_total   = total_gelir - total_gider

            html  = self._rh('📊', 'Aylık Karşılaştırma Raporu',
                             f'{start_date} — {end_date}', '#2196F3')
            html += self._kpi_row([
                ('📅', 'Toplam Gelir',  f'{total_gelir:,.2f} ₺', '#4CAF50'),
                ('📉', 'Toplam Gider',  f'{total_gider:,.2f} ₺', '#f44336'),
                ('📊', 'Net Toplam',    f'{net_total:,.2f} ₺',   '#4CAF50' if net_total >= 0 else '#f44336'),
                ('🗓', 'Ay Sayısı',     str(len(sorted_months)),  '#2196F3'),
            ])
            html += self._section('Aylık Detay', '#2196F3')
            html += self._table_header(['Ay', 'Gelir (₺)', 'Gider (₺)', 'Net (₺)', 'Değişim'], '#2196F3')

            prev_net = None
            for i, month in enumerate(sorted_months):
                d   = monthly_data[month]
                net = d['gelir'] - d['gider']
                if prev_net is not None:
                    ch  = ((net - prev_net) / prev_net * 100) if prev_net != 0 else 0
                    chstr = f'{"▲" if ch > 0 else "▼"} {abs(ch):.1f}%'
                    chc   = '#4CAF50' if ch > 0 else '#f44336'
                else:
                    chstr, chc = '—', '#888'
                netc = '#4CAF50' if net >= 0 else '#f44336'
                bg   = '#1e2a3a' if i % 2 == 0 else '#1a2332'
                html += self._tr(
                    [month,
                     f'{d["gelir"]:,.2f}', f'{d["gider"]:,.2f}',
                     f'{net:,.2f}', chstr],
                    ['#e0e0e0', '#4CAF50', '#f44336', netc, chc],
                    bold=[False, False, False, True, False],
                    bg=bg
                )
                prev_net = net
            html += '</table>'
            return html
        finally:
            session.close()

    def _generate_top_caris_report(self, start_date, end_date, detail_level):
        """En çok işlem yapılan cariler raporu - modern stil"""
        from src.database.db import SessionLocal
        from sqlalchemy import func
        session = SessionLocal()
        try:
            cari_stats = session.query(
                Transaction.customer_name,
                func.count(Transaction.id).label('count'),
                func.sum(Transaction.amount).label('total')
            ).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date,
                Transaction.customer_name != None
            ).group_by(Transaction.customer_name)\
             .order_by(func.sum(Transaction.amount).desc()).limit(20).all()

            grand_total = sum(r[2] for r in cari_stats) or 1

            html  = self._rh('👥', 'En Çok İşlem Yapılan Cariler',
                             f'{start_date} — {end_date}', '#FF9800')
            html += self._kpi_row([
                ('👥', 'Toplam Cari',       str(len(cari_stats)),          '#FF9800'),
                ('💰', 'Toplam Tutar',      f'{grand_total:,.2f} ₺',       '#4CAF50'),
                ('🏆', 'En Yüksek',         f'{cari_stats[0][2]:,.2f} ₺' if cari_stats else '—', '#2196F3'),
                ('📅', 'Dönem',             f'{start_date} / {end_date}',  '#9C27B0'),
            ])
            html += self._section('Cari Sıralaması', '#FF9800')
            html += self._table_header(['#', 'Cari Adı', 'İşlem Sayısı', 'Toplam Tutar (₺)', 'Pay %'], '#FF9800')

            for i, (customer, count, total) in enumerate(cari_stats, 1):
                pct = total / grand_total * 100
                bg  = '#1e2a3a' if i % 2 == 0 else '#1a2332'
                html += self._tr(
                    [str(i), customer or '—', str(count), f'{total:,.2f}',
                     self._progress_bar_html(pct)],
                    ['#FF9800', '#e0e0e0', '#aaa', '#4CAF50', '#e0e0e0'],
                    bold=[True, True, False, True, False],
                    bg=bg
                )
            html += '</table>'

            if detail_level == 'Detaylı':
                html += self._section('Detaylı İşlem Listesi (İlk 5 Cari)', '#FF9800')
                for customer, _, _ in cari_stats[:5]:
                    txns = session.query(Transaction).filter(
                        Transaction.user_id == self.user.id,
                        Transaction.customer_name == customer,
                        Transaction.transaction_date >= start_date,
                        Transaction.transaction_date <= end_date
                    ).order_by(Transaction.transaction_date.desc()).limit(10).all()
                    html += f"<p style='color:#FF9800;font-weight:bold;margin:12px 0 4px;'>{customer}</p>"
                    html += self._table_header(['Tarih', 'Tür', 'Açıklama', 'Tutar (₺)'], '#555')
                    for j, t in enumerate(txns):
                        is_in = t.transaction_type in [TransactionType.GELIR, TransactionType.KESILEN_FATURA]
                        bg = '#1e2a3a' if j % 2 == 0 else '#1a2332'
                        html += self._tr(
                            [str(t.transaction_date)[:10], t.transaction_type.value,
                             t.description or '—', f'{t.amount:,.2f}'],
                            ['#aaa', '#4CAF50' if is_in else '#f44336',
                             '#ccc', '#4CAF50' if is_in else '#f44336'],
                            bold=[False, False, False, True], bg=bg
                        )
                    html += '</table>'
            return html
        finally:
            session.close()

    def _generate_payment_method_distribution(self, start_date, end_date):
        """Ödeme yöntemi dağılımı raporu - modern stil"""
        from src.database.db import SessionLocal
        from sqlalchemy import func
        session = SessionLocal()
        try:
            payment_stats = session.query(
                Transaction.payment_method,
                func.count(Transaction.id).label('count'),
                func.sum(Transaction.amount).label('total')
            ).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).group_by(Transaction.payment_method).all()

            total_amount = sum(s[2] for s in payment_stats) or 1
            total_count  = sum(s[1] for s in payment_stats)

            html  = self._rh('💳', 'Ödeme Yöntemi Dağılımı',
                             f'{start_date} — {end_date}', '#9C27B0')
            html += self._kpi_row([
                ('💳', 'Toplam Tutar',    f'{total_amount:,.2f} ₺',     '#9C27B0'),
                ('🔢', 'Toplam İşlem',    str(total_count),              '#2196F3'),
                ('📋', 'Yöntem Sayısı',   str(len(payment_stats)),       '#FF9800'),
                ('📅', 'Dönem',           f'{start_date} / {end_date}',  '#4CAF50'),
            ])
            html += self._section('Ödeme Yöntemi Breakdown', '#9C27B0')
            html += self._table_header(
                ['Ödeme Yöntemi', 'İşlem Sayısı', 'Toplam Tutar (₺)', 'Pay %', 'Dağılım'],
                '#9C27B0')

            COLORS = ['#9C27B0', '#2196F3', '#FF9800', '#4CAF50', '#f44336', '#00BCD4']
            for i, (method, count, total) in enumerate(payment_stats):
                pct  = total / total_amount * 100
                name = method.value if method else 'Belirtilmemiş'
                clr  = COLORS[i % len(COLORS)]
                bg   = '#1e2a3a' if i % 2 == 0 else '#1a2332'
                html += self._tr(
                    [name, str(count), f'{total:,.2f}',
                     f'{pct:.1f}%', self._progress_bar_html(pct)],
                    [clr, '#aaa', '#4CAF50', clr, '#e0e0e0'],
                    bold=[True, False, True, True, False], bg=bg
                )
            html += '</table>'
            return html
        finally:
            session.close()

    def _generate_weekly_trend_report(self, start_date, end_date):
        """Haftalık trend analizi - modern stil"""
        from src.database.db import SessionLocal
        session = SessionLocal()
        try:
            weekly_data = defaultdict(lambda: {'gelir': 0, 'gider': 0})
            transactions = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).all()
            for t in transactions:
                ws = t.transaction_date - timedelta(days=t.transaction_date.weekday())
                wk = ws.strftime('%Y-%m-%d')
                if t.transaction_type in [TransactionType.GELIR, TransactionType.KESILEN_FATURA]:
                    weekly_data[wk]['gelir'] += t.amount
                elif t.transaction_type in [TransactionType.GIDER, TransactionType.GELEN_FATURA]:
                    weekly_data[wk]['gider'] += t.amount

            sorted_weeks = sorted(weekly_data.keys())
            total_gelir  = sum(weekly_data[w]['gelir'] for w in sorted_weeks)
            total_gider  = sum(weekly_data[w]['gider'] for w in sorted_weeks)
            net_total    = total_gelir - total_gider

            html  = self._rh('📈', 'Haftalık Trend Analizi',
                             f'{start_date} — {end_date}', '#4CAF50')
            html += self._kpi_row([
                ('📈', 'Toplam Gelir',  f'{total_gelir:,.2f} ₺', '#4CAF50'),
                ('📉', 'Toplam Gider',  f'{total_gider:,.2f} ₺', '#f44336'),
                ('📊', 'Net',           f'{net_total:,.2f} ₺',   '#4CAF50' if net_total >= 0 else '#f44336'),
                ('🗓', 'Hafta Sayısı',  str(len(sorted_weeks)),  '#2196F3'),
            ])
            html += self._section('Haftalık Detay', '#4CAF50')
            html += self._table_header(
                ['Hafta Başlangıcı', 'Gelir (₺)', 'Gider (₺)', 'Net (₺)', 'Trend'],
                '#4CAF50')

            prev_net = None
            for i, week in enumerate(sorted_weeks):
                d   = weekly_data[week]
                net = d['gelir'] - d['gider']
                if prev_net is not None:
                    if net > prev_net:   trend, tc = '▲ Yükseliş', '#4CAF50'
                    elif net < prev_net: trend, tc = '▼ Düşüş',    '#f44336'
                    else:                trend, tc = '● Sabit',     '#888'
                else:
                    trend, tc = '—', '#888'
                netc = '#4CAF50' if net >= 0 else '#f44336'
                bg   = '#1e2a3a' if i % 2 == 0 else '#1a2332'
                html += self._tr(
                    [week, f'{d["gelir"]:,.2f}', f'{d["gider"]:,.2f}',
                     f'{net:,.2f}', trend],
                    ['#e0e0e0', '#4CAF50', '#f44336', netc, tc],
                    bold=[False, False, False, True, True],
                    bg=bg
                )
                prev_net = net
            html += '</table>'
            return html
        finally:
            session.close()

    def _generate_payroll_report(self, start_date, end_date):
        """Maaş ödemeleri raporu — çalışanlar DB + kaydedilen bordrolar"""
        from src.database.db import SessionLocal
        from src.database.models import Employee

        # ── Çalışanları DB'den al ────────────────────────────────────────
        session = SessionLocal()
        try:
            employees = session.query(Employee).filter(
                Employee.is_active == True
            ).order_by(Employee.first_name, Employee.last_name).all()
            emp_info = {
                e.get_full_name(): {
                    'brut': e.gross_salary,
                    'sgk_rate': e.sgk_rate,
                    'baslangic': str(e.start_date) if e.start_date else '—',
                }
                for e in employees
            }
        finally:
            session.close()

        # ── Kaydedilen bordrolardan ay/yıl filtreli kayıtlar ─────────────
        all_records = self._load_saved_payroll_records()

        filtered = []
        for r in all_records:
            try:
                rec_year  = int(r.get('year', 0))
                rec_month = int(r.get('month', 0))
                from datetime import date as _date
                rec_date = _date(rec_year, rec_month, 1)
                if start_date <= rec_date <= end_date:
                    filtered.append(r)
            except Exception:
                continue

        # ── Çalışan bazında topla ────────────────────────────────────────
        from collections import defaultdict
        emp_totals = defaultdict(lambda: {
            'kayit': 0, 'brut': 0.0, 'net': 0.0,
            'sgk': 0.0, 'issizlik': 0.0, 'gelir_vergisi': 0.0,
            'damga_vergisi': 0.0, 'toplam_kesinti': 0.0,
            'donemler': []
        })

        for r in filtered:
            name = str(r.get('employee', 'Belirtilmedi')).strip()
            et = emp_totals[name]
            et['kayit']         += 1
            et['brut']          += float(r.get('gross_total', 0) or 0)
            et['net']           += float(r.get('net_salary', 0) or 0)
            et['sgk']           += float(r.get('sgk_deduction', 0) or 0)
            et['issizlik']      += float(r.get('unemployment_deduction', 0) or 0)
            et['gelir_vergisi'] += float(r.get('income_tax', 0) or 0)
            et['damga_vergisi'] += float(r.get('stamp_tax', 0) or 0)
            et['toplam_kesinti']+= float(r.get('total_deductions', 0) or 0)
            donem = f"{int(r.get('month',0)):02d}/{r.get('year','')}"
            if donem not in et['donemler']:
                et['donemler'].append(donem)

        toplam_brut    = sum(v['brut']  for v in emp_totals.values())
        toplam_net     = sum(v['net']   for v in emp_totals.values())
        toplam_sgk     = sum(v['sgk']   for v in emp_totals.values())
        toplam_kesinti = sum(v['toplam_kesinti'] for v in emp_totals.values())
        kisi_sayisi    = len(emp_totals)

        # ── HTML ─────────────────────────────────────────────────────────
        html  = self._rh('👷', 'Maaş Ödemeleri Raporu',
                         f'{start_date} — {end_date}', '#1565C0')
        html += self._kpi_row([
            ('👷', 'Çalışan Sayısı',    str(kisi_sayisi),           '#1565C0'),
            ('💵', 'Toplam Brüt',       f'{toplam_brut:,.2f} ₺',    '#4CAF50'),
            ('💰', 'Toplam Net',        f'{toplam_net:,.2f} ₺',      '#2196F3'),
            ('📉', 'Toplam Kesinti',    f'{toplam_kesinti:,.2f} ₺',  '#f44336'),
        ])

        # Aktif çalışanlar listesi (DB'den — kaydedilmiş bordro olmasa bile gösterilsin)
        html += self._section('👥 Aktif Çalışanlar (Tanımlı Maaş)', '#1565C0')
        html += self._table_header(
            ['Çalışan', 'Brüt Maaş (₺)', 'SGK %', 'İşe Başlama', 'Kaydedilen Bordro'],
            '#1565C0')
        for i, (name, info) in enumerate(emp_info.items()):
            et = emp_totals.get(name, {})
            kayit_sayisi = et.get('kayit', 0)
            kayit_str = f'{kayit_sayisi} dönem' if kayit_sayisi else '—'
            bg = '#1e2a3a' if i % 2 == 0 else '#1a2332'
            html += self._tr(
                [name, f'{info["brut"]:,.2f}', f'{info["sgk_rate"]}%',
                 info['baslangic'], kayit_str],
                ['#e0e0e0', '#4CAF50', '#aaa', '#aaa',
                 '#2196F3' if kayit_sayisi else '#666'],
                bold=[True, True, False, False, False], bg=bg
            )
        html += '</table>'

        if not filtered:
            html += f"""
            <div style='margin:20px; padding:16px; background:#1e2a3a;
                        border-left:4px solid #FF9800; border-radius:4px;'>
                <span style='color:#FF9800; font-weight:bold;'>ℹ️ Bu tarih aralığında kaydedilmiş bordro bulunamadı.</span><br>
                <span style='color:#aaa; font-size:9pt;'>Maaş Bordro sayfasında hesaplama yapıp
                "Bordroyu Kaydet" butonuna basarak kayıt ekleyebilirsiniz.</span>
            </div>"""
            return html

        # Dönem bazında özet tablo
        html += self._section(f'📋 Çalışan Bazında Özet ({start_date} — {end_date})', '#1565C0')
        html += self._table_header(
            ['Çalışan', 'Dönem Sayısı', 'Toplam Brüt (₺)', 'Toplam Net (₺)',
             'SGK Kesinti (₺)', 'Gelir Vergisi (₺)', 'Dönemler'],
            '#1565C0')
        for i, (name, et) in enumerate(sorted(emp_totals.items())):
            pct_net = (et['net'] / et['brut'] * 100) if et['brut'] else 0
            donem_str = ', '.join(sorted(et['donemler']))
            bg = '#1e2a3a' if i % 2 == 0 else '#1a2332'
            html += self._tr(
                [name, str(et['kayit']),
                 f'{et["brut"]:,.2f}', f'{et["net"]:,.2f}',
                 f'{et["sgk"]:,.2f}', f'{et["gelir_vergisi"]:,.2f}',
                 donem_str],
                ['#e0e0e0', '#aaa', '#4CAF50', '#2196F3',
                 '#f44336', '#FF9800', '#888'],
                bold=[True, False, True, True, False, False, False], bg=bg
            )
        html += '</table>'

        # Dönem dönem detay tablosu
        html += self._section('📅 Dönem / Çalışan Bazında Detay', '#34495e')
        html += self._table_header(
            ['Dönem', 'Çalışan', 'Brüt (₺)', 'Net (₺)',
             'SGK (₺)', 'Gelir V. (₺)', 'Damga V. (₺)', 'Toplam Kesinti (₺)'],
            '#34495e')
        for i, r in enumerate(sorted(filtered,
                key=lambda x: (int(x.get('year',0)), int(x.get('month',0)),
                               str(x.get('employee',''))))):
            donem = f"{int(r.get('month',0)):02d}/{r.get('year','')}"
            bg = '#1e2a3a' if i % 2 == 0 else '#1a2332'
            html += self._tr(
                [donem, str(r.get('employee','—')),
                 f'{float(r.get("gross_total",0) or 0):,.2f}',
                 f'{float(r.get("net_salary",0) or 0):,.2f}',
                 f'{float(r.get("sgk_deduction",0) or 0):,.2f}',
                 f'{float(r.get("income_tax",0) or 0):,.2f}',
                 f'{float(r.get("stamp_tax",0) or 0):,.2f}',
                 f'{float(r.get("total_deductions",0) or 0):,.2f}'],
                ['#aaa', '#e0e0e0', '#4CAF50', '#2196F3',
                 '#f44336', '#FF9800', '#FF9800', '#f44336'],
                bold=[False, True, True, True, False, False, False, False], bg=bg
            )
        html += '</table>'
        return html

    def _generate_konu_gider_report(self, start_date, end_date):
        """Konuya göre gider raporu — işlemler sayfasındaki konu alanına göre filtreli"""
        from src.database.db import SessionLocal
        from src.database.models import Transaction, TransactionType

        GIDER_TYPES = [
            TransactionType.GIDER,
            TransactionType.KREDI_DOSYA_MASRAFI,
            TransactionType.KREDI_ODEME,
            TransactionType.KREDI_KARTI_ODEME,
            TransactionType.EKSPERTIZ_UCRETI,
        ]

        session = SessionLocal()
        try:
            # Seçilen tarih aralığındaki tüm gider tipi işlemlerini çek
            transactions = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_type.in_(GIDER_TYPES),
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date,
            ).order_by(Transaction.transaction_date.desc()).all()

            # Tüm zamanlardaki benzersiz konuları çek (yeni eklenen konular da görünsün)
            all_time_subjects = session.query(Transaction.subject).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_type.in_(GIDER_TYPES),
                Transaction.subject != None,
                Transaction.subject != "",
            ).distinct().all()
            all_subjects = sorted(set(
                list(set((t.subject or "—") for t in transactions)) +
                [row[0] for row in all_time_subjects if row[0]]
            ))
        finally:
            session.close()

        # Konu filtre combosu'nu doldur (sinyalleri blokla)
        combo = self.report_konu_filter
        combo.blockSignals(True)
        prev = combo.currentText()
        combo.clear()
        combo.addItem("🔍 Tümü")
        for s in all_subjects:
            combo.addItem(s)
        idx = combo.findText(prev)
        combo.setCurrentIndex(idx if idx >= 0 else 0)
        combo.blockSignals(False)

        # Seçili konuya göre filtrele
        selected = combo.currentText()
        if selected and selected != "🔍 Tümü":
            filtered = [t for t in transactions if (t.subject or "—") == selected]
            subtitle_konu = f" — Konu: {selected}"
        else:
            filtered = transactions
            subtitle_konu = " — Tüm Konular"

        toplam = sum(t.amount for t in filtered)
        sayi   = len(filtered)

        # KPI istatistikleri: konuya göre toplam
        from collections import defaultdict
        konu_totals = defaultdict(lambda: {"sayi": 0, "tutar": 0.0})
        for t in filtered:
            k = t.subject or "—"
            konu_totals[k]["sayi"]  += 1
            konu_totals[k]["tutar"] += t.amount

        html  = self._rh("🏷️", "Konuya Göre Gider Raporu",
                         f"{start_date} — {end_date}{subtitle_konu}", "#B71C1C")
        html += self._kpi_row([
            ("📋", "İşlem Sayısı",  str(sayi),               "#546E7A"),
            ("📉", "Toplam Gider",  f"{toplam:,.2f} ₺",      "#B71C1C"),
            ("🏷️", "Konu Adedi",   str(len(konu_totals)),    "#37474F"),
        ])

        # Konu başlıklı toplam tutar KPI kartları
        if konu_totals:
            KONU_COLORS = [
                "#1565C0", "#2E7D32", "#6A1B9A", "#E65100",
                "#00695C", "#AD1457", "#4527A0", "#283593",
                "#558B2F", "#BF360C",
            ]
            sorted_konular = sorted(konu_totals.items(), key=lambda x: -x[1]["tutar"])
            chunk_size = 4
            for chunk_start in range(0, len(sorted_konular), chunk_size):
                chunk = sorted_konular[chunk_start:chunk_start + chunk_size]
                kpi_items = []
                for ci, (konu, info) in enumerate(chunk):
                    color = KONU_COLORS[(chunk_start + ci) % len(KONU_COLORS)]
                    konu_label = konu if len(konu) <= 22 else konu[:20] + "…"
                    kpi_items.append(
                        ("🏷️", konu_label, f"{info['tutar']:,.2f} ₺", color)
                    )
                html += self._kpi_row(kpi_items)

        # Konu bazında özet tablo
        html += self._section("🏷️ Konu Bazında Özet", "#B71C1C")
        html += self._table_header(["Konu", "İşlem Sayısı", "Toplam Tutar (₺)", "Oran"])
        for i, (konu, info) in enumerate(
                sorted(konu_totals.items(), key=lambda x: -x[1]["tutar"])):
            pct = (info["tutar"] / toplam * 100) if toplam else 0
            bg  = "#FAFAFA" if i % 2 else "white"
            html += self._tr(
                [konu, str(info["sayi"]),
                 f'{info["tutar"]:,.2f}',
                 self._progress_bar_html(pct)],
                [None, "#546E7A", "#C62828", None],
                bg=bg
            )
        if konu_totals:
            html += self._tr(
                ["<b>TOPLAM</b>", f"<b>{sayi}</b>", f"<b>{toplam:,.2f}</b>", ""],
                [None, None, "#B71C1C", None], bold=True, bg="#FFEBEE"
            )
        html += "</table><br>"

        # Detay tablosu
        if filtered:
            html += self._section(
                f"📋 İşlem Detayları ({len(filtered)} kayıt)", "#37474F")
            html += self._table_header(
                ["Tarih", "Müşteri Ünvanı", "Açıklama", "Konu",
                 "Ödeme Şekli", "Ödeyen Kişi", "Tutar (₺)"])
            for i, t in enumerate(filtered):
                bg = "#FAFAFA" if i % 2 else "white"
                html += self._tr(
                    [str(t.transaction_date),
                     t.customer_name or "—",
                     (t.description or "")[:60],
                     t.subject or "—",
                     t.payment_type or (t.payment_method.value if t.payment_method else "—"),
                     t.person or "—",
                     f'{t.amount:,.2f}'],
                    [None, None, "#546E7A", "#B71C1C", None, None, "#C62828"],
                    bg=bg
                )
            html += "</table>"
        else:
            html += f"""
            <div style='margin:20px; padding:16px; background:#FFF8F8;
                        border-left:4px solid #EF9A9A; border-radius:4px;'>
                <span style='color:#C62828; font-weight:bold;'>
                    ℹ️ Seçilen tarih aralığında ve konuda gider işlemi bulunamadı.</span>
            </div>"""

        return html

    def _generate_kira_takip_report(self):
        """Kira Takip raporu — JSON veri dosyasından okur ve tüm sekmeleri özetler."""
        import json
        from pathlib import Path
        from datetime import date, datetime

        AYLAR = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
                 "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]

        data_file = Path("data") / f"kira_takip_data_{self.user.id}.json"
        if not data_file.exists():
            html  = self._rh("🏠", "Kira Takip Raporu", "Henüz veri yok", "#1565C0")
            html += """
            <div style='margin:20px; padding:16px; background:#E3F2FD;
                        border-left:4px solid #1565C0; border-radius:4px;'>
                <span style='color:#0D47A1; font-weight:bold;'>
                    ℹ️ Kira Takip sekmesine gidip kiracı ve sekme ekleyiniz.</span>
            </div>"""
            return html

        try:
            raw = json.loads(data_file.read_text(encoding="utf-8"))
        except Exception as e:
            return f"<p style='color:red; padding:20px;'>Veri yüklenirken hata: {e}</p>"

        tabs = raw.get("tabs", [])
        if not tabs:
            html  = self._rh("🏠", "Kira Takip Raporu", "Sekme bulunamadı", "#1565C0")
            html += "<p style='padding:20px; color:#555;'>Kira Takip sekmesinde henüz sekme oluşturulmamış.</p>"
            return html

        today = date.today()
        cur_year  = today.year
        cur_month = today.month

        # ── Genel KPI hesapla ─────────────────────────────────────────────
        total_tenants    = 0
        grand_expected   = 0.0
        grand_collected  = 0.0
        grand_pending    = 0.0
        cur_month_expected  = 0.0
        cur_month_collected = 0.0

        for tab in tabs:
            contracts = tab.get("contracts", [])
            payments  = tab.get("payments", {})
            year      = tab.get("year", cur_year)

            for c in contracts:
                total_tenants += 1
                cid   = c["id"]
                tutar = float(c.get("tutar", 0))
                pays  = {int(m): v for m, v in payments.get(str(cid), {}).items()}

                try:
                    bas = datetime.strptime(c["bas"], "%d.%m.%Y").date()
                    bit = datetime.strptime(c["bit"], "%d.%m.%Y").date()
                except Exception:
                    continue

                for ay in range(1, 13):
                    try:
                        ay_date = date(year, ay, 1)
                    except ValueError:
                        continue
                    bas_first = bas.replace(day=1)
                    bit_first = bit.replace(day=1)
                    if not (bas_first <= ay_date <= bit_first):
                        continue

                    grand_expected += tutar
                    val = pays.get(ay)
                    odendi = isinstance(val, dict) or val == "ODENDI"
                    if odendi:
                        grand_collected += tutar
                    else:
                        grand_pending   += tutar

                    if year == cur_year and ay == cur_month:
                        cur_month_expected += tutar
                        if odendi:
                            cur_month_collected += tutar

        cur_remaining = max(0.0, cur_month_expected - cur_month_collected)
        oran_pct      = (grand_collected / grand_expected * 100) if grand_expected > 0 else 0.0

        # ── HTML başlık + KPIs ────────────────────────────────────────────
        tab_years = sorted({tab.get("year", cur_year) for tab in tabs})
        year_str  = ", ".join(str(y) for y in tab_years) if tab_years else str(cur_year)
        html  = self._rh("🏠", "Kira Takip Raporu", f"Tüm Sekmeler — Yıl(lar): {year_str}", "#1565C0")
        html += self._kpi_row([
            ("👥", "Toplam Kiracı",    str(total_tenants),                                 "#3f51b5"),
            ("✅", "Toplam Tahsilat",  f"{grand_collected:,.0f} ₺".replace(",", "."),       "#4caf50"),
            ("⏳", "Toplam Bekliyor",  f"{grand_pending:,.0f} ₺".replace(",", "."),         "#f44336"),
            ("💰", "Yıllık Beklenen",  f"{grand_expected:,.0f} ₺".replace(",", "."),        "#ff9800"),
        ])
        html += self._kpi_row([
            ("📅", "Bu Ay Beklenen",   f"{cur_month_expected:,.0f} ₺".replace(",", "."),   "#1976D2"),
            ("📅", "Bu Ay Tahsilat",   f"{cur_month_collected:,.0f} ₺".replace(",", "."),  "#2e7d32"),
            ("📅", "Bu Ay Kalan",      f"{cur_remaining:,.0f} ₺".replace(",", "."),        "#c62828"),
            ("📊", "Tahsilat Oranı",   f"%{oran_pct:.1f}",                                 "#6a1b9a"),
        ])

        # ── Sekme bazlı kiracı detay tabloları ────────────────────────────
        for tab in tabs:
            tab_name  = tab.get("tab_name", "Sekme")
            title     = tab.get("title", tab_name)
            hdr_color = tab.get("hdr_color", "#1565C0")
            contracts = tab.get("contracts", [])
            payments  = tab.get("payments", {})
            year      = tab.get("year", cur_year)

            if not contracts:
                continue

            html += self._section(f"🏠 {tab_name}  —  {title}  ({year})", hdr_color)
            html += self._table_header(
                ["Kiracı", "Sözleşme Dönemi", "Aylık Kira (₺)", "Ödeme Günü",
                 "Ödenmiş Ay", "Tahsilat (₺)", "Kalan (₺)"],
                color="#BBDEFB"
            )

            tab_collected = 0.0
            tab_expected  = 0.0

            for i, c in enumerate(contracts):
                cid   = c["id"]
                tutar = float(c.get("tutar", 0))
                pays  = {int(m): v for m, v in payments.get(str(cid), {}).items()}

                try:
                    bas = datetime.strptime(c["bas"], "%d.%m.%Y").date()
                    bit = datetime.strptime(c["bit"], "%d.%m.%Y").date()
                    contract_str = f"{c['bas']} – {c['bit']}"
                except Exception:
                    contract_str = f"{c.get('bas','')} – {c.get('bit','')}"
                    bas = bit = None

                paid_count     = 0
                possible_count = 0
                c_collected    = 0.0

                for ay in range(1, 13):
                    if bas and bit:
                        try:
                            ay_date = date(year, ay, 1)
                        except ValueError:
                            continue
                        if not (bas.replace(day=1) <= ay_date <= bit.replace(day=1)):
                            continue
                    possible_count += 1
                    val = pays.get(ay)
                    if isinstance(val, dict) or val == "ODENDI":
                        paid_count  += 1
                        c_collected += tutar

                c_expected   = possible_count * tutar
                c_kalan      = max(0.0, c_expected - c_collected)
                tab_collected += c_collected
                tab_expected  += c_expected

                bg = "#FAFAFA" if i % 2 == 0 else "white"
                html += self._tr(
                    [
                        f"<b>{c['kiraci']}</b>",
                        contract_str,
                        f"{tutar:,.0f}".replace(",", "."),
                        c.get("odeme_gunu", "—"),
                        f"{paid_count}/{possible_count}",
                        f"{c_collected:,.0f}".replace(",", "."),
                        f"{c_kalan:,.0f}".replace(",", "."),
                    ],
                    colors=[None, "#546E7A", "#1565C0", "#6A1B9A",
                            "#2E7D32" if paid_count == possible_count else "#B71C1C",
                            "#2E7D32", "#C62828" if c_kalan > 0 else "#2E7D32"],
                    bg=bg
                )

            # Sekme toplamı satırı
            tab_kalan = max(0.0, tab_expected - tab_collected)
            html += self._tr(
                ["<b>SEKME TOPLAMI</b>", "", "", "",
                 "",
                 f"<b>{tab_collected:,.0f}</b>".replace(",", "."),
                 f"<b>{tab_kalan:,.0f}</b>".replace(",", ".")],
                colors=[None, None, None, None, None, "#2E7D32", "#C62828"],
                bold=True, bg="#E3F2FD"
            )
            html += "</table><br>"

        # ── Aylık tahsilat özet tablosu (tüm sekmeler, cur_year) ──────────
        html += self._section(f"📅 Aylık Tahsilat Özeti — {cur_year} Yılı", "#1565C0")
        html += self._table_header(
            ["Ay", "Beklenen (₺)", "Tahsil Edilen (₺)", "Bekleyen (₺)", "Tahsilat Oranı"],
            color="#BBDEFB"
        )

        monthly_expected  = [0.0] * 12
        monthly_collected = [0.0] * 12

        for tab in tabs:
            if tab.get("year", cur_year) != cur_year:
                continue
            contracts = tab.get("contracts", [])
            payments  = tab.get("payments", {})

            for c in contracts:
                cid   = c["id"]
                tutar = float(c.get("tutar", 0))
                pays  = {int(m): v for m, v in payments.get(str(cid), {}).items()}

                try:
                    bas = datetime.strptime(c["bas"], "%d.%m.%Y").date()
                    bit = datetime.strptime(c["bit"], "%d.%m.%Y").date()
                except Exception:
                    bas = bit = None

                for ay in range(1, 13):
                    if bas and bit:
                        try:
                            ay_date = date(cur_year, ay, 1)
                        except ValueError:
                            continue
                        if not (bas.replace(day=1) <= ay_date <= bit.replace(day=1)):
                            continue
                    monthly_expected[ay - 1]  += tutar
                    val = pays.get(ay)
                    if isinstance(val, dict) or val == "ODENDI":
                        monthly_collected[ay - 1] += tutar

        for i, ay_ad in enumerate(AYLAR):
            exp = monthly_expected[i]
            col = monthly_collected[i]
            pen = max(0.0, exp - col)
            oran_str = f"%{(col / exp * 100):.0f}" if exp > 0 else "—"
            bg = "#FAFAFA" if i % 2 == 0 else "white"

            # Gelecek aylar için farklı renk
            is_future = (i + 1) > cur_month
            renk_exp  = "#546E7A" if is_future else "#1565C0"
            renk_col  = "#9E9E9E" if is_future else "#2E7D32"
            renk_pen  = "#9E9E9E" if is_future else ("#C62828" if pen > 0 else "#2E7D32")

            html += self._tr(
                [
                    f"<b>{ay_ad}</b>" + (" <i>(gelecek)</i>" if is_future else ""),
                    f"{exp:,.0f}".replace(",", "."),
                    f"{col:,.0f}".replace(",", "."),
                    f"{pen:,.0f}".replace(",", "."),
                    oran_str,
                ],
                colors=[None, renk_exp, renk_col, renk_pen, "#6A1B9A"],
                bg=bg
            )

        # Yıl toplamı
        yr_exp = sum(monthly_expected)
        yr_col = sum(monthly_collected)
        yr_pen = max(0.0, yr_exp - yr_col)
        yr_oran = f"%{(yr_col / yr_exp * 100):.0f}" if yr_exp > 0 else "—"
        html += self._tr(
            [f"<b>YILLIK TOPLAM ({cur_year})</b>",
             f"<b>{yr_exp:,.0f}</b>".replace(",", "."),
             f"<b>{yr_col:,.0f}</b>".replace(",", "."),
             f"<b>{yr_pen:,.0f}</b>".replace(",", "."),
             f"<b>{yr_oran}</b>"],
            colors=[None, "#1565C0", "#2E7D32", "#C62828", "#6A1B9A"],
            bold=True, bg="#E3F2FD"
        )
        html += "</table><br>"

        return html

    def _generate_fatura_vade_report(self):
        """Fatura Vade Takibi raporu — gecikmiş, bugün, yaklaşan, kısmi, ödenmiş."""
        from src.database.db import SessionLocal
        from src.database.models import Transaction, TransactionType
        from datetime import date as _date

        today = _date.today()
        session = SessionLocal()
        try:
            rows = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_type == TransactionType.KESILEN_FATURA
            ).order_by(Transaction.transaction_date.desc()).all()

            # Session açıkken tüm ilişkili alanları dict'e çek
            invoices = []
            for inv in rows:
                invoices.append({
                    'id': inv.id,
                    'invoice_number': getattr(inv, 'invoice_number', None),
                    'cari_name': inv.cari.name if inv.cari else (getattr(inv, 'customer_name', None) or "-"),
                    'amount': inv.amount,
                    'transaction_date': inv.transaction_date,
                    'due_date': getattr(inv, 'due_date', None),
                    'is_paid': getattr(inv, 'is_paid', False) or False,
                    'paid_date': getattr(inv, 'paid_date', None),
                    'paid_amount': getattr(inv, 'paid_amount', 0.0) or 0.0,
                })
        finally:
            session.close()

        # Kategorize et
        paid_list, partial_list = [], []
        overdue, due_today, upcoming_5, upcoming_more = [], [], [], []
        for inv in invoices:
            paid_amt  = inv['paid_amount']
            total_amt = inv['amount']
            fully_paid = paid_amt >= total_amt and total_amt > 0
            is_partial = 0 < paid_amt < total_amt
            if fully_paid:
                paid_list.append(inv)
                continue
            if is_partial:
                partial_list.append(inv)
            due = inv['due_date']
            if not due:
                continue
            days_left = (due - today).days
            entry = (inv, days_left, due)
            if days_left < 0:
                overdue.append(entry)
            elif days_left == 0:
                due_today.append(entry)
            elif days_left <= 5:
                upcoming_5.append(entry)
            else:
                upcoming_more.append(entry)

        def fmt_amount(v):
            return f"{v:,.2f} \u20ba".replace(",", "X").replace(".", ",").replace("X", ".")

        def fatura_row(inv, days_left, due, bg, idx):
            invoice_no = inv['invoice_number'] or f"F-{inv['id']}"
            paid_amt = inv['paid_amount']
            if paid_amt > 0 and paid_amt < inv['amount']:
                tutar_str = (
                    f"{fmt_amount(inv['amount'])}"
                    f"<br><small style='color:#E65100;'>\U0001f7e0 \u00d6d: {fmt_amount(paid_amt)}"
                    f" / Kalan: {fmt_amount(inv['amount']-paid_amt)}</small>"
                )
            else:
                tutar_str = fmt_amount(inv['amount'])
            if days_left < 0:
                status_cell = f"<span style='color:#B71C1C;'>\u26d4 {abs(days_left)} g\u00fcn ge\u00e7</span>"
            elif days_left == 0:
                status_cell = "<span style='color:#E65100;'>\U0001f534 Bug\u00fcn!</span>"
            else:
                status_cell = f"<span style='color:#F57F17;'>\U0001f7e1 {days_left} g\u00fcn</span>"
            return (
                f"<tr style='background:{bg};'>"
                f"<td style='padding:6px 8px;'>{invoice_no}</td>"
                f"<td style='padding:6px 8px;'>{inv['cari_name']}</td>"
                f"<td style='padding:6px 8px; text-align:right;'>{tutar_str}</td>"
                f"<td style='padding:6px 8px;'>{inv['transaction_date']}</td>"
                f"<td style='padding:6px 8px;'>{due}</td>"
                f"<td style='padding:6px 8px; font-weight:bold;'>{status_cell}</td></tr>"
            )

        def section_table(title, color, entries):
            if not entries:
                return ""
            s = self._section(title, color)
            s += self._table_header(
                ["Fatura No", "Cari", "Tutar", "Fatura Tarihi", "Vade Tarihi", "Durum"],
                color=color
            )
            for idx, (inv, dl, due) in enumerate(entries):
                s += fatura_row(inv, dl, due, "#FAFAFA" if idx % 2 == 0 else "white", idx)
            s += "</table><br>"
            return s

        # KPI özet
        total_overdue_amount = sum(inv['amount'] - inv['paid_amount'] for inv, _, _ in overdue)
        total_due_today  = sum(inv['amount'] for inv, _, _ in due_today)
        total_upcoming_5 = sum(inv['amount'] for inv, _, _ in upcoming_5)
        total_all = total_overdue_amount + total_due_today + total_upcoming_5
        total_paid_amount    = sum(inv['amount']     for inv in paid_list)
        total_partial_paid   = sum(inv['paid_amount'] for inv in partial_list)
        total_partial_amount = sum(inv['amount']      for inv in partial_list)
        total_partial_kalan  = total_partial_amount - total_partial_paid

        html = self._rh("\U0001f4c6", "Fatura Vade Takibi",
                        f"Kesilen Faturalar\u0131n Vade Durumu \u2014 {today.strftime('%d.%m.%Y')} itibar\u0131yla",
                        "#1565C0")
        html += self._kpi_row([
            ("\u26d4",   "Gecikmi\u015f Fatura",       f"{len(overdue)} adet",         "#B71C1C"),
            ("\U0001f4b8", "Gecikmi\u015f Toplam \u20ba", fmt_amount(total_overdue_amount), "#B71C1C"),
            ("\U0001f534", "Bug\u00fcn Vadeli",           f"{len(due_today)} adet",       "#E65100"),
            ("\U0001f7e1", "5 G\u00fcn \u0130\u00e7inde", f"{len(upcoming_5)} adet",      "#F57F17"),
        ])
        html += self._kpi_row([
            ("\u26a0\ufe0f", "Yakla\u015fan Toplam \u20ba", fmt_amount(total_all),           "#6A1B9A"),
            ("\U0001f4ca",   "Toplam A\u00e7\u0131k",       f"{len(overdue)+len(due_today)+len(upcoming_5)+len(upcoming_more)} adet", "#37474F"),
            ("\U0001f7e0",   "K\u0131smi \u00d6deme",       f"{len(partial_list)} adet",    "#E65100"),
            ("\U0001f4b0",   "K\u0131smi \u00d6denen \u20ba", fmt_amount(total_partial_paid), "#2E7D32"),
        ])
        html += self._kpi_row([
            ("\U0001f4b8",   "K\u0131smi Kalan \u20ba",     fmt_amount(total_partial_kalan), "#B71C1C"),
            ("\u2705",       "\u00d6denmi\u015f Fatura",    f"{len(paid_list)} adet",       "#1B5E20"),
            ("\U0001f4b0",   "\u00d6denmi\u015f Toplam \u20ba", fmt_amount(total_paid_amount), "#1B5E20"),
        ])

        if not invoices:
            html += "<p style='padding:20px; color:#555;'>Hen\u00fcz kesilen fatura bulunmuyor.</p>"
            return html

        if not (overdue or due_today or upcoming_5 or upcoming_more or paid_list or partial_list):
            html += "<p style='padding:20px; color:#555;'>Vade tarihi girilmi\u015f fatura bulunamad\u0131.</p>"
            return html

        html += section_table("\u26d4 Gecikmi\u015f \u00d6demeler",      "#B71C1C", overdue)
        html += section_table("\U0001f534 Bug\u00fcn \u00d6deme G\u00fcn\u00fc", "#E65100", due_today)
        html += section_table("\U0001f7e1 5 G\u00fcn \u0130\u00e7inde \u00d6denecek", "#F57F17", upcoming_5)
        html += section_table("\U0001f7e2 \u0130lerideki Vadeler",        "#2E7D32", upcoming_more)

        # Kısmi ödeme tablosu
        if partial_list:
            html += self._section("\U0001f7e0 K\u0131smi \u00d6demeli Faturalar", "#E65100")
            html += self._table_header(
                ["Fatura No", "Cari", "Toplam Tutar", "\u00d6denen", "Kalan", "Fatura Tarihi", "Vade Tarihi"],
                color="#FFE0B2"
            )
            for idx, inv in enumerate(partial_list):
                invoice_no = inv['invoice_number'] or f"F-{inv['id']}"
                bg = "#FFF8E1" if idx % 2 == 0 else "white"
                kalan = inv['amount'] - inv['paid_amount']
                html += self._tr(
                    [invoice_no, inv['cari_name'],
                     fmt_amount(inv['amount']),
                     f"<span style='color:#2E7D32;font-weight:bold;'>{fmt_amount(inv['paid_amount'])}</span>",
                     f"<span style='color:#B71C1C;font-weight:bold;'>{fmt_amount(kalan)}</span>",
                     str(inv['transaction_date']),
                     str(inv['due_date']) if inv['due_date'] else "-"],
                    bg=bg
                )
            html += "</table><br>"

        # Ödenmiş faturalar tablosu
        if paid_list:
            html += self._section("\u2705 \u00d6denmi\u015f Faturalar", "#1B5E20")
            html += self._table_header(
                ["Fatura No", "Cari", "Tutar", "Fatura Tarihi", "Vade Tarihi", "\u00d6deme Tarihi"],
                color="#C8E6C9"
            )
            for idx, inv in enumerate(paid_list):
                invoice_no = inv['invoice_number'] or f"F-{inv['id']}"
                bg = "#F1F8E9" if idx % 2 == 0 else "white"
                paid_date_str = (
                    f"<span style='color:#1B5E20;font-weight:bold;'>\u2705 {inv['paid_date']}</span>"
                    if inv['paid_date'] else
                    "<span style='color:#1B5E20;'>\u2705 \u00d6dendi</span>"
                )
                html += self._tr(
                    [invoice_no, inv['cari_name'], fmt_amount(inv['amount']),
                     str(inv['transaction_date']),
                     str(inv['due_date']) if inv['due_date'] else "-",
                     paid_date_str],
                    bg=bg
                )
            html += "</table><br>"

        return html

    def _generate_kredi_bu_ay_report(self):
        """Bu ay ödeme günü gelen aktif kredileri listeleyen rapor."""
        from src.database.db import SessionLocal
        from src.database.models import Loan, Transaction, TransactionType
        from datetime import date as _date
        import calendar

        today = _date.today()
        month_start = _date(today.year, today.month, 1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        month_end = _date(today.year, today.month, last_day)

        session = SessionLocal()
        try:
            loans = session.query(Loan).filter(
                Loan.user_id == self.user.id,
                Loan.is_active == True,
                Loan.status == 'AKTIF',
            ).all()

            # Bu ay yapılan kredi ödemeleri
            this_month_payments = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_type == TransactionType.KREDI_ODEME,
                Transaction.transaction_date >= month_start,
                Transaction.transaction_date <= month_end,
            ).all()
        finally:
            session.close()

        # Hangi loan_id'ler bu ay ödenmiş?
        paid_loan_ids = set()
        for t in this_month_payments:
            lid = self._extract_loan_id_from_notes(t.notes)
            if lid:
                paid_loan_ids.add(lid)

        def fmt(v):
            return f"{v:,.2f} ₺".replace(",", "X").replace(".", ",").replace("X", ".")

        rows = []
        for l in loans:
            due_day = l.due_day or 15
            pay_day = min(due_day, last_day)
            payment_date = _date(today.year, today.month, pay_day)
            monthly = float(l.monthly_payment or 0)
            remaining = float(l.remaining_balance or 0)
            paid_inst = l.paid_installments or 0
            total_inst = l.total_installments
            rows.append({
                'loan_id': l.id,
                'loan_name': l.loan_name or '-',
                'bank_name': l.bank_name or '-',
                'company_name': l.company_name or '-',
                'monthly_payment': monthly,
                'remaining_balance': remaining,
                'payment_date': payment_date,
                'paid_installments': paid_inst,
                'total_installments': total_inst,
                'is_paid_this_month': l.id in paid_loan_ids,
            })

        rows.sort(key=lambda r: r['payment_date'])
        toplam = sum(r['monthly_payment'] for r in rows)
        odendi = sum(r['monthly_payment'] for r in rows if r['is_paid_this_month'])
        bekleyen = toplam - odendi

        html = self._rh("💳", "Bu Ay Ödenecek Krediler",
                        f"{today.strftime('%B %Y')} — {today.strftime('%d.%m.%Y')} itibarıyla",
                        "#1A237E")
        html += self._kpi_row([
            ("📋", "Aktif Kredi Sayısı", str(len(rows)), "#37474F"),
            ("💰", "Bu Ay Toplam Ödeme", fmt(toplam), "#1565C0"),
            ("✅", "Ödenmiş", fmt(odendi), "#2E7D32"),
            ("⏳", "Bekleyen Ödeme", fmt(bekleyen), "#E65100"),
        ])

        if not rows:
            html += "<p style='padding:20px; color:#888;'>Bu ay için aktif kredi bulunamadı.</p>"
            return html

        html += """<table style='width:100%; border-collapse:collapse; margin-top:16px; table-layout:fixed; font-size:12px;'>
            <colgroup>
                <col style='width:20%;'>
                <col style='width:13%;'>
                <col style='width:10%;'>
                <col style='width:10%;'>
                <col style='width:7%;'>
                <col style='width:14%;'>
                <col style='width:14%;'>
                <col style='width:12%;'>
            </colgroup>
            <tr style='background:#1A237E; color:#fff; font-size:12px;'>
                <th style='padding:8px 6px; text-align:left; overflow:hidden; white-space:nowrap;'>Kredi Adı</th>
                <th style='padding:8px 6px; text-align:left; overflow:hidden; white-space:nowrap;'>Banka</th>
                <th style='padding:8px 6px; text-align:left; overflow:hidden; white-space:nowrap;'>Firma</th>
                <th style='padding:8px 6px; text-align:center; overflow:hidden; white-space:nowrap;'>Ödeme Günü</th>
                <th style='padding:8px 6px; text-align:center; overflow:hidden; white-space:nowrap;'>Taksit</th>
                <th style='padding:8px 6px; text-align:right; overflow:hidden; white-space:nowrap;'>Aylık Tutar</th>
                <th style='padding:8px 6px; text-align:right; overflow:hidden; white-space:nowrap;'>Kalan Bakiye</th>
                <th style='padding:8px 6px; text-align:center; overflow:hidden; white-space:nowrap;'>Durum</th>
            </tr>"""

        for i, r in enumerate(rows):
            bg = "#fff" if i % 2 == 0 else "#F3F4F6"
            pay_str = r['payment_date'].strftime('%d.%m.%Y')
            days_to = (r['payment_date'] - today).days
            if r['is_paid_this_month']:
                durum = "<span style='color:#2E7D32; font-weight:bold;'>✅ Ödendi</span>"
                gun_str = f"<span style='color:#888;'>{pay_str}</span>"
            elif days_to == 0:
                durum = "<span style='color:#E65100; font-weight:bold;'>🔔 Bugün</span>"
                gun_str = f"<span style='color:#E65100; font-weight:bold;'>{pay_str}</span>"
            elif days_to <= 3:
                durum = f"<span style='color:#E65100; font-weight:bold;'>⚠️ {days_to} gün</span>"
                gun_str = f"<span style='color:#E65100;'>{pay_str}</span>"
            else:
                durum = f"<span style='color:#1565C0;'>🕐 {days_to} gün</span>"
                gun_str = f"<span style='color:#555;'>{pay_str}</span>"

            taksit_str = f"{r['paid_installments']}/{r['total_installments']}" if r['total_installments'] else f"{r['paid_installments']}/—"

            html += f"""<tr style='background:{bg};'>
                <td style='padding:7px 6px; font-weight:bold; overflow:hidden; white-space:nowrap; text-overflow:ellipsis;' title='{r["loan_name"]}'>{r['loan_name']}</td>
                <td style='padding:7px 6px; overflow:hidden; white-space:nowrap; text-overflow:ellipsis;'>{r['bank_name']}</td>
                <td style='padding:7px 6px; overflow:hidden; white-space:nowrap; text-overflow:ellipsis;'>{r['company_name']}</td>
                <td style='padding:7px 6px; text-align:center; white-space:nowrap;'>{gun_str}</td>
                <td style='padding:7px 6px; text-align:center; color:#555; white-space:nowrap;'>{taksit_str}</td>
                <td style='padding:7px 6px; text-align:right; font-weight:bold; white-space:nowrap;'>{fmt(r['monthly_payment'])}</td>
                <td style='padding:7px 6px; text-align:right; color:#B71C1C; white-space:nowrap;'>{fmt(r['remaining_balance'])}</td>
                <td style='padding:7px 6px; text-align:center; white-space:nowrap;'>{durum}</td>
            </tr>"""

        html += f"""<tr style='background:#1A237E; color:#fff; font-weight:bold;'>
            <td colspan='5' style='padding:10px 8px; text-align:right;'>TOPLAM:</td>
            <td style='padding:10px 8px; text-align:right;'>{fmt(toplam)}</td>
            <td colspan='2'></td>
        </tr>"""
        html += "</table>"
        return html

    def _generate_kredi_bitis_report(self):
        """Tüm kredileri bitiş tarihine göre sıralayan rapor."""
        from src.database.db import SessionLocal
        from src.database.models import Loan
        from datetime import date as _date

        today = _date.today()
        session = SessionLocal()
        try:
            loans = session.query(Loan).filter(
                Loan.user_id == self.user.id,
                Loan.is_active == True
            ).all()

            loan_rows = []
            for l in loans:
                remaining = max(0.0, max(float(l.remaining_balance or 0), float(l.loan_amount or 0)) - float(l.total_paid or 0))
                loan_rows.append({
                    'id': l.id,
                    'loan_name': l.loan_name or '-',
                    'bank_name': l.bank_name or '-',
                    'loan_type': l.loan_type or '-',
                    'monthly_payment': float(l.monthly_payment or 0),
                    'remaining_balance': remaining,
                    'status': l.status or 'AKTIF',
                    'start_date': l.start_date,
                    'end_date': l.end_date,
                    'total_installments': l.total_installments,
                    'paid_installments': l.paid_installments or 0,
                    'interest_rate': float(l.interest_rate or 0),
                })
        finally:
            session.close()

        # Bitiş tarihine göre sırala: önce tarihi olanlar (yakın → uzak), sonra tarihi olmayanlar
        with_date    = sorted([r for r in loan_rows if r['end_date']], key=lambda r: r['end_date'])
        without_date = [r for r in loan_rows if not r['end_date']]
        sorted_loans = with_date + without_date

        def fmt(v):
            return f"{v:,.2f} ₺".replace(",", "X").replace(".", ",").replace("X", ".")

        def date_cell(end_date):
            if not end_date:
                return "<span style='color:#888;'>—</span>"
            days_left = (end_date - today).days
            date_str = end_date.strftime('%d.%m.%Y')
            if days_left < 0:
                return f"<span style='color:#B71C1C; font-weight:bold;'>⛔ {date_str}<br><small>{abs(days_left)} gün geçti</small></span>"
            elif days_left <= 30:
                return f"<span style='color:#E65100; font-weight:bold;'>🔴 {date_str}<br><small>{days_left} gün kaldı</small></span>"
            elif days_left <= 90:
                return f"<span style='color:#F57F17; font-weight:bold;'>🟡 {date_str}<br><small>{days_left} gün kaldı</small></span>"
            else:
                return f"<span style='color:#2E7D32;'>🟢 {date_str}<br><small>{days_left} gün kaldı</small></span>"

        def status_badge(status):
            colors = {'AKTIF': '#1565C0', 'KAPATILDI': '#388E3C', 'IPTAL': '#B71C1C'}
            c = colors.get(status, '#555')
            return f"<span style='color:{c}; font-weight:bold;'>{status}</span>"

        # KPI hesapla
        active_count   = sum(1 for r in sorted_loans if r['status'] == 'AKTIF')
        expired_count  = sum(1 for r in sorted_loans if r['end_date'] and (r['end_date'] - today).days < 0 and r['status'] == 'AKTIF')
        soon_30_count  = sum(1 for r in sorted_loans if r['end_date'] and 0 <= (r['end_date'] - today).days <= 30 and r['status'] == 'AKTIF')
        total_remaining = sum(r['remaining_balance'] for r in sorted_loans)
        total_monthly   = sum(r['monthly_payment'] for r in sorted_loans if r['status'] == 'AKTIF')

        html = self._rh("⏰", "Kredi Bitiş Sıralaması",
                        f"Tüm krediler bitiş tarihine göre — {today.strftime('%d.%m.%Y')} itibarıyla",
                        "#4A148C")
        html += self._kpi_row([
            ("📋", "Toplam Kredi",          str(len(sorted_loans)),       "#37474F"),
            ("✅", "Aktif",                  str(active_count),            "#1565C0"),
            ("⛔", "Vadesi Geçmiş",          str(expired_count),           "#B71C1C"),
            ("🔴", "30 Gün İçinde Bitiyor",  str(soon_30_count),           "#E65100"),
        ])
        html += self._kpi_row([
            ("⏳", "Toplam Kalan Borç",      fmt(total_remaining),         "#6A1B9A"),
            ("💸", "Aylık Toplam Taksit",    fmt(total_monthly),           "#0D47A1"),
        ])

        if not sorted_loans:
            html += "<p style='padding:20px; color:#555;'>Kayıtlı kredi bulunamadı.</p>"
            return html

        html += self._section("⏰ Kredi Bitiş Sıralaması (Yakından Uzağa)", "#4A148C")
        html += self._table_header(
            ["Kredi Adı", "Banka", "Tip", "Faiz %", "Aylık Taksit", "Kalan Borç", "Bitiş Tarihi", "Durum"],
            color="#EDE7F6"
        )
        for i, r in enumerate(sorted_loans):
            bg = "white" if i % 2 == 0 else "#FAFAFA"
            taksit_str = (
                f"{r['paid_installments']}/{r['total_installments']}"
                if r['total_installments'] else f"{r['paid_installments']} ödendi"
            )
            html += (
                f"<tr style='background:{bg};'>"
                f"<td style='padding:7px 8px; font-weight:bold;'>{r['loan_name']}</td>"
                f"<td style='padding:7px 8px;'>{r['bank_name']}</td>"
                f"<td style='padding:7px 8px;'>{r['loan_type']}</td>"
                f"<td style='padding:7px 8px; text-align:center;'>{r['interest_rate']:.2f}%</td>"
                f"<td style='padding:7px 8px; text-align:right;'>{fmt(r['monthly_payment'])}</td>"
                f"<td style='padding:7px 8px; text-align:right; color:#B71C1C; font-weight:bold;'>{fmt(r['remaining_balance'])}</td>"
                f"<td style='padding:7px 8px;'>{date_cell(r['end_date'])}</td>"
                f"<td style='padding:7px 8px;'>{status_badge(r['status'])}</td>"
                f"</tr>"
            )
        html += "</table><br>"
        return html

    def _generate_nakit_kasasi_report(self, start_date, end_date):
        """Nakit Kasası raporu — nakit giriş/çıkış ve güncel bakiye"""
        from src.database.db import SessionLocal
        from src.database.models import Transaction, TransactionType, PaymentMethod

        session = SessionLocal()
        try:
            # ── TÜM ZAMANLAR: Güncel nakit bakiye hesapla ─────────────────
            all_tx = session.query(Transaction).filter(
                Transaction.user_id == self.user.id
            ).order_by(Transaction.transaction_date.asc()).all()

            NAKIT_GIRIS_TYPES  = {TransactionType.GELIR, TransactionType.KESILEN_FATURA}
            NAKIT_CIKIS_TYPES  = {TransactionType.GIDER, TransactionType.GELEN_FATURA,
                                  TransactionType.KREDI_ODEME, TransactionType.KREDI_KARTI_ODEME,
                                  TransactionType.EK_HESAP_FAIZLERI, TransactionType.KREDI_DOSYA_MASRAFI,
                                  TransactionType.EKSPERTIZ_UCRETI}

            toplam_nakit_giris  = 0.0
            toplam_nakit_cikis  = 0.0
            toplam_cekilis      = 0.0  # bankadan çekilen
            toplam_yatirim      = 0.0  # bankaya yatırılan

            for tx in all_tx:
                if tx.transaction_type == TransactionType.NAKIT_CEKIMI:
                    toplam_cekilis += tx.amount
                elif tx.transaction_type == TransactionType.NAKIT_YATIRIMI:
                    toplam_yatirim += tx.amount
                elif tx.payment_method == PaymentMethod.NAKIT:
                    if tx.transaction_type in NAKIT_GIRIS_TYPES:
                        toplam_nakit_giris += tx.amount
                    elif tx.transaction_type in NAKIT_CIKIS_TYPES:
                        toplam_nakit_cikis += tx.amount

            nakit_bakiye = toplam_nakit_giris + toplam_cekilis - toplam_nakit_cikis - toplam_yatirim

            # ── TARİH ARALIĞI: Seçili döneme ait işlemler ─────────────────
            donem_tx = session.query(Transaction).filter(
                Transaction.user_id == self.user.id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).order_by(Transaction.transaction_date.desc()).all()

            donem_giris  = 0.0
            donem_cikis  = 0.0
            donem_cekilis = 0.0
            donem_yatirim = 0.0
            nakit_islemler = []

            for tx in donem_tx:
                if tx.transaction_type == TransactionType.NAKIT_CEKIMI:
                    donem_cekilis += tx.amount
                    nakit_islemler.append((tx, '+', tx.amount, '#4CAF50', 'Nakit Çekim (Bankadan)'))
                elif tx.transaction_type == TransactionType.NAKIT_YATIRIMI:
                    donem_yatirim += tx.amount
                    nakit_islemler.append((tx, '-', tx.amount, '#f44336', 'Nakit Yatırım (Bankaya)'))
                elif tx.payment_method == PaymentMethod.NAKIT:
                    if tx.transaction_type in NAKIT_GIRIS_TYPES:
                        donem_giris += tx.amount
                        nakit_islemler.append((tx, '+', tx.amount, '#4CAF50',
                                               tx.transaction_type.value))
                    elif tx.transaction_type in NAKIT_CIKIS_TYPES:
                        donem_cikis += tx.amount
                        nakit_islemler.append((tx, '-', tx.amount, '#f44336',
                                               tx.transaction_type.value))

            donem_net = donem_giris + donem_cekilis - donem_cikis - donem_yatirim

        finally:
            session.close()

        # ── HTML ───────────────────────────────────────────────────────────
        bakiye_renk = '#4CAF50' if nakit_bakiye >= 0 else '#f44336'
        donem_renk  = '#4CAF50' if donem_net >= 0 else '#f44336'

        html  = self._rh('💵', 'Nakit Kasası Raporu',
                         f'Dönem: {start_date} — {end_date}', '#2E7D32')
        html += self._kpi_row([
            ('💵', 'Güncel Nakit Bakiye', f'{nakit_bakiye:,.2f} ₺', bakiye_renk),
            ('📥', 'Toplam Banka Çekim',  f'{toplam_cekilis:,.2f} ₺', '#1565C0'),
            ('📤', 'Toplam Banka Yatırım',f'{toplam_yatirim:,.2f} ₺', '#FF6F00'),
            ('📊', 'Dönem Net Değişim',   f'{donem_net:,.2f} ₺', donem_renk),
        ])

        html += self._kpi_row([
            ('📈', 'Dönem Nakit Giriş',   f'{donem_giris + donem_cekilis:,.2f} ₺', '#388E3C'),
            ('📉', 'Dönem Nakit Çıkış',   f'{donem_cikis + donem_yatirim:,.2f} ₺', '#C62828'),
            ('🏦', 'Tüm Zamanlarda Çekim',f'{toplam_cekilis:,.2f} ₺', '#1976D2'),
            ('🏦', 'Tüm Zamanlarda Yatırım',f'{toplam_yatirim:,.2f} ₺', '#E65100'),
        ])

        # Nakit kasası özeti kutusu
        html += f"""
        <table width='100%' cellpadding='0' cellspacing='0' style='margin-top:12px;'>
        <tr><td style='background-color:#1B5E20; padding:14px 18px; border-radius:6px;'>
            <span style='font-size:12pt; font-weight:bold; color:white;'>💵 Nakit Kasası Güncel Durumu</span><br>
            <span style='font-size:10pt; color:#A5D6A7;'>
                Nakit Gelir: <b>{toplam_nakit_giris:,.2f} ₺</b> &nbsp;+&nbsp;
                Bankadan Çekim: <b>{toplam_cekilis:,.2f} ₺</b> &nbsp;−&nbsp;
                Nakit Gider: <b>{toplam_nakit_cikis:,.2f} ₺</b> &nbsp;−&nbsp;
                Bankaya Yatırım: <b>{toplam_yatirim:,.2f} ₺</b> &nbsp;=&nbsp;
                <span style='color:#69F0AE; font-size:12pt;'><b>Kasa: {nakit_bakiye:,.2f} ₺</b></span>
            </span>
        </td></tr></table><br>
        """

        # Seçili dönemdeki nakit işlemler tablosu
        html += self._section(f'📋 Dönem Nakit İşlemleri ({start_date} — {end_date})', '#2E7D32')

        if not nakit_islemler:
            html += """
            <div style='margin:16px; padding:14px; background:#1e2a3a;
                        border-left:4px solid #FF9800; border-radius:4px;'>
                <span style='color:#FF9800; font-weight:bold;'>ℹ️ Bu tarih aralığında nakit işlem bulunamadı.</span>
            </div>"""
        else:
            html += self._table_header(
                ['Tarih', 'G/Ç', 'Tür', 'Müşteri/Açıklama', 'Tutar (₺)'],
                '#2E7D32')
            for i, (tx, yon, amount, renk, tur) in enumerate(nakit_islemler):
                aciklama = (tx.customer_name or '') + (' — ' + (tx.description or '')[:50] if tx.description else '')
                bg = '#1e2a3a' if i % 2 == 0 else '#1a2332'
                html += self._tr(
                    [str(tx.transaction_date), yon, tur, aciklama, f'{amount:,.2f}'],
                    ['#aaa', renk, '#e0e0e0', '#ccc', renk],
                    bold=False, bg=bg
                )
            html += '</table>'

        return html

    def create_payroll_tab(self) -> QWidget:
        """Maaş bordro hesaplama sekmesi"""
        outer_widget = QWidget()
        outer_layout = QVBoxLayout(outer_widget)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        # Scroll area – küçük çözünürlüklerde kaydırılabilir
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        inner_widget = QWidget()
        layout = QVBoxLayout(inner_widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        title = QLabel("💼 Maaş Bordro Hesaplama")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        layout.addWidget(title)

        # ── Bordro Bilgileri GroupBox ─────────────────────────────────────
        form_group = QGroupBox("Bordro Bilgileri")
        form_grid = QGridLayout()
        form_grid.setSpacing(8)
        form_grid.setColumnStretch(1, 1)
        form_grid.setColumnStretch(3, 1)
        LABEL_MIN = 210

        # Personel Adı – tam genişlik
        lbl_emp = QLabel("Personel Adı:")
        lbl_emp.setMinimumWidth(LABEL_MIN)
        self.payroll_employee_input = QLineEdit()
        self.payroll_employee_input.setPlaceholderText("Örn: Ahmet Yılmaz")
        form_grid.addWidget(lbl_emp, 0, 0)
        form_grid.addWidget(self.payroll_employee_input, 0, 1, 1, 3)

        # Bordro Dönemi
        lbl_period = QLabel("Bordro Dönemi:")
        lbl_period.setMinimumWidth(LABEL_MIN)
        self.payroll_period_date = QDateEdit()
        self.payroll_period_date.setDate(QDate.currentDate())
        self.payroll_period_date.setDisplayFormat("MM.yyyy")
        self.payroll_period_date.setMaximumWidth(120)
        form_grid.addWidget(lbl_period, 1, 0)
        form_grid.addWidget(self.payroll_period_date, 1, 1)

        # Alan tanımları – (key, etiket, varsayılan)
        self.payroll_fields = {}
        field_defs = [
            ("gross_salary",      "Brüt Maaş (₺)",                 "0"),
            ("month_days",        "Ay Gün Sayısı",                  "30"),
            ("worked_days",       "Çalışılan Gün",                  "30"),
            ("paid_leave_days",   "Ücretli İzin Gün",               "0"),
            ("unpaid_leave_days", "Ücretsiz İzin Gün",              "0"),
            ("child_count",       "Bakmakla Yükümlü Çocuk Sayısı",  "0"),
            ("overtime_hours",    "Fazla Mesai (Saat)",              "0"),
            ("overtime_rate",     "Fazla Mesai Saat Ücreti (₺)",    "0"),
            ("bonus",             "Prim / İkramiye (₺)",            "0"),
            ("meal",              "Yemek Yardımı (₺)",              "0"),
            ("transport",         "Yol Yardımı (₺)",                "0"),
            ("advance",           "Avans Kesintisi (₺)",            "0"),
            ("private_insurance", "Özel Sigorta Kesintisi (₺)",     "0"),
            ("other_deductions",  "Diğer Kesintiler (₺)",           "0"),
            ("sgk_rate",          "SGK İşçi Oranı (%)",             "14"),
            ("unemployment_rate", "İşsizlik Sigortası (%)",         "1"),
            ("income_tax_rate",   "Gelir Vergisi Oranı (%)",        "15"),
            ("stamp_tax_rate",    "Damga Vergisi Oranı (%)",        "0.759"),
        ]
        self.payroll_field_defaults = {key: default for key, _, default in field_defs}

        # Grid'e ekle: çift index → sol sütun, tek index → sağ sütun
        for idx, (key, label, default) in enumerate(field_defs):
            row_offset = idx // 2 + 2          # 2'den başla (0,1 yukarıda dolu)
            col_offset = (idx % 2) * 2         # 0 veya 2

            lbl = QLabel(label + ":")
            lbl.setMinimumWidth(LABEL_MIN)
            inp = QLineEdit(default)
            inp.setMinimumWidth(110)

            form_grid.addWidget(lbl, row_offset, col_offset)
            form_grid.addWidget(inp, row_offset, col_offset + 1)
            self.payroll_fields[key] = inp

        form_group.setLayout(form_grid)
        layout.addWidget(form_group)

        # ── Ayarlar GroupBox ──────────────────────────────────────────────
        opt_group = QGroupBox("Ayarlar")
        opt_grid = QGridLayout()
        opt_grid.setSpacing(8)
        opt_grid.setColumnStretch(1, 1)
        opt_grid.setColumnStretch(3, 1)

        # Asgari ücret modu butonu – tam genişlik
        self.asgari_mode_button = QPushButton("🎯 Asgari Ücret Modu: Kapalı")
        self.asgari_mode_button.setCheckable(True)
        self.asgari_mode_button.setMinimumHeight(34)
        self.asgari_mode_button.toggled.connect(self.toggle_asgari_ucret_mode)
        opt_grid.addWidget(self.asgari_mode_button, 0, 0, 1, 4)

        # Sol: asgari istisna + asgari ücret brüt + kümülatif dilim
        self.apply_min_wage_exemption_check = QCheckBox("Asgari vergi istisnasını uygula")
        self.apply_min_wage_exemption_check.setChecked(True)
        opt_grid.addWidget(self.apply_min_wage_exemption_check, 1, 0, 1, 2)

        lbl_min_wage = QLabel("Asgari Ücret Brüt (₺):")
        lbl_min_wage.setMinimumWidth(LABEL_MIN)
        self.min_wage_gross_input = QLineEdit("33030")
        self.min_wage_gross_input.setMinimumWidth(110)
        opt_grid.addWidget(lbl_min_wage, 2, 0)
        opt_grid.addWidget(self.min_wage_gross_input, 2, 1)

        self.progressive_tax_check = QCheckBox("Kümülatif vergi dilimi hesabı uygula")
        self.progressive_tax_check.setChecked(True)
        opt_grid.addWidget(self.progressive_tax_check, 3, 0, 1, 2)

        # Sağ: vergi dilimleri + kümülatif matrah + manuel matrah
        lbl_brackets = QLabel("Vergi Dilimleri (üst:oran%):")
        lbl_brackets.setMinimumWidth(LABEL_MIN)
        self.tax_brackets_input = QLineEdit("158000:15,330000:20,1200000:27,4300000:35,sonsuz:40")
        self.tax_brackets_input.setToolTip("Örn: 158000:15,330000:20,1200000:27,4300000:35,sonsuz:40")
        opt_grid.addWidget(lbl_brackets, 1, 2)
        opt_grid.addWidget(self.tax_brackets_input, 1, 3)

        self.use_saved_cumulative_check = QCheckBox("Kümülatif matrahı sistemde kaydedilen bordrolardan al")
        self.use_saved_cumulative_check.setChecked(True)
        opt_grid.addWidget(self.use_saved_cumulative_check, 2, 2, 1, 2)

        lbl_prev = QLabel("Manuel Önceki Kümülatif Matrah (₺):")
        lbl_prev.setMinimumWidth(LABEL_MIN)
        self.manual_prev_tax_base_input = QLineEdit("0")
        self.manual_prev_tax_base_input.setMinimumWidth(110)
        opt_grid.addWidget(lbl_prev, 3, 2)
        opt_grid.addWidget(self.manual_prev_tax_base_input, 3, 3)

        opt_group.setLayout(opt_grid)
        layout.addWidget(opt_group)

        # ── Net → Brüt GroupBox ───────────────────────────────────────────
        reverse_group = QGroupBox("🔄 Net Maaştan Brütü Hesapla (Ters Hesaplama)")
        rev_grid = QGridLayout()
        rev_grid.setSpacing(8)
        rev_grid.setColumnStretch(1, 1)
        rev_grid.setColumnStretch(3, 1)

        rev_grid.addWidget(QLabel("Net Maaş (₺):"), 0, 0)
        self.reverse_net_input = QLineEdit("0")
        self.reverse_net_input.setMinimumWidth(110)
        rev_grid.addWidget(self.reverse_net_input, 0, 1)

        rev_grid.addWidget(QLabel("Çocuk Sayısı:"), 0, 2)
        self.reverse_child_count_input = QLineEdit("0")
        self.reverse_child_count_input.setMinimumWidth(110)
        rev_grid.addWidget(self.reverse_child_count_input, 0, 3)

        btn_reverse = QPushButton("🔢 Net'ten Brütü Hesapla")
        btn_reverse.setMinimumHeight(34)
        btn_reverse.clicked.connect(self.calculate_gross_from_net)
        rev_grid.addWidget(btn_reverse, 1, 0, 1, 4)

        reverse_group.setLayout(rev_grid)
        layout.addWidget(reverse_group)

        # ── Aksiyon butonları ─────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        btn_calculate = QPushButton("🧮 Bordroyu Hesapla")
        btn_calculate.setMinimumHeight(38)
        btn_calculate.clicked.connect(self.calculate_payroll)
        btn_row.addWidget(btn_calculate)

        btn_export_pdf = QPushButton("📄 Bordro PDF Çıktısı Al")
        btn_export_pdf.setMinimumHeight(38)
        btn_export_pdf.clicked.connect(self.export_payroll_to_pdf)
        btn_row.addWidget(btn_export_pdf)

        btn_save_payroll = QPushButton("💾 Bordroyu Sisteme Kaydet")
        btn_save_payroll.setMinimumHeight(38)
        btn_save_payroll.clicked.connect(self.save_last_payroll_record)
        btn_row.addWidget(btn_save_payroll)

        btn_reset = QPushButton("🧹 Formu Sıfırla")
        btn_reset.setMinimumHeight(38)
        btn_reset.clicked.connect(self.reset_payroll_form)
        btn_row.addWidget(btn_reset)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        # ── Sonuç alanı ───────────────────────────────────────────────────
        self.payroll_result_view = QTextBrowser()
        # Kendi scroll'unu kapat — dış QScrollArea halleder
        self.payroll_result_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.payroll_result_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.payroll_result_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.payroll_result_view.setMinimumHeight(40)
        # İçerik değişince yüksekliği otomatik ayarla
        self.payroll_result_view.document().contentsChanged.connect(self._adjust_payroll_result_height)
        self.payroll_result_view.setHtml(
            "<p><i>Bordro hesaplamak için bilgileri girip 'Bordroyu Hesapla' butonuna basın.</i></p>"
        )
        layout.addWidget(self.payroll_result_view)

        self.last_payroll_html = ""
        self.last_payroll_employee = ""
        self.last_payroll_period = ""
        self.last_payroll_data = {}

        scroll.setWidget(inner_widget)
        outer_layout.addWidget(scroll)
        self.payroll_scroll = scroll
        return outer_widget

    def create_payroll_records_tab(self) -> QWidget:
        """Kaydedilen bordroları listeleme ve arama sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        title = QLabel("📚 Kaydedilen Bordrolar")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        layout.addWidget(title)

        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("Personel Ara:"))
        self.payroll_search_input = QLineEdit()
        self.payroll_search_input.setPlaceholderText("Örn: Burak Tekin")
        self.payroll_search_input.setMinimumWidth(240)
        search_row.addWidget(self.payroll_search_input)

        btn_search = QPushButton("🔎 Ara")
        btn_search.setMinimumHeight(34)
        btn_search.clicked.connect(self.search_saved_payroll_records)
        search_row.addWidget(btn_search)

        btn_show_all = QPushButton("📋 Tümünü Göster")
        btn_show_all.setMinimumHeight(34)
        btn_show_all.clicked.connect(self.show_all_saved_payroll_records)
        search_row.addWidget(btn_show_all)
        search_row.addStretch()
        layout.addLayout(search_row)

        action_row = QHBoxLayout()
        btn_refresh_saved = QPushButton("🔄 Kayıtları Yenile")
        btn_refresh_saved.setMinimumHeight(34)
        btn_refresh_saved.clicked.connect(self.refresh_saved_payroll_records_table)
        action_row.addWidget(btn_refresh_saved)

        btn_preview_saved_pdf = QPushButton("👁️ Ön İzleme")
        btn_preview_saved_pdf.setMinimumHeight(34)
        btn_preview_saved_pdf.clicked.connect(self.preview_selected_payroll_record_pdf)
        action_row.addWidget(btn_preview_saved_pdf)

        btn_export_saved_pdf = QPushButton("📄 Seçili Kaydı PDF Çıktı Al")
        btn_export_saved_pdf.setMinimumHeight(34)
        btn_export_saved_pdf.clicked.connect(self.export_selected_payroll_record_pdf)
        action_row.addWidget(btn_export_saved_pdf)

        btn_preview_all_pdf = QPushButton("👁️ Toplu Ön İzleme")
        btn_preview_all_pdf.setMinimumHeight(34)
        btn_preview_all_pdf.clicked.connect(self.preview_all_payroll_records_pdf)
        action_row.addWidget(btn_preview_all_pdf)

        btn_export_all_pdf = QPushButton("📑 Tüm Bordroları PDF Yap")
        btn_export_all_pdf.setMinimumHeight(34)
        btn_export_all_pdf.clicked.connect(self.export_all_payroll_records_to_pdf)
        action_row.addWidget(btn_export_all_pdf)

        btn_delete_saved = QPushButton("🗑️ Seçili Kaydı Sil")
        btn_delete_saved.setMinimumHeight(34)
        btn_delete_saved.clicked.connect(self.delete_selected_payroll_record)
        action_row.addWidget(btn_delete_saved)
        action_row.addStretch()
        layout.addLayout(action_row)

        self.saved_payroll_table = QTableWidget()
        self.saved_payroll_table.setColumnCount(8)
        self.saved_payroll_table.setHorizontalHeaderLabels([
            "Personel", "Dönem", "Brüt", "Vergi Matrahı", "Gelir Vergisi", "Damga Vergisi", "Net", "Kaydedilme"
        ])
        self.saved_payroll_table.horizontalHeader().setStretchLastSection(True)
        self.saved_payroll_table.setMinimumHeight(500)
        layout.addWidget(self.saved_payroll_table)

        self.payroll_records_cache = []
        self.refresh_saved_payroll_records_table()

        widget.setLayout(layout)
        return widget

    def _parse_payroll_float(self, input_widget: QLineEdit, field_name: str) -> float:
        raw_value = input_widget.text().strip()
        if not raw_value:
            return 0.0

        if "," in raw_value and "." in raw_value:
            if raw_value.rfind(",") > raw_value.rfind("."):
                normalized = raw_value.replace(".", "").replace(",", ".")
            else:
                normalized = raw_value.replace(",", "")
        elif "," in raw_value:
            normalized = raw_value.replace(",", ".")
        else:
            normalized = raw_value

        try:
            return float(normalized)
        except ValueError:
            raise ValueError(f"Geçersiz sayı: {field_name}")

    def _adjust_payroll_result_height(self):
        """QTextBrowser'ı içerik yüksekliğine göre boyutlandır, dış scroll ile görünür kıl."""
        if not hasattr(self, 'payroll_result_view'):
            return
        doc = self.payroll_result_view.document()
        # Döküman genişliğini widget genişliğine sabitle
        doc.setTextWidth(self.payroll_result_view.viewport().width())
        content_height = int(doc.size().height()) + 20
        self.payroll_result_view.setMinimumHeight(content_height)
        self.payroll_result_view.setMaximumHeight(content_height)

    def _scroll_to_payroll_result(self):
        """Dış scroll'u sonuç kutusuna kaydır."""
        if hasattr(self, 'payroll_scroll') and hasattr(self, 'payroll_result_view'):
            self.payroll_scroll.ensureWidgetVisible(self.payroll_result_view, 0, 0)

    def toggle_asgari_ucret_mode(self, enabled: bool):
        """Asgari ücret modunda gelir ve damga vergisini sıfırla"""
        income_input = self.payroll_fields.get("income_tax_rate")
        stamp_input = self.payroll_fields.get("stamp_tax_rate")

        if not income_input or not stamp_input:
            return

        if enabled:
            self._prev_income_tax_rate = income_input.text().strip() or "15"
            self._prev_stamp_tax_rate = stamp_input.text().strip() or "0.759"
            income_input.setText("0")
            stamp_input.setText("0")
            income_input.setEnabled(False)
            stamp_input.setEnabled(False)
            self.asgari_mode_button.setText("🎯 Asgari Ücret Modu: Açık")
        else:
            income_input.setEnabled(True)
            stamp_input.setEnabled(True)
            income_input.setText(getattr(self, "_prev_income_tax_rate", "15"))
            stamp_input.setText(getattr(self, "_prev_stamp_tax_rate", "0.759"))
            self.asgari_mode_button.setText("🎯 Asgari Ücret Modu: Kapalı")

    def reset_payroll_form(self):
        """Bordro alanlarını varsayılan değerlere sıfırla"""
        self.payroll_employee_input.clear()
        self.payroll_period_date.setDate(QDate.currentDate())

        if hasattr(self, "asgari_mode_button") and self.asgari_mode_button.isChecked():
            self.asgari_mode_button.setChecked(False)

        if hasattr(self, "apply_min_wage_exemption_check"):
            self.apply_min_wage_exemption_check.setChecked(True)

        if hasattr(self, "min_wage_gross_input"):
            self.min_wage_gross_input.setText("33030")

        if hasattr(self, "progressive_tax_check"):
            self.progressive_tax_check.setChecked(True)

        if hasattr(self, "tax_brackets_input"):
            self.tax_brackets_input.setText("158000:15,330000:20,1200000:27,4300000:35,sonsuz:40")

        if hasattr(self, "use_saved_cumulative_check"):
            self.use_saved_cumulative_check.setChecked(True)

        if hasattr(self, "manual_prev_tax_base_input"):
            self.manual_prev_tax_base_input.setText("0")

        defaults = getattr(self, "payroll_field_defaults", {})
        for key, inp in self.payroll_fields.items():
            inp.setEnabled(True)
            inp.setText(defaults.get(key, "0"))

        self.last_payroll_html = ""
        self.last_payroll_employee = ""
        self.last_payroll_period = ""
        self.last_payroll_data = {}
        self.payroll_result_view.setHtml("<p><i>Bordro hesaplamak için bilgileri girip 'Bordroyu Hesapla' butonuna basın.</i></p>")

    def _parse_tax_brackets(self, raw_text: str):
        brackets = []
        raw_parts = [part.strip() for part in (raw_text or "").split(",") if part.strip()]
        if not raw_parts:
            raise ValueError("Vergi dilimleri boş olamaz")

        for part in raw_parts:
            if ":" not in part:
                raise ValueError(f"Geçersiz dilim formatı: {part}")
            limit_raw, rate_raw = part.split(":", 1)
            limit_raw = limit_raw.strip().lower()
            rate = float(rate_raw.strip().replace(",", ".")) / 100

            if limit_raw in ["sonsuz", "inf", "infinity", "max"]:
                limit = None
            else:
                limit = float(limit_raw.replace(".", "").replace(",", "."))
            brackets.append((limit, rate))

        finite_brackets = [b for b in brackets if b[0] is not None]
        finite_brackets.sort(key=lambda item: item[0])
        infinite_brackets = [b for b in brackets if b[0] is None]
        if len(infinite_brackets) > 1:
            raise ValueError("Sadece bir adet sonsuz dilim tanımlanabilir")
        if infinite_brackets:
            finite_brackets.append(infinite_brackets[0])
        return finite_brackets

    def _compute_progressive_tax(self, taxable_amount: float, brackets):
        if taxable_amount <= 0:
            return 0.0

        total_tax = 0.0
        lower_limit = 0.0
        remaining = taxable_amount

        for upper_limit, rate in brackets:
            if upper_limit is None:
                total_tax += remaining * rate
                break

            if taxable_amount <= lower_limit:
                break

            band_width = upper_limit - lower_limit
            taxable_in_band = min(remaining, band_width)
            if taxable_in_band > 0:
                total_tax += taxable_in_band * rate
                remaining -= taxable_in_band

            lower_limit = upper_limit
            if remaining <= 0:
                break

        return total_tax

    def _load_saved_payroll_records(self):
        records = UserSettingsService.get_json_setting(self.user.id, "payroll_records", [])
        if isinstance(records, list):
            return records
        return []

    def _save_payroll_records(self, records):
        UserSettingsService.set_json_setting(self.user.id, "payroll_records", records)

    def refresh_saved_payroll_records_table(self):
        """Kaydedilen bordroları tabloda göster"""
        self._render_saved_payroll_records(self._load_saved_payroll_records())

    def _filter_payroll_payment_transactions(self, transactions, employee_name, month, year):
        name = (employee_name or "").casefold().strip()
        if not name:
            return []

        results = []
        for trans in transactions:
            desc = (getattr(trans, "description", "") or "").casefold()
            if "maaş ödemesi" not in desc and "maas odemesi" not in desc:
                continue

            customer_name = (getattr(trans, "customer_name", "") or "").casefold()
            subject = (getattr(trans, "subject", "") or "").casefold()
            if name not in desc and name not in customer_name and name not in subject:
                continue

            if month and year and getattr(trans, "transaction_date", None):
                if trans.transaction_date.month != month or trans.transaction_date.year != year:
                    continue

            results.append(trans)

        return results

    def show_payroll_payment_statement(self, employee_name, month, year):
        try:
            from src.database.db import SessionLocal
            from src.database.models import Transaction

            session = SessionLocal()
            transactions = session.query(Transaction).filter(
                Transaction.user_id == self.user.id
            ).all()
            session.close()

            matches = self._filter_payroll_payment_transactions(
                transactions, employee_name, month, year
            )

            dialog = QDialog(self)
            period = f"{month:02d}.{year}" if month and year else "-"
            dialog.setWindowTitle(f"Maaş Dökümü - {employee_name} ({period})")
            dialog.setMinimumSize(700, 420)

            layout = QVBoxLayout(dialog)
            info_label = QLabel(f"Personel: {employee_name}   |   Dönem: {period}")
            info_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
            layout.addWidget(info_label)

            table = QTableWidget()
            table.setColumnCount(3)
            table.setHorizontalHeaderLabels(["Tarih", "Açıklama", "Tutar"])
            table.horizontalHeader().setStretchLastSection(True)
            table.setRowCount(len(matches))

            total_paid = 0.0
            matches = sorted(matches, key=lambda t: t.transaction_date)
            for row, trans in enumerate(matches):
                date_text = str(trans.transaction_date) if trans.transaction_date else "-"
                desc = trans.description or ""
                amount = float(trans.amount or 0.0)
                total_paid += amount

                table.setItem(row, 0, QTableWidgetItem(date_text))
                table.setItem(row, 1, QTableWidgetItem(desc))
                table.setItem(row, 2, QTableWidgetItem(f"{format_tr(amount)} ₺"))

            layout.addWidget(table)

            total_label = QLabel(f"Toplam Ödenen: {format_tr(total_paid)} ₺")
            total_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
            total_label.setAlignment(Qt.AlignRight)
            layout.addWidget(total_label)

            btn_close = QPushButton("Kapat")
            btn_close.clicked.connect(dialog.close)
            btn_row = QHBoxLayout()
            btn_row.addStretch()
            btn_row.addWidget(btn_close)
            layout.addLayout(btn_row)

            dialog.exec_()
        except Exception as e:
            QMessageBox.warning(self, "Uyarı", f"Maaş dökümü açılamadı: {e}")

    def _render_saved_payroll_records(self, records):
        """Verilen bordro kayıtlarını tabloya bas"""
        records = list(records or [])
        records.sort(key=lambda r: (int(r.get("year", 0)), int(r.get("month", 0)), str(r.get("employee", "")).lower()))
        self.payroll_records_cache = records


        if not hasattr(self, "saved_payroll_table"):
            return

        self.saved_payroll_table.setRowCount(len(records))

        # Helper function for Turkish number formatting
        def format_tr(val):
            """Türkçe sayı formatı: 7.521,00"""
            try:
                val_float = float(val)
            except (ValueError, TypeError):
                return str(val)
            s = f"{val_float:.2f}"
            parts = s.split('.')
            integer_part = parts[0]
            decimal_part = parts[1] if len(parts) > 1 else "00"
            result = ""
            for i, digit in enumerate(reversed(integer_part)):
                if i > 0 and i % 3 == 0:
                    result = "." + result
                result = digit + result
            return f"{result},{decimal_part}"

        for row, rec in enumerate(records):
            period = f"{int(rec.get('month', 0)):02d}.{int(rec.get('year', 0))}" if rec.get('month') and rec.get('year') else "-"
            saved_at = str(rec.get("saved_at", ""))[:19].replace("T", " ") if rec.get("saved_at") else "-"

            gross = float(rec.get('gross_total', 0.0) or 0.0)
            sgk = float(rec.get('sgk_deduction', 0.0) or 0.0)
            unemployment = float(rec.get('unemployment_deduction', 0.0) or 0.0)
            income_tax = float(rec.get('income_tax', 0.0) or 0.0)
            stamp_tax = float(rec.get('stamp_tax', 0.0) or 0.0)
            net = float(rec.get('net_salary', 0.0) or 0.0)
            
            # Calculate tax base properly: gross - sgk - unemployment
            tax_base = gross - sgk - unemployment

            values = [
                str(rec.get("employee", "-")),
                period,
                format_tr(gross),
                format_tr(tax_base),
                format_tr(income_tax),
                format_tr(stamp_tax),
                format_tr(net),
                saved_at,
            ]

            for col, value in enumerate(values):
                self.saved_payroll_table.setItem(row, col, QTableWidgetItem(value))

    def search_saved_payroll_records(self):
        """Personel adına göre kaydedilen bordroları filtrele"""
        query = self.payroll_search_input.text().strip().lower() if hasattr(self, "payroll_search_input") else ""
        records = self._load_saved_payroll_records()
        if query:
            records = [
                r for r in records
                if query in str(r.get("employee", "")).strip().lower()
            ]
        self._render_saved_payroll_records(records)

    def show_all_saved_payroll_records(self):
        """Tüm kayıtlı bordroları göster"""
        if hasattr(self, "payroll_search_input"):
            self.payroll_search_input.clear()
        self.refresh_saved_payroll_records_table()

    def export_all_payroll_records_to_pdf(self):
        """Tüm kayıtlı bordroları resmi vergi formatında PDF oluştur"""
        records = self._load_saved_payroll_records()
        if not records:
            QMessageBox.warning(self, "Uyarı", "Kaydedilmiş bordro bulunamadı")
            return

        try:
            from PyQt5.QtWidgets import QFileDialog
            from PyQt5.QtPrintSupport import QPrinter
            from PyQt5.QtGui import QTextDocument
            from datetime import datetime

            num_records = len(records)
            default_name = f"toplu_bordro_{num_records}_kayit.pdf"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Toplu Bordro PDF Kaydet",
                default_name,
                "PDF Dosyası (*.pdf)"
            )

            if not file_path:
                return

            if not file_path.lower().endswith(".pdf"):
                file_path += ".pdf"

            def get_float(record, key, default=0.0):
                return float(record.get(key, default) or default)

            def get_int(record, key, default=0):
                return int(record.get(key, default) or default)

            def format_money(val):
                return format_tr(val)

            today = datetime.now()
            year = records[0].get("year", today.year) if records else today.year
            month = records[0].get("month", today.month) if records else today.month
            
            # Ay adları
            months_tr = {
                1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
                7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
            }
            month_name = months_tr.get(month, str(month))

            # Veri hazırla
            table_data = []
            totals = {
                "worked_days": 0, "paid_leave": 0, "unpaid_leave": 0,
                "gross": 0.0, "sgk": 0.0, "unemployment": 0.0, 
                "income_tax": 0.0, "stamp_tax": 0.0, "net": 0.0
            }

            for idx, record in enumerate(records, 1):
                employee = str(record.get("employee", "-"))
                
                worked = get_int(record, "worked_days")
                paid_leave = get_int(record, "paid_leave_days")
                unpaid = get_int(record, "unpaid_leave_days")
                
                gross = get_float(record, "gross_total")
                sgk = get_float(record, "sgk_deduction")
                unemployment = get_float(record, "unemployment_deduction")
                income_tax = get_float(record, "income_tax")
                stamp_tax = get_float(record, "stamp_tax")
                net = get_float(record, "net_salary")
                
                totals["worked_days"] += worked
                totals["paid_leave"] += paid_leave
                totals["unpaid_leave"] += unpaid
                totals["gross"] += gross
                totals["sgk"] += sgk
                totals["unemployment"] += unemployment
                totals["income_tax"] += income_tax
                totals["stamp_tax"] += stamp_tax
                totals["net"] += net

                table_data.append({
                    "no": idx,
                    "name": employee,
                    "worked": worked,
                    "paid_leave": paid_leave,
                    "unpaid": unpaid,
                    "gross": gross,
                    "sgk": sgk,
                    "unemployment": unemployment,
                    "income_tax": income_tax,
                    "stamp_tax": stamp_tax,
                    "net": net
                })

            # HTML Oluştur - RESMİ FORMATTA
            html = f"""
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    * {{ margin: 0; padding: 0; }}
                    body {{ 
                        font-family: Arial, sans-serif;
                        font-size: 11pt;
                        padding: 10px;
                        line-height: 1.4;
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 15px;
                        border-top: 3px solid #000;
                        border-bottom: 2px solid #000;
                        padding: 10px 0;
                    }}
                    .title {{
                        font-size: 16pt;
                        font-weight: bold;
                        margin: 8px 0;
                    }}
                    .subtitle {{
                        font-size: 12pt;
                        font-weight: bold;
                    }}
                    .info-row {{
                        display: table;
                        width: 100%;
                        margin: 3px 0;
                        font-size: 9pt;
                    }}
                    .info-col {{
                        display: table-cell;
                        width: 33%;
                        padding: 2px 5px;
                    }}
                    .info-label {{
                        font-weight: bold;
                    }}
                    table {
                        border-collapse: collapse;
                        width: 100%;
                        margin-left: 0;
                        margin-top: 12px;
                        font-size: 12pt;
                        table-layout: auto;
                    }
                    thead tr {{
                        background-color: #ddd;
                    }}
                    th {
                        border: 1px solid #000;
                        padding: 8px 8px;
                        font-weight: bold;
                        text-align: center;
                        font-size: 12pt;
                        white-space: normal;
                    }
                    td {
                        border: 1px solid #000;
                        padding: 8px 8px;
                        text-align: right;
                        font-family: 'Courier New';
                        white-space: normal;
                    }
                    td.name {{
                        text-align: left;
                    }}
                    td.center {{
                        text-align: center;
                    }}
                    tr.total td {{
                        background-color: #ddd;
                        font-weight: bold;
                        border-top: 2px solid #000;
                    }}
                    .footer {{
                        margin-top: 15px;
                        font-size: 9pt;
                        text-align: center;
                        border-top: 1px solid #000;
                        padding-top: 8px;
                        padding-bottom: 10px;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="title">{month_name} {year} AYLIK ÜCRET BORDROSU</div>
                    <div class="info-row">
                        <div class="info-col"><span class="info-label">Dönem:</span> {month:02d}/{year}</div>
                        <div class="info-col"><span class="info-label">Personel Sayısı:</span> {num_records}</div>
                        <div class="info-col"><span class="info-label">Tarih:</span> {today.strftime('%d.%m.%Y')}</div>
                    </div>
                </div>

                <table>
                    <thead>
                        <tr>
                            <th>S.No</th>
                            <th>Personel Adı</th>
                            <th>Çalış. Gün</th>
                            <th>Ücretli İzin</th>
                            <th>Ücretsiz İzin</th>
                            <th>Brüt Maaş</th>
                            <th>SGK %14</th>
                            <th>İşsizlik %1</th>
                            <th>Gelir Vergisi</th>
                            <th>Damga Vergisi</th>
                            <th>Net Maaş</th>
                        </tr>
                    </thead>
                    <tbody>
            """

            # Satırları ekle
            for row in table_data:
                html += f"""
                        <tr>
                            <td class="center">{row['no']}</td>
                            <td class="name">{row['name']}</td>
                            <td class="center">{row['worked']}</td>
                            <td class="center">{row['paid_leave']}</td>
                            <td class="center">{row['unpaid']}</td>
                            <td>{format_money(row['gross'])}</td>
                            <td>{format_money(row['sgk'])}</td>
                            <td>{format_money(row['unemployment'])}</td>
                            <td>{format_money(row['income_tax'])}</td>
                            <td>{format_money(row['stamp_tax'])}</td>
                            <td><strong>{format_money(row['net'])}</strong></td>
                        </tr>
                """

            # Toplam satırı
            html += f"""
                        <tr class="total">
                            <td class="center" colspan="2">TOPLAM</td>
                            <td class="center">{totals['worked_days']}</td>
                            <td class="center">{totals['paid_leave']}</td>
                            <td class="center">{totals['unpaid_leave']}</td>
                            <td>{format_money(totals['gross'])}</td>
                            <td>{format_money(totals['sgk'])}</td>
                            <td>{format_money(totals['unemployment'])}</td>
                            <td>{format_money(totals['income_tax'])}</td>
                            <td>{format_money(totals['stamp_tax'])}</td>
                            <td><strong>{format_money(totals['net'])}</strong></td>
                        </tr>
                    </tbody>
                </table>

                <div class="footer">
                    <p><strong>Bu bordro resmi nitelik taşımamakta olup bilgilendirme amaçlıdır.</strong></p>
                    <p>Muhasebe müdürü/İşletmeci tarafından onanmalıdır.</p>
                    <p style="margin-top: 15px; font-size: 11pt; font-weight: bold;">
                        TOPLAM ÖDENECEK TUTARI: <span style="color: #000; text-decoration: underline;">{format_money(totals['net'])} TL</span>
                    </p>
                </div>
            </body>
            </html>
            """

            # PDF oluştur
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageSize(QPrinter.A4)
            printer.setOrientation(QPrinter.Landscape)
            printer.setPageMargins(8, 8, 8, 8, QPrinter.Millimeter)

            document = QTextDocument()
            document.setHtml(html)
            document.print_(printer)

            QMessageBox.information(self, "Başarılı", 
                f"Toplu bordro ({num_records} personel) PDF olarak kaydedildi:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Toplu PDF oluşturulamadı:\n{str(e)}")

    def delete_selected_payroll_record(self):
        """Seçilen bordro kaydını sil"""
        if not hasattr(self, "saved_payroll_table"):
            return

        row = self.saved_payroll_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Silmek için tablodan bir kayıt seçin")
            return

        if row >= len(self.payroll_records_cache):
            QMessageBox.warning(self, "Uyarı", "Seçilen kayıt bulunamadı")
            return

        selected_record = self.payroll_records_cache[row]
        confirm = QMessageBox.question(
            self,
            "Kayıt Sil",
            f"Seçili bordro kaydı silinsin mi?\n\nPersonel: {selected_record.get('employee', '-')}\nDönem: {int(selected_record.get('month', 0)):02d}.{int(selected_record.get('year', 0))}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if confirm != QMessageBox.Yes:
            return

        records = self._load_saved_payroll_records()
        records = [
            r for r in records
            if not (
                str(r.get("employee", "")).strip().lower() == str(selected_record.get("employee", "")).strip().lower()
                and int(r.get("year", 0)) == int(selected_record.get("year", 0))
                and int(r.get("month", 0)) == int(selected_record.get("month", 0))
            )
        ]

        self._save_payroll_records(records)
        self.refresh_saved_payroll_records_table()
        QMessageBox.information(self, "Başarılı", "Seçili bordro kaydı silindi")

    def preview_selected_payroll_record_pdf(self):
        """Seçilen bordronun ön izlemesini göster"""
        if not hasattr(self, "saved_payroll_table"):
            return

        row = self.saved_payroll_table.currentRow()
        if row < 0 or row >= len(self.payroll_records_cache):
            QMessageBox.warning(self, "Uyarı", "Ön izleme için tablodan bir kayıt seçin")
            return

        record = self.payroll_records_cache[row]

        try:
            from PyQt5.QtPrintSupport import QPrinter, QPrintPreviewDialog
            from PyQt5.QtGui import QTextDocument

            employee = str(record.get("employee", "Belirtilmedi"))
            year = int(record.get("year", 0) or 0)
            month = int(record.get("month", 0) or 0)
            saved_at = str(record.get("saved_at", "-")).replace("T", " ")[:19]

            def get_float(key, default=0.0):
                return float(record.get(key, default) or default)
            
            def get_int(key, default=0):
                return int(record.get(key, default) or default)

            gross_salary = get_float("gross_salary")
            month_days = get_int("month_days", 30)
            worked_days = get_int("worked_days")
            paid_leave_days = get_int("paid_leave_days")
            unpaid_leave_days = get_int("unpaid_leave_days")
            child_count = get_int("child_count")
            paid_days = get_int("paid_days")
            day_ratio = get_float("day_ratio", 1.0)
            overtime_amount = get_float("overtime_amount")
            bonus = get_float("bonus")
            meal = get_float("meal")
            transport = get_float("transport")
            gross_total = get_float("gross_total")
            sgk_rate = get_float("sgk_rate")
            sgk_deduction = get_float("sgk_deduction")
            unemployment_rate = get_float("unemployment_rate")
            unemployment_deduction = get_float("unemployment_deduction")
            tax_base = get_float("tax_base")
            child_allowance = get_float("child_allowance")
            income_tax_rate = get_float("income_tax_rate")
            income_tax = get_float("income_tax")
            stamp_tax_rate = get_float("stamp_tax_rate")
            stamp_tax = get_float("stamp_tax")
            income_tax_exemption = get_float("income_tax_exemption")
            stamp_tax_exemption = get_float("stamp_tax_exemption")
            total_deductions = get_float("total_deductions")
            advance = get_float("advance")
            private_insurance = get_float("private_insurance")
            other_deductions = get_float("other_deductions")
            additional_deductions = get_float("additional_deductions")
            total_exemptions = get_float("total_exemptions")
            net_salary = get_float("net_salary")

            def format_money(val):
                return format_tr(val)

            months_tr = {
                1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
                7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
            }
            month_name = months_tr.get(month, str(month))

            html = f"""
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    * {{ margin: 0; padding: 0; }}
                    body {{ 
                        font-family: Arial, sans-serif;
                        font-size: 9pt;
                        padding: 8px;
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 12px;
                        border-top: 2px solid #000;
                        border-bottom: 2px solid #000;
                        padding: 6px 0;
                    }}
                    .title {{
                        font-size: 12pt;
                        font-weight: bold;
                        margin: 3px 0;
                    }}
                    .info {{
                        font-size: 9pt;
                        display: flex;
                        justify-content: space-around;
                        padding: 2px 0;
                    }}
                    .info-item {{
                        flex: 1;
                        text-align: center;
                    }}
                    .section-title {{
                        font-weight: bold;
                        background-color: #ddd;
                        padding: 2px 4px;
                        margin-top: 6px;
                        margin-bottom: 3px;
                    }}
                    table {{
                        border-collapse: collapse;
                        width: 100%;
                        border: 1px solid #000;
                        margin-bottom: 3px;
                        font-size: 8pt;
                    }}
                    th, td {{
                        border: 1px solid #000;
                        padding: 2px 3px;
                        text-align: center;
                    }}
                    th {{
                        background-color: #e0e0e0;
                        font-weight: bold;
                    }}
                    td {{
                        height: 16px;
                    }}
                    .net-section {{
                        text-align: center;
                        margin-top: 10px;
                        font-weight: bold;
                        font-size: 11pt;
                        padding: 8px;
                        border: 2px solid #000;
                        background-color: #f0f0f0;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="title">{month_name} {year} AYLIK ÜCRET BORDROSU</div>
                    <div class="info">
                        <div class="info-item"><strong>Personel:</strong> {employee}</div>
                        <div class="info-item"><strong>Dönem:</strong> {month:02d}.{year}</div>
                        <div class="info-item"><strong>Tarih:</strong> {saved_at}</div>
                    </div>
                </div>

                <div class="section-title">ÇALIŞMA BİLGİLERİ</div>
                <table>
                    <tr>
                        <th>Ay Gün</th>
                        <th>Çalış.</th>
                        <th>Ücretli İzin</th>
                        <th>Ücretsiz İzin</th>
                        <th>Ücretl. Gün</th>
                        <th>Gün Oranı</th>
                        <th>Çocuk Sayısı</th>
                        <th>Çocuk İnd.</th>
                    </tr>
                    <tr>
                        <td>{month_days:.0f}</td>
                        <td>{worked_days:.0f}</td>
                        <td>{paid_leave_days:.0f}</td>
                        <td>{unpaid_leave_days:.0f}</td>
                        <td>{paid_days:.0f}</td>
                        <td>{day_ratio*100:.1f}%</td>
                        <td>{child_count:.0f}</td>
                        <td>{format_money(child_allowance)}</td>
                    </tr>
                </table>

                <div class="section-title">BRÜT ÜCRET BİLEŞENLERİ</div>
                <table>
                    <tr>
                        <th>Brüt (Tam)</th>
                        <th>Brüt (Gün Oranlı)</th>
                        <th>Fazla Mesai</th>
                        <th>Prim/İkramiye</th>
                        <th>Yemek Yardımı</th>
                        <th>Yol Yardımı</th>
                        <th><strong>Toplam Brüt</strong></th>
                    </tr>
                    <tr>
                        <td>{format_money(gross_salary)}</td>
                        <td>{format_money(gross_salary * day_ratio)}</td>
                        <td>{format_money(overtime_amount)}</td>
                        <td>{format_money(bonus)}</td>
                        <td>{format_money(meal)}</td>
                        <td>{format_money(transport)}</td>
                        <td><strong>{format_money(gross_total)}</strong></td>
                    </tr>
                </table>

                <div class="section-title">VERGİ MATRAHÍ HESAPLAMASI</div>
                <table>
                    <tr>
                        <th>SGK İşçi (%{sgk_rate:.3g})</th>
                        <th>İşsizlik (%{unemployment_rate:.3g})</th>
                        <th>Orijinal Matrah</th>
                        <th>Çocuk İnd. (-)</th>
                        <th><strong>İndirimli Matrah</strong></th>
                    </tr>
                    <tr>
                        <td>{format_money(sgk_deduction)}</td>
                        <td>{format_money(unemployment_deduction)}</td>
                        <td>{format_money(tax_base + child_allowance)}</td>
                        <td>{format_money(child_allowance)}</td>
                        <td><strong>{format_money(tax_base)}</strong></td>
                    </tr>
                </table>

                <div class="section-title">KESİNTİLER</div>
                <table>
                    <tr>
                        <th>Gelir Vergisi (%{income_tax_rate:.3g})</th>
                        <th>Damga Vergisi (%{stamp_tax_rate:.3g})</th>
                        <th>Gel.Vergi İst.</th>
                        <th>Damga İst.</th>
                        <th>Avans</th>
                        <th>Özel Sigorta</th>
                        <th>Diğer Kes.</th>
                        <th><strong>Toplam Kes.</strong></th>
                    </tr>
                    <tr>
                        <td>{format_money(income_tax)}</td>
                        <td>{format_money(stamp_tax)}</td>
                        <td>-{format_money(income_tax_exemption)}</td>
                        <td>-{format_money(stamp_tax_exemption)}</td>
                        <td>{format_money(advance)}</td>
                        <td>{format_money(private_insurance)}</td>
                        <td>{format_money(other_deductions)}</td>
                        <td><strong>{format_money(total_deductions + additional_deductions)}</strong></td>
                    </tr>
                </table>

                <div class="net-section">
                    <div style="margin-bottom: 8px;">NET ÜCRET: <strong>{format_money(net_salary)} TL</strong></div>
                </div>
            </body>
            </html>
            """

            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPrinter.A4)
            printer.setOrientation(QPrinter.Landscape)
            printer.setPageMargins(6, 6, 6, 6, QPrinter.Millimeter)

            document = QTextDocument()
            document.setHtml(html)

            dialog = QPrintPreviewDialog(printer, self)
            dialog.paintRequested.connect(lambda p: document.print_(p))
            dialog.exec_()

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ön izleme oluşturulamadı:\n{str(e)}")

    def preview_all_payroll_records_pdf(self):
        """Tüm bordroların ön izlemesini göster"""
        records = self._load_saved_payroll_records()
        if not records:
            QMessageBox.warning(self, "Uyarı", "Kaydedilmiş bordro bulunamadı")
            return

        try:
            from PyQt5.QtPrintSupport import QPrinter, QPrintPreviewDialog
            from PyQt5.QtGui import QTextDocument
            from datetime import datetime

            num_records = len(records)

            def get_float(record, key, default=0.0):
                return float(record.get(key, default) or default)

            def get_int(record, key, default=0):
                return int(record.get(key, default) or default)

            def format_money(val):
                return format_tr(val)

            today = datetime.now()
            year = records[0].get("year", today.year) if records else today.year
            month = records[0].get("month", today.month) if records else today.month
            
            months_tr = {
                1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
                7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
            }
            month_name = months_tr.get(month, str(month))

            # Veri hazırla
            table_data = []
            totals = {
                "worked_days": 0, "paid_leave": 0, "unpaid_leave": 0,
                "gross": 0.0, "sgk": 0.0, "unemployment": 0.0, 
                "income_tax": 0.0, "stamp_tax": 0.0, "net": 0.0
            }

            for idx, record in enumerate(records, 1):
                employee = str(record.get("employee", "-"))
                
                worked = get_int(record, "worked_days")
                paid_leave = get_int(record, "paid_leave_days")
                unpaid = get_int(record, "unpaid_leave_days")
                
                gross = get_float(record, "gross_total")
                sgk = get_float(record, "sgk_deduction")
                unemployment = get_float(record, "unemployment_deduction")
                income_tax = get_float(record, "income_tax")
                stamp_tax = get_float(record, "stamp_tax")
                net = get_float(record, "net_salary")
                
                totals["worked_days"] += worked
                totals["paid_leave"] += paid_leave
                totals["unpaid_leave"] += unpaid
                totals["gross"] += gross
                totals["sgk"] += sgk
                totals["unemployment"] += unemployment
                totals["income_tax"] += income_tax
                totals["stamp_tax"] += stamp_tax
                totals["net"] += net

                table_data.append({
                    "no": idx,
                    "name": employee,
                    "worked": worked,
                    "paid_leave": paid_leave,
                    "unpaid": unpaid,
                    "gross": gross,
                    "sgk": sgk,
                    "unemployment": unemployment,
                    "income_tax": income_tax,
                    "stamp_tax": stamp_tax,
                    "net": net
                })

            # HTML Oluştur
            html = f"""
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    * {{ margin: 0; padding: 0; }}
                    body {{ 
                        font-family: Arial, sans-serif;
                        font-size: 10pt;
                        padding: 8px;
                        line-height: 1.2;
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 10px;
                        border-top: 3px solid #000;
                        border-bottom: 2px solid #000;
                        padding: 8px 0;
                    }}
                    .title {{
                        font-size: 14pt;
                        font-weight: bold;
                        margin: 5px 0;
                    }}
                    .subtitle {{
                        font-size: 11pt;
                        font-weight: bold;
                    }}
                    .info-row {{
                        display: table;
                        width: 100%;
                        margin: 5px 0;
                        font-size: 11pt;
                    }}
                    .info-col {{
                        display: table-cell;
                        width: 33%;
                        padding: 4px 10px;
                    }}
                    .info-label {{
                        font-weight: bold;
                    }}
                    table {{
                        border-collapse: collapse;
                        width: 100%;
                        margin: 12px 0;
                        font-size: 10pt;
                    }}
                    thead tr {{
                        background-color: #ddd;
                    }}
                    th {{
                        border: 1px solid #000;
                        padding: 4px 2px;
                        font-weight: bold;
                        text-align: center;
                        font-size: 8pt;
                    }}
                    td {{
                        border: 1px solid #000;
                        padding: 2px 4px;
                        text-align: right;
                        font-family: 'Courier New';
                    }}
                    td.name {{
                        text-align: left;
                    }}
                    td.center {{
                        text-align: center;
                    }}
                    tr.total td {{
                        background-color: #ddd;
                        font-weight: bold;
                        border-top: 2px solid #000;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="title">{month_name} {year} AYLIK ÜCRET BORDROSU</div>
                    <div class="info-row">
                        <div class="info-col"><span class="info-label">Dönem:</span> {month:02d}/{year}</div>
                        <div class="info-col"><span class="info-label">Personel Sayısı:</span> {num_records}</div>
                        <div class="info-col"><span class="info-label">Tarih:</span> {today.strftime('%d.%m.%Y')}</div>
                    </div>
                </div>

                <table>
                    <thead>
                        <tr>
                            <th style="width:4%">S.No</th>
                            <th style="width:18%">Personel Adı</th>
                            <th style="width:6%">Çalış. Gün</th>
                            <th style="width:6%">Ücretli İzin</th>
                            <th style="width:6%">Ücretsiz İzin</th>
                            <th style="width:10%">Brüt Maaş</th>
                            <th style="width:10%">SGK %14</th>
                            <th style="width:10%">İşsizlik %1</th>
                            <th style="width:10%">Gelir Vergisi</th>
                            <th style="width:10%">Damga Vergisi</th>
                            <th style="width:10%">Net Maaş</th>
                        </tr>
                    </thead>
                    <tbody>
            """

            # Satırları ekle
            for row in table_data:
                html += f"""
                        <tr>
                            <td class="center">{row['no']}</td>
                            <td class="name">{row['name']}</td>
                            <td class="center">{row['worked']}</td>
                            <td class="center">{row['paid_leave']}</td>
                            <td class="center">{row['unpaid']}</td>
                            <td>{format_money(row['gross'])}</td>
                            <td>{format_money(row['sgk'])}</td>
                            <td>{format_money(row['unemployment'])}</td>
                            <td>{format_money(row['income_tax'])}</td>
                            <td>{format_money(row['stamp_tax'])}</td>
                            <td><strong>{format_money(row['net'])}</strong></td>
                        </tr>
                """

            # Toplam satırı
            html += f"""
                        <tr class="total">
                            <td class="center" colspan="2">TOPLAM</td>
                            <td class="center">{totals['worked_days']}</td>
                            <td class="center">{totals['paid_leave']}</td>
                            <td class="center">{totals['unpaid_leave']}</td>
                            <td>{format_money(totals['gross'])}</td>
                            <td>{format_money(totals['sgk'])}</td>
                            <td>{format_money(totals['unemployment'])}</td>
                            <td>{format_money(totals['income_tax'])}</td>
                            <td>{format_money(totals['stamp_tax'])}</td>
                            <td><strong>{format_money(totals['net'])}</strong></td>
                        </tr>
                    </tbody>
                </table>

                <div style="margin-top: 12px; font-size: 9pt; text-align: center; border-top: 1px solid #000; padding-top: 8px;">
                    <p><strong>Bu bordro resmi nitelik taşımamakta olup bilgilendirme amaçlıdır.</strong></p>
                    <p>Muhasebe müdürü/İşletmeci tarafından onanmalıdır.</p>
                    <p style="margin-top: 15px; font-size: 11pt; font-weight: bold;">
                        TOPLAM ÖDENECEK TUTARI: <span style="text-decoration: underline;">{format_money(totals['net'])} TL</span>
                    </p>
                </div>
            </body>
            </html>
            """

            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPrinter.A4)
            printer.setOrientation(QPrinter.Landscape)
            printer.setPageMargins(8, 8, 8, 8, QPrinter.Millimeter)

            document = QTextDocument()
            document.setHtml(html)

            dialog = QPrintPreviewDialog(printer, self)
            dialog.paintRequested.connect(lambda p: document.print_(p))
            dialog.exec_()

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ön izleme oluşturulamadı:\n{str(e)}")

    def export_selected_payroll_record_pdf(self):
        """Seçilen kayıtlı bordroyu PDF çıktısı olarak kaydet (Landscape - Yan Yana)"""
        if not hasattr(self, "saved_payroll_table"):
            return

        row = self.saved_payroll_table.currentRow()
        if row < 0 or row >= len(self.payroll_records_cache):
            QMessageBox.warning(self, "Uyarı", "PDF çıktısı için tablodan bir kayıt seçin")
            return

        record = self.payroll_records_cache[row]

        try:
            from PyQt5.QtWidgets import QFileDialog
            from PyQt5.QtPrintSupport import QPrinter
            from PyQt5.QtGui import QTextDocument

            employee = str(record.get("employee", "Belirtilmedi"))
            year = int(record.get("year", 0) or 0)
            month = int(record.get("month", 0) or 0)
            
            period_text = f"{month:02d}.{year}" if month and year else "-"
            saved_at = str(record.get("saved_at", "-")).replace("T", " ")[:19]

            # Tüm detayları al (eski kayıtlarda olmayabilir - varsayılan değer set et)
            def get_float(key, default=0.0):
                return float(record.get(key, default) or default)
            
            def get_int(key, default=0):
                return int(record.get(key, default) or default)

            gross_salary = get_float("gross_salary")
            month_days = get_int("month_days", 30)
            worked_days = get_int("worked_days")
            paid_leave_days = get_int("paid_leave_days")
            unpaid_leave_days = get_int("unpaid_leave_days")
            child_count = get_int("child_count")
            paid_days = get_int("paid_days")
            day_ratio = get_float("day_ratio", 1.0)
            overtime_hours = get_float("overtime_hours")
            overtime_rate = get_float("overtime_rate")
            overtime_amount = get_float("overtime_amount")
            bonus = get_float("bonus")
            meal = get_float("meal")
            transport = get_float("transport")
            gross_total = get_float("gross_total")
            sgk_rate = get_float("sgk_rate")
            sgk_deduction = get_float("sgk_deduction")
            unemployment_rate = get_float("unemployment_rate")
            unemployment_deduction = get_float("unemployment_deduction")
            tax_base = get_float("tax_base")
            child_allowance = get_float("child_allowance")
            income_tax_rate = get_float("income_tax_rate")
            income_tax = get_float("income_tax")
            stamp_tax_rate = get_float("stamp_tax_rate")
            stamp_tax = get_float("stamp_tax")
            income_tax_exemption = get_float("income_tax_exemption")
            stamp_tax_exemption = get_float("stamp_tax_exemption")
            total_deductions = get_float("total_deductions")
            advance = get_float("advance")
            private_insurance = get_float("private_insurance")
            other_deductions = get_float("other_deductions")
            additional_deductions = get_float("additional_deductions")
            total_exemptions = get_float("total_exemptions")
            net_salary = get_float("net_salary")

            def format_money(val):
                return format_tr(val)

            # Ay adları
            months_tr = {
                1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
                7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
            }
            month_name = months_tr.get(month, str(month))

            html = f"""
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    * {{ margin: 0; padding: 0; }}
                    body {{ 
                        font-family: Arial, sans-serif;
                        font-size: 9pt;
                        padding: 8px;
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 12px;
                        border-top: 2px solid #000;
                        border-bottom: 2px solid #000;
                        padding: 6px 0;
                    }}
                    .title {{
                        font-size: 12pt;
                        font-weight: bold;
                        margin: 3px 0;
                    }}
                    .info {{
                        font-size: 9pt;
                        display: flex;
                        justify-content: space-around;
                        padding: 2px 0;
                    }}
                    .info-item {{
                        flex: 1;
                        text-align: center;
                    }}
                    .section-title {{
                        font-weight: bold;
                        background-color: #ddd;
                        padding: 2px 4px;
                        margin-top: 6px;
                        margin-bottom: 3px;
                    }}
                    table {{
                        border-collapse: collapse;
                        width: 100%;
                        border: 1px solid #000;
                        margin-bottom: 3px;
                        font-size: 8pt;
                    }}
                    th, td {{
                        border: 1px solid #000;
                        padding: 2px 3px;
                        text-align: center;
                    }}
                    th {{
                        background-color: #e0e0e0;
                        font-weight: bold;
                    }}
                    td {{
                        height: 16px;
                    }}
                    tbody tr:first-child td {{
                        border-top: 1px solid #000;
                    }}
                    .label {{
                        text-align: left;
                        font-weight: bold;
                        background-color: #f5f5f5;
                    }}
                    .footer {{
                        margin-top: 12px;
                        font-size: 9pt;
                        text-align: center;
                        border-top: 1px solid #000;
                        padding-top: 8px;
                    }}
                    .net-section {{
                        text-align: center;
                        margin-top: 10px;
                        font-weight: bold;
                        font-size: 11pt;
                        padding: 8px;
                        border: 2px solid #000;
                        background-color: #f0f0f0;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="title">{month_name} {year} AYLIK ÜCRET BORDROSU</div>
                    <div class="info">
                        <div class="info-item"><strong>Personel:</strong> {employee}</div>
                        <div class="info-item"><strong>Dönem:</strong> {period_text}</div>
                        <div class="info-item"><strong>Tarih:</strong> {saved_at}</div>
                    </div>
                </div>

                <div class="section-title">ÇALIŞMA BİLGİLERİ</div>
                <table>
                    <tr>
                        <th>Ay Gün</th>
                        <th>Çalış.</th>
                        <th>Ücretli İzin</th>
                        <th>Ücretsiz İzin</th>
                        <th>Ücretl. Gün</th>
                        <th>Gün Oranı</th>
                        <th>Çocuk Sayısı</th>
                        <th>Çocuk İnd.</th>
                    </tr>
                    <tr>
                        <td>{month_days:.0f}</td>
                        <td>{worked_days:.0f}</td>
                        <td>{paid_leave_days:.0f}</td>
                        <td>{unpaid_leave_days:.0f}</td>
                        <td>{paid_days:.0f}</td>
                        <td>{day_ratio*100:.1f}%</td>
                        <td>{child_count:.0f}</td>
                        <td>{format_money(child_allowance)}</td>
                    </tr>
                </table>

                <div class="section-title">BRÜT ÜCRET BİLEŞENLERİ</div>
                <table>
                    <tr>
                        <th>Brüt (Tam)</th>
                        <th>Brüt (Gün Oranlı)</th>
                        <th>Fazla Mesai</th>
                        <th>Prim/İkramiye</th>
                        <th>Yemek Yardımı</th>
                        <th>Yol Yardımı</th>
                        <th><strong>Toplam Brüt</strong></th>
                    </tr>
                    <tr>
                        <td>{format_money(gross_salary)}</td>
                        <td>{format_money(gross_salary * day_ratio)}</td>
                        <td>{format_money(overtime_amount)}</td>
                        <td>{format_money(bonus)}</td>
                        <td>{format_money(meal)}</td>
                        <td>{format_money(transport)}</td>
                        <td><strong>{format_money(gross_total)}</strong></td>
                    </tr>
                </table>

                <div class="section-title">VERGİ MATRAHÍ HESAPLAMASI</div>
                <table>
                    <tr>
                        <th>SGK İşçi (%{sgk_rate:.3g})</th>
                        <th>İşsizlik (%{unemployment_rate:.3g})</th>
                        <th>Orijinal Matrah</th>
                        <th>Çocuk İnd. (-)</th>
                        <th><strong>İndirimli Matrah</strong></th>
                    </tr>
                    <tr>
                        <td>{format_money(sgk_deduction)}</td>
                        <td>{format_money(unemployment_deduction)}</td>
                        <td>{format_money(tax_base + child_allowance)}</td>
                        <td>{format_money(child_allowance)}</td>
                        <td><strong>{format_money(tax_base)}</strong></td>
                    </tr>
                </table>

                <div class="section-title">KESİNTİLER</div>
                <table>
                    <tr>
                        <th>Gelir Vergisi (%{income_tax_rate:.3g})</th>
                        <th>Damga Vergisi (%{stamp_tax_rate:.3g})</th>
                        <th>Gel.Vergi İst.</th>
                        <th>Damga İst.</th>
                        <th>Avans</th>
                        <th>Özel Sigorta</th>
                        <th>Diğer Kes.</th>
                        <th><strong>Toplam Kes.</strong></th>
                    </tr>
                    <tr>
                        <td>{format_money(income_tax)}</td>
                        <td>{format_money(stamp_tax)}</td>
                        <td>-{format_money(income_tax_exemption)}</td>
                        <td>-{format_money(stamp_tax_exemption)}</td>
                        <td>{format_money(advance)}</td>
                        <td>{format_money(private_insurance)}</td>
                        <td>{format_money(other_deductions)}</td>
                        <td><strong>{format_money(total_deductions + additional_deductions)}</strong></td>
                    </tr>
                </table>

                <div class="net-section">
                    <div style="margin-bottom: 8px;">NET ÜCRET: <strong>{format_money(net_salary)} TL</strong></div>
                </div>

                <div class="footer">
                    <p><strong>Bu bordro resmi nitelik taşımamakta olup bilgilendirme amaçlıdır.</strong></p>
                    <p>Muhasebe müdürü/İşletmeci tarafından onanmalıdır.</p>
                </div>
            </body>
            </html>
            """

            safe_employee = employee.replace(" ", "_")
            default_name = f"kayitli_bordro_{safe_employee}_{period_text.replace('.', '_')}.pdf"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Kayıtlı Bordro PDF Kaydet",
                default_name,
                "PDF Dosyası (*.pdf)"
            )

            if not file_path:
                return

            if not file_path.lower().endswith(".pdf"):
                file_path += ".pdf"

            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageSize(QPrinter.A4)
            printer.setOrientation(QPrinter.Landscape)
            printer.setPageMargins(6, 6, 6, 6, QPrinter.Millimeter)

            document = QTextDocument()
            document.setHtml(html)
            document.print_(printer)

            QMessageBox.information(self, "Başarılı", f"Kayıtlı bordro PDF olarak kaydedildi:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{str(e)}")

    def _get_saved_cumulative_tax_base(self, employee_name: str, year: int, month: int):
        normalized_name = (employee_name or "").strip().lower()
        if not normalized_name:
            return 0.0

        total = 0.0
        for record in self._load_saved_payroll_records():
            if str(record.get("employee", "")).strip().lower() != normalized_name:
                continue
            if int(record.get("year", 0)) != int(year):
                continue
            if int(record.get("month", 0)) >= int(month):
                continue
            total += float(record.get("tax_base", 0.0) or 0.0)
        return total

    def save_last_payroll_record(self):
        """Son hesaplanan bordroyu sistem verisine kaydet"""
        if not self.last_payroll_data:
            QMessageBox.warning(self, "Uyarı", "Önce bordroyu hesaplamalısınız")
            return

        try:
            record = self.last_payroll_data.copy()
            record["saved_at"] = datetime.now().isoformat()

            records = self._load_saved_payroll_records()
            records = [
                r for r in records
                if not (
                    str(r.get("employee", "")).strip().lower() == str(record["employee"]).strip().lower()
                    and int(r.get("year", 0)) == int(record["year"])
                    and int(r.get("month", 0)) == int(record["month"])
                )
            ]
            records.append(record)
            records.sort(key=lambda r: (str(r.get("employee", "")).lower(), int(r.get("year", 0)), int(r.get("month", 0))))
            self._save_payroll_records(records)
            self.refresh_saved_payroll_records_table()

            QMessageBox.information(self, "Başarılı", "Bordro sisteme kaydedildi. Sonraki aylarda kümülatif matrah hesabında kullanılacak.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Bordro kaydedilemedi:\n{str(e)}")

    def calculate_payroll(self):
        """Maaş bordrosunu hesapla ve göster"""
        try:
            def money(value):
                return round(float(value) + 1e-9, 2)

            def fmt_money(value):
                return format_tr(money(value))

            employee_name = self.payroll_employee_input.text().strip() or "Belirtilmedi"
            period_qdate = self.payroll_period_date.date()
            period_text = period_qdate.toString("MM.yyyy")
            period_year = period_qdate.year()
            period_month = period_qdate.month()

            gross_salary = self._parse_payroll_float(self.payroll_fields["gross_salary"], "Brüt Maaş")
            month_days = max(self._parse_payroll_float(self.payroll_fields["month_days"], "Ay Gün Sayısı"), 1)
            worked_days = max(self._parse_payroll_float(self.payroll_fields["worked_days"], "Çalışılan Gün"), 0)
            paid_leave_days = max(self._parse_payroll_float(self.payroll_fields["paid_leave_days"], "Ücretli İzin Gün"), 0)
            unpaid_leave_days = max(self._parse_payroll_float(self.payroll_fields["unpaid_leave_days"], "Ücretsiz İzin Gün"), 0)
            child_count = max(int(self._parse_payroll_float(self.payroll_fields["child_count"], "Çocuk Sayısı")), 0)
            overtime_hours = self._parse_payroll_float(self.payroll_fields["overtime_hours"], "Fazla Mesai Saat")
            overtime_rate = self._parse_payroll_float(self.payroll_fields["overtime_rate"], "Fazla Mesai Ücreti")
            bonus = self._parse_payroll_float(self.payroll_fields["bonus"], "Prim")
            meal = self._parse_payroll_float(self.payroll_fields["meal"], "Yemek Yardımı")
            transport = self._parse_payroll_float(self.payroll_fields["transport"], "Yol Yardımı")
            advance = max(self._parse_payroll_float(self.payroll_fields["advance"], "Avans Kesintisi"), 0)
            private_insurance = max(self._parse_payroll_float(self.payroll_fields["private_insurance"], "Özel Sigorta Kesintisi"), 0)
            other_deductions = max(self._parse_payroll_float(self.payroll_fields["other_deductions"], "Diğer Kesintiler"), 0)

            sgk_rate = self._parse_payroll_float(self.payroll_fields["sgk_rate"], "SGK Oranı")
            unemployment_rate = self._parse_payroll_float(self.payroll_fields["unemployment_rate"], "İşsizlik Sigortası Oranı")
            income_tax_rate = self._parse_payroll_float(self.payroll_fields["income_tax_rate"], "Gelir Vergisi Oranı")
            stamp_tax_rate = self._parse_payroll_float(self.payroll_fields["stamp_tax_rate"], "Damga Vergisi Oranı")
            min_wage_gross = self._parse_payroll_float(self.min_wage_gross_input, "Asgari Ücret Brüt")
            manual_prev_base = self._parse_payroll_float(self.manual_prev_tax_base_input, "Manuel Önceki Kümülatif Matrah")

            if hasattr(self, "asgari_mode_button") and self.asgari_mode_button.isChecked():
                income_tax_rate = 0.0
                stamp_tax_rate = 0.0

            paid_days_input = worked_days + paid_leave_days
            if unpaid_leave_days > 0:
                paid_days_input = month_days - unpaid_leave_days
            paid_days = min(max(paid_days_input, 0), month_days)
            day_ratio = paid_days / month_days if month_days > 0 else 1.0

            overtime_amount_raw = overtime_hours * overtime_rate
            prorated_gross_salary_raw = gross_salary * day_ratio
            gross_total_raw = prorated_gross_salary_raw + overtime_amount_raw + bonus + meal + transport

            deduction_base_raw = gross_total_raw
            sgk_deduction_raw = deduction_base_raw * (sgk_rate / 100)
            unemployment_deduction_raw = deduction_base_raw * (unemployment_rate / 100)
            tax_base_raw = max(deduction_base_raw - sgk_deduction_raw - unemployment_deduction_raw, 0)

            overtime_amount = money(overtime_amount_raw)
            gross_total = money(gross_total_raw)
            sgk_deduction = money(sgk_deduction_raw)
            unemployment_deduction = money(unemployment_deduction_raw)
            tax_base = money(tax_base_raw)

            progressive_enabled = self.progressive_tax_check.isChecked()
            child_allowance = child_count * 346.5
            
            saved_prev_base = 0.0
            if self.use_saved_cumulative_check.isChecked() and employee_name != "Belirtilmedi":
                saved_prev_base = self._get_saved_cumulative_tax_base(employee_name, period_year, period_month)

            reduced_tax_base_raw = max(tax_base_raw - child_allowance, 0)
            cumulative_prev_base = money(saved_prev_base + manual_prev_base)
            cumulative_current_base = money(cumulative_prev_base + reduced_tax_base_raw)
            cumulative_prev_base_raw = float(saved_prev_base + manual_prev_base)
            cumulative_current_base_raw = float(cumulative_prev_base_raw + reduced_tax_base_raw)

            if progressive_enabled and not (hasattr(self, "asgari_mode_button") and self.asgari_mode_button.isChecked()):
                brackets = self._parse_tax_brackets(self.tax_brackets_input.text())
                income_tax_raw = (
                    self._compute_progressive_tax(cumulative_current_base_raw, brackets)
                    - self._compute_progressive_tax(cumulative_prev_base_raw, brackets)
                )
                income_tax = money(income_tax_raw)
                income_tax_rate_label = "Kümülatif Dilim"
                min_wage_income_tax_rate = brackets[0][1] * 100 if brackets else income_tax_rate
            else:
                income_tax_raw = reduced_tax_base_raw * (income_tax_rate / 100)
                income_tax = money(income_tax_raw)
                income_tax_rate_label = f"%{income_tax_rate:.3g}"
                min_wage_income_tax_rate = income_tax_rate

            stamp_tax_raw = deduction_base_raw * (stamp_tax_rate / 100)
            stamp_tax = money(stamp_tax_raw)

            income_tax_exemption = 0.0
            stamp_tax_exemption = 0.0
            income_tax_exemption_raw = 0.0
            stamp_tax_exemption_raw = 0.0
            if self.apply_min_wage_exemption_check.isChecked() and min_wage_gross > 0:
                min_sgk_raw = min_wage_gross * (sgk_rate / 100)
                min_unemployment_raw = min_wage_gross * (unemployment_rate / 100)
                min_tax_base_raw = max(min_wage_gross - min_sgk_raw - min_unemployment_raw, 0)
                income_tax_exemption_raw = min(min_tax_base_raw * (min_wage_income_tax_rate / 100), income_tax_raw)
                stamp_tax_exemption_raw = min(min_wage_gross * (stamp_tax_rate / 100), stamp_tax_raw)
                income_tax_exemption = money(income_tax_exemption_raw)
                stamp_tax_exemption = money(stamp_tax_exemption_raw)

            total_deductions_raw = sgk_deduction_raw + unemployment_deduction_raw + income_tax_raw + stamp_tax_raw
            additional_deductions_raw = advance + private_insurance + other_deductions
            total_exemptions_raw = income_tax_exemption_raw + stamp_tax_exemption_raw

            total_deductions = money(total_deductions_raw)
            additional_deductions = money(additional_deductions_raw)
            total_exemptions = money(total_exemptions_raw)
            net_salary = money(gross_total_raw - total_deductions_raw - additional_deductions_raw + total_exemptions_raw)

            total_final_deduction = money(total_deductions_raw + additional_deductions_raw)

            html = f"""
            <h2 style='color:#1976D2;'>💼 MAAŞ BORDROSU</h2>
            <p><strong>Personel:</strong> {employee_name}</p>
            <p><strong>Dönem:</strong> {period_text}</p>
            <p><strong>Asgari Ücret Modu:</strong> {'Açık' if hasattr(self, 'asgari_mode_button') and self.asgari_mode_button.isChecked() else 'Kapalı'}</p>
            <p><strong>Kümülatif Vergi Hesabı:</strong> {'Açık' if progressive_enabled else 'Kapalı'}</p>
            <p><strong>Önceki Kümülatif Matrah:</strong> {fmt_money(cumulative_prev_base)} ₺</p>
            <p><strong>Yeni Kümülatif Matrah:</strong> {fmt_money(cumulative_current_base)} ₺</p>
            <hr>
            <h3 style='color:#455A64;'>Çalışma Bilgileri</h3>
            <table border='1' cellpadding='6' style='border-collapse:collapse; width:100%;'>
                <tr style='background:#f5f5f5;'><th align='left'>Alan</th><th>Değer</th></tr>
                <tr><td>Ay Gün Sayısı</td><td>{month_days:.0f}</td></tr>
                <tr><td>Çalışılan Gün</td><td>{worked_days:.0f}</td></tr>
                <tr><td>Ücretli İzin</td><td>{paid_leave_days:.0f}</td></tr>
                <tr><td>Ücretsiz İzin</td><td>{unpaid_leave_days:.0f}</td></tr>
                <tr><td>Ücretlendirilen Gün</td><td>{paid_days:.0f}</td></tr>
                <tr><td>Gün Oranı</td><td>{day_ratio*100:.2f}%</td></tr>
                <tr><td>Bakmakla Yükümlü Çocuk</td><td>{child_count:.0f}</td></tr>
                <tr><td>Çocuk İndirimi (₺)</td><td>{fmt_money(child_allowance)}</td></tr>
            </table>
            <br>
            <table border='1' cellpadding='6' style='border-collapse:collapse; width:100%;'>
                <tr style='background:#f5f5f5;'><th align='left'>Kalem</th><th>Tutar (₺)</th></tr>
                <tr><td>Brüt Maaş (Tam)</td><td>{fmt_money(gross_salary)}</td></tr>
                <tr><td>Brüt Maaş (Gün Oranlı)</td><td>{fmt_money(prorated_gross_salary_raw)}</td></tr>
                <tr><td>Fazla Mesai</td><td>{fmt_money(overtime_amount)}</td></tr>
                <tr><td>Prim / İkramiye</td><td>{fmt_money(bonus)}</td></tr>
                <tr><td>Yemek Yardımı</td><td>{fmt_money(meal)}</td></tr>
                <tr><td>Yol Yardımı</td><td>{fmt_money(transport)}</td></tr>
                <tr style='font-weight:bold; background:#E3F2FD;'><td>Toplam Brüt</td><td>{fmt_money(gross_total)}</td></tr>
            </table>
            <br>
            <table border='1' cellpadding='6' style='border-collapse:collapse; width:100%;'>
                <tr style='background:#f5f5f5;'><th align='left'>Vergi Matrahı</th><th>Tutar (₺)</th></tr>
                <tr><td>Orijinal Vergi Matrahı</td><td>{fmt_money(tax_base)}</td></tr>
                <tr><td>Çocuk İndirimi</td><td>-{fmt_money(child_allowance)}</td></tr>
                <tr style='background:#FFF3E0;'><td>İndirimli Vergi Matrahı</td><td>{fmt_money(reduced_tax_base_raw)}</td></tr>
            </table>
            <br>
            <table border='1' cellpadding='6' style='border-collapse:collapse; width:100%;'>
                <tr style='background:#f5f5f5;'><th align='left'>Kesinti</th><th>Tutar (₺)</th></tr>
                <tr><td>SGK İşçi Payı (%{sgk_rate:.3g})</td><td>{fmt_money(sgk_deduction)}</td></tr>
                <tr><td>İşsizlik Sigortası (%{unemployment_rate:.3g})</td><td>{fmt_money(unemployment_deduction)}</td></tr>
                <tr><td>Gelir Vergisi ({income_tax_rate_label})</td><td>{fmt_money(income_tax)}</td></tr>
                <tr><td>Damga Vergisi (%{stamp_tax_rate:.3g})</td><td>{fmt_money(stamp_tax)}</td></tr>
                <tr><td>Asgari Ücret Gelir Vergisi İstisnası</td><td>-{fmt_money(income_tax_exemption)}</td></tr>
                <tr><td>Asgari Ücret Damga Vergisi İstisnası</td><td>-{fmt_money(stamp_tax_exemption)}</td></tr>
                <tr style='font-weight:bold; background:#FFEBEE;'><td>Toplam Kesinti</td><td>{fmt_money(total_deductions)}</td></tr>
                <tr><td>Avans Kesintisi</td><td>{fmt_money(advance)}</td></tr>
                <tr><td>Özel Sigorta Kesintisi</td><td>{fmt_money(private_insurance)}</td></tr>
                <tr><td>Diğer Kesintiler</td><td>{fmt_money(other_deductions)}</td></tr>
                <tr style='font-weight:bold; background:#FCE4EC;'><td>Toplam Ek Kesinti</td><td>{fmt_money(additional_deductions)}</td></tr>
            </table>
            <br>
            <p><strong>Toplam Nihai Kesinti:</strong> {fmt_money(total_final_deduction)} ₺</p>
            <p><strong>Toplam İstisna:</strong> {fmt_money(total_exemptions)} ₺</p>
            <h3 style='color:#2E7D32;'>Net Maaş: {fmt_money(net_salary)} ₺</h3>
            <p style='font-size:9pt; color:#666;'>Not: Bu bordro hesaplaması bilgilendirme amaçlıdır.</p>
            """

            self.last_payroll_html = html
            self.last_payroll_employee = employee_name
            self.last_payroll_period = period_text
            self.last_payroll_data = {
                "employee": employee_name,
                "year": period_year,
                "month": period_month,
                "gross_salary": gross_salary,
                "month_days": month_days,
                "worked_days": worked_days,
                "paid_leave_days": paid_leave_days,
                "unpaid_leave_days": unpaid_leave_days,
                "child_count": child_count,
                "paid_days": paid_days,
                "day_ratio": day_ratio,
                "overtime_hours": overtime_hours,
                "overtime_rate": overtime_rate,
                "overtime_amount": overtime_amount,
                "bonus": bonus,
                "meal": meal,
                "transport": transport,
                "gross_total": gross_total,
                "sgk_rate": sgk_rate,
                "sgk_deduction": sgk_deduction,
                "unemployment_rate": unemployment_rate,
                "unemployment_deduction": unemployment_deduction,
                "tax_base": money(reduced_tax_base_raw),
                "child_allowance": child_allowance,
                "income_tax_rate": income_tax_rate,
                "income_tax": income_tax,
                "stamp_tax_rate": stamp_tax_rate,
                "stamp_tax": stamp_tax,
                "income_tax_exemption": income_tax_exemption,
                "stamp_tax_exemption": stamp_tax_exemption,
                "total_deductions": total_deductions,
                "advance": money(advance),
                "private_insurance": money(private_insurance),
                "other_deductions": money(other_deductions),
                "additional_deductions": additional_deductions,
                "total_exemptions": total_exemptions,
                "net_salary": net_salary,
                "savings_time": datetime.now().isoformat()
            }
            self.payroll_result_view.setHtml(html)
            QTimer.singleShot(50, self._adjust_payroll_result_height)
            QTimer.singleShot(150, self._scroll_to_payroll_result)

        except Exception as e:
            QMessageBox.warning(self, "Uyarı", f"Bordro hesaplanamadı: {str(e)}") 

    def _calculate_net_for_gross_value(self, gross_value, s, u, t, d, child_allowance, min_wage_gross, use_exemption, money_fn):
        """Verilen brüt için net'i hesapla (Binary search için helper)"""
        # Temel kesintiler
        sgk = gross_value * s
        unemployment = gross_value * u
        tax_base = gross_value - sgk - unemployment
        reduced_tax_base = max(tax_base - child_allowance, 0)
        income_tax = reduced_tax_base * t
        stamp_tax = gross_value * d
        
        # Muafiyet hesabı (her brüt için uygulanabilir - min ücret tutarı kadar)
        income_tax_exemption = 0.0
        stamp_tax_exemption = 0.0
        if use_exemption and min_wage_gross > 0:
            # Minimum ücretin vergi tutarı hesapla
            min_sgk = min_wage_gross * s
            min_unemployment = min_wage_gross * u
            min_tax_base = max(min_wage_gross - min_sgk - min_unemployment, 0)
            # Muafiyet: min ücretin vergisi, ama ödenmesi gereken vergiden fazla olamaz
            income_tax_exemption = min(min_tax_base * t, income_tax)
            stamp_tax_exemption = min(min_wage_gross * d, stamp_tax)
        
        # Net hesapla
        net = gross_value - sgk - unemployment - income_tax - stamp_tax + income_tax_exemption + stamp_tax_exemption
        return money_fn(net)
    def calculate_gross_from_net(self):
        """Net maaştan brüt maaşı hesapla (Binary Search)"""
        try:
            def money(value):
                return round(float(value) + 1e-9, 2)

            net_salary_input = self._parse_payroll_float(self.reverse_net_input, "Net Maaş")
            child_count = max(int(self._parse_payroll_float(self.reverse_child_count_input, "Çocuk Sayısı")), 0)
            
            # Parametrelerden oranları oku
            sgk_rate = self._parse_payroll_float(self.payroll_fields["sgk_rate"], "SGK Oranı")
            unemployment_rate = self._parse_payroll_float(self.payroll_fields["unemployment_rate"], "İşsizlik Sigortası")
            income_tax_rate = self._parse_payroll_float(self.payroll_fields["income_tax_rate"], "Gelir Vergisi")
            stamp_tax_rate = self._parse_payroll_float(self.payroll_fields["stamp_tax_rate"], "Damga Vergisi")
            min_wage_gross = self._parse_payroll_float(self.min_wage_gross_input, "Asgari Ücret Brüt")
            
            # Ondalık sayılara dönüştür
            s = sgk_rate / 100.0  # SGK oranı
            u = unemployment_rate / 100.0  # İşsizlik oranı
            t = income_tax_rate / 100.0  # Gelir vergisi oranı
            d = stamp_tax_rate / 100.0  # Damga vergisi oranı
            
            # Çocuk indirimi
            child_allowance = child_count * 346.5
            
            # Binary search ile brüt bul
            low = net_salary_input * 0.9
            high = net_salary_input * 2.5
            
            gross_calculated = None
            for iteration in range(100):
                mid = (low + high) / 2
                
                # Bu brüt için net hesapla
                calculated_net = self._calculate_net_for_gross_value(
                    mid, s, u, t, d, child_allowance, min_wage_gross, 
                    self.apply_min_wage_exemption_check.isChecked(), money
                )
                
                # Yakınlık kontrolü
                if abs(calculated_net - net_salary_input) < 0.01:
                    gross_calculated = mid
                    break
                
                # Binary search güncelle
                if calculated_net < net_salary_input:
                    low = mid  # Net düşük, gross'u artır
                else:
                    high = mid  # Net yüksek, gross'u azalt
            
            if gross_calculated is None:
                gross_calculated = (low + high) / 2
            
            gross_calculated = money(gross_calculated)
            
            # Hesaplanan brüt ile doğrulama yaparak net'i kontrol et
            if gross_calculated > 0:
                # Verifikasyon: Hesaplanan brütün net'ini hesapla
                net_verification = self._calculate_net_for_gross_value(
                    gross_calculated, s, u, t, d, child_allowance, min_wage_gross,
                    self.apply_min_wage_exemption_check.isChecked(), money
                )
                
                # Doğrulama detayları hesapla
                sgk_ded = gross_calculated * s
                unemployment_ded = gross_calculated * u
                tax_base = gross_calculated - sgk_ded - unemployment_ded
                reduced_tax_base = max(tax_base - child_allowance, 0)
                income_tax_val = reduced_tax_base * t
                stamp_tax_val = gross_calculated * d
                
                # Muafiyet hesabı
                income_tax_exemption = 0.0
                stamp_tax_exemption = 0.0
                if self.apply_min_wage_exemption_check.isChecked() and min_wage_gross > 0 and gross_calculated < min_wage_gross:
                    min_sgk = min_wage_gross * s
                    min_unemployment = min_wage_gross * u
                    min_tax_base = max(min_wage_gross - min_sgk - min_unemployment, 0)
                    income_tax_exemption = money(min(min_tax_base * t, income_tax_val))
                    stamp_tax_exemption = money(min(min_wage_gross * d, stamp_tax_val))
                
                # Toplam kesinti (muafiyet uygulanmış)
                total_deductions = money(
                    sgk_ded + unemployment_ded + 
                    (income_tax_val - income_tax_exemption) + 
                    (stamp_tax_val - stamp_tax_exemption)
                )
                net_verification = money(gross_calculated - total_deductions)
                
                # Sonuçları göster
                html = f"""
                <h2 style='color:#2E7D32;'>✅ NET'TEN BRÜT HESAPLAMA SONUCU</h2>
                <hr>
                <table border='1' cellpadding='6' style='border-collapse:collapse; width:100%;'>
                    <tr style='background:#f5f5f5;'><th align='left'>Alan</th><th>Tutar (₺)</th></tr>
                    <tr><td><strong>Hedef Net Maaş</strong></td><td><strong>{format_tr(net_salary_input)}</strong></td></tr>
                    <tr><td><strong>Çocuk Sayısı</strong></td><td><strong>{child_count:.0f}</strong></td></tr>
                    <tr style='background:#E3F2FD;'><td><strong>Hesaplanan Brüt Maaş</strong></td><td><strong>{format_tr(gross_calculated)}</strong></td></tr>
                </table>
                <br>
                <h3 style='color:#1976D2;'>Doğrulama Hesaplaması</h3>
                <table border='1' cellpadding='6' style='border-collapse:collapse; width:100%;'>
                    <tr style='background:#f5f5f5;'><th align='left'>Kesinti</th><th>Tutar (₺)</th></tr>
                    <tr><td>SGK İşçi Payı (%{sgk_rate:.3g})</td><td>{format_tr(money(sgk_ded))}</td></tr>
                    <tr><td>İşsizlik Sigortası (%{unemployment_rate:.3g})</td><td>{format_tr(money(unemployment_ded))}</td></tr>
                    <tr><td>Vergi Matrahı (Çocuk İndirimi Sonrası)</td><td>{format_tr(money(reduced_tax_base))}</td></tr>
                    <tr><td>Gelir Vergisi (%{income_tax_rate:.3g})</td><td>{format_tr(money(income_tax_val))}</td></tr>
                    <tr><td>Gelir Vergisi İstisnası</td><td>-{format_tr(money(income_tax_exemption))}</td></tr>
                    <tr><td>Damga Vergisi (%{stamp_tax_rate:.3g})</td><td>{format_tr(money(stamp_tax_val))}</td></tr>
                    <tr><td>Damga Vergisi İstisnası</td><td>-{format_tr(money(stamp_tax_exemption))}</td></tr>
                    <tr style='font-weight:bold; background:#FFEBEE;'><td>Toplam Kesinti</td><td>{format_tr(total_deductions)}</td></tr>
                </table>
                <br>
                <p><strong>Doğrulama - Hesaplanan Net Maaş:</strong> {format_tr(net_verification)} ₺</p>
                <p style='color:#1976D2;'><strong>Fark:</strong> {format_tr(money(abs(net_verification - net_salary_input)))} ₺ {'✓ Doğru' if money(abs(net_verification - net_salary_input)) < 0.01 else '⚠ Rounding farkı'}</p>
                <br>
                <div style='background:#FFF9C4; padding:10px; border-left:4px solid #FBC02D;'>
                    <strong>💡 Kullanım:</strong> Hesaplanan <strong>{format_tr(gross_calculated)} ₺</strong> brüt maaşını 
                    'Brüt Maaş' alanına kopyalayıp 'Bordroyu Hesapla' butonuna basınız.
                </div>
                <p style='font-size:9pt; color:#666;'>Not: Rounding nedeniyle doğrulama netinde küçük farklar oluşabilir.</p>
                """
                self.payroll_result_view.setHtml(html)
                QTimer.singleShot(50, self._adjust_payroll_result_height)
                QTimer.singleShot(150, self._scroll_to_payroll_result)
                
                # Brütü otomatik olarak Brüt Maaş alanına yerleştir
                self.payroll_fields["gross_salary"].setText(format_tr(gross_calculated))
                
        except Exception as e:
            QMessageBox.warning(self, "Uyarı", f"Ters hesaplama başarısız: {str(e)}")

    def export_payroll_to_pdf(self):
        """Hesaplanan bordroyu PDF çıktısı olarak kaydet"""
        if not self.last_payroll_html:
            QMessageBox.warning(self, "Uyarı", "Önce bordroyu hesaplamalısınız")
            return

        try:
            from PyQt5.QtWidgets import QFileDialog
            from PyQt5.QtPrintSupport import QPrinter
            from PyQt5.QtGui import QTextDocument

            safe_employee = self.last_payroll_employee.replace(" ", "_") if self.last_payroll_employee else "personel"
            default_file = f"bordro_{safe_employee}_{self.last_payroll_period.replace('.', '_')}.pdf"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Bordro PDF Kaydet",
                default_file,
                "PDF Dosyası (*.pdf)"
            )

            if not file_path:
                return

            if not file_path.lower().endswith(".pdf"):
                file_path += ".pdf"

            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)

            document = QTextDocument()
            document.setHtml(self.last_payroll_html)
            document.print_(printer)
            
            QMessageBox.information(self, "Başarılı", f"Bordro PDF'ye kaydedildi:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{str(e)}")

    def create_employees_tab(self) -> QWidget:
        """Çalışanlar yönetim sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # İşlem butonları
        button_layout = QHBoxLayout()
        
        btn_add_emp = QPushButton("➕ Yeni Çalışan Ekle")
        btn_add_emp.setMinimumHeight(34)
        btn_add_emp.clicked.connect(self.show_add_employee_dialog)
        button_layout.addWidget(btn_add_emp)
        
        btn_edit_emp = QPushButton("✏️ Seçileni Düzenle")
        btn_edit_emp.setMinimumHeight(34)
        btn_edit_emp.clicked.connect(self.show_edit_employee_dialog)
        button_layout.addWidget(btn_edit_emp)
        
        btn_delete_emp = QPushButton("🗑️ Seçileni Sil")
        btn_delete_emp.setMinimumHeight(34)
        btn_delete_emp.clicked.connect(self.delete_employee)
        button_layout.addWidget(btn_delete_emp)
        
        search_label = QLabel("Ara:")
        search_label.setMinimumWidth(50)
        button_layout.addWidget(search_label)
        
        self.emp_search_field = QLineEdit()
        self.emp_search_field.setPlaceholderText("Ad, Soyad veya E-posta...")
        self.emp_search_field.setMaximumWidth(250)
        self.emp_search_field.textChanged.connect(self.search_employees)
        button_layout.addWidget(self.emp_search_field)
        
        btn_refresh_emp = QPushButton("🔄 Yenile")
        btn_refresh_emp.setMinimumHeight(34)
        btn_refresh_emp.clicked.connect(self.refresh_employees_table)
        button_layout.addWidget(btn_refresh_emp)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # Çalışanlar tablosu
        self.employees_table = QTableWidget()
        self.employees_table.setColumnCount(13)
        self.employees_table.setHorizontalHeaderLabels([
            "Ad", "Soyad", "E-posta", "Telefon", "Başlama Tarihi",
            "Brüt Maaş", "Net Maaş", "Kesilmiş Bordro", "Ödenen Maaş", "Kalan Maaş", "Döküm", "Durum", "ID"
        ])
        self.employees_table.horizontalHeader().setStretchLastSection(True)
        self.employees_table.setColumnHidden(12, True)
        layout.addWidget(self.employees_table)
        
        widget.setLayout(layout)
        self.refresh_employees_table()
        return widget
    
    def refresh_employees_table(self):
        """Çalışanlar tablosunu yenile"""
        try:
            from src.services.employee_service import EmployeeService
            from src.database.db import get_db
            
            db = get_db()
            emp_service = EmployeeService(db)
            employees = emp_service.get_all_employees(active_only=False)
            self._render_employees_table(employees)
            db.close()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Çalışanlar yüklenemedi:\n{str(e)}")
    
    def search_employees(self):
        """Çalışan ara"""
        search_text = self.emp_search_field.text().strip()
        
        try:
            from src.services.employee_service import EmployeeService
            from src.database.db import get_db
            
            db = get_db()
            emp_service = EmployeeService(db)
            
            if search_text:
                employees = emp_service.search_employees(search_text, active_only=False)
            else:
                employees = emp_service.get_all_employees(active_only=False)
            self._render_employees_table(employees)
            db.close()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Arama yapılamadı:\n{str(e)}")

    def _render_employees_table(self, employees):
        if not hasattr(self, "employees_table"):
            return

        records = self._load_saved_payroll_records()
        payroll_map = {}
        for rec in records:
            name = str(rec.get("employee", "")).strip()
            if not name:
                continue
            key = name.casefold()
            year = int(rec.get("year", 0) or 0)
            month = int(rec.get("month", 0) or 0)
            current = payroll_map.get(key)
            if not current or (year, month) > (current["year"], current["month"]):
                payroll_map[key] = {
                    "record": rec,
                    "year": year,
                    "month": month,
                    "name": name,
                }

        try:
            from src.database.db import SessionLocal
            from src.database.models import Transaction

            session = SessionLocal()
            payroll_transactions = session.query(Transaction).filter(
                Transaction.user_id == self.user.id
            ).all()
            session.close()
        except Exception:
            payroll_transactions = []

        self.employees_table.setRowCount(0)
        for employee in employees:
            row = self.employees_table.rowCount()
            self.employees_table.insertRow(row)

            full_name = f"{employee.first_name} {employee.last_name}".strip()
            key = full_name.casefold()
            payroll_info = payroll_map.get(key)
            if not payroll_info:
                for stored_key, info in payroll_map.items():
                    if stored_key in key or key in stored_key:
                        payroll_info = info
                        break

            net_text = "-"
            bordro_text = "-"
            paid_text = "-"
            remaining_text = "-"
            month = year = 0

            if payroll_info:
                rec = payroll_info["record"]
                month = payroll_info["month"]
                year = payroll_info["year"]
                net = float(rec.get("net_salary", 0.0) or 0.0)
                bordro_amount = net
                paid_transactions = self._filter_payroll_payment_transactions(
                    payroll_transactions, payroll_info["name"], month, year
                )
                paid_total = sum(float(t.amount or 0.0) for t in paid_transactions)
                remaining = bordro_amount - paid_total
                net_text = format_tr(net)
                bordro_text = format_tr(bordro_amount)
                paid_text = format_tr(paid_total)
                remaining_text = format_tr(remaining)

            self.employees_table.setItem(row, 0, QTableWidgetItem(employee.first_name))
            self.employees_table.setItem(row, 1, QTableWidgetItem(employee.last_name))
            self.employees_table.setItem(row, 2, QTableWidgetItem(employee.email or "-"))
            self.employees_table.setItem(row, 3, QTableWidgetItem(employee.phone or "-"))
            self.employees_table.setItem(row, 4, QTableWidgetItem(
                employee.start_date.strftime("%d.%m.%Y") if employee.start_date else "-"
            ))
            self.employees_table.setItem(row, 5, QTableWidgetItem(format_tr(employee.gross_salary)))
            self.employees_table.setItem(row, 6, QTableWidgetItem(net_text))
            self.employees_table.setItem(row, 7, QTableWidgetItem(bordro_text))
            self.employees_table.setItem(row, 8, QTableWidgetItem(paid_text))

            remaining_item = QTableWidgetItem(remaining_text)
            if remaining_text != "-":
                if remaining > 0:
                    remaining_item.setBackground(QColor("#fff9c4"))
                elif remaining < 0:
                    remaining_item.setBackground(QColor("#ffcdd2"))
                else:
                    remaining_item.setBackground(QColor("#c8e6c9"))
            self.employees_table.setItem(row, 9, remaining_item)

            btn_statement = QPushButton("📑 Döküm Aç")
            btn_statement.setMinimumHeight(24)
            btn_statement.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    padding: 4px 8px;
                    font-size: 9pt;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #45a049; }
            """)
            if payroll_info and month and year:
                btn_statement.clicked.connect(
                    lambda checked, n=payroll_info["name"], m=month, y=year: self.show_payroll_payment_statement(n, m, y)
                )
            else:
                btn_statement.setEnabled(False)
            self.employees_table.setCellWidget(row, 10, btn_statement)

            status = "Aktif" if employee.is_active else "Pasif"
            self.employees_table.setItem(row, 11, QTableWidgetItem(status))
            self.employees_table.setItem(row, 12, QTableWidgetItem(str(employee.id)))
    
    def show_add_employee_dialog(self):
        """Yeni çalışan ekle diyaloğu"""
        from PyQt5.QtWidgets import QDialog, QLabel, QLineEdit, QSpinBox, QDoubleSpinBox, QDateEdit
        from datetime import date
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Yeni Çalışan Ekle")
        dialog.setGeometry(100, 100, 400, 500)
        
        layout = QVBoxLayout()
        
        # Ad
        layout.addWidget(QLabel("Ad:"))
        first_name_field = QLineEdit()
        layout.addWidget(first_name_field)
        
        # Soyad
        layout.addWidget(QLabel("Soyad:"))
        last_name_field = QLineEdit()
        layout.addWidget(last_name_field)
        
        # E-posta
        layout.addWidget(QLabel("E-posta:"))
        email_field = QLineEdit()
        layout.addWidget(email_field)
        
        # Telefon
        layout.addWidget(QLabel("Telefon:"))
        phone_field = QLineEdit()
        layout.addWidget(phone_field)
        
        # Başlama Tarihi
        layout.addWidget(QLabel("Başlama Tarihi:"))
        start_date_field = QDateEdit()
        start_date_field.setDisplayFormat("dd.MM.yyyy")
        start_date_field.setDate(QDate.currentDate())
        layout.addWidget(start_date_field)
        
        # Brüt Maaş
        layout.addWidget(QLabel("Brüt Maaş (₺):"))
        gross_salary_field = QDoubleSpinBox()
        gross_salary_field.setMaximum(999999.99)
        gross_salary_field.setValue(30000.00)
        layout.addWidget(gross_salary_field)
        
        # SGK Oranı
        layout.addWidget(QLabel("SGK Oranı (%):"))
        sgk_rate_field = QDoubleSpinBox()
        sgk_rate_field.setValue(14.0)
        layout.addWidget(sgk_rate_field)
        
        # İşsizlik Sigortası Oranı
        layout.addWidget(QLabel("İşsizlik Sig. Oranı (%):"))
        unemp_rate_field = QDoubleSpinBox()
        unemp_rate_field.setValue(1.0)
        layout.addWidget(unemp_rate_field)
        
        # Gelir Vergisi Oranı
        layout.addWidget(QLabel("Gelir Vergisi Oranı (%):"))
        income_tax_rate_field = QDoubleSpinBox()
        income_tax_rate_field.setValue(15.0)
        layout.addWidget(income_tax_rate_field)
        
        # Bakmakla Yükümlü Çocuk Sayısı
        layout.addWidget(QLabel("Bakmakla Yükümlü Çocuk Sayısı:"))
        child_count_field = QSpinBox()
        child_count_field.setMaximum(10)
        child_count_field.setValue(0)
        layout.addWidget(child_count_field)
        
        # Butonlar
        button_layout = QHBoxLayout()
        
        def save_employee():
            try:
                from src.services.employee_service import EmployeeService
                from src.database.db import get_db
                
                db = get_db()
                emp_service = EmployeeService(db)
                
                emp_service.create_employee(
                    first_name=first_name_field.text().strip(),
                    last_name=last_name_field.text().strip(),
                    email=email_field.text().strip() or None,
                    phone=phone_field.text().strip() or None,
                    start_date=start_date_field.date().toPyDate(),
                    gross_salary=gross_salary_field.value(),
                    sgk_rate=sgk_rate_field.value(),
                    unemployment_rate=unemp_rate_field.value(),
                    income_tax_rate=income_tax_rate_field.value(),
                    child_count=child_count_field.value()
                )
                db.close()
                QMessageBox.information(dialog, "Başarılı", "Çalışan kaydedildi")
                dialog.accept()
                self.refresh_employees_table()
            except Exception as e:
                QMessageBox.critical(dialog, "Hata", f"Çalışan kaydedilemedi:\n{str(e)}")
        
        btn_save = QPushButton("💾 Kaydet")
        btn_save.clicked.connect(save_employee)
        button_layout.addWidget(btn_save)
        
        btn_cancel = QPushButton("❌ İptal")
        btn_cancel.clicked.connect(dialog.reject)
        button_layout.addWidget(btn_cancel)
        
        layout.addLayout(button_layout)
        dialog.setLayout(layout)
        dialog.exec_()
    
    def show_edit_employee_dialog(self):
        """Çalışan düzenle diyaloğu"""
        row = self.employees_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Düzenlemek için bir çalışan seçin")
            return
        
        employee_id = int(self.employees_table.item(row, 7).text())
        
        try:
            from src.services.employee_service import EmployeeService
            from src.database.db import get_db
            from PyQt5.QtWidgets import QDialog, QDoubleSpinBox, QSpinBox
            
            db = get_db()
            emp_service = EmployeeService(db)
            employee = emp_service.get_employee(employee_id)
            
            if not employee:
                db.close()
                QMessageBox.warning(self, "Hata", "Çalışan bulunamadı")
                return
            
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Çalışan Düzenle: {employee.get_full_name()}")
            dialog.setGeometry(100, 100, 400, 500)
            
            layout = QVBoxLayout()
            
            layout.addWidget(QLabel("Ad:"))
            first_name_field = QLineEdit()
            first_name_field.setText(employee.first_name)
            layout.addWidget(first_name_field)
            
            layout.addWidget(QLabel("Soyad:"))
            last_name_field = QLineEdit()
            last_name_field.setText(employee.last_name)
            layout.addWidget(last_name_field)
            
            layout.addWidget(QLabel("E-posta:"))
            email_field = QLineEdit()
            email_field.setText(employee.email or "")
            layout.addWidget(email_field)
            
            layout.addWidget(QLabel("Telefon:"))
            phone_field = QLineEdit()
            phone_field.setText(employee.phone or "")
            layout.addWidget(phone_field)
            
            layout.addWidget(QLabel("Brüt Maaş (₺):"))
            gross_salary_field = QDoubleSpinBox()
            gross_salary_field.setMaximum(999999.99)
            gross_salary_field.setValue(employee.gross_salary)
            layout.addWidget(gross_salary_field)
            
            layout.addWidget(QLabel("SGK Oranı (%):"))
            sgk_rate_field = QDoubleSpinBox()
            sgk_rate_field.setValue(employee.sgk_rate)
            layout.addWidget(sgk_rate_field)
            
            layout.addWidget(QLabel("İşsizlik Sig. Oranı (%):"))
            unemp_rate_field = QDoubleSpinBox()
            unemp_rate_field.setValue(employee.unemployment_rate)
            layout.addWidget(unemp_rate_field)
            
            layout.addWidget(QLabel("Gelir Vergisi Oranı (%):"))
            income_tax_rate_field = QDoubleSpinBox()
            income_tax_rate_field.setValue(employee.income_tax_rate)
            layout.addWidget(income_tax_rate_field)
            
            layout.addWidget(QLabel("Bakmakla Yükümlü Çocuk Sayısı:"))
            child_count_field = QSpinBox()
            child_count_field.setMaximum(10)
            child_count_field.setValue(getattr(employee, 'child_count', 0))
            layout.addWidget(child_count_field)
            
            button_layout = QHBoxLayout()
            
            def update_employee():
                try:
                    emp_service.update_employee(
                        employee_id,
                        first_name=first_name_field.text().strip(),
                        last_name=last_name_field.text().strip(),
                        email=email_field.text().strip() or None,
                        phone=phone_field.text().strip() or None,
                        gross_salary=gross_salary_field.value(),
                        sgk_rate=sgk_rate_field.value(),
                        unemployment_rate=unemp_rate_field.value(),
                        income_tax_rate=income_tax_rate_field.value(),
                        child_count=child_count_field.value()
                    )
                    QMessageBox.information(dialog, "Başarılı", "Çalışan güncellendi")
                    dialog.accept()
                except Exception as e:
                    QMessageBox.critical(dialog, "Hata", f"Güncelleme başarısız:\n{str(e)}")
            
            btn_save = QPushButton("💾 Kaydet")
            btn_save.clicked.connect(update_employee)
            button_layout.addWidget(btn_save)
            
            btn_cancel = QPushButton("❌ İptal")
            btn_cancel.clicked.connect(dialog.reject)
            button_layout.addWidget(btn_cancel)
            
            layout.addLayout(button_layout)
            dialog.setLayout(layout)
            dialog.exec_()
            
            # İşlem bitikten sonra db kapat ve tabloyu yenile
            db.close()
            self.refresh_employees_table()
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Düzenlenemedi:\n{str(e)}")
    
    def delete_employee(self):
        """Çalışan sil"""
        row = self.employees_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Silmek için bir çalışan seçin")
            return
        
        employee_id = int(self.employees_table.item(row, 7).text())
        emp_name = self.employees_table.item(row, 0).text() + " " + self.employees_table.item(row, 1).text()
        
        reply = QMessageBox.question(self, "Onayla", 
            f"'{emp_name}' çalışanını silmek istediğinize emin misiniz?",
            QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                from src.services.employee_service import EmployeeService
                from src.database.db import get_db
                
                db = get_db()
                emp_service = EmployeeService(db)
                emp_service.delete_employee(employee_id)
                db.close()
                QMessageBox.information(self, "Başarılı", "Çalışan silindi")
                self.refresh_employees_table()
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Silme işlemi başarısız:\n{str(e)}")
    
    def create_bulk_payroll_tab(self) -> QWidget:
        """Toplu bordro hesaplama sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        title = QLabel("⚙️ Toplu Bordro Hesaplama")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        layout.addWidget(title)

        # Ay-Yıl Seçimi Grubu
        select_group = QGroupBox("Bordro Tarihi ve Genel Ayarlar")
        select_layout = QHBoxLayout()
        
        select_layout.addWidget(QLabel("Ay:"))
        self.bulk_payroll_month = QComboBox()
        months = {
            1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
            7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
        }
        for i in range(1, 13):
            self.bulk_payroll_month.addItem(months[i], i)
        self.bulk_payroll_month.setCurrentIndex(datetime.now().month - 1)
        select_layout.addWidget(self.bulk_payroll_month)
        
        select_layout.addWidget(QLabel("Yıl:"))
        self.bulk_payroll_year = QComboBox()
        current_year = datetime.now().year
        for year in range(current_year - 2, current_year + 3):
            self.bulk_payroll_year.addItem(str(year), year)
        self.bulk_payroll_year.setCurrentIndex(2)
        select_layout.addWidget(self.bulk_payroll_year)
        
        select_layout.addStretch()
        select_group.setLayout(select_layout)
        layout.addWidget(select_group)
        
        # Vergi Seçenekleri Grubu
        options_group = QGroupBox("Vergi Seçenekleri")
        options_layout = QVBoxLayout()
        options_two_col = QHBoxLayout()
        
        options_left = QVBoxLayout()
        options_right = QVBoxLayout()
        
        self.bulk_asgari_mode_button = QPushButton("🎯 Asgari Ücret Modu: Kapalı")
        self.bulk_asgari_mode_button.setCheckable(True)
        self.bulk_asgari_mode_button.setMinimumHeight(34)
        self.bulk_asgari_mode_button.toggled.connect(self.toggle_bulk_asgari_ucret_mode)
        options_left.addWidget(self.bulk_asgari_mode_button)
        
        self.bulk_apply_min_wage_exemption_check = QCheckBox("Asgari vergi istisnasını uygula")
        self.bulk_apply_min_wage_exemption_check.setChecked(True)
        options_left.addWidget(self.bulk_apply_min_wage_exemption_check)
        
        min_wage_row = QHBoxLayout()
        min_wage_row.addWidget(QLabel("Asgari Ücret Brüt (₺):"))
        self.bulk_min_wage_gross_input = QLineEdit("33030")
        self.bulk_min_wage_gross_input.setMinimumWidth(160)
        min_wage_row.addWidget(self.bulk_min_wage_gross_input)
        options_left.addLayout(min_wage_row)
        
        self.bulk_progressive_tax_check = QCheckBox("Kümülatif vergi dilimi hesabı uygula")
        self.bulk_progressive_tax_check.setChecked(False)  # Default kapalı toplu için
        options_right.addWidget(self.bulk_progressive_tax_check)
        
        brackets_row = QHBoxLayout()
        brackets_row.addWidget(QLabel("Vergi Dilimleri:"))
        self.bulk_tax_brackets_input = QLineEdit("158000:15,330000:20,1200000:27,4300000:35,sonsuz:40")
        self.bulk_tax_brackets_input.setToolTip("Örn: 158000:15,330000:20 (üst:oran%)")
        brackets_row.addWidget(self.bulk_tax_brackets_input)
        options_right.addLayout(brackets_row)
        
        options_left.addStretch()
        options_right.addStretch()
        options_two_col.addLayout(options_left)
        options_two_col.addLayout(options_right)
        options_layout.addLayout(options_two_col)
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        # Hesapla Butonu
        btn_row = QHBoxLayout()
        btn_calculate = QPushButton("🧮 Toplu Bordroyu Hesapla")
        btn_calculate.setMinimumHeight(38)
        btn_calculate.clicked.connect(self.calculate_bulk_payroll)
        btn_row.addWidget(btn_calculate)
        btn_row.addStretch()
        layout.addLayout(btn_row)
        
        # Sonuç tablosu
        self.bulk_payroll_table = QTableWidget()
        self.bulk_payroll_table.setColumnCount(13)  # +1 Seç checkbox'ı için
        self.bulk_payroll_table.setHorizontalHeaderLabels([
            "Seç", "Personel", "Ay Gün", "Çalış.", "Ücretli İzin", "Ücretsiz İz.", 
            "Brüt", "SGK", "İşsizlik", "Gel.Vergi", "Damga V.", "Muafiyet", "Net"
        ])
        self.bulk_payroll_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.bulk_payroll_table)
        
        # Özet
        summary_layout = QHBoxLayout()
        self.bulk_summary_label = QLabel("Hesaplama sonuçları burada gösterilecektir")
        summary_layout.addWidget(self.bulk_summary_label)
        summary_layout.addStretch()
        
        # Tümünü Seç / Seçimi Kaldır butonları
        btn_select_all = QPushButton("✓ Tümünü Seç")
        btn_select_all.setMaximumWidth(120)
        btn_select_all.clicked.connect(self.bulk_select_all_employees)
        summary_layout.addWidget(btn_select_all)
        
        btn_deselect_all = QPushButton("✗ Seçimi Kaldır")
        btn_deselect_all.setMaximumWidth(120)
        btn_deselect_all.clicked.connect(self.bulk_deselect_all_employees)
        summary_layout.addWidget(btn_deselect_all)
        
        # Bordroyu Kaydet butonu
        btn_save_payroll = QPushButton("💾 Bordroyu Kaydet")
        btn_save_payroll.setMinimumHeight(36)
        btn_save_payroll.setMaximumWidth(180)
        btn_save_payroll.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        btn_save_payroll.clicked.connect(self.save_bulk_payroll)
        summary_layout.addWidget(btn_save_payroll)
        
        layout.addLayout(summary_layout)

        widget.setLayout(layout)
        return widget
    
    def toggle_bulk_asgari_ucret_mode(self, checked):
        """Toplu hesaplama için asgari ücret modu toggle et"""
        text = "🎯 Asgari Ücret Modu: Açık" if checked else "🎯 Asgari Ücret Modu: Kapalı"
        self.bulk_asgari_mode_button.setText(text)
    
    def calculate_bulk_payroll(self):
        """Tüm çalışanlar için bordro hesapla (TÜM VERGİ SEÇENEKLERİ desteği)"""
        try:
            from src.services.employee_service import EmployeeService
            from src.database.db import get_db
            
            def money(value):
                """Tutarları hassas hesapla ve yuvarla"""
                return round(float(value) + 1e-9, 2)
            
            db = get_db()
            emp_service = EmployeeService(db)
            employees = emp_service.get_all_employees(active_only=True)
            
            if not employees:
                QMessageBox.warning(self, "Uyarı", "Aktif çalışan bulunamadı")
                return
            
            month = self.bulk_payroll_month.currentData()
            year = self.bulk_payroll_year.currentData()
            min_wage_gross = self._parse_payroll_float(self.bulk_min_wage_gross_input, "Asgari Ücret Brüt")
            progressive_enabled = self.bulk_progressive_tax_check.isChecked()
            apply_exemption = self.bulk_apply_min_wage_exemption_check.isChecked()
            asgari_mode = self.bulk_asgari_mode_button.isChecked()
            
            brackets = self._parse_tax_brackets(self.bulk_tax_brackets_input.text()) if progressive_enabled else []
            
            # Standart değerler - tüm gün çalışmayı varsayıyoruz
            month_days = 30
            worked_days = 30
            paid_leave_days = 0
            unpaid_leave_days = 0
            paid_days = 30
            day_ratio = 1.0
            
            self.bulk_payroll_table.setRowCount(0)
            total_gross = 0.0
            total_sgk = 0.0
            total_unemp = 0.0
            total_income_tax = 0.0
            total_stamp_tax = 0.0
            total_exemptions_amount = 0.0
            total_net = 0.0
            
            # Kümülatif vergi bazı
            cumulative_base = 0.0
            
            for employee in employees:
                # BRÜT HESAPLAMA
                gross_salary = employee.gross_salary
                prorated_gross_salary_raw = gross_salary * day_ratio
                gross_total_raw = prorated_gross_salary_raw
                
                # TAX BASE HESAPLAMA
                deduction_base_raw = gross_total_raw
                sgk_deduction_raw = deduction_base_raw * (employee.sgk_rate / 100.0)
                unemployment_deduction_raw = deduction_base_raw * (employee.unemployment_rate / 100.0)
                tax_base_raw = max(deduction_base_raw - sgk_deduction_raw - unemployment_deduction_raw, 0)
                
                # ÇOCUK İNDİRİMİ
                child_allowance = getattr(employee, 'child_count', 0) * 346.5
                reduced_tax_base_raw = max(tax_base_raw - child_allowance, 0)
                
                # VERGİ HESAPLAMA - Progressive veya Normal
                if progressive_enabled and not asgari_mode:
                    # Kümülatif vergi dilimi kullan
                    cumulative_base_old = cumulative_base
                    cumulative_base_new = cumulative_base + reduced_tax_base_raw
                    income_tax_raw = (
                        self._compute_progressive_tax(cumulative_base_new, brackets)
                        - self._compute_progressive_tax(cumulative_base_old, brackets)
                    )
                    cumulative_base = cumulative_base_new
                else:
                    # Normal vergi hesabı
                    if asgari_mode:
                        income_tax_raw = 0.0
                    else:
                        income_tax_raw = reduced_tax_base_raw * (employee.income_tax_rate / 100.0)
                
                stamp_tax_raw = deduction_base_raw * (employee.stamp_tax_rate / 100.0)
                
                # MUAFIYET KONTROLLERI (Asgari Ücret Muafiyeti - her brüt için uygulanabilir!)
                income_tax_exemption_raw = 0.0
                stamp_tax_exemption_raw = 0.0
                
                if apply_exemption and not asgari_mode:
                    # Muafiyet her brüt için uygulanabilir - min ücret tutarı kadar
                    min_sgk_raw = min_wage_gross * (employee.sgk_rate / 100.0)
                    min_unemployment_raw = min_wage_gross * (employee.unemployment_rate / 100.0)
                    min_tax_base_raw = max(min_wage_gross - min_sgk_raw - min_unemployment_raw, 0)
                    
                    if progressive_enabled:
                        # Progressive modda muafiyet hesapla
                        min_wage_income_tax_rate = brackets[0][1] * 100 if brackets else employee.income_tax_rate
                    else:
                        min_wage_income_tax_rate = employee.income_tax_rate
                    
                    income_tax_exemption_raw = min(
                        min_tax_base_raw * (min_wage_income_tax_rate / 100.0),
                        income_tax_raw
                    )
                    stamp_tax_exemption_raw = min(
                        min_wage_gross * (employee.stamp_tax_rate / 100.0),
                        stamp_tax_raw
                    )
                
                # KESİNTİLER
                advance = 0.0
                private_insurance = 0.0
                other_deductions = 0.0
                
                # Vergileri muafiyet sonrası hesapla
                income_tax_after_exemption = max(income_tax_raw - income_tax_exemption_raw, 0)
                stamp_tax_after_exemption = max(stamp_tax_raw - stamp_tax_exemption_raw, 0)
                
                total_deductions_raw = sgk_deduction_raw + unemployment_deduction_raw + income_tax_after_exemption + stamp_tax_after_exemption
                additional_deductions_raw = advance + private_insurance + other_deductions
                total_exemptions_raw = income_tax_exemption_raw + stamp_tax_exemption_raw
                
                # SONUÇLAR
                gross_total = money(gross_total_raw)
                sgk_deduction = money(sgk_deduction_raw)
                unemployment_deduction = money(unemployment_deduction_raw)
                income_tax = money(income_tax_after_exemption)  # MUAFIYET SONRASI
                stamp_tax = money(stamp_tax_after_exemption)  # MUAFIYET SONRASI
                total_deductions = money(total_deductions_raw)
                additional_deductions = money(additional_deductions_raw)
                total_exemptions = money(total_exemptions_raw)
                net_salary = money(gross_total_raw - total_deductions_raw - additional_deductions_raw)
                
                # TABLOYA EKLE
                row = self.bulk_payroll_table.rowCount()
                self.bulk_payroll_table.insertRow(row)
                
                # Seç checkbox'ı
                checkbox_item = QTableWidgetItem()
                checkbox_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                checkbox_item.setCheckState(Qt.CheckState.Checked)  # Varsayılan seçili
                checkbox_item.setData(Qt.ItemDataRole.UserRole, employee.id)  # Employee ID sakla
                self.bulk_payroll_table.setItem(row, 0, checkbox_item)
                
                self.bulk_payroll_table.setItem(row, 1, QTableWidgetItem(employee.get_full_name()))
                self.bulk_payroll_table.setItem(row, 2, QTableWidgetItem("30"))
                self.bulk_payroll_table.setItem(row, 3, QTableWidgetItem("30"))
                self.bulk_payroll_table.setItem(row, 4, QTableWidgetItem("0"))
                self.bulk_payroll_table.setItem(row, 5, QTableWidgetItem("0"))
                self.bulk_payroll_table.setItem(row, 6, QTableWidgetItem(format_tr(gross_total)))
                self.bulk_payroll_table.setItem(row, 7, QTableWidgetItem(format_tr(sgk_deduction)))
                self.bulk_payroll_table.setItem(row, 8, QTableWidgetItem(format_tr(unemployment_deduction)))
                self.bulk_payroll_table.setItem(row, 9, QTableWidgetItem(format_tr(income_tax)))
                self.bulk_payroll_table.setItem(row, 10, QTableWidgetItem(format_tr(stamp_tax)))
                self.bulk_payroll_table.setItem(row, 11, QTableWidgetItem(format_tr(total_exemptions)))
                self.bulk_payroll_table.setItem(row, 12, QTableWidgetItem(format_tr(net_salary)))

                
                # TOPLAMLAR
                total_gross += gross_total
                total_sgk += sgk_deduction
                total_unemp += unemployment_deduction
                total_income_tax += income_tax
                total_stamp_tax += stamp_tax
                total_exemptions_amount += total_exemptions
                total_net += net_salary
            
            # ÖZETİ GÜNCELLE
            self.bulk_summary_label.setText(
                f"Toplu Brüt: {format_tr(total_gross)} TL | "
                f"Total Kesinti: {format_tr(money(total_sgk + total_unemp + total_income_tax + total_stamp_tax))} TL | "
                f"Total Muafiyet: {format_tr(total_exemptions_amount)} TL | "
                f"Net Toplam (ÖDENECEK): {format_tr(total_net)} TL"
            )
            
            db.close()
            QMessageBox.information(self, "Başarılı", 
                f"{len(employees)} çalışan için {month:02d}/{year} bordrosu hesaplandı\n\nNet Toplam: {format_tr(total_net)} TL")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Toplu bordro hesaplaması başarısız:\n{str(e)}")
    
    def _save_bulk_payroll_to_db(self, employees, month, year):
        """Hesaplanan bordroları veritabanına kaydet"""
        try:
            from src.database.db import get_db
            from datetime import datetime
            
            db = get_db()
            
            for row in range(self.bulk_payroll_table.rowCount()):
                employee_name = self.bulk_payroll_table.item(row, 1).text()  # +1 checkbox sütunu nedeniyle
                
                # Çalışanı bul
                employee = None
                for emp in employees:
                    if emp.get_full_name() == employee_name:
                        employee = emp
                        break
                
                if not employee:
                    continue
                
                # Tablo hücrelerinden veriler al
                def get_table_value(col):
                    item = self.bulk_payroll_table.item(row, col)
                    if item:
                        # Türkçe virgülü ve noktayı düzelt
                        text = item.text().replace(".", "").replace(",", ".")
                        try:
                            return float(text)
                        except:
                            return 0.0
                    return 0.0
                
                gross_total = get_table_value(6)
                sgk_deduction = get_table_value(7)
                unemployment_deduction = get_table_value(8)
                income_tax = get_table_value(9)
                stamp_tax = get_table_value(10)
                total_exemptions = get_table_value(11)
                net_salary = get_table_value(12)
                
                # Bordro kaydı oluştur - Maaş Bordro sistemiyle uyumlu
                payroll_record = {
                    "employee": employee_name,
                    "year": year,
                    "month": month,
                    "month_days": 30,
                    "worked_days": 30,
                    "paid_leave_days": 0,
                    "unpaid_leave_days": 0,
                    "child_count": 0,
                    "paid_days": 30,
                    "day_ratio": 1.0,
                    "gross_salary": employee.gross_salary,
                    "overtime_hours": 0.0,
                    "overtime_rate": employee.overtime_rate,
                    "overtime_amount": 0.0,
                    "bonus": 0.0,
                    "meal": 0.0,
                    "transport": 0.0,
                    "gross_total": gross_total,
                    "sgk_rate": employee.sgk_rate,
                    "sgk_deduction": sgk_deduction,
                    "unemployment_rate": employee.unemployment_rate,
                    "unemployment_deduction": unemployment_deduction,
                    "tax_base": gross_total - sgk_deduction - unemployment_deduction,
                    "child_allowance": 0.0,
                    "income_tax_rate": employee.income_tax_rate,
                    "income_tax": income_tax,
                    "stamp_tax_rate": employee.stamp_tax_rate,
                    "stamp_tax": stamp_tax,
                    "income_tax_exemption": total_exemptions if income_tax > 0 else 0.0,
                    "stamp_tax_exemption": 0.0,
                    "total_deductions": sgk_deduction + unemployment_deduction + income_tax + stamp_tax,
                    "advance": 0.0,
                    "private_insurance": 0.0,
                    "other_deductions": 0.0,
                    "additional_deductions": 0.0,
                    "total_exemptions": total_exemptions,
                    "net_salary": net_salary,
                    "saved_at": datetime.now().isoformat()
                }
                
                # Mevcut kayıtları yükle
                records = self._load_saved_payroll_records()
                
                # Aynı ay/yıl için eski kaydı sil
                records = [
                    r for r in records
                    if not (
                        str(r.get("employee", "")).strip().lower() == employee_name.strip().lower()
                        and int(r.get("year", 0)) == year
                        and int(r.get("month", 0)) == month
                    )
                ]
                
                # Yeni kaydı ekle
                records.append(payroll_record)
                self._save_payroll_records(records)
            
            db.close()
        except Exception as e:
            print(f"Toplu bordro kaydı hatası: {str(e)}")

    def bulk_select_all_employees(self):
        """Toplu bordro tablosundaki tüm çalışanları seç"""
        for row in range(self.bulk_payroll_table.rowCount()):
            item = self.bulk_payroll_table.item(row, 0)
            if item:
                item.setCheckState(Qt.CheckState.Checked)

    def bulk_deselect_all_employees(self):
        """Toplu bordro tablosundaki tüm seçimleri kaldır"""
        for row in range(self.bulk_payroll_table.rowCount()):
            item = self.bulk_payroll_table.item(row, 0)
            if item:
                item.setCheckState(Qt.CheckState.Unchecked)

    def save_bulk_payroll(self):
        """Toplu bordro tablosunda seçili olan çalışanların bordrolarını kaydet"""
        try:
            from src.database.db import get_db
            from datetime import datetime
            
            db = get_db()
            selected_rows = []
            
            # Seçili çalışanları bul
            for row in range(self.bulk_payroll_table.rowCount()):
                checkbox_item = self.bulk_payroll_table.item(row, 0)
                if checkbox_item and checkbox_item.checkState() == Qt.CheckState.Checked:
                    selected_rows.append(row)
            
            if not selected_rows:
                QMessageBox.warning(self, "Uyarı", "Kaydetmek için lütfen çalışan seçiniz!")
                return
            
            # Seçili çalışanlar için bordro kaydı yap
            month = self.bulk_payroll_month.currentData()
            year = self.bulk_payroll_year.currentData()
            
            # Kaydedilen önceki kayıtları al
            records = self._load_saved_payroll_records()
            
            saved_count = 0
            for row in selected_rows:
                employee_name = self.bulk_payroll_table.item(row, 1).text()
                
                # Aynı ay/yıl için eski kaydı sil
                records = [r for r in records if not (r.get("employee") == employee_name and 
                                                      str(r.get("month")) == str(month) and
                                                      str(r.get("year")) == str(year))]
                
                # Tablo hücrelerinden değerleri al
                def get_value(col):
                    text = self.bulk_payroll_table.item(row, col).text()
                    text = text.replace(".", "").replace(",", ".")
                    try:
                        return float(text)
                    except:
                        return 0.0
                
                # Yeni kayıt oluştur
                payroll_record = {
                    "employee": employee_name,
                    "month": int(month),
                    "year": int(year),
                    "worked_days": get_value(2),
                    "paid_leave_days": get_value(3),
                    "unpaid_leave_days": get_value(4),
                    "gross_total": get_value(6),
                    "tax_base": get_value(6),  # Toplu bordro için brüt = tax base
                    "sgk_deduction": get_value(7),
                    "unemployment_deduction": get_value(8),
                    "income_tax": get_value(9),
                    "stamp_tax": get_value(10),
                    "net_salary": get_value(12),
                    "saved_at": datetime.now().isoformat()
                }
                records.append(payroll_record)
                saved_count += 1
            
            self._save_payroll_records(records)
            db.close()
            QMessageBox.information(self, "Başarılı", f"{saved_count} çalışanın bordrosu kaydedildi!")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Bordro kaydı başarısız:\n{str(e)}")

            QMessageBox.information(self, "Başarılı", f"Bordro PDF olarak kaydedildi:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{str(e)}")    
    def save_google_sheets_settings(self):
        """Google Sheets ayarlarını kaydet (credentials gerektirmeden)"""
        try:
            url = self.gsheets_url_input.text().strip()

            caris_sheet = self.gsheets_caris_sheet_input.text().strip() or "Cariler"
            trans_sheet = self.gsheets_trans_sheet_input.text().strip() or "İşlemler"
            gelen_fatura_sheet = self.gsheets_gelen_fatura_sheet_input.text().strip() or "Gelen Fatura"
            kesilen_fatura_sheet = self.gsheets_kesilen_fatura_sheet_input.text().strip() or "Kesilen Fatura"
            gider_sheet = self.gsheets_gider_sheet_input.text().strip() or "Gider"
            gelir_sheet = self.gsheets_gelir_sheet_input.text().strip() or "Gelir"
            auto_sync = self.gsheets_auto_sync_check.isChecked()

            # Ayarları kaydet
            UserSettingsService.set_setting(self.user.id, "google_sheets_caris_sheet", caris_sheet)
            UserSettingsService.set_setting(self.user.id, "google_sheets_trans_sheet", trans_sheet)
            UserSettingsService.set_setting(self.user.id, "google_sheets_gelen_fatura_sheet", gelen_fatura_sheet)
            UserSettingsService.set_setting(self.user.id, "google_sheets_kesilen_fatura_sheet", kesilen_fatura_sheet)
            UserSettingsService.set_setting(self.user.id, "google_sheets_gider_sheet", gider_sheet)
            UserSettingsService.set_setting(self.user.id, "google_sheets_gelir_sheet", gelir_sheet)
            UserSettingsService.set_json_setting(self.user.id, "google_sheets_auto_sync", auto_sync)
            
            if url:
                # URL'den Spreadsheet ID'yi çıkar
                spreadsheet_id = GoogleSheetsService.get_spreadsheet_id_from_url(url)
                if not spreadsheet_id:
                    spreadsheet_id = url  # Belki direkt ID verilmiş
                
                UserSettingsService.set_setting(self.user.id, "google_sheets_url", url)
                UserSettingsService.set_setting(self.user.id, "google_sheets_id", spreadsheet_id)
                
                # Otomatik senkronizasyonu ayarla
                if auto_sync:
                    self.setup_google_sheets_timer()
                
                QMessageBox.information(self, "Başarılı", 
                    "Google Sheets ayarları kaydedildi!\n\n"
                    "'Şimdi Senkronize Et' butonuna basarak verileri çekebilirsiniz.")
            else:
                # URL silinmişse timer'ı durdur
                if hasattr(self, 'gsheets_timer') and self.gsheets_timer:
                    self.gsheets_timer.stop()
                    self.gsheets_timer = None
                
                QMessageBox.information(self, "Başarılı", "Ayarlar kaydedildi!")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Hata", f"Ayarlar kaydedilemedi:\n{str(e)}")
    
    def test_google_sheets_connection(self):
        """Google Sheets bağlantısını test et"""
        try:
            if self.gsheets_worker and self.gsheets_worker.isRunning():
                QMessageBox.information(self, "Bilgi", "Google Sheets işlemi devam ediyor, lütfen bekleyin.")
                return

            url = self.gsheets_url_input.text().strip()
            if not url:
                QMessageBox.warning(self, "Uyarı", "Lütfen Google Sheets URL'si girin!")
                return
            
            spreadsheet_id = GoogleSheetsService.get_spreadsheet_id_from_url(url)
            if not spreadsheet_id:
                spreadsheet_id = url

            self._set_gsheets_busy(True)
            self.gsheets_worker = GoogleSheetsWorker('test', self.user.id, spreadsheet_id)
            self.gsheets_worker.finished_signal.connect(self._on_gsheets_worker_finished)
            self.gsheets_worker.start()
                
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Bağlantı testi başarısız:\n{str(e)}")
    
    def sync_from_google_sheets(self):
        """Google Sheets'ten tek yönlü veri aktar"""
        try:
            if self.gsheets_worker and self.gsheets_worker.isRunning():
                QMessageBox.information(self, "Bilgi", "Google Sheets işlemi devam ediyor, lütfen bekleyin.")
                return

            url = self.gsheets_url_input.text().strip()
            if not url:
                QMessageBox.warning(self, "Uyarı", "Lütfen Google Sheets URL'si girin!")
                return
            
            spreadsheet_id = GoogleSheetsService.get_spreadsheet_id_from_url(url)
            if not spreadsheet_id:
                spreadsheet_id = url
            

            caris_sheet = self.gsheets_caris_sheet_input.text().strip() or "Cariler"
            trans_sheet = self.gsheets_trans_sheet_input.text().strip() or "İşlemler"
            gelen_fatura_sheet = self.gsheets_gelen_fatura_sheet_input.text().strip() or "Gelen Fatura"
            kesilen_fatura_sheet = self.gsheets_kesilen_fatura_sheet_input.text().strip() or "Kesilen Fatura"
            gider_sheet = self.gsheets_gider_sheet_input.text().strip() or "Gider"
            gelir_sheet = self.gsheets_gelir_sheet_input.text().strip() or "Gelir"

            sheet_mappings = {
                'caris': caris_sheet,
                'transactions': trans_sheet,
                'gelen_fatura': gelen_fatura_sheet,
                'kesilen_fatura': kesilen_fatura_sheet,
                'gider': gider_sheet,
                'gelir': gelir_sheet
            }

            self._set_gsheets_busy(True)
            self.gsheets_worker = GoogleSheetsWorker('sync', self.user.id, spreadsheet_id, sheet_mappings)
            self.gsheets_worker.finished_signal.connect(self._on_gsheets_worker_finished)
            self.gsheets_worker.start()
                
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Senkronizasyon başarısız:\n{str(e)}")

    def _set_gsheets_busy(self, busy: bool):
        """Google Sheets butonlarını işlem sırasında kilitle"""
        if hasattr(self, 'gsheets_test_button') and self.gsheets_test_button:
            self.gsheets_test_button.setEnabled(not busy)
            self.gsheets_test_button.setText("⏳ İşleniyor..." if busy else "🔗 Bağlantıyı Test Et")

        if hasattr(self, 'gsheets_sync_button') and self.gsheets_sync_button:
            self.gsheets_sync_button.setEnabled(not busy)
            self.gsheets_sync_button.setText("⏳ İşleniyor..." if busy else "🔄 Şimdi Google Sheets'ten Aktar")

    def _on_gsheets_worker_finished(self, operation: str, success: bool, message: str, stats: dict):
        """Arka plan Google Sheets işlemi tamamlandı"""
        self._set_gsheets_busy(False)

        if operation == 'test':
            if success:
                QMessageBox.information(self, "Başarılı", message)
            else:
                if "credentials.json" in message:
                    QMessageBox.information(self, "Bilgi",
                        "Google Sheets özelliğini kullanmak için:\n\n"
                        "1. Google Cloud Console'dan credentials.json indirin\n"
                        "2. data/google_credentials/ klasörüne koyun\n\n"
                        "Detaylı bilgi için GOOGLE_SHEETS_SETUP.md dosyasına bakın.\n\n"
                        "Not: Bu özellik isteğe bağlıdır.")
                else:
                    QMessageBox.warning(self, "Bağlantı Hatası", message)
            return

        if operation == 'sync':
            if success:
                last_sync = UserSettingsService.get_setting(self.user.id, "google_sheets_last_sync", None)
                if last_sync:
                    last_sync_dt = datetime.fromisoformat(last_sync)
                    self.gsheets_last_sync_label.setText(f"Son senkronizasyon: {last_sync_dt.strftime('%d.%m.%Y %H:%M')}")

                self.refresh_dashboard()
                duplicate_items = stats.get('duplicate_transactions', []) if stats else []
                added_duplicates = 0
                skipped_duplicates = len(duplicate_items)

                if duplicate_items:
                    from src.ui.dialogs.duplicate_transactions_dialog import DuplicateTransactionsDialog

                    dialog = DuplicateTransactionsDialog(duplicate_items, self)
                    if dialog.exec_() == QDialog.Accepted:
                        selected = dialog.get_selected_row_ids()
                        for item in duplicate_items:
                            if item.get('row_key') not in selected:
                                continue
                            transaction, _ = TransactionService.create_transaction(
                                user_id=self.user.id,
                                transaction_date=item.get('date'),
                                transaction_type=item.get('transaction_type'),
                                payment_method=item.get('payment_method'),
                                customer_name=item.get('customer_name') or "",
                                description=item.get('description') or "",
                                amount=float(item.get('amount') or 0.0),
                                cari_id=item.get('cari_id'),
                                payment_type=item.get('payment_type')
                            )
                            if transaction:
                                added_duplicates += 1

                        skipped_duplicates = len(duplicate_items) - len(selected)
                        if added_duplicates:
                            self.refresh_dashboard()

                if duplicate_items:
                    message = (
                        f"{message}\nMükerrer: {len(duplicate_items)}"
                        f" (eklendi: {added_duplicates}, atlandı: {skipped_duplicates})"
                    )
                QMessageBox.information(self, "Başarılı", message)
            else:
                QMessageBox.warning(self, "Senkronizasyon Hatası", message)
    
    def setup_google_sheets_timer(self):
        """Google Sheets otomatik senkronizasyon zamanlayıcısı"""
        try:
            if hasattr(self, 'gsheets_timer') and self.gsheets_timer:
                self.gsheets_timer.stop()
            
            auto_sync = UserSettingsService.get_json_setting(self.user.id, "google_sheets_auto_sync", False)
            if not auto_sync:
                return
            
            url = UserSettingsService.get_setting(self.user.id, "google_sheets_url", None)
            if not url:
                return
            
            # credentials.json yoksa sessizce atla
            credentials_file = GoogleSheetsService.CREDENTIALS_FILE
            if not credentials_file.exists():
                print("Google Sheets credentials.json bulunamadı, otomatik senkronizasyon devre dışı")
                return
            
            # 10 dakikada bir senkronize et
            self.gsheets_timer = QTimer(self)
            self.gsheets_timer.timeout.connect(self.auto_sync_google_sheets)
            self.gsheets_timer.start(10 * 60 * 1000)  # 10 dakika = 600,000 ms
            
        except Exception as e:
            print(f"Google Sheets timer kurulumu hatası: {e}")
    
    def auto_sync_google_sheets(self):
        """Otomatik Google Sheets senkronizasyonu (sessiz)"""
        try:
            # credentials.json yoksa sessizce atla
            credentials_file = GoogleSheetsService.CREDENTIALS_FILE
            if not credentials_file.exists():
                return
            
            url = UserSettingsService.get_setting(self.user.id, "google_sheets_url", None)
            if not url:
                return
            
            spreadsheet_id = GoogleSheetsService.get_spreadsheet_id_from_url(url)
            if not spreadsheet_id:
                spreadsheet_id = url
            

            caris_sheet = UserSettingsService.get_setting(self.user.id, "google_sheets_caris_sheet", "Cariler")
            trans_sheet = UserSettingsService.get_setting(self.user.id, "google_sheets_trans_sheet", "İşlemler")
            gelen_fatura_sheet = UserSettingsService.get_setting(self.user.id, "google_sheets_gelen_fatura_sheet", "Gelen Fatura")
            kesilen_fatura_sheet = UserSettingsService.get_setting(self.user.id, "google_sheets_kesilen_fatura_sheet", "Kesilen Fatura")
            gider_sheet = UserSettingsService.get_setting(self.user.id, "google_sheets_gider_sheet", "Gider")
            gelir_sheet = UserSettingsService.get_setting(self.user.id, "google_sheets_gelir_sheet", "Gelir")

            sheet_mappings = {
                'caris': caris_sheet,
                'transactions': trans_sheet,
                'gelen_fatura': gelen_fatura_sheet,
                'kesilen_fatura': kesilen_fatura_sheet,
                'gider': gider_sheet,
                'gelir': gelir_sheet
            }
            
            service = GoogleSheetsService()
            success, message, stats = service.sync_from_sheets(self.user.id, spreadsheet_id, sheet_mappings)
            
            if success:
                # Dashboard'u sessizce yenile
                self.refresh_dashboard()
                print(f"Otomatik senkronizasyon başarılı: {message}")
            else:
                # Sadece log'a yaz, kullanıcıyı rahatsız etme
                print(f"Otomatik senkronizasyon hatası: {message}")
                
        except Exception as e:
            print(f"Otomatik senkronizasyon hatası: {e}")
