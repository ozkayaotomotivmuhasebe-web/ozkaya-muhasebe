"""
Google Sheets için örnek Excel dosyası oluştur
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

def create_sample_excel():
    """Örnek Excel dosyası oluştur"""
    wb = Workbook()
    
    # İlk sayfayı sil
    wb.remove(wb.active)
    
    # 1. Cariler Sayfası
    ws_caris = wb.create_sheet("Cariler")
    
    # Başlıklar
    headers_caris = ["Ad/Ünvan", "Tür", "Telefon", "Adres", "Vergi No/TC"]
    ws_caris.append(headers_caris)
    
    # Başlık stilini ayarla
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws_caris[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Örnek cariler
    sample_caris = [
        ["ABC Tekstil Ltd. Şti.", "MÜŞTERİ", "0555 123 4567", "Atatürk Cad. No:45 Beyoğlu/İSTANBUL", "1234567890"],
        ["XYZ Yapı Malzemeleri A.Ş.", "TEDARİKÇİ", "0532 987 6543", "Cumhuriyet Mah. İnönü Sok. No:12 Çankaya/ANKARA", "9876543210"],
        ["Mehmet Yılmaz", "MÜŞTERİ", "0505 111 2233", "Barbaros Bulvarı No:78 Beşiktaş/İSTANBUL", "12345678901"],
        ["DEF Elektronik Tic. Ltd.", "MÜŞTERİ", "0543 444 5566", "Konak Meydanı No:23 Konak/İZMİR", "3456789012"],
        ["GHI Gıda San. ve Tic. A.Ş.", "TEDARİKÇİ", "0533 777 8899", "OSB 1. Cad. No:56 OSB/BURSA", "6789012345"],
        ["Ayşe Demir Danışmanlık", "MÜŞTERİ", "0542 222 3344", "Bağdat Cad. No:234 Kadıköy/İSTANBUL", "23456789012"],
        ["JKL Otomotiv Yedek Parça", "TEDARİKÇİ", "0555 888 9900", "Sanayi Sitesi 45. Sok. No:12 Tuzla/İSTANBUL", "4567890123"],
        ["MNO Reklam Ajansı", "MÜŞTERİ", "0544 555 6677", "Tunalı Hilmi Cad. No:89 Çankaya/ANKARA", "34567890123"],
        ["PQR İnşaat Taahhüt Ltd.", "MÜŞTERİ", "0536 666 7788", "Kemeraltı Cad. No:156 Karşıyaka/İZMİR", "5678901234"],
        ["STU Teknoloji A.Ş.", "TEDARİKÇİ", "0545 999 0011", "Teknokent Binası Kat:3 Gebze/KOCAELİ", "7890123456"]
    ]
    
    for row_data in sample_caris:
        ws_caris.append(row_data)
    
    # Sütun genişliklerini ayarla
    ws_caris.column_dimensions['A'].width = 35
    ws_caris.column_dimensions['B'].width = 15
    ws_caris.column_dimensions['C'].width = 18
    ws_caris.column_dimensions['D'].width = 45
    ws_caris.column_dimensions['E'].width = 18
    
    # 2. İşlemler Sayfası
    ws_trans = wb.create_sheet("İşlemler")
    
    # Başlıklar
    headers_trans = ["Tarih", "Tür", "Tutar", "Açıklama", "Cari", "Ödeme"]
    ws_trans.append(headers_trans)
    
    # Başlık stilini ayarla
    for cell in ws_trans[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Örnek işlemler (son 30 günden)
    today = datetime.now().date()
    sample_trans = [
        [(today - timedelta(days=25)).strftime("%d.%m.%Y"), "GELİR", 15000, "Ürün satışı - Şubat dönemi", "ABC Tekstil Ltd. Şti.", "BANKA"],
        [(today - timedelta(days=24)).strftime("%d.%m.%Y"), "GİDER", 8500, "Hammadde alımı", "XYZ Yapı Malzemeleri A.Ş.", "BANKA"],
        [(today - timedelta(days=22)).strftime("%d.%m.%Y"), "GELİR", 3200, "Danışmanlık ücreti", "Mehmet Yılmaz", "NAKİT"],
        [(today - timedelta(days=20)).strftime("%d.%m.%Y"), "GELİR", 12500, "Yazılım geliştirme projesi", "DEF Elektronik Tic. Ltd.", "BANKA"],
        [(today - timedelta(days=18)).strftime("%d.%m.%Y"), "GİDER", 4750, "Ofis malzemeleri", "GHI Gıda San. ve Tic. A.Ş.", "KART"],
        [(today - timedelta(days=15)).strftime("%d.%m.%Y"), "GELİR", 8900, "Eğitim semineri", "Ayşe Demir Danışmanlık", "BANKA"],
        [(today - timedelta(days=12)).strftime("%d.%m.%Y"), "GİDER", 6200, "Araç bakım ve onarım", "JKL Otomotiv Yedek Parça", "BANKA"],
        [(today - timedelta(days=10)).strftime("%d.%m.%Y"), "GELİR", 18000, "Web tasarım projesi", "MNO Reklam Ajansı", "BANKA"],
        [(today - timedelta(days=8)).strftime("%d.%m.%Y"), "GELİR", 22500, "İnşaat danışmanlığı", "PQR İnşaat Taahhüt Ltd.", "BANKA"],
        [(today - timedelta(days=6)).strftime("%d.%m.%Y"), "GİDER", 11200, "Sunucu ve hosting hizmetleri", "STU Teknoloji A.Ş.", "BANKA"],
        [(today - timedelta(days=5)).strftime("%d.%m.%Y"), "GELİR", 5400, "Küçük ölçekli proje", "ABC Tekstil Ltd. Şti.", "NAKİT"],
        [(today - timedelta(days=4)).strftime("%d.%m.%Y"), "GİDER", 2800, "Kırtasiye malzemeleri", "XYZ Yapı Malzemeleri A.Ş.", "NAKİT"],
        [(today - timedelta(days=3)).strftime("%d.%m.%Y"), "GELİR", 9600, "Aylık bakım ücreti", "DEF Elektronik Tic. Ltd.", "BANKA"],
        [(today - timedelta(days=2)).strftime("%d.%m.%Y"), "GİDER", 3500, "Elektrik ve su faturaları", "", "BANKA"],
        [(today - timedelta(days=1)).strftime("%d.%m.%Y"), "GELİR", 14800, "Sosyal medya yönetimi", "MNO Reklam Ajansı", "BANKA"],
        [today.strftime("%d.%m.%Y"), "GELİR", 7200, "Danışmanlık hizmeti", "Ayşe Demir Danışmanlık", "BANKA"],
        [today.strftime("%d.%m.%Y"), "GİDER", 1850, "Akaryakıt gideri", "", "KART"],
    ]
    
    for row_data in sample_trans:
        ws_trans.append(row_data)
    
    # Sütun genişliklerini ayarla
    ws_trans.column_dimensions['A'].width = 15
    ws_trans.column_dimensions['B'].width = 12
    ws_trans.column_dimensions['C'].width = 12
    ws_trans.column_dimensions['D'].width = 40
    ws_trans.column_dimensions['E'].width = 30
    ws_trans.column_dimensions['F'].width = 12
    
    # Tutar sütununu sayı formatına çevir
    for row in range(2, len(sample_trans) + 2):
        cell = ws_trans.cell(row=row, column=3)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
    
    # Dosyayı kaydet
    filename = "Google_Sheets_Ornek_Veri.xlsx"
    wb.save(filename)
    print(f"✓ Örnek dosya oluşturuldu: {filename}")
    print(f"\nDosya içeriği:")
    print(f"  • Cariler sayfası: {len(sample_caris)} cari hesap")
    print(f"  • İşlemler sayfası: {len(sample_trans)} işlem kaydı")
    print(f"\nBu dosyayı Google Drive'a yükleyip Google Sheets ile açabilirsiniz.")

if __name__ == "__main__":
    create_sample_excel()
