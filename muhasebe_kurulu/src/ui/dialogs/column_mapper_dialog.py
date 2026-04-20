from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
                           QFileDialog, QTableWidget, QTableWidgetItem, QComboBox,
                           QSpinBox, QMessageBox, QGroupBox, QFormLayout, QHeaderView)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor
import openpyxl
from datetime import datetime


class ColumnMapperDialog(QDialog):
    """Excel sütunlarını dinamik olarak eşleştir"""
    
    def __init__(self, parent=None, file_path=None):
        super().__init__(parent)
        self.file_path = file_path
        self.file_data = []
        self.column_mapping = {}
        self._auto_detect_done = False
        
        self.setWindowTitle("📋 Sütun Eşleştirmesi")
        self.setGeometry(100, 100, 1000, 600)
        self.setMinimumSize(1000, 600)
        self.setModal(True)
        self.init_ui()
        if self.file_path:
            self.file_label.setText(self.file_path.split("\\")[-1])
            self.file_label.setStyleSheet("color: #333;")
            self.reload_preview()
    
    def init_ui(self):
        """Arayüz elemanlarını oluştur"""
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Başlık
        title = QLabel("📋 Excel Sütunlarını Eşleştir")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        layout.addWidget(title)
        
        # Dosya seçimi
        file_layout = QHBoxLayout()
        self.file_label = QLabel("Dosya seçilmedi...")
        self.file_label.setStyleSheet("color: #666;")
        file_layout.addWidget(self.file_label)
        
        btn_browse = QPushButton("📁 Excel Dosyası Seç")
        btn_browse.setMinimumHeight(35)
        btn_browse.clicked.connect(self.select_file)
        file_layout.addWidget(btn_browse)
        file_layout.addStretch()
        layout.addLayout(file_layout)
        layout.addSpacing(10)
        
        # Seçenekler
        options_group = QGroupBox("İthal Seçenekleri")
        options_layout = QFormLayout()
        
        self.sheet_spin = QSpinBox()
        self.sheet_spin.setMinimum(1)
        self.sheet_spin.setValue(1)
        self.sheet_spin.valueChanged.connect(self.reload_preview)
        options_layout.addRow("Excel Sayfası Numarası:", self.sheet_spin)
        
        self.start_row_spin = QSpinBox()
        self.start_row_spin.setMinimum(1)
        self.start_row_spin.setValue(2)
        self.start_row_spin.valueChanged.connect(self.reload_preview)
        options_layout.addRow("Başlangıç Satırı:", self.start_row_spin)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        layout.addSpacing(10)
        
        # Preview tabelosu
        preview_label = QLabel("Verileri Önizle ve Sütunları Eşleştir:")
        preview_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
        layout.addWidget(preview_label)
        
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(0)
        self.preview_table.setRowCount(0)
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.preview_table)
        
        layout.addSpacing(10)
        
        # Eşleştirme seçenekleri
        mapping_group = QGroupBox("Sütun Eşleştirmesi (Tümü İsteğe Bağlı)")
        mapping_layout = QFormLayout()
        
        self.date_combo = QComboBox()
        self.date_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("📅 Tarih:", self.date_combo)
        
        self.name_combo = QComboBox()
        self.name_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("📝 İşlem Adı:", self.name_combo)
        
        self.amount_combo = QComboBox()
        self.amount_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("💰 Tutar:", self.amount_combo)

        self.type_combo = QComboBox()
        self.type_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("🏷️ Tür:", self.type_combo)

        self.customer_title_combo = QComboBox()
        self.customer_title_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("👤 Müşteri Ünvanı:", self.customer_title_combo)
        
        self.description_combo = QComboBox()
        self.description_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("📄 Açıklama:", self.description_combo)

        self.payment_type_combo = QComboBox()
        self.payment_type_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("💳 Ödeme Şekli:", self.payment_type_combo)

        self.subject_combo = QComboBox()
        self.subject_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("📌 Konu:", self.subject_combo)

        self.company_combo = QComboBox()
        self.company_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("🏢 Firma Adı:", self.company_combo)

        self.loan_bank_combo = QComboBox()
        self.loan_bank_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("🏦 Kredi Bankası:", self.loan_bank_combo)

        self.loan_combo = QComboBox()
        self.loan_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("🏦 Ödenecek Kredi:", self.loan_combo)

        self.person_combo = QComboBox()
        self.person_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("🙍 Ödeyen Kişi:", self.person_combo)
        
        self.reference_combo = QComboBox()
        self.reference_combo.addItem("-- Seçiniz --", -1)
        mapping_layout.addRow("🔗 Referans/Belge No:", self.reference_combo)
        
        mapping_group.setLayout(mapping_layout)
        layout.addWidget(mapping_group)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        
        btn_ok = QPushButton("✅ Onayla ve Devam Et")
        btn_ok.setMinimumHeight(40)
        btn_ok.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
        """)
        btn_ok.clicked.connect(self.accept_mapping)
        btn_layout.addWidget(btn_ok)
        
        btn_cancel = QPushButton("❌ İptal")
        btn_cancel.setMinimumHeight(40)
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #9E9E9E;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #757575; }
        """)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def select_file(self):
        """Excel dosyasını seç ve ön izlemeyi yükle"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", "", "Excel Dosyaları (*.xlsx *.xls)"
        )
        if file_path:
            self.file_path = file_path
            self.file_label.setText(file_path.split("\\")[-1])
            self.file_label.setStyleSheet("color: #333;")
            self.reload_preview()
    
    def _detect_header_row_xlrd(self, ws, max_scan=50):
        keywords = ["tarih", "islem", "işlem", "tutar", "aciklama", "açıklama", "valör", "bakiye"]
        for r in range(min(max_scan, ws.nrows)):
            row = [str(ws.cell_value(r, c)).strip().lower() for c in range(ws.ncols)]
            if any(k in cell for cell in row for k in keywords):
                return r
        return None

    def _detect_header_row_openpyxl(self, ws, max_scan=50):
        keywords = ["tarih", "islem", "işlem", "tutar", "aciklama", "açıklama", "valör", "bakiye"]
        for r in range(1, min(max_scan, ws.max_row) + 1):
            row = [str(cell.value or "").strip().lower() for cell in ws[r]]
            if any(k in cell for cell in row for k in keywords):
                return r
        return None

    def reload_preview(self):
        """Preview tablosunu yenile"""
        if not self.file_path:
            return
        
        try:
            # .xls ve .xlsx desteği
            if self.file_path.lower().endswith('.xls'):
                try:
                    import xlrd
                    wb = xlrd.open_workbook(self.file_path, on_demand=True)
                    sheet_idx = self.sheet_spin.value() - 1
                    if sheet_idx >= len(wb.sheet_names()):
                        QMessageBox.warning(self, "Uyarı", "Belirtilen sayfa bulunamadı!")
                        return
                    ws = wb.sheet_by_index(sheet_idx)
                    start_row = self.start_row_spin.value()

                    if not self._auto_detect_done:
                        header_row = self._detect_header_row_xlrd(ws)
                        if header_row is not None:
                            self.start_row_spin.setValue(header_row + 2)
                            start_row = self.start_row_spin.value()
                        self._auto_detect_done = True
                    
                    # xlrd ile başlıkları oku
                    header_row = start_row - 1
                    if header_row < 0:
                        header_row = 0
                    headers = []
                    for col_idx in range(ws.ncols):
                        cell_value = ws.cell_value(header_row, col_idx)
                        if cell_value:
                            headers.append(str(cell_value))
                        else:
                            headers.append(f"Sütun {col_idx + 1}")
                    
                    # Veri satırlarını oku (ilk 10)
                    data_rows = []
                    for row_idx in range(start_row, min(start_row + 10, ws.nrows)):
                        row = []
                        for col_idx in range(ws.ncols):
                            row.append(ws.cell_value(row_idx, col_idx))
                        data_rows.append(row)
                except ImportError:
                    QMessageBox.critical(self, "Hata", "xlrd kitaplığı yüklü değil. .xlsx formatını kullanın veya 'pip install xlrd' çalıştırın")
                    return
            else:
                wb = openpyxl.load_workbook(self.file_path, data_only=True)
                sheet_idx = self.sheet_spin.value() - 1
                
                if sheet_idx >= len(wb.sheetnames):
                    QMessageBox.warning(self, "Uyarı", "Belirtilen sayfa bulunamadı!")
                    return
                
                ws = wb.worksheets[sheet_idx]
                start_row = self.start_row_spin.value()

                if not self._auto_detect_done:
                    header_row = self._detect_header_row_openpyxl(ws)
                    if header_row is not None:
                        self.start_row_spin.setValue(header_row + 1)
                        start_row = self.start_row_spin.value()
                    self._auto_detect_done = True
                
                # İlk satırı başlık olarak oku (sütun isimlerini al)
                header_row = start_row - 1
                if header_row < 1:
                    header_row = 1
                headers = []
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(str(cell.value))
                    else:
                        headers.append(f"Sütun {cell.column}")
            
                # Veri satırlarını oku (ilk 10) - openpyxl için
                data_rows = []
                for row_idx, row in enumerate(ws.iter_rows(
                    min_row=start_row, 
                    values_only=True
                ), start=start_row):
                    if row_idx >= start_row + 9:  # Max 10 satır preview
                        break
                    data_rows.append(row)
            
            self.file_data = data_rows
            
            # Tabloyu güncelle
            self.preview_table.setColumnCount(len(headers))
            self.preview_table.setHorizontalHeaderLabels(headers)
            self.preview_table.setRowCount(len(data_rows))
            
            for row_idx, row in enumerate(data_rows):
                for col_idx, cell_value in enumerate(row):
                    item = QTableWidgetItem(str(cell_value) if cell_value else "")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.preview_table.setItem(row_idx, col_idx, item)
            
            # Combo box'ları güncelle (sütun numaralarıyla)
            for combo in [
                self.date_combo, self.name_combo, self.amount_combo,
                self.type_combo, self.customer_title_combo, self.description_combo,
                self.payment_type_combo, self.subject_combo, self.company_combo,
                self.loan_bank_combo, self.loan_combo, self.person_combo, self.reference_combo
            ]:
                current_data = combo.currentData()
                combo.blockSignals(True)  # Sinyal blokla
                combo.clear()
                combo.addItem("-- Seçiniz --", -1)
                for idx, header in enumerate(headers):
                    combo.addItem(f"{idx + 1}. {header}", idx)
                
                # Önceki seçimi geri yükle
                if current_data >= 0 and current_data < len(headers):
                    combo.setCurrentIndex(current_data + 1)
                combo.blockSignals(False)  # Sinyal aç

            # Başlıklara göre otomatik seç
            def find_header_index(keys):
                for i, h in enumerate(headers):
                    text = h.lower()
                    if any(k in text for k in keys):
                        return i
                return -1

            if self.date_combo.currentData() in (-1, None):
                idx = find_header_index(["tarih", "tarih/saat", "valör", "valor"])
                if idx >= 0:
                    self.date_combo.setCurrentIndex(idx + 1)
            if self.name_combo.currentData() in (-1, None):
                idx = find_header_index(["işlem", "islem", "açıklama", "aciklama"])
                if idx >= 0:
                    self.name_combo.setCurrentIndex(idx + 1)
            if self.amount_combo.currentData() in (-1, None):
                idx = find_header_index(["tutar", "işlem tutarı", "islem tutari", "borç", "borc", "alacak"])
                if idx >= 0:
                    self.amount_combo.setCurrentIndex(idx + 1)
            if self.description_combo.currentData() in (-1, None):
                idx = find_header_index(["açıklama", "aciklama", "işlem", "islem"])
                if idx >= 0:
                    self.description_combo.setCurrentIndex(idx + 1)
            if self.type_combo.currentData() in (-1, None):
                idx = find_header_index(["tür", "tur", "işlem türü", "islem turu", "borç/alacak", "borc/alacak"])
                if idx >= 0:
                    self.type_combo.setCurrentIndex(idx + 1)
            if self.customer_title_combo.currentData() in (-1, None):
                idx = find_header_index(["müşteri ünvanı", "musteri unvani", "müşteri", "musteri", "cari", "ünvan", "unvan"])
                if idx >= 0:
                    self.customer_title_combo.setCurrentIndex(idx + 1)
            if self.payment_type_combo.currentData() in (-1, None):
                idx = find_header_index(["ödeme şekli", "odeme sekli", "ödeme", "odeme", "payment"])
                if idx >= 0:
                    self.payment_type_combo.setCurrentIndex(idx + 1)
            if self.subject_combo.currentData() in (-1, None):
                idx = find_header_index(["konu", "subject"])
                if idx >= 0:
                    self.subject_combo.setCurrentIndex(idx + 1)
            if self.company_combo.currentData() in (-1, None):
                idx = find_header_index(["firma", "firma adı", "firma adi", "şirket", "sirket", "company"])
                if idx >= 0:
                    self.company_combo.setCurrentIndex(idx + 1)
            if self.loan_bank_combo.currentData() in (-1, None):
                idx = find_header_index(["kredi bankası", "kredi bankasi", "banka adı", "banka adi", "kredi banka"])
                if idx >= 0:
                    self.loan_bank_combo.setCurrentIndex(idx + 1)
            if self.loan_combo.currentData() in (-1, None):
                idx = find_header_index(["ödenecek kredi", "odenecek kredi", "kredi no", "kredi numarası", "kredi numarasi", "kredi adi", "kredi adı", "loan"])
                if idx >= 0:
                    self.loan_combo.setCurrentIndex(idx + 1)
            if self.person_combo.currentData() in (-1, None):
                idx = find_header_index(["ödeyen kişi", "odeyen kisi", "kişi", "kisi", "person"])
                if idx >= 0:
                    self.person_combo.setCurrentIndex(idx + 1)
        
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel okuma hatası: {str(e)}")
    
    def accept_mapping(self):
        """Sütun eşleştirmesini doğrula ve onayla"""
        date_col = self.date_combo.currentData()
        name_col = self.name_combo.currentData()
        amount_col = self.amount_combo.currentData()
        
        # Eşleştirmeyi kaydet
        self.column_mapping = {
            'file_path': self.file_path,
            'sheet': self.sheet_spin.value() - 1,
            'start_row': self.start_row_spin.value(),
            'date_column': date_col,
            'name_column': name_col,
            'amount_column': amount_col,
            'type_column': self.type_combo.currentData(),
            'customer_title_column': self.customer_title_combo.currentData(),
            'description_column': self.description_combo.currentData(),
            'payment_type_column': self.payment_type_combo.currentData(),
            'subject_column': self.subject_combo.currentData(),
            'company_column': self.company_combo.currentData(),
            'loan_bank_column': self.loan_bank_combo.currentData(),
            'loan_column': self.loan_combo.currentData(),
            'person_column': self.person_combo.currentData(),
            'reference_column': self.reference_combo.currentData(),
        }
        
        self.accept()


# Teste için
if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    
    app = QApplication(sys.argv)
    dialog = ColumnMapperDialog()
    if dialog.exec_():
        print("Mapping:", dialog.column_mapping)
